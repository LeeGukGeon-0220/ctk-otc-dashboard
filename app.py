# -*- coding: utf-8 -*-
"""
CTK OTC LAB 실적 대시보드
--------------------------------------------------------------
v7 (2026-08-20) 전면 개편
  * 모든 숫자를 '원장(raw)'에서 직접 계산합니다. 엑셀 수식에 의존하지 않습니다.
  * 거래일 DD/MM/YYYY 텍스트를 명시적으로 파싱합니다 (일/월 뒤바뀜 방지).
  * 매출원가 / 판관비 / 영업이익 계산 오류를 바로잡았습니다.
  * '계정과목 상세' 페이지를 새로 만들었습니다.
--------------------------------------------------------------
"""
import io
import os
import re
import math
import zipfile
import posixpath
import collections
import json
import calendar
import datetime
import urllib.request

import pandas as pd
import numpy as np
import altair as alt
import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(page_title='CTK OTC LAB 실적보고', layout='wide')

# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# 0. 테마 / 스타일  (밝은 화면 · 어두운 화면 두 가지)
# ══════════════════════════════════════════════════════════════
if 'ui_theme' not in st.session_state:
    st.session_state['ui_theme'] = '밝은 화면'
밝게 = st.session_state['ui_theme'] == '밝은 화면'

# ── 색 기준 ────────────────────────────────────────────────────
#  남색(navy)   : 화면의 뼈대 — 사이드바 · 표 머리글 · 제목
#  주황(accent) : 시선을 끌 곳 하나만 — 현재 메뉴 · 제목 밑줄 · 당해연도 막대
#  숫자는 검정, 음수는 (빨강 괄호) — 이 규칙은 두 화면에서 같습니다.
if 밝게:
    T = dict(bg='#F6F7F9', panel='#FFFFFF', panel2='#FAFBFD', line='#E7E7E1',
             ink='#0B0B0B', ink2='#4A4A46', ink3='#8A8880',
             accent='#E2611C', accentbg='rgba(226,97,28,0.10)', dot='#C7C6BC',
             grid='rgba(11,11,11,0.07)', axis='#8A8880',
             ok='#0A6B3D', warn='#D03B3B', shadow='0 1px 2px rgba(13,27,51,0.05)',
             head='#12294D', zebra='#FAFBFD',
             side='#12294D', upc='#D03B3B', downc='#1D3E70',
             navy='#12294D', navy7='#1D3E70', navysoft='#F1F4F9', navyline='#D8E0EC',
             headink='#0D1B33', strongline='#12294D',
             sideink='#FFFFFF', sideink2='rgba(255,255,255,0.76)',
             sideink3='rgba(255,255,255,0.50)', sideline='rgba(255,255,255,0.13)',
             sidehover='rgba(255,255,255,0.09)')
else:
    T = dict(bg='#0B1220', panel='#111C30', panel2='#0E1829', line='rgba(255,255,255,0.10)',
             ink='#F2F5FA', ink2='#B9C4D6', ink3='#8798AE',
             accent='#F0834F', accentbg='rgba(240,131,79,0.16)', dot='#3C4E69',
             grid='rgba(255,255,255,0.07)', axis='#8798AE',
             ok='#3DD68C', warn='#F06A6A', shadow='0 1px 3px rgba(0,0,0,0.35)',
             head='#1D3E70', zebra='rgba(255,255,255,0.035)',
             side='#0A1526', upc='#F06A6A', downc='#7FA8E0',
             navy='#16305C', navy7='#2B5490', navysoft='rgba(255,255,255,0.05)',
             navyline='rgba(255,255,255,0.14)',
             headink='#F2F5FA', strongline='rgba(255,255,255,0.28)',
             sideink='#F2F5FA', sideink2='rgba(255,255,255,0.74)',
             sideink3='rgba(255,255,255,0.48)', sideline='rgba(255,255,255,0.11)',
             sidehover='rgba(255,255,255,0.08)')

# 그래프 계열 색 — 1·2·3번은 색약(色弱)에서도 서로 구분되도록 검증된 조합입니다.
#   TEAL  = 당해연도 · 대표 계열 (주황)      SLATE = 전년 · 비교용 (회색)
#   BLUE  = 두 번째 계열 (파랑)              AMBER = 세 번째 계열 (청록)
#   ROSE  = 경고 · 마이너스 (빨강)
if 밝게:
    TEAL, ROSE, BLUE, SLATE, AMBER = '#eb6834', '#D03B3B', '#2a78d6', '#898781', '#1baf7a'
else:
    TEAL, ROSE, BLUE, SLATE, AMBER = '#d95926', '#F06A6A', '#3987e5', '#8E8C84', '#199e70'


def _ctk_theme():
    return {
      'padding': {'left': 46, 'top': 6, 'right': 18, 'bottom': 6},
      'config': {
        'background': 'transparent',
        'view': {'stroke': 'transparent'},
        'font': "'Malgun Gothic','맑은 고딕','Apple SD Gothic Neo',sans-serif",
        'axis': {'domainColor': T['navyline'], 'gridColor': T['grid'],
                 'labelColor': T['ink3'], 'titleColor': T['ink2'], 'tickColor': T['navyline'],
                 'labelFontSize': 13, 'titleFontSize': 13, 'labelPadding': 5,
                 'labelFontWeight': 500, 'grid': True, 'domain': True},
        'legend': {'labelColor': T['ink2'], 'titleColor': T['ink2'], 'labelFontWeight': 600,
                   'labelFontSize': 13, 'titleFontSize': 13, 'symbolSize': 118,
                   'symbolType': 'square'},
        'text': {'color': T['headink'], 'fontWeight': 700, 'fontSize': 12.5},
        'title': {'color': T['headink'], 'fontSize': 16.5, 'fontWeight': 800, 'anchor': 'start',
                  'offset': 12},
    }}


alt.themes.register('ctk_theme', _ctk_theme)
alt.themes.enable('ctk_theme')

st.markdown(f"""
<style>
/* ── 글꼴 : 화면 전체를 맑은 고딕으로 통일 ────────────────── */
html, body,
[data-testid="stAppViewContainer"] *,
[data-testid="stSidebar"] * {{
    font-family: 'Malgun Gothic', '맑은 고딕', 'Apple SD Gothic Neo',
                 system-ui, -apple-system, sans-serif !important; }}
/* ★아이콘은 글꼴이 곧 그림이라 반드시 되돌려 놓습니다 — 안 그러면 화살표가
   'keyboard_double_arrow_left' 같은 글자로 보입니다. (위 규칙보다 뒤에 둡니다) */
[data-testid="stIconMaterial"], [data-testid="stIconMaterial"] *,
[data-testid="stTooltipIcon"], [data-testid="stTooltipIcon"] *,
.stTooltipIcon, .stTooltipIcon * {{
    font-family: 'Material Symbols Rounded' !important; }}
[data-testid="stAppViewContainer"], [data-testid="stHeader"] {{ background-color: {T['bg']}; }}
.main .block-container, [data-testid="stMainBlockContainer"], .block-container {{
    color: {T['ink']};
    padding-top: 1.6rem !important; padding-bottom: 2.5rem !important;
    padding-left: 0.9rem !important; padding-right: 1.2rem !important;
    max-width: none !important; }}
/* 오른쪽 위 Streamlit 메뉴(점 세 개 ⋮)는 다시 보이게 둡니다 — 공유·인쇄·설정이 거기 있습니다.
   맨 위 무지개 줄만 감추고, 머리줄은 배경 없이 띄워 화면은 그대로 깔끔하게 둡니다. */
[data-testid="stDecoration"] {{ display: none !important; }}
[data-testid="stHeader"] {{ background: transparent !important; }}
[data-testid="stToolbar"] {{ display: flex !important; right: 0.6rem !important; }}
[data-testid="stToolbar"] button {{ color: {T['ink3']} !important; }}
/* 요소 사이 세로 간격을 조금 좁힙니다 */
.main [data-testid="stVerticalBlock"] {{ gap: 0.7rem !important; }}
[data-testid="stSidebar"] {{ background-color: {T['side']};
    border-right: 1px solid {T['line']}; }}
[data-testid="stSidebar"] * {{ color: {T['sideink']} !important; }}
.brand-box {{ padding: 10px 4px 18px 4px; border-bottom: 1px solid {T['sideline']}; margin-bottom: 10px; }}
.brand-title {{ font-size: 17.5px; font-weight: 800; letter-spacing: -0.03em;
    color: {T['sideink']} !important; }}
.brand-sub {{ font-size: 11px; font-weight: 700; color: {T['sideink3']} !important;
    letter-spacing: 0.12em; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button {{
    background: transparent !important; border: none !important; border-radius: 9px !important;
    box-shadow: none !important; padding: 10px 12px !important; min-height: 0 !important;
    display: flex !important; align-items: center !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button > div {{
    justify-content: flex-start !important; width: 100% !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button,
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button * {{
    color: {T['sideink2']} !important; font-size: 14.5px !important; font-weight: 600 !important;
    text-align: left !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button::before {{
    content: ""; display: inline-block; width: 6px; height: 6px; border-radius: 50%;
    background: rgba(255,255,255,0.34); margin-right: 11px; flex-shrink: 0; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button:hover {{
    background: {T['sidehover']} !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button:hover,
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button:hover * {{
    color: {T['sideink']} !important; }}
/* 지금 보고 있는 메뉴 — 왼쪽에 주황 막대를 세워 색을 못 알아봐도 구분되게 합니다 */
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button[kind="primary"] {{
    background: {T['sidehover']} !important;
    border-left: 3px solid {T['accent']} !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button[kind="primary"],
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button[kind="primary"] * {{
    color: {T['sideink']} !important; font-weight: 800 !important; }}
[data-testid="stSidebar"] .st-key-nav_container [data-testid="stButton"] > button[kind="primary"]::before {{
    background: {T['accent']} !important; }}
[data-testid="stSidebar"] .st-key-pl_sub [data-testid="stButton"] > button,
[data-testid="stSidebar"] .st-key-rep_sub [data-testid="stButton"] > button {{
    padding-left: 30px !important; }}
[data-testid="stSidebar"] .st-key-pl_sub [data-testid="stButton"] > button::before,
[data-testid="stSidebar"] .st-key-rep_sub [data-testid="stButton"] > button::before {{
    content: "-"; width: auto; height: auto; border-radius: 0;
    background: transparent !important; color: {T['ink3']};
    font-weight: 700; margin-right: 9px; line-height: 1; }}
[data-testid="stSidebar"] .st-key-pl_sub [data-testid="stButton"]
    > button[kind="primary"]::before,
[data-testid="stSidebar"] .st-key-rep_sub [data-testid="stButton"]
    > button[kind="primary"]::before {{
    color: {T['accent']} !important; background: transparent !important;
    width: auto !important; height: auto !important; border-radius: 0 !important; }}
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {{
    font-size: 12.5px !important; color: {T['sideink3']} !important; }}
[data-testid="stSidebar"] hr {{ border-color: {T['sideline']} !important;
    margin: 14px 0 10px !important; }}
.side-foot {{ margin-top: 6px; padding-top: 13px; border-top: 1px solid {T['sideline']};
    font-size: 12.5px; line-height: 1.9; color: {T['sideink3']}; }}
.side-foot b {{ color: {T['sideink']}; font-weight: 700; }}
/* 설정 접이칸 — 평소엔 조용히 있다가 누르면 펴집니다 */
[data-testid="stSidebar"] [data-testid="stExpander"] {{
    border: 1px solid {T['sideline']} !important; border-radius: 9px !important;
    background: transparent !important; }}
[data-testid="stSidebar"] [data-testid="stExpander"] summary {{
    padding: 8px 12px !important; }}
/* 펼쳤을 때 제목줄이 흰 바탕이 되어 글씨가 안 보이던 것을 막습니다 */
[data-testid="stSidebar"] [data-testid="stExpander"] details,
[data-testid="stSidebar"] [data-testid="stExpander"] summary,
[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover,
[data-testid="stSidebar"] [data-testid="stExpanderDetails"] {{
    background: transparent !important; background-color: transparent !important; }}
[data-testid="stSidebar"] [data-testid="stExpander"] summary svg {{
    fill: {T['sideink2']} !important; color: {T['sideink2']} !important; }}
[data-testid="stSidebar"] [data-testid="stExpander"] summary p {{
    font-size: 12.5px !important; font-weight: 700 !important;
    color: {T['sideink2']} !important; }}
/* 사이드바가 남색이라 안에 들어가는 입력칸도 어둡게 맞춥니다 (안 그러면 흰 글씨가 안 보입니다) */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {{
    background: {T['sidehover']} !important; border: 1px dashed {T['sideink3']} !important; }}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] * {{
    color: {T['sideink2']} !important; }}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {{
    background: {T['accent']} !important; border: none !important; }}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button * {{
    color: #FFFFFF !important; font-weight: 700 !important; }}
[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] {{
    background: {T['sidehover']} !important; border-radius: 8px; padding: 6px 8px; }}
[data-testid="stSidebar"] input, [data-testid="stSidebar"] [data-baseweb="input"],
[data-testid="stSidebar"] [data-baseweb="base-input"],
[data-testid="stSidebar"] [data-testid="stNumberInputContainer"],
[data-testid="stSidebar"] [data-baseweb="select"] > div {{
    background: {T['sidehover']} !important; color: {T['sideink']} !important;
    border-color: {T['sideline']} !important; }}
[data-testid="stSidebar"] [data-testid="stNumberInputStepUp"],
[data-testid="stSidebar"] [data-testid="stNumberInputStepDown"] {{
    background: {T['sidehover']} !important; }}
[data-testid="stSidebar"] [data-testid="stNumberInputStepUp"] svg,
[data-testid="stSidebar"] [data-testid="stNumberInputStepDown"] svg {{
    fill: {T['sideink2']} !important; }}
[data-testid="stSidebar"] [data-baseweb="radio"] div[aria-checked="false"] {{
    border-color: {T['sideink3']} !important; background: transparent !important; }}
h1, h2, h3, h4 {{ color: {T['headink']} !important; font-weight: 800 !important;
    letter-spacing: -0.03em !important; }}
.main .block-container p, .main .block-container li,
.main .block-container [data-testid="stMarkdownContainer"] {{ font-size: 16px; }}
[data-testid="stCaptionContainer"] p {{ font-size: 14.5px !important; color: {T['ink3']} !important; }}
.chart-head {{ display:flex; align-items:baseline; justify-content:space-between;
    margin: 6px 0 2px; }}
.chart-head .t {{ font-size:17.5px; font-weight:800; letter-spacing:-.025em; color:{T['headink']};
    padding-bottom:7px; position:relative; }}
.chart-head .t::after {{ content:''; position:absolute; left:0; bottom:0; width:26px; height:3px;
    background:{T['accent']}; border-radius:2px; }}
.chart-head .u {{ font-size:14px; font-weight:600; color:{T['ink3']}; }}
/* 어두운 화면에서 선택상자·라디오가 혼자 하얗게 남지 않도록 맞춥니다 */
{'' if 밝게 else f'''
.main [data-baseweb="select"] > div, .main [data-testid="stNumberInputContainer"],
.main [data-baseweb="popover"] li, .main [data-baseweb="menu"] {{
    background: {T['panel']} !important; color: {T['ink']} !important;
    border-color: {T['line']} !important; }}
.main [data-baseweb="select"] svg {{ fill: {T['ink2']} !important; }}
.main [data-testid="stRadio"] label p {{ color: {T['ink2']} !important; }}
/* 본문 파일 업로드 칸 (실적보고 엑셀작성 화면) 도 함께 어둡게 */
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderDropzone"] {{
    background: {T['panel2']} !important; border: 1px dashed {T['line']} !important; }}
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderDropzone"] * {{
    color: {T['ink2']} !important; }}
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderDropzone"] button {{
    background: {T['accent']} !important; border: none !important; }}
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderDropzone"] button,
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderDropzone"] button * {{
    color: #FFFFFF !important; font-weight: 700 !important; }}
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderFile"] {{
    background: {T['panel2']} !important; border-radius: 8px; padding: 6px 8px; }}
[data-testid="stMainBlockContainer"] [data-testid="stFileUploaderFile"] * {{
    color: {T['ink']} !important; }}
'''}
</style>
""", unsafe_allow_html=True)

카드스타일 = f"""
<style>
  .wrap {{ margin:0 0 6px; padding:2px; color:{T['ink']};
          font-family:'Malgun Gothic','맑은 고딕','Apple SD Gothic Neo',
                      system-ui,-apple-system,sans-serif; }}
  .page-head {{ display:flex; align-items:center; gap:12px; margin:0 0 3px; }}
  .page-head .t {{ font-size:26px; font-weight:800; letter-spacing:-.035em;
      color:{T['headink']}; line-height:1.2; }}
  .page-head .pill {{ background:{T['navy']}; color:#FFFFFF; font-size:13px; font-weight:800;
      letter-spacing:-.01em; padding:6px 15px; border-radius:22px; white-space:nowrap; }}
  .page-cap {{ font-size:14px; color:{T['ink3']}; margin:0 0 16px; }}
  .wrap .note {{ font-size:13.5px; color:{T['ink3']}; font-weight:600; margin-bottom:12px; }}
  .wrap .kpi-row {{ display:grid; grid-template-columns:repeat(var(--n,4),1fr); gap:12px; margin-bottom:16px; }}
  .wrap .kpi-card {{ background:{T['panel']}; border:1px solid {T['line']}; border-radius:12px;
              padding:0 17px 15px; box-shadow:{T['shadow']}; overflow:hidden; }}
  /* 카드 위에 얇은 띠 — 첫 칸만 주황으로 눈에 띄게 합니다 */
  .wrap .kpi-card::before {{ content:''; display:block; height:3px; margin:0 -17px 13px;
              background:{T['navyline']}; }}
  .wrap .kpi-row .kpi-card:first-child::before {{ background:{T['accent']}; }}
  .wrap .kpi-label {{ font-size:13.5px; font-weight:700; color:{T['headink']}; margin-bottom:10px;
              letter-spacing:-.01em; }}
  .wrap .kpi-value {{ font-size:34px; font-weight:800; color:{T['ink']};
               letter-spacing:-.035em; line-height:1.05; }}
  .wrap .kpi-value.ok {{ color:{T['ok']}; }} .kpi-value.warn {{ color:{T['warn']}; }}
  .wrap .kpi-delta {{ margin-top:9px; font-size:14px; }}
  .wrap .unit {{ font-size:13.5px; font-weight:600; color:{T['ink3']}; }}
  .wrap .up {{ color:{T['upc']}; }} .down {{ color:{T['downc']}; }} .muted {{ color:{T['ink3']}; }}
  .wrap .card {{ background:{T['panel']}; border:1px solid {T['line']}; border-radius:12px;
          padding:15px 20px 16px; margin-bottom:12px; box-shadow:{T['shadow']}; }}
  /* 제목 아래 짧은 주황 밑줄 — 어디를 보는 화면인지 바로 알 수 있게 */
  .wrap .card h3 {{ font-size:18px; font-weight:800; color:{T['headink']}; margin:0 0 3px;
          letter-spacing:-.025em; padding-bottom:7px; position:relative; }}
  .wrap .card h3::after {{ content:''; position:absolute; left:0; bottom:0; width:26px; height:3px;
          background:{T['accent']}; border-radius:2px; }}
  .wrap .card .sub {{ font-size:13.5px; color:{T['ink3']}; margin:8px 0 11px; }}
  .wrap table {{ width:100%; border-collapse:separate; border-spacing:0; }}
  /* 남색 머리칸 글자는 모두 가운데 정렬합니다 (표·세부내역 공통) */
  .wrap th {{ text-align:center; font-size:13.5px; color:#FFFFFF; background:{T['navy']};
       font-weight:800; padding:9px 11px; letter-spacing:-.01em; white-space:nowrap; }}
  .wrap thead tr:first-child th:first-child {{ border-radius:8px 0 0 0; }}
  .wrap thead tr:first-child th:last-child {{ border-radius:0 8px 0 0; }}
  .wrap td:first-child {{ text-align:left; }}
  .wrap td {{ padding:8px 11px; font-size:15px; line-height:1.3; border-bottom:1px solid {T['line']};
       text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap;
       color:{T['ink2']}; }}
  .wrap td:first-child {{ color:{T['ink']}; font-weight:500; }}
  .wrap tbody tr:last-child td {{ border-bottom:none; }}
  .wrap tbody tr:last-child td:first-child {{ border-radius:0 0 0 8px; }}
  .wrap tbody tr:last-child td:last-child {{ border-radius:0 0 8px 0; }}
  .wrap td.lft {{ text-align:left; }}   /* 머리칸(th)은 가운데, 내용칸만 왼쪽 */
  .wrap tbody tr:hover td {{ background:{T['zebra']}; }}
  .wrap tr.total td {{ font-weight:800; background:{T['zebra']}; color:{T['ink']}; }}
  .wrap tr.sub td:first-child {{ padding-left:22px; color:{T['ink2']}; font-weight:400; }}
  .wrap tr.sub2 td:first-child {{ padding-left:44px; color:{T['ink3']}; }}
  .wrap .tag {{ display:inline-block; font-size:12.5px; padding:1px 8px; border-radius:20px; margin-left:8px; font-weight:700; }}
  .wrap .tag.rev {{ color:{T['accent']}; background:{T['accentbg']}; }}
  .wrap .tag.cogs {{ color:{AMBER}; background:rgba(217,138,23,0.15); }}
  .wrap .tag.sga {{ color:{BLUE}; background:rgba(59,130,246,0.15); }}
  .wrap .tag.non {{ color:#8B7BE8; background:rgba(139,123,232,0.15); }}
  .wrap .bars {{ display:flex; align-items:flex-end; gap:18px; height:178px; padding:0 4px; }}
  .wrap .bar-group {{ display:flex; flex-direction:column; align-items:center; gap:8px; flex:1; height:100%; justify-content:flex-end; }}
  .wrap .bar-pair {{ display:flex; gap:7px; align-items:flex-end; height:152px; }}
  .wrap .bar-col {{ display:flex; flex-direction:column; align-items:center; justify-content:flex-end; }}
  .wrap .bar-val {{ font-size:13px; font-weight:600; color:{T['ink2']}; margin-bottom:4px;
             font-variant-numeric:tabular-nums; white-space:nowrap; }}
  .wrap .bar {{ width:17px; border-radius:4px 4px 0 0; }}
  .wrap .bar.cur {{ background:{TEAL}; }} .bar.prev {{ background:{SLATE}; }}
  .wrap .bar-month {{ font-size:14px; color:{T['ink3']}; }}
  .wrap .bar-sep {{ width:1px; align-self:stretch; margin:0 6px 26px;
             background:{T['line']}; flex:0 0 1px; }}
  .wrap .bar-group.cum .bar-month {{ font-weight:700; color:{T['ink']}; }}
  .wrap .bar-group.cum .bar-val {{ color:{T['ink']}; }}
  .wrap .unitbadge {{ float:right; font-size:13px; font-weight:600; color:{T['ink3']}; }}
  .wrap .mix-total {{ display:flex; justify-content:space-between; font-size:15.5px; font-weight:800;
               color:{T['headink']}; border-top:2px solid {T['strongline']};
               padding-top:11px; margin-top:14px; }}
  /* ── 셋트로 나뉜 표 (당월·월별 실적집계) ─────────────────── */
  .wrap table.fixed {{ width:auto; table-layout:fixed; }}
  .wrap table.fixed th, .wrap table.fixed td {{ width:128px; }}
  /* 머리글은 <br> 로만 줄을 나누고, 위아래 가운데에 놓습니다 */
  .wrap table.fixed thead th {{ white-space:nowrap; line-height:1.5;
      vertical-align:middle; padding-top:7px; padding-bottom:7px; }}
  .wrap table.fixed {{ border:none; }}
  .wrap table.fixed th {{ text-align:center; }}
  .wrap table.fixed td.name {{ text-align:left; }}   /* 머리칸은 가운데 */
  .wrap table.fixed th.name, .wrap table.fixed td.name {{ width:210px; }}
  .wrap table.fixed th.grp {{ padding-bottom:6px; }}
  /* 셋트 경계선 */
  .wrap table.fixed th.s1, .wrap table.fixed td.s1 {{ border-left:1px solid {T['navyline']}; }}
  .wrap table.fixed th.gh {{ text-align:center; font-size:13px; font-weight:800;
      color:#FFFFFF; background:{T['navy7']}; padding:7px 9px; letter-spacing:-.01em; }}
  .wrap table.fixed th.gh0 {{ background:{T['navy']}; vertical-align:middle; }}
  /* 증감 칸만 주황 — 강조색은 한 곳에만 씁니다 */
  .wrap table.fixed th.d {{ background:{T['accent']}; color:#FFFFFF; }}
  .wrap table.fixed td.d {{ background:{T['accentbg']}; font-weight:700; }}
  .wrap table.fixed td.d.plus {{ color:{T['upc']}; }}
  .wrap table.fixed td.d.minus {{ color:{T['downc']}; }}
  .wrap table.fixed td.d.zero {{ color:{T['ink3']}; font-weight:400; }}
  /* 표가 화면보다 넓으면 옆으로 밀어 볼 수 있게 하고, 스크롤 막대를 눈에 보이게 둡니다 */
  .wrap .scrollx {{ overflow-x:auto; scrollbar-width:thin; padding-bottom:3px; }}
  .wrap .scrollx::-webkit-scrollbar {{ height:9px; }}
  .wrap .scrollx::-webkit-scrollbar-track {{ background:transparent; }}
  .wrap .scrollx::-webkit-scrollbar-thumb {{ background:{T['line']}; border-radius:5px; }}
  .wrap .scrollx:hover::-webkit-scrollbar-thumb {{ background:{T['ink3']}; }}
  /* 긴 표 — 아래로 내려도 남색 머리줄(계정·연도)이 늘 맨 위에 붙어 있게 합니다 */
  .wrap .stick {{ overflow:auto; max-height:74vh; scrollbar-width:thin; }}
  .wrap .stick::-webkit-scrollbar {{ width:9px; height:9px; }}
  .wrap .stick::-webkit-scrollbar-track {{ background:transparent; }}
  .wrap .stick::-webkit-scrollbar-thumb {{ background:{T['line']}; border-radius:5px; }}
  .wrap .stick:hover::-webkit-scrollbar-thumb {{ background:{T['ink3']}; }}
  .wrap .stick thead th {{ position:sticky; top:0; z-index:4; line-height:1.35;
      background:{T['navy']}; box-shadow:inset 0 -1px 0 {T['navy']}; }}
  .wrap .stick thead th .th2 {{ font-size:11.5px; font-weight:600; opacity:0.85; }}
  .wrap .stick thead th:last-child {{ background:{T['navy7']}; }}
  /* 현금흐름 — 앞으로 올 달(예상) 칸은 색을 달리해 실적과 구분합니다 */
  .wrap table.lined th.fc {{ background:{T['navy7']}; line-height:1.25; }}
  .wrap table.lined th.fc .fcs {{ font-size:11px; font-weight:600; opacity:0.85; }}
  .wrap table.lined td.fc {{ background:{T['navysoft']}; }}
  /* 국문 ↔ 영문 대사 알림 */
  .wrap .reconc {{ margin:10px 0 2px; padding:10px 13px; border-radius:8px;
      background:{T['navysoft']}; font-size:13px; line-height:1.7; color:{T['ink2']}; }}
  .wrap .reconc b {{ color:{T['headink']}; font-weight:800; }}
  .wrap .chks {{ display:flex; flex-wrap:wrap; gap:6px; margin-top:6px; }}
  .wrap .chk {{ display:inline-block; padding:3px 9px; border-radius:999px;
      font-size:12.5px; font-weight:700; }}
  .wrap .chk.ok {{ background:{T['panel2']}; color:{T['ink']}; }}
  .wrap .chk.no {{ background:{T['warn']}; color:#FFFFFF; }}
  .wrap tr.head td {{ font-weight:700; color:{T['headink']}; }}
  .wrap .neg {{ color:{T['warn']}; }}
  .wrap .memo {{ display:flex; justify-content:space-between; gap:10px;
      font-size:12.5px; line-height:1.6; padding:5px 0;
      border-bottom:1px dashed {T['line']}; }}
  .wrap .memo b {{ color:{T['ink']}; font-weight:600; flex:0 0 auto; max-width:52%;
      overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
  .wrap .memo span {{ color:{T['ink3']}; text-align:right; }}
  /* 거래 내용 — 거래처 × 달 표. 옆(달)·아래(거래처) 어느 쪽으로도 밀어 볼 수 있고,
     맨 윗줄과 맨 왼쪽 칸은 밀어도 제자리에 붙어 있습니다. */
  .wrap .mscroll {{ overflow:auto; max-height:330px; scrollbar-width:thin;
      border:1px solid {T['line']}; border-radius:8px; }}
  .wrap .mscroll::-webkit-scrollbar {{ width:9px; height:9px; }}
  .wrap .mscroll::-webkit-scrollbar-track {{ background:transparent; }}
  .wrap .mscroll::-webkit-scrollbar-thumb {{ background:{T['line']}; border-radius:5px; }}
  .wrap .mscroll:hover::-webkit-scrollbar-thumb {{ background:{T['ink3']}; }}
  .wrap table.mtab {{ border-collapse:separate; border-spacing:0; font-size:12.5px;
      width:100%; }}
  .wrap table.mtab th {{ position:sticky; top:0; z-index:3; background:{T['navy']};
      color:#FFFFFF; font-weight:700; padding:7px 7px; text-align:right;
      white-space:nowrap; }}
  .wrap table.mtab td {{ padding:6px 7px; text-align:right; white-space:nowrap;
      color:{T['ink2']}; border-top:1px solid {T['line']}; }}
  .wrap table.mtab th.lft, .wrap table.mtab td.lft {{ position:sticky; left:0;
      text-align:left; width:124px; min-width:124px; max-width:124px;
      white-space:normal; }}
  .wrap table.mtab th.lft {{ z-index:5; }}
  .wrap table.mtab td.lft {{ z-index:2; background:{T['panel']};
      box-shadow:1px 0 0 {T['line']}; }}
  .wrap table.mtab td.lft b {{ color:{T['ink']}; font-weight:600; font-size:11.5px;
      display:block; line-height:1.35; }}
  .wrap table.mtab td.lft .cap {{ display:block; font-size:11px; font-weight:400;
      color:{T['ink3']}; line-height:1.4; white-space:nowrap;
      overflow:hidden; text-overflow:ellipsis; }}
  .wrap table.mtab td.sum, .wrap table.mtab th.sum {{ font-weight:700;
      color:{T['ink']}; background:{T['navysoft']}; }}
  .wrap table.mtab th.sum {{ color:#FFFFFF; background:{T['navy7']}; }}
  .wrap table.mtab tr.tot td {{ font-weight:700; color:{T['ink']};
      background:{T['panel2']}; border-top:2px solid {T['line']}; }}
  .wrap table.mtab tr.tot td.lft {{ background:{T['panel2']}; }}
  .wrap .calc {{ font-size:13px; font-weight:400; color:{T['ink3']}; letter-spacing:0; }}
  .wrap .calcnote {{ margin-top:13px; padding:12px 14px; border-radius:8px;
      background:{T['navysoft']}; border:none;
      font-size:13.5px; line-height:1.75; color:{T['ink2']}; }}
  .wrap .calcnote b {{ color:{T['headink']}; font-weight:800; }}
  /* ── 접었다 펴는 「보는 방법」 — 평소에는 접혀 있고 누르면 펼쳐집니다 ── */
  .wrap details.calcnote {{ padding:0; overflow:hidden; }}
  .wrap details.calcnote summary {{ cursor:pointer; padding:11px 14px; font-weight:800;
      color:{T['headink']}; list-style:none; user-select:none; }}
  .wrap details.calcnote summary::-webkit-details-marker {{ display:none; }}
  .wrap details.calcnote summary::before {{ content:'▸'; display:inline-block; margin-right:7px;
      color:{T['accent']}; transition:transform .15s; }}
  .wrap details.calcnote[open] summary::before {{ transform:rotate(90deg); }}
  .wrap details.calcnote summary .hint {{ font-size:12.5px; font-weight:600;
      color:{T['ink3']}; margin-left:8px; }}
  .wrap details.calcnote .body {{ padding:0 14px 12px; }}
  /* ── 공헌이익 표 ── */
  .wrap table.cmtab {{ table-layout:fixed; min-width:1050px; }}
  .wrap table.cmtab th {{ vertical-align:middle; line-height:1.35; }}
  .wrap table.cmtab th:last-child {{ background:{T['navy7']}; }}
  .wrap table.cmtab td:last-child {{ background:{T['navysoft']}; color:{T['ink']}; font-weight:800; }}
  .wrap table.cmtab tr.cmhl td {{ background:{T['accentbg']}; font-weight:800; color:{T['ink']};
      border-top:2px solid {T['accent']}; border-bottom:2px solid {T['accent']}; }}
  .wrap table.cmtab tr.cmrate td {{ font-size:12px; color:{T['ink3']};
      padding-top:2px; padding-bottom:8px; }}
  .wrap table.cmtab tr.cmrate td:last-child {{ color:{T['accent']}; font-weight:800; }}
  /* 거래처별 공헌이익 표의 공헌이익률 — 파란색 · 본문과 같은 글자 크기 · 위아래 라인 */
  .wrap table.cmtab tr.cmblue td {{ color:{BLUE}; font-weight:700; font-size:15px;
      border-top:2px solid {T['strongline']}; border-bottom:2px solid {T['strongline']}; }}
  .wrap table.cmtab tr.cmblue td:first-child {{ color:{BLUE}; }}
  .wrap table.cmtab tr.cmblue td:last-child {{ color:{BLUE}; }}
  .wrap .cmfoot {{ font-size:12.5px; color:{T['ink3']}; margin:10px 2px 0; line-height:1.7; }}
  .wrap .cmfoot b {{ color:{T['headink']}; }}
  /* ── 차입금·리스부채 표 — 마지막 칸(비고)은 강조 없이 왼쪽 정렬 글씨 칸 ── */
  .wrap table.bigo td:last-child {{ background:transparent; font-weight:400; color:{T['ink2']};
      text-align:left; white-space:normal; min-width:150px; }}
  .wrap table.bigo th:last-child {{ background:{T['navy']}; }}
  .wrap table.bigo tr.total td:last-child {{ background:{T['zebra']}; }}
  .wrap tr.total td .calc {{ font-weight:400; }}
  /* 손익계산서 · 현금흐름 — 화면이 좁아지면 열도 같이 줄어듭니다 (열너비는 %로 잡습니다).
     두 표가 같은 %를 쓰기 때문에 화면 크기가 달라져도 위아래 칸이 항상 맞습니다. */
  .wrap table.lined {{ width:100%; table-layout:fixed; }}
  .wrap table.lined th, .wrap table.lined td {{ padding-left:7px; padding-right:7px; }}
  .wrap table.lined th:first-child, .wrap table.lined td:first-child {{ white-space:normal; line-height:1.4; }}
  .wrap table.lined th {{ text-align:center; }}
  /* 맨 오른쪽 「누적」 칸은 제일 자주 보는 곳이라 따로 세웁니다 */
  .wrap table.lined th:last-child {{ background:{T['navy7']}; }}
  .wrap table.lined td:last-child {{ background:{T['navysoft']}; color:{T['ink']}; font-weight:800; }}
  .wrap table.lined tr.total td:last-child {{ background:{T['navysoft']};
      filter:brightness(0.97); }}
  /* 오른쪽 세부내역 — 거래처 이름이 길어도 카드 밖으로 삐져나가지 않게 */
  .wrap .card table {{ max-width:100%; }}
  /* 현금흐름 카드 안의 월말 잔액 꺾은선 — 아래 표의 월 칸과 가로 위치를 맞춥니다.
     선은 viewBox 를 100 으로 두고 preserveAspectRatio="none" 으로 늘려 %와 정확히 맞추고,
     굵기가 같이 늘어나지 않도록 vector-effect="non-scaling-stroke" 를 씁니다. */
  .wrap .cashline {{ position:relative; width:100%; height:126px; margin:2px 0 12px; }}
  .wrap .cashline .cgrid {{ position:absolute; height:0; border-top:1px solid {T['grid']}; }}
  .wrap .cashline .cgrid.zero {{ border-top-color:{T['navyline']}; }}
  /* 세로 눈금 글자는 그림 맨 왼쪽에서 축선 바로 앞까지 오른쪽 맞춤으로 앉힙니다
     (기초 잔액 점이 축선 위에 서므로 겹치지 않게 왼쪽으로 밀어 둡니다) */
  .wrap .cashline .cylab {{ position:absolute; left:0; transform:translateY(-50%);
      font-size:12.5px; font-weight:500;
      color:{T['ink3']}; text-align:right; font-variant-numeric:tabular-nums; }}
  .wrap .cashline .cdomain {{ position:absolute; top:0; bottom:0; width:0;
      border-left:1px solid {T['navyline']}; }}
  /* ※ Streamlit 이 <svg> 를 지워 버려서, 선과 면을 칸(div) + clip-path 로 그립니다.
       칸 하나가 점과 점 사이 한 구간이고, 네 꼭짓점을 %로 찍어 사다리꼴을 오려 냅니다. */
  .wrap .cashline .cseg, .wrap .cashline .cfill {{ position:absolute; top:0; height:100%; }}
  .wrap .cashline .cseg {{ background:{T['accent']}; }}
  .wrap .cashline .cdot {{ position:absolute; width:9px; height:9px; border-radius:50%;
      background:{T['accent']}; transform:translate(-50%,-50%);
      box-shadow:0 0 0 2px {T['panel']}; }}
  .wrap .cashline .cdot.last {{ width:11px; height:11px; }}
  /* 기초 잔액 점은 속을 비워, 「출발점이지 이번 달 실적이 아니다」를 눈으로 구분합니다 */
  .wrap .cashline .cdot.first {{ background:{T['panel']};
      box-shadow:0 0 0 2px {T['accent']}; }}
  .wrap .cashline .cval {{ position:absolute; transform:translate(-50%,-100%);
      font-size:12.5px; font-weight:800; color:{T['ink']}; white-space:nowrap;
      font-variant-numeric:tabular-nums; }}
  .wrap .cashline .ctag {{ position:absolute; transform:translate(-50%,0);
      margin-top:9px; font-size:11.5px; font-weight:700; color:{T['ink3']};
      white-space:nowrap; }}
  .wrap .cashline .cbase {{ position:absolute; height:0;
      border-top:1px solid {T['navyline']}; }}
  /* 그림 제목은 그림 바로 위 한 줄로 둡니다 (왼쪽 자리는 기초 점과 눈금 글자가 씁니다) */
  .wrap .cashcap {{ margin:4px 0 2px; display:flex; align-items:baseline; gap:10px;
      flex-wrap:wrap; }}
  .wrap .cashcap .t {{ font-size:15px; font-weight:800;
      color:{T['headink']}; letter-spacing:-.02em; line-height:1.35; }}
  .wrap .cashcap .u {{ font-size:12.5px; font-weight:600; color:{T['ink3']}; }}
  .wrap table.detail {{ table-layout:fixed; }}
  .wrap table.detail th, .wrap table.detail td {{ padding-left:5px; padding-right:5px; }}
  .wrap td.lft {{ white-space:normal; word-break:break-word; }}
  /* 화면이 아주 좁을 때만 글자를 줄입니다.
     ※ 예전에는 1500px 이하에서 바로 줄여 버려, 보통 크기 창에서도 글자가 작게 나왔습니다.
        기준을 1280px 로 낮춰 웬만한 화면에서는 제 크기(15px)가 그대로 보이게 했습니다. */
  @media (max-width:1280px) {{
    .wrap table.lined th, .wrap table.lined td {{ font-size:14px; padding-left:6px; padding-right:6px; }}
    .wrap .card table {{ font-size:14px; }}
    .wrap table.fixed th, .wrap table.fixed td {{ width:106px; font-size:14px; padding-left:6px; padding-right:6px; }}
    .wrap table.fixed th.name, .wrap table.fixed td.name {{ width:178px; }}
  }}
  @media (max-width:1080px) {{
    .wrap table.lined th, .wrap table.lined td {{ font-size:13px; padding-left:5px; padding-right:5px; }}
    .wrap table.fixed th, .wrap table.fixed td {{ width:94px; font-size:13px; padding-left:5px; padding-right:5px; }}
    .wrap table.fixed th.name, .wrap table.fixed td.name {{ width:162px; }}
  }}
  .wrap .legend {{ display:flex; gap:16px; margin-top:12px; font-size:14px; color:{T['ink2']}; }}
  .wrap .legend span {{ display:flex; align-items:center; gap:5px; }}
  .wrap .sw {{ width:10px; height:10px; border-radius:2px; display:inline-block; }}
  .wrap .mix-item .top {{ display:flex; justify-content:space-between; font-size:15.5px; margin-bottom:5px; }}
  .wrap .mix-track {{ height:8px; background:{T['grid']}; border-radius:5px; overflow:hidden; margin-bottom:12px; }}
  .wrap .mix-fill {{ height:100%; background:{TEAL}; border-radius:5px; }}
  .wrap .warnbox {{ background:{T['accentbg']}; border-left:4px solid {T['accent']};
             border-radius:9px; padding:13px 15px; font-size:14px; color:{T['ink2']};
             font-weight:600; margin-bottom:14px; }}
  /* 데이터 점검 — 한 줄에 아이콘 · 항목 · 값 */
  .wrap .chk {{ display:flex; align-items:center; gap:12px; padding:12px 14px; border-radius:9px;
             margin-bottom:8px; border:1px solid {T['line']}; background:{T['panel2']}; }}
  .wrap .chk.ok {{ border-left:4px solid {T['ok']}; }}
  .wrap .chk.wn {{ border-left:4px solid {T['accent']}; background:{T['accentbg']}; }}
  .wrap .chk .ic {{ width:21px; height:21px; flex:0 0 21px; border-radius:50%; color:#FFFFFF;
             display:flex; align-items:center; justify-content:center;
             font-size:12px; font-weight:800; }}
  .wrap .chk.ok .ic {{ background:{T['ok']}; }}
  .wrap .chk.wn .ic {{ background:{T['accent']}; }}
  .wrap .chk .t {{ flex:1; font-size:14.5px; font-weight:600; color:{T['ink2']}; }}
  .wrap .chk .v {{ font-size:14.5px; font-weight:800; color:{T['ink']};
             font-variant-numeric:tabular-nums; }}
  /* 정상 대 장기미수 비율 띠 */
  .wrap .split {{ display:flex; height:40px; border-radius:8px; overflow:hidden;
             gap:2px; margin:2px 0 16px; }}
  .wrap .split i {{ display:flex; align-items:center; justify-content:center; color:#FFFFFF;
             font-size:12.5px; font-weight:800; font-style:normal; min-width:6px; }}
  .wrap .split i.ok {{ background:{T['navy7']}; }}
  .wrap .split i.bad {{ background:{ROSE}; }}
  .wrap td .sw {{ width:10px; height:10px; border-radius:3px; display:inline-block;
             margin-right:8px; vertical-align:middle; }}
</style>
"""

st.markdown(카드스타일, unsafe_allow_html=True)   # 카드 스타일은 여기서 한 번만
CARD_CSS = ''  # (아래 f-string 들과의 호환을 위해 빈 값으로 둡니다)

매출계정 = ['용역매출', '상품매출', '제품매출', '패키지매출', '기타매출']
매출원가계정 = ['용역매출원가', '상품매출원가', '제품매출원가', '패키지매출원가', '기타매출원가']
영업외계정 = ['이자수익', '이자비용', '잡이익', '잡손실',
              '외환차익', '외환차손', '외화환산이익', '외화환산손실',
              '유형자산처분이익', '유형자산처분손실', '유형자산폐기손실',
              '무형자산처분손실', '무형자산폐기손실', '기타의대손상각비',
              '기부금', '대손충당금환입']
법인세계정 = ['법인세비용']
판관비항목 = ['인건비성 항목', '커미션 비용', '외부용역', '건물관리비', '차량유지비', 'ICT',
             '출장비', '운반 및 창고료', '소모성경비', '상각비', '기타관리경비']
ACT2ITEM = {'인건비성 항목': '인건비성 항목', '커미션 비용': '커미션 비용', '외부용역': '외부용역',
            '건물': '건물관리비', '차량': '차량유지비', 'ICT': 'ICT', '출장비': '출장비',
            '운반 및 창고료': '운반 및 창고료', '소모성경비': '소모성경비'}


# ══════════════════════════════════════════════════════════════
# 1. 데이터 읽기 (v7 신규 양식 / v6 이전 양식 자동 판별)
# ══════════════════════════════════════════════════════════════
def _parse_dmy(v):
    """QuickBooks 원장의 '일/월/연도' 텍스트를 진짜 날짜로 바꿉니다."""
    # ※ 빈 칸은 float(nan) 으로 올 때도 있고 NaT 으로 올 때도 있습니다.
    #   NaT 은 datetime 의 한 종류라서 아래 isinstance 에 걸려 버리므로 여기서 먼저 걸러 냅니다.
    if v is None:
        return pd.NaT
    try:
        if pd.isna(v):
            return pd.NaT
    except (TypeError, ValueError):
        pass
    if isinstance(v, (datetime.datetime, pd.Timestamp)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, datetime.date):
        return pd.Timestamp(v)
    parts = str(v).strip().replace('-', '/').split('/')
    if len(parts) != 3:
        return pd.NaT
    try:
        return pd.Timestamp(int(parts[2]), int(parts[1]), int(parts[0]))
    except Exception:
        return pd.NaT


def 재분류읽기(xls):
    """「재분류」 시트 — 특정 전표만 다른 계정으로 보이게 하는 규칙.

    한 줄 = 「전표번호 + 계정(영문)」 이 같은 원장 줄의 계정과목·활동분류·활동세부를
    바꿔서 봅니다. 원장 원본과 BS_IS_매핑 은 건드리지 않습니다.
    """
    if '재분류' not in xls.sheet_names:
        return {}
    try:
        d = pd.read_excel(xls, sheet_name='재분류', header=None)
    except Exception:
        return {}
    머리 = next((i for i in range(min(20, len(d)))
                 if any(str(v).strip().startswith('전표번호') for v in d.iloc[i].tolist())), None)
    if 머리 is None:
        return {}
    규칙 = {}
    for _i, 줄 in d.iloc[머리 + 1:].iterrows():
        칸 = [('' if pd.isna(v) else str(v).strip()) for v in 줄.tolist()] + [''] * 5
        try:
            전표 = int(float(칸[0]))
        except (TypeError, ValueError):
            continue
        계정, 과목 = 칸[1].lower(), 칸[2]
        if not 계정 or not 과목:
            continue
        규칙[(전표, 계정)] = (과목, 칸[3], 칸[4])
    return 규칙


def _표준화(df, 매핑표, 연도힌트, 재분류=None):
    """v6 / v7 어느 양식이든 같은 컬럼 이름으로 맞춥니다."""
    ren = {'Distribution account': '계정영문', 'Transaction date': '거래일',
           'Transaction ID': 'TransactionID', '#': '전표번호',
           '계정분류(BS/IS)': '계정분류'}
    df = df.rename(columns={k: v for k, v in ren.items() if k in df.columns})
    if '계정영문' not in df.columns:
        return None
    df = df[df['계정영문'].notna()].copy()
    df['계정영문'] = df['계정영문'].astype(str).str.strip()
    df = df[~df['계정영문'].isin(['', 'nan', 'Beginning Balance'])]

    df['거래일'] = df['거래일'].apply(_parse_dmy)
    df = df[df['거래일'].notna()].copy()

    for c in ('Debit', 'Credit'):
        df[c] = pd.to_numeric(df.get(c), errors='coerce').fillna(0)
    df['금액'] = df['Credit'] - df['Debit']          # 수익 +, 비용 −
    df['년'] = df['거래일'].dt.year
    df['월'] = df['거래일'].dt.month

    # 계정분류/계정과목은 항상 매핑표 기준으로 다시 계산 (원본 수식 오류 방지)
    k = df['계정영문'].str.lower()
    df['계정분류'] = k.map(매핑표['bsis'])
    df['계정과목'] = k.map(매핑표['kor'])

    # 「재분류」 시트 — 전표번호+계정이 맞는 줄만 계정과목을 바꿔서 봅니다
    바꿀활동 = {}
    if 재분류:
        tid = pd.to_numeric(df.get('TransactionID'), errors='coerce')
        for i in df.index:
            t = tid.at[i] if tid is not None and i in tid.index else None
            if t is None or pd.isna(t):
                continue
            규 = 재분류.get((int(t), k.at[i]))
            if 규:
                df.at[i, '계정과목'] = 규[0]
                if 규[1] or 규[2]:
                    바꿀활동[i] = 규

    def _분류(a):
        if pd.isna(a):
            return np.nan
        if a in 매출계정:
            return '매출'
        if a in 매출원가계정:
            return '매출원가'
        if a in 영업외계정:
            return '영업외손익'
        if a in 법인세계정:
            return '법인세'
        return '판관비'

    df['분류'] = np.where(df['계정분류'].eq('IS'), df['계정과목'].map(_분류), None)

    # 원장 양식에 따라 열 이름이 「활동분류」이기도 하고 「활동분류(대분류)」이기도 합니다.
    # ★둘 중 어느 쪽이 와도 같게 읽어야 합니다★ — 안 그러면 판관비가 전부 「기타관리경비」로 잡힙니다.
    if '활동분류(대분류)' not in df.columns and '활동분류' in df.columns:
        df['활동분류(대분류)'] = df['활동분류']
    for c in ('활동분류(대분류)', '활동분류(소분류)', '활동세부', '분류상태', 'Description', 'Name', 'Balance'):
        if c not in df.columns:
            df[c] = np.nan

    def _보고항목(r):
        if r['분류'] != '판관비':
            return np.nan
        if r['계정과목'] in ('감가상각비', '무형자산상각비'):
            return '상각비'
        return ACT2ITEM.get(r['활동분류(대분류)'], '기타관리경비')

    for i, 규 in 바꿀활동.items():          # 재분류 — 활동분류·활동세부도 함께
        if 규[1]:
            df.at[i, '활동분류(대분류)'] = 규[1]
        if 규[2]:
            df.at[i, '활동세부'] = 규[2]

    df['보고항목'] = df.apply(_보고항목, axis=1)
    if 연도힌트:
        df = df[df['년'] == 연도힌트]
    return df.reset_index(drop=True)


def 월간보고입력읽기(xls):
    """실적파일의 「월간보고_입력」 시트를 읽습니다. 없으면 빈 값으로 돌려줍니다."""
    빈값 = {'단일': {}, '월별': {}, '진척': {}, '서술': {}, '고객목표': {}}
    if '월간보고_입력' not in xls.sheet_names:
        return 빈값
    try:
        d = pd.read_excel(xls, sheet_name='월간보고_입력', header=None)
    except Exception:
        return 빈값
    구역 = None
    for _, row in d.iterrows():
        vals = list(row.values)
        라벨 = str(vals[1]).strip() if len(vals) > 1 and pd.notna(vals[1]) else ''
        if not 라벨 or 라벨.lower() == 'nan':
            continue
        if 라벨.startswith('■'):
            t = 라벨
            구역 = ('단일' if '기본' in t else '월별' if '월별' in t else
                    '고객목표' if '고객' in t else '진척' if '진척' in t else
                    '서술' if '서술' in t else None)
            continue
        if 구역 is None or 라벨.startswith('※') or 라벨 in ('항목', '구분', '고객사'):
            continue
        v2 = vals[2] if len(vals) > 2 else None
        if 구역 == '월별':
            빈값['월별'][라벨] = [float(vals[2 + i]) if len(vals) > 2 + i
                                  and isinstance(vals[2 + i], (int, float))
                                  and pd.notna(vals[2 + i]) else 0.0 for i in range(12)]
        elif 구역 == '서술':
            빈값['서술'][라벨] = '' if pd.isna(v2) else str(v2)
        else:
            if pd.isna(v2):
                continue
            try:
                빈값[구역][라벨] = float(v2)
            except (TypeError, ValueError):
                빈값[구역][라벨] = v2
    return 빈값


def _매핑표만들기(m):
    """BS_IS_매핑 시트 → 계정 영문명으로 찾아 쓰는 사전. 원장읽기와 엑셀작성이 함께 씁니다."""
    m = m.copy()
    m.columns = [str(c).strip() for c in m.columns]
    영문col = [c for c in m.columns if '영문' in c or 'account' in c.lower()][0]
    bsiscol = [c for c in m.columns if 'BS' in c][0]
    korcol = [c for c in m.columns if '한글' in c][0]
    key = m[영문col].astype(str).str.strip().str.lower()
    return {'bsis': dict(zip(key, m[bsiscol])), 'kor': dict(zip(key, m[korcol])),
            'table': m.rename(columns={영문col: '계정(영문)', bsiscol: 'BS/IS',
                                       korcol: '계정과목(한글)'})}


def _기초현금계산(raw, 매핑):
    """원장 원본(머리글 5행)에서 현금·예금 계정의 Beginning Balance 합계를 구합니다.

    Beginning Balance 줄은 거래가 아니라 기초잔액이라 _표준화()에서 걸러지므로
    여기서 따로 읽습니다. 계정 이름은 A열에 한 번만 적혀 있어 ffill 로 채웁니다.
    """
    현금계정 = {k for k, v in 매핑['kor'].items()
                if 매핑['bsis'].get(k) == 'BS' and isinstance(v, str)
                and ('현금' in v or '예금' in v)}
    try:
        if 'Distribution account' not in [str(c) for c in raw.columns]:
            return 0.0
        그룹 = raw.iloc[:, 0].astype('object').ffill().astype(str).str.strip().str.lower()
        계정 = raw['Distribution account'].astype(str).str.strip()
        잔액 = pd.to_numeric(raw.get('Balance'), errors='coerce').fillna(0)
        return float(잔액[계정.eq('Beginning Balance') & 그룹.isin(현금계정)].sum())
    except Exception:
        return 0.0


# 차입금 줄에 붙일 한글 이름 — 원장 계정 코드(앞 4자리)로 찾고, 없으면 영문 이름 그대로 씁니다
차입이름 = {'2750': 'CTK 본사 차입금 (Loan from HQ)', '2710': 'Open Bank 한도대출 (LOC #0060)',
            '2810': 'Open Bank Term Loan 31900232', '2820': 'Open Bank Term Loan 30160061',
            '2830': 'Chase 451928870001', '2880': 'EIDL 대출', '2890': 'PPP 대출',
            '2910': '주주 차입금'}
# 리스부채 줄 이름 (표시 순서대로) — 계정 영문 이름에 든 낱말로 나눕니다
# ※ OTC 법인은 2026년 6월 현재 사용권자산·리스부채 계정이 없습니다.
#   리스를 도입하면 아래 목록에 계정 영문 이름의 낱말을 넣으면 됩니다.
리스이름 = [('lease_building', '건물 리스', 'building'),
            ('lease_equip', '설비 리스', 'equipment'),
            ('lease_office', '사무기기 리스', 'office equipment'),
            ('lease_auto', '차량 리스', 'automobile')]


def _이자분류(적요):
    """이자비용(9100) 전표를 적요 낱말로 나눕니다. 차입금·리스부채 표에서 씁니다."""
    s = str(적요).lower().strip()
    if re.search(r'\baro\b', s):
        return 'aro'
    if 'hamner lease' in s or 'lease interest accrual' in s:
        return 'lease_building'
    if s.startswith('rent for '):
        return 'rent'                      # 건물 리스이자로 추정 — 확인 필요 (건물 줄에 합산)
    if ('warehouse equipment' in s or 'big joe' in s or 'liberty' in s
            or s.startswith('first monthly payment')):
        return 'lease_equip'
    if 'tesla' in s or 'audi' in s:
        return 'lease_auto'
    if 'camry' in s:
        return 'auto'                      # 차량 할부금(대출) 이자
    if 'interest on loan payable' in s:
        return '2420'
    if (re.match(r'interest for (?!line of credit|cd)', s)
            or s.startswith('accrued interest')):
        return 'unknown'                   # 차입금 이자인데 적요에 은행명이 없음
    return 'etc'


def _차입리스현황(xls, 매핑):
    """차입금·리스부채의 기초/증감/기말 잔액과 두 해 지급이자.

    잔액 : 26년 원장_raw 의 Beginning Balance(기초) + 거래 Credit−Debit(증감)
    이자 : 25·26년 원장의 이자비용 전표를 적요 기준으로 나눔 (_이자분류)
    """
    차입계정 = {k for k, v in 매핑['kor'].items()
                if isinstance(v, str) and v in ('단기차입금', '장기차입금')}
    복구계정 = {k for k, v in 매핑['kor'].items()
                if isinstance(v, str) and v == '복구충당부채'}
    리스계정 = {k for k, v in 매핑['kor'].items()
                if isinstance(v, str) and v.startswith('리스부채')}
    이자계정 = {k for k, v in 매핑['kor'].items()
                if isinstance(v, str) and v == '이자비용'}
    if not 차입계정:
        return None

    def 묶음키(영문소문자):
        if 영문소문자 in 차입계정:
            if 'auto loan' in 영문소문자:
                return 'auto'
            코드 = 영문소문자.split()[0][:4]
            return 코드 if 코드 in 차입이름 else 'loan_' + 코드
        if 영문소문자 in 복구계정:
            return 'aro'
        if 영문소문자 in 리스계정:
            for key, _, 낱말 in 리스이름:
                if 낱말 in 영문소문자:
                    return key
            return 'lease_building'
        return None

    # ── 잔액 (26년 원장_raw 기준) ──────────────────────────
    raw = pd.read_excel(xls, sheet_name='26년 원장_raw', header=4)
    그룹 = raw.iloc[:, 0].astype('object').ffill().astype(str).str.strip().str.lower()
    계정 = raw['Distribution account'].astype(str).str.strip()
    잔액 = pd.to_numeric(raw.get('Balance'), errors='coerce').fillna(0)
    차 = pd.to_numeric(raw.get('Debit'), errors='coerce').fillna(0)
    대 = pd.to_numeric(raw.get('Credit'), errors='coerce').fillna(0)
    잔고 = {}                                    # key → [기초, 증감]
    for i in raw.index:
        key = 묶음키(그룹.iat[i])
        if key is None:
            continue
        칸 = 잔고.setdefault(key, [0.0, 0.0])
        if 계정.iat[i] == 'Beginning Balance':
            칸[0] += float(잔액.iat[i])
        else:
            칸[1] += float(대.iat[i] - 차.iat[i])   # 부채는 대변이 늘어나는 쪽

    # ── 지급이자 (두 해 원장의 이자비용 전표) ───────────────
    이자 = {'전년': {}, '당해': {}, '렌트': 0.0}
    for 시트, 해 in (('25년 원장_raw', '전년'), ('26년 원장_raw', '당해')):
        if 시트 not in xls.sheet_names:
            continue
        r = pd.read_excel(xls, sheet_name=시트, header=4)
        g = r.iloc[:, 0].astype('object').ffill().astype(str).str.strip().str.lower()
        a = r['Distribution account'].astype(str).str.strip()
        d = pd.to_numeric(r.get('Debit'), errors='coerce').fillna(0)
        c = pd.to_numeric(r.get('Credit'), errors='coerce').fillna(0)
        표적 = g.isin(이자계정) & ~a.eq('Beginning Balance')
        이름칸 = r['Name'] if 'Name' in r.columns else None
        for i in r.index[표적]:
            금액 = float(d.iat[i] - c.iat[i])       # 비용은 차변이 +
            if abs(금액) < 0.005:
                continue
            key = _이자분류(r['Description'].iat[i] if 'Description' in r.columns else '')
            # 적요에 은행명이 없는 이자는 전표의 거래처(Name) 이름으로 은행을 찾습니다
            if key == 'unknown' and 이름칸 is not None:
                nm = str(이름칸.iat[i]).lower()
                for 낱말, 코드 in (('shinhan', '2450'), ('hanmi', '2460'), ('woori', '2410'),
                                   ('citibank', '2440'), ('sba', '2430')):
                    if 낱말 in nm:
                        key = 코드
                        break
            if key == 'rent':                       # 건물 리스이자로 추정해 합산 (확인 필요)
                이자[해]['lease_building'] = 이자[해].get('lease_building', 0.0) + 금액
                if 해 == '전년':
                    이자['렌트'] += 금액
                continue
            이자[해][key] = 이자[해].get(key, 0.0) + 금액

    # ── 「차입금 비고」 시트 — 엑셀에 넣어 두면 웹 대시보드에서도 비고가 보입니다.
    #    A열 = 코드(2450 처럼), 마지막 열 = 비고 글. 내 PC 에서 키인한 비고가 있으면 그쪽이 우선.
    엑셀비고 = {}
    if '차입금 비고' in xls.sheet_names:
        try:
            b = pd.read_excel(xls, sheet_name='차입금 비고', header=0)
            for _, r in b.iterrows():
                k = str(r.iloc[0]).strip()
                if re.fullmatch(r'\d+\.0', k):
                    k = k[:-2]
                v = str(r.iloc[-1]).strip()
                if k and k.lower() != 'nan' and v and v.lower() != 'nan':
                    엑셀비고[k] = v
        except Exception:
            pass
    return {'잔고': 잔고, '이자': 이자, '엑셀비고': 엑셀비고}



def _IS밑자료(xls, 매핑):
    """연결패키지 손익계산서에 넣을 「국문 계정과목별 금액」과, 계정(영문)별 금액.

    당기 : 26년 원장_raw 1~보고월 (수익은 대변−차변, 그 밖은 차변−대변)
    전기 : 25년 원장_raw 한 해 전체
    """
    def 한해(시트):
        if 시트 not in xls.sheet_names:
            return {}, {}
        raw = pd.read_excel(xls, sheet_name=시트, header=4)
        이름 = _글자칸(raw, 'Distribution account')
        차 = pd.to_numeric(raw.get('Debit'), errors='coerce').fillna(0)
        대 = pd.to_numeric(raw.get('Credit'), errors='coerce').fillna(0)
        날 = pd.to_datetime(raw['Transaction date'], errors='coerce')
        과목별, 계정별 = {}, {}
        for i in range(len(raw)):
            n = 이름[i]
            if not n or n == 'Beginning Balance' or pd.isna(날.iat[i]):
                continue
            키 = n.strip().lower()
            if 매핑['bsis'].get(키) != 'IS':
                continue
            과목 = str(매핑['kor'].get(키, '')).strip()
            매출 = 과목 in ('제품매출', '기타매출', '잡이익', '외환차익', '외화환산이익')
            v = (float(대.iat[i]) - float(차.iat[i])) if 매출 else (float(차.iat[i]) - float(대.iat[i]))
            계정별[n] = 계정별.get(n, 0.0) + v
            if 과목:
                과목별[과목] = 과목별.get(과목, 0.0) + v
        return 과목별, 계정별

    당기과목, 당기계정 = 한해('26년 원장_raw')
    전기과목, 전기계정 = 한해('25년 원장_raw')
    return {'당기': 당기과목, '전기': 전기과목,
            '당기계정': 당기계정, '전기계정': 전기계정}


def _영문BS자료(xls, 매핑):
    """퀵북 계정(영문) 이름 그대로의 재무상태표 — 전기 · 당기 · 증감."""
    전기, 차례 = {}, []
    if '26년 BS기초' in xls.sheet_names:
        d = pd.read_excel(xls, sheet_name='26년 BS기초', header=None)
        for _i, 줄 in d.iterrows():
            n = 줄.iloc[0]
            if n is None or (isinstance(n, float) and pd.isna(n)):
                continue
            n = str(n).strip()
            if not n:
                continue
            v = pd.to_numeric(pd.Series([줄.iloc[1]]), errors='coerce').fillna(0).iat[0]
            if n not in 전기:
                차례.append(n)
            전기[n] = 전기.get(n, 0.0) + float(v)

    raw = pd.read_excel(xls, sheet_name='26년 원장_raw', header=4)
    머리 = _글자칸(raw, raw.columns[0])
    그룹, 지금 = [], ''
    for v in 머리:
        if v:
            지금 = v
        그룹.append(지금)
    이름 = _글자칸(raw, 'Distribution account')
    차 = pd.to_numeric(raw.get('Debit'), errors='coerce').fillna(0)
    대 = pd.to_numeric(raw.get('Credit'), errors='coerce').fillna(0)
    잔 = pd.to_numeric(raw.get('Balance'), errors='coerce').fillna(0)
    날 = pd.to_datetime(raw['Transaction date'], errors='coerce')
    기초, 증감 = {}, {}
    for i in range(len(raw)):
        if 이름[i] == 'Beginning Balance':
            기초[그룹[i]] = 기초.get(그룹[i], 0.0) + float(잔.iat[i])
        elif 이름[i] and pd.notna(날.iat[i]):
            증감[이름[i]] = 증감.get(이름[i], 0.0) + float(차.iat[i]) - float(대.iat[i])
    당기 = {}
    for 계정 in set(기초) | set(증감):
        키 = 계정.strip().lower()
        과목 = str(매핑['kor'].get(키, '')).strip()
        if 매핑['bsis'].get(키) != 'BS':
            continue
        b, m = 기초.get(계정, 0.0), 증감.get(계정, 0.0)
        if 과목 in BS자산칸부채:
            당기[계정] = m - b
        elif 과목 in BS부채자본:
            당기[계정] = b - m
        else:
            당기[계정] = b + m
    for 계정 in 당기:
        if 계정 not in 전기:
            차례.append(계정)
    # 퀵북 표 차례대로 그릴 때 이름으로 찾아 쓰는 전기 값 (빈 칸은 빈 칸 그대로)
    전기맞춤 = {}
    if '26년 BS기초' in xls.sheet_names:
        d = pd.read_excel(xls, sheet_name='26년 BS기초', header=None)
        for _i, 줄 in d.iterrows():
            n = 줄.iloc[0]
            if n is None or (isinstance(n, float) and pd.isna(n)):
                continue
            키 = _맞춤이름(n)
            if not 키:
                continue
            v = pd.to_numeric(pd.Series([줄.iloc[1]]), errors='coerce').iat[0]
            if pd.notna(v):
                전기맞춤[키] = float(v)
            else:
                전기맞춤.setdefault(키, None)
    return {'차례': 차례, '전기': 전기, '당기': 당기, '전기맞춤': 전기맞춤}


꼬리줄 = re.compile(r'(accrual|cash)\s+basis|\w+day,\s', re.I)


def _총계이름(n):
    """「Total for Income」 · 「Total Income」 → 'income' (아니면 None)."""
    m = re.match(r'^\s*total\s+(?:for\s+)?(.+?)\s*$', str(n or ''), re.I)
    return re.sub(r'\s+', ' ', m.group(1)).lower() if m else None


def _맞춤이름(n):
    return re.sub(r'^total for\s+', 'total ',
                  re.sub(r'\s+', ' ', str(n or '')).strip().lower())


@st.cache_data(show_spinner=False)
def 퀵북표읽기(_도장, 종류):
    """보관해 둔 원장의 「○월(BS)」·「○월(IS)」 시트를 차례·들여쓰기 그대로 읽습니다.

      · 줄 차례는 퀵북이 뽑아 준 그대로입니다 (구역 머리글·Total 줄 포함)
      · 들여쓰기는 「뒤에 Total 줄이 있는 이름 = 구역을 여는 줄」 로 되살립니다
    """
    d = pd.read_excel(원장보관, sheet_name=None, header=None)
    시트, 달 = None, 0
    for s in d:
        m = re.search(r'(\d+)\s*월\s*\(\s*(BS|IS)\s*\)', str(s), re.I)
        if m and m.group(2).upper() == 종류:
            시트, 달 = s, int(m.group(1))
    if 시트 is None:
        return None
    표 = d[시트]
    제목, 이름들, 금액들 = [], [], []
    for i, 줄 in 표.iterrows():
        n = 줄.iloc[0]
        n = '' if n is None or (isinstance(n, float) and pd.isna(n)) else str(n).strip()
        v = pd.to_numeric(pd.Series([줄.iloc[1] if len(줄) > 1 else None]),
                          errors='coerce').iat[0]
        if i < 5:
            if n and n.lower() != 'total':
                제목.append(n)
            continue
        if not n or 꼬리줄.search(n):
            continue
        이름들.append(n)
        금액들.append(None if pd.isna(v) else float(v))
    남은 = {}
    for n in 이름들:
        t = _총계이름(n)
        if t:
            남은[t] = 남은.get(t, 0) + 1
    깊이, 쌓기 = [], []
    for n in 이름들:
        t = _총계이름(n)
        키 = re.sub(r'\s+', ' ', n).lower()
        if t and t in 쌓기:
            while 쌓기 and 쌓기[-1] != t:
                쌓기.pop()
            쌓기.pop()
            깊이.append(len(쌓기))
        else:
            깊이.append(len(쌓기))
            if 남은.get(키, 0) > 0 and 키 not in 쌓기:
                쌓기.append(키)
                남은[키] -= 1
    while len(제목) < 3:
        제목.append('')
    return {'제목': 제목, '시트': 시트, '월': 달,
            '줄': [{'이름': n, '깊이': g, '당기': v, '총계': bool(_총계이름(n))}
                   for n, g, v in zip(이름들, 깊이, 금액들)]}


# 「26년 자금」 시트의 활동성항목 → 현금흐름 표의 줄
예상항목줄 = {'매출채권 회수': 'AR', '기타입금': 'ETCIN', '매입채무 지급': 'AP',
              '급여': 'PAY', '인건비성 항목': 'PAY', '기타 비용': 'OTHOP',
              '설비 투자': 'CAPEX', '유상증자': 'EQ', '차입증감': '차입증감',
              '리스료 지급': '리스', '환율변동 효과': 'FX'}
# 돈이 「들어오는」 항목 — 나머지는 나가는 것으로 봅니다.
#  ※ 차변/대변 어느 칸에 적으셔도 항목 이름으로 방향을 정합니다.
예상입금항목 = {'매출채권 회수', '기타입금', '유상증자', '차입증감'}


@st.cache_data(show_spinner=False)
def 예상자금읽기(데이터바이트):
    """「26년 자금」 시트의 2번(예상 출금) · 3번(예상 입금)을 월별로 모읍니다.

      · 칸 자리를 미리 정해 두지 않고 머리글(일자·차변·대변·활동성항목)을 보고 찾습니다.
        그래서 「기말」 칸을 빼시거나 자리를 옮기셔도 그대로 읽습니다.
      돌려주는 값 : {줄이름: {월: 금액}} — 들어온 돈 +, 나간 돈 −
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(데이터바이트))
        시트 = next((s for s in xls.sheet_names if '자금' in str(s)), None)
        if 시트 is None:
            return None
        d = pd.read_excel(xls, sheet_name=시트, header=None)
    except Exception:
        return None

    def 글자(i, j):
        v = d.iat[i, j] if 0 <= i < len(d) and 0 <= j < d.shape[1] else None
        return '' if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()

    머리행 = next((i for i in range(min(12, len(d)))
                   if any(글자(i, j) == '일자' for j in range(d.shape[1]))), None)
    if 머리행 is None:
        return {}
    일자열 = [j for j in range(d.shape[1]) if 글자(머리행, j) == '일자']
    모음 = {}
    for k, 첫 in enumerate(일자열):
        끝 = 일자열[k + 1] if k + 1 < len(일자열) else d.shape[1]
        자리 = {}
        for j in range(첫, 끝):
            이름 = 글자(머리행, j)
            if 이름 in ('차변', '대변', '활동성항목') and 이름 not in 자리:
                자리[이름] = j
        # 이 칸이 몇 번 표인지 (머리글 위쪽에서 「2.」·「3.」 을 찾습니다)
        번호 = ''
        for i in range(머리행):
            for j in range(max(0, 첫 - 1), 끝):
                m = re.match(r'^\s*([123])\s*\.', 글자(i, j))
                if m:
                    번호 = m.group(1)
        if 번호 not in ('2', '3') or '활동성항목' not in 자리:
            continue                      # 1번(실제)은 원장에서 이미 읽었습니다
        날 = pd.to_datetime(d.iloc[:, 첫], errors='coerce')
        차 = (pd.to_numeric(d.iloc[:, 자리['차변']], errors='coerce').fillna(0)
              if '차변' in 자리 else pd.Series(0.0, index=d.index))
        대 = (pd.to_numeric(d.iloc[:, 자리['대변']], errors='coerce').fillna(0)
              if '대변' in 자리 else pd.Series(0.0, index=d.index))
        이름칸 = d.iloc[:, 자리['활동성항목']]
        for i in range(len(d)):
            if pd.isna(날.iat[i]):
                continue
            항목 = 이름칸.iat[i]
            항목 = '' if 항목 is None or (isinstance(항목, float) and pd.isna(항목)) \
                else str(항목).strip()
            줄 = 예상항목줄.get(항목, 'OTHOP')
            # 차변·대변 어느 칸에 넣으셔도 됩니다 — 방향은 활동성항목으로 정합니다
            금액 = float(차.iat[i]) + float(대.iat[i])
            v = 금액 if 항목 in 예상입금항목 else -금액
            if not v:
                continue
            모음.setdefault(줄, {})
            m = int(날.iat[i].month)
            모음[줄][m] = 모음[줄].get(m, 0.0) + v
    return 모음


@st.cache_data(show_spinner=False)
def 예상출금기본(데이터바이트, 보고월):
    """자금 시트에 예상 출금 금액이 안 들어 있을 때 쓰는 값 —
       「월별 실적집계」 관리비 항목의 1~보고월 월평균을 그대로 씁니다."""
    if 보고월 >= 12:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(데이터바이트))
        시트 = next((s for s in xls.sheet_names
                     if '월별' in str(s) and '실적집계' in str(s)), None)
        if 시트 is None:
            return {}
        d = pd.read_excel(xls, sheet_name=시트, header=None)
    except Exception:
        return {}
    # ※ .astype(str) 은 빈 칸을 float nan 으로 내놓을 때가 있어 한 칸씩 바꿉니다
    글 = d.iloc[:, 2].apply(lambda v: '' if v is None or (isinstance(v, float)
                                                         and pd.isna(v)) else str(v).strip())
    시작 = next((i for i in range(len(글)) if 글.iat[i].startswith('관리비 합계')), None)
    끝 = next((i for i in range(시작 + 1, len(글))
               if 글.iat[i].startswith('영업이익')), len(글)) if 시작 is not None else None
    if 시작 is None:
        return {}
    나감 = {}
    for i in range(시작 + 1, 끝):
        m = re.match(r'^(\d+)\)\s*(.+)$', 글.iat[i])
        if not m:
            continue
        이름 = m.group(2).strip()
        if '상각비' in 이름:                            # 현금이 나가지 않습니다
            continue
        값 = pd.to_numeric(d.iloc[i, 23:23 + 보고월], errors='coerce').fillna(0)
        평균 = float(값.sum()) / max(보고월, 1)
        if abs(평균) < 0.005:
            continue
        줄 = 'PAY' if 이름 == '인건비성 항목' else 'OTHOP'
        나감.setdefault(줄, {})
        for 월 in range(보고월 + 1, 13):
            나감[줄][월] = 나감[줄].get(월, 0.0) - 평균     # 나가는 돈이라 −
    return 나감


@st.cache_data(show_spinner='원장을 읽는 중입니다...')
def 원장읽기(데이터바이트):
    xls = pd.ExcelFile(io.BytesIO(데이터바이트))
    시트 = xls.sheet_names
    매핑표 = _매핑표만들기(pd.read_excel(xls, sheet_name='BS_IS_매핑'))
    재분류 = 재분류읽기(xls)

    def _read(sheet, 연도):
        if sheet not in 시트:
            return None
        for header in (0, 4):
            try:
                d = pd.read_excel(xls, sheet_name=sheet, header=header)
            except Exception:
                continue
            cols = [str(c) for c in d.columns]
            if any(c in cols for c in ('계정영문', 'Distribution account')):
                out = _표준화(d, 매핑표, 연도, 재분류)
                if out is not None and len(out):
                    return out
        return None

    g26 = _read('26년 원장_raw', None)
    g25 = _read('25년 원장_raw', None)
    if g26 is None:
        raise ValueError('26년 원장_raw 시트를 읽을 수 없습니다.')
    당해 = int(g26['년'].mode().iat[0])
    g26 = g26[g26['년'] == 당해].reset_index(drop=True)
    if g25 is not None:
        g25 = g25[g25['년'] == 당해 - 1].reset_index(drop=True)
    else:
        g25 = g26.iloc[0:0].copy()

    # ── 현금 계정의 기초잔액
    try:
        기초현금 = _기초현금계산(pd.read_excel(xls, sheet_name='26년 원장_raw', header=4), 매핑표)
    except Exception:
        기초현금 = 0.0

    입력값 = 월간보고입력읽기(xls)
    # ── 연결패키지 재무상태표에 쓸 밑자료 (전기 잔액 · 당기 계정별 잔액)
    try:
        BS자료 = _BS밑자료(xls, 매핑표)
        BS자료['IS'] = _IS밑자료(xls, 매핑표)
        BS자료['영문BS'] = _영문BS자료(xls, 매핑표)
    except Exception:
        BS자료 = None
    # ── 차입금·리스부채 현황 (현금흐름 화면 아래 표)
    try:
        차입리스 = _차입리스현황(xls, 매핑표)
    except Exception:
        차입리스 = None
    return g26, g25, 당해, 매핑표, 기초현금, 입력값, BS자료, 차입리스


glob_소계 = re.compile(r'^\s*Total\b', re.I)
# 자산 자리에 적혀 있지만 실제로는 부채라 부호를 한 번 더 뒤집는 계정과목
BS자산칸부채 = {'부가세대급금'}


def _BS밑자료(xls, 매핑):
    """연결패키지 BS 에 넣을 「국문 계정과목별 금액」을 만듭니다.

    전기 : 26년 BS기초 시트 (F열 CTK 연결 계정과목 · B열 최종확정 금액)
    당기 : 26년 원장_raw 에서 계정별로 기초 + 증감
           ※ 자산은 기초+(차변−대변), 부채·자본은 기초−(차변−대변) —
             퀵북 재무상태표가 부채·자본을 양수로 보여 주는 방식에 맞춥니다.
    """
    전기 = {}
    if '26년 BS기초' in xls.sheet_names:
        d = pd.read_excel(xls, sheet_name='26년 BS기초', header=None)
        for _i, 줄 in d.iterrows():
            # ※ 시트의 「CTK 연결 계정과목」 칸이 아니라, 당기와 똑같이 BS_IS_매핑을 태웁니다.
            #   (그 칸은 매출채권을 미수금으로 적어 두어 당기와 이름이 어긋납니다)
            계정 = 줄.iloc[0]
            if 계정 is None or (isinstance(계정, float) and pd.isna(계정)):
                continue
            글 = str(계정).strip()
            if glob_소계.match(글):      # 「Total …」 줄은 밑줄과 겹쳐 두 번 세집니다
                continue
            키 = 글.lower()
            과목 = str(매핑['kor'].get(키, '')).strip()
            if not 과목 or 매핑['bsis'].get(키) != 'BS':
                continue
            v = pd.to_numeric(pd.Series([줄.iloc[1]]), errors='coerce').fillna(0).iat[0]
            전기[과목] = 전기.get(과목, 0.0) + float(v)

    raw = pd.read_excel(xls, sheet_name='26년 원장_raw', header=4)
    계정칸 = _글자칸(raw, raw.columns[0])
    그룹, 지금 = [], ''
    for v in 계정칸:                      # 계정 이름은 A열 머리줄에 한 번만 적혀 있습니다
        if v:
            지금 = v
        그룹.append(지금)
    이름 = _글자칸(raw, 'Distribution account')
    차 = pd.to_numeric(raw.get('Debit'), errors='coerce').fillna(0)
    대 = pd.to_numeric(raw.get('Credit'), errors='coerce').fillna(0)
    잔 = pd.to_numeric(raw.get('Balance'), errors='coerce').fillna(0)
    날 = pd.to_datetime(raw['Transaction date'], errors='coerce')
    기초, 증감 = {}, {}
    for i in range(len(raw)):
        if 이름[i] == 'Beginning Balance':
            기초[그룹[i]] = 기초.get(그룹[i], 0.0) + float(잔.iat[i])
        elif 이름[i] and pd.notna(날.iat[i]):
            증감[이름[i]] = 증감.get(이름[i], 0.0) + float(차.iat[i]) - float(대.iat[i])

    한글 = {k: v for k, v in 매핑['kor'].items()}
    비에스 = 매핑['bsis']
    당기 = {}
    for 계정 in set(기초) | set(증감):
        키 = 계정.strip().lower()
        과목 = str(한글.get(키, '')).strip()
        if not 과목 or 비에스.get(키) != 'BS':
            continue
        b, m = 기초.get(계정, 0.0), 증감.get(계정, 0.0)
        # 부채·자본은 대변이 늘어나는 쪽이라 부호를 뒤집어 양수로 보여 줍니다
        # (어느 쪽인지는 BS 서식에서 「부채」 머리줄 아래에 있는지로 가립니다)
        if 과목 in BS자산칸부채:      # 자산 자리에 있지만 실제로는 부채인 계정
            값 = m - b
        elif 과목 in BS부채자본:
            값 = b - m
        else:
            값 = b + m
        당기[과목] = 당기.get(과목, 0.0) + 값
    return {'전기': 전기, '당기': 당기}


def 손익표(df, 기준='손익', 상위=7):
    """월별 손익계산서(1~12월). 비용은 양수로 표시.

    기준='손익' : 판관비를 손익계산서 계정과목으로 나눔 (금액 큰 순 상위 N개 + 기타)
    기준='활동' : 판관비를 활동분류(보고항목) 14개로 나눔
    어느 쪽이든 매출액·영업이익·EBITDA 같은 합계 줄은 똑같습니다.
    """
    d = df[df['계정분류'].eq('IS')]
    수준 = {}                      # 줄마다 들여쓰기 단계 (0 굵게 / 1 / 2)

    def S(mask):
        return d[mask].groupby('월')['금액'].sum().reindex(range(1, 13), fill_value=0.0)

    o = {}
    o['매출액'] = S(d['분류'].eq('매출'))
    o['제품 매출'] = S(d['계정과목'].eq('제품매출'))
    o['상품 매출'] = S(d['계정과목'].eq('상품매출'))
    o['용역 매출'] = S(d['계정과목'].eq('용역매출'))
    o['매출원가'] = -S(d['분류'].eq('매출원가'))
    o['매출총이익'] = o['매출액'] - o['매출원가']
    판관 = d['분류'].eq('판관비')
    o['판매관리비'] = -S(판관)
    리스계정 = d['계정영문'].astype(str).str.strip().str.lower().str.startswith('6623')

    def 상각쪼개기(이름, 마스크, 부모유지):
        """EBITDA 에서 무엇을 더하고 뺐는지 보이도록 상각비를 둘로 나눕니다.

        부모유지=True (활동 기준) : 「상각비」 아래에 감가상각비 · 리스상각비를 답니다.
        부모유지=False (손익 기준): 부모 이름이 「감가상각비」와 겹치므로,
                                    부모 자리를 감가상각비(리스 제외분)로 바꾸고
                                    리스상각비를 같은 단계로 나란히 놓습니다.
        """
        리스분 = -S(마스크 & 리스계정)
        본체 = o[이름] - 리스분
        깊이 = 2 if 부모유지 else 1
        if 부모유지:
            o['감가상각비'] = 본체
        else:
            o[이름] = 본체                 # 부모 줄을 그대로 재사용 (이름 중복 방지)
            수준[이름] = 1
        o['리스상각비'] = 리스분
        수준.setdefault('감가상각비', 깊이)
        수준['리스상각비'] = 깊이

    if 기준 == '활동':
        for it in 판관비항목:
            o[it] = -S(d['보고항목'].eq(it))
            수준[it] = 1
            if it == '상각비':
                상각쪼개기(it, d['보고항목'].eq(it), 부모유지=True)
    else:
        큰순 = (-d[판관].groupby('계정과목')['금액'].sum()).sort_values(ascending=False)
        고른계정 = [str(a) for a in 큰순.index[:상위] if str(a) != 'nan']
        for acc in 고른계정:
            마스크 = 판관 & d['계정과목'].eq(acc)
            o[acc] = -S(마스크)
            수준[acc] = 1
            if acc in ('감가상각비', '무형자산상각비'):
                상각쪼개기(acc, 마스크, 부모유지=False)
        # 「기타」는 상위 계정을 뺀 나머지 — 리스상각비는 상위 계정에서 갈라져 나온 것이라 함께 뺍니다
        고른합 = sum((o[a] for a in 고른계정), pd.Series(0.0, index=range(1, 13)))
        if '리스상각비' in o:
            고른합 = 고른합 + o['리스상각비']
        o['기타'] = o['판매관리비'] - 고른합
        수준['기타'] = 1
    # EBITDA 에서 리스(사용권자산) 상각은 제외합니다
    o['영업이익(손실)'] = o['매출총이익'] - o['판매관리비']
    o['영업외손익'] = S(d['분류'].eq('영업외손익'))
    o['세전이익'] = o['영업이익(손실)'] + o['영업외손익']
    o['법인세비용'] = -S(d['분류'].eq('법인세'))
    o['당기순이익'] = o['세전이익'] - o['법인세비용']
    상각계정 = d['계정과목'].isin(['감가상각비', '무형자산상각비'])
    상각 = -S(상각계정)
    리스상각 = -S(상각계정 & 리스계정)
    o['EBITDA 이익(손실)'] = o['영업이익(손실)'] + 상각 - 리스상각
    t = pd.DataFrame(o).T
    t.columns = list(range(1, 13))
    기본수준 = {'매출액': 0, '제품 매출': 1, '상품 매출': 1, '용역 매출': 1,
                '매출원가': 0, '매출총이익': 0,
                '판매관리비': 0, '영업이익(손실)': 0, '영업외손익': 0, '세전이익': 0,
                '법인세비용': 0, '당기순이익': 0, 'EBITDA 이익(손실)': 0}
    t.attrs['수준'] = {**기본수준, **수준}
    return t


# 줄 들여쓰기 단계는 손익표()가 표마다 t.attrs['수준'] 으로 함께 돌려줍니다.

# 실적집계 두 화면에서는 빼고 보여줄 줄 (다른 화면 계산에는 계속 씁니다)
집계제외 = ['영업외손익', '세전이익', '법인세비용', '당기순이익']

# 오른쪽 세부내역에서 고를 수 없게 할 줄 — 합계·계산으로 만들어진 줄입니다
드릴제외 = ('매출총이익', '영업이익(손실)', 'EBITDA 이익(손실)', '세전이익', '당기순이익')


def 최하위줄(줄들, 수준표, 제외=()):
    """표에서 「더 이상 쪼개지지 않는」 줄만 골라 냅니다.

    바로 다음 줄이 더 깊이 들여쓰기되어 있으면 그 줄은 하위 항목을 거느린
    상위 줄이므로 (매출액 · 판매관리비 · 감가상각비처럼) 오른쪽 세부내역
    목록에서는 빼고, 하위 줄만 남깁니다. 하위가 없는 줄은 그대로 씁니다.
    """
    골라낸 = []
    for i, 이름 in enumerate(줄들):
        내수준 = 수준표.get(이름, 1)
        다음수준 = 수준표.get(줄들[i + 1], 1) if i + 1 < len(줄들) else -1
        if 다음수준 > 내수준 or 이름 in 제외:
            continue
        골라낸.append(이름)
    return 골라낸
# ══════════════════════════════════════════════════════════════
# 2-2. 월간실적보고용 공통 계산 (통화 · 현금)
# ══════════════════════════════════════════════════════════════
def _미사용_정부지원월별(df):     # OTC법인 미사용 (캐나다 전용이던 함수)
    """(사용하지 않습니다)"""
    m = df['계정영문'].astype(str).str.strip().str.lower().eq('sr&ed claim')
    return df[m].groupby('월')['금액'].sum().reindex(range(1, 13), fill_value=0.0)


def 현금잔고월별(df, 기초, 매핑=None):
    매핑 = 매핑 or 매핑표      # 평소엔 화면이 읽어 둔 매핑표, 엑셀작성 화면에선 넘겨받은 것
    현금 = {k for k, v in 매핑['kor'].items()
            if 매핑['bsis'].get(k) == 'BS' and isinstance(v, str)
            and ('현금' in v or '예금' in v)}
    m = df['계정영문'].astype(str).str.strip().str.lower().isin(현금)
    증감 = (-df[m].groupby('월')['금액'].sum()).reindex(range(1, 13), fill_value=0.0)
    return 기초 + 증감.cumsum(), 증감


# ── 자금 변동(현금흐름) 분류 규칙 — 필요하면 여기만 고치면 됩니다 ─────
자금분류 = {
    'AR':    {'매출채권', '미수금', '선수금'},
    'AP':    {'매입채무', '미지급금', '미지급비용'},
    'PAY':   {'예수금', '예수금(보험료)', '예수금(401k)', '미지급급여',
              '급여', '복리후생비', '퇴직급여'},
    'EQ':    {'보통주자본금', '주식발행초과금', '자본잉여금', '기타자본'},
    'DEBT':  {'단기차입금', '장기차입금', '전환사채', '단기대여금', '장기대여금'},
    'LEASE': {'리스부채(유동)', '리스부채(비유동)'},
    'CAPEX': {'기계장치', '건설중인자산', '시설장치', '비품', '공구와기구',
              '차량운반구', '소프트웨어', '기타의무형자산', '장기선급금',
              '무형자산', '특허권', '사용권자산', '사용권자산(비유동)'},
    'FX':    {'외환차손', '외화환산이익', '외환차익', '외화환산손실',
              '외환차손익', '외화환산손익'},
}
# 매출채권 회수가 아닌 특별 입금을 「기타입금」으로 가려내던 자리입니다.
# 캐나다는 SR&ED(정부 R&D 세액공제) 환급이 여기 해당했지만, OTC법인에는 없습니다.
# 나중에 그런 입금이 생기면 아래 목록에 적요 단어를 넣으면 됩니다.
기타입금단어 = []          # 예) ['grant', 'refund']


def _기타입금줄(df, 현금행, 차변):
    """현금이 들어온 줄 가운데 매출채권 회수가 아닌 것을 가려냅니다."""
    if not 기타입금단어:
        return pd.Series(False, index=df.index)
    글 = pd.Series('', index=df.index)
    for c in ('Description', '#'):
        if c in df.columns:
            글 = 글 + ' ' + df[c].fillna('').astype(str)
    들어옴 = pd.to_numeric(차변, errors='coerce').fillna(0) > 0
    말 = re.compile('|'.join(기타입금단어), re.I)
    return 현금행 & 들어옴 & 글.str.contains(말, regex=True, na=False)


채권구간 = ['within Due', '~30', '~60', '~90', '~120', '121~']


def _만기일(거래일):
    """만기 = 매출한 달의 「다음 달 말일」."""
    해, 달 = 거래일.year + (거래일.month == 12), 거래일.month % 12 + 1
    return pd.Timestamp(해, 달, calendar.monthrange(해, 달)[1])


def _채권구간(경과일):
    if 경과일 <= 0:
        return 채권구간[0]
    return (채권구간[1] if 경과일 <= 30 else 채권구간[2] if 경과일 <= 60
            else 채권구간[3] if 경과일 <= 90 else 채권구간[4] if 경과일 <= 120
            else 채권구간[5])


def 채권나이표(ar, 기초채권, 기준일):
    """거래처별 매출채권을 선입선출로 남겨 경과 구간에 나눕니다.

      · 만기 = 매출한 달의 다음 달 말일 (예: 6월 매출 → 7월 31일)
      · 회수는 오래된 매출부터 차례로 지웁니다 (선입선출)
      · 그 거래처의 올해 매출보다 더 받은 돈은 지난해에서 넘어온
        「기초채권」을 갚은 것으로 보고 기초채권에서 뺍니다
    """
    d = ar.sort_values(['거래일', 'TransactionID'], kind='stable')
    이름 = d['Name'].apply(lambda v: '' if v is None or (isinstance(v, float) and pd.isna(v))
                           else str(v).strip())
    차 = pd.to_numeric(d.get('Debit'), errors='coerce').fillna(0).tolist()
    대 = pd.to_numeric(d.get('Credit'), errors='coerce').fillna(0).tolist()
    날 = list(d['거래일'])
    이름 = 이름.tolist()

    남은, 초과 = {}, 0.0
    for i in range(len(d)):
        n = 이름[i] or '(거래처 미상)'
        큐 = 남은.setdefault(n, [])
        if 차[i] > 0:
            큐.append([날[i], float(차[i])])
        돈 = float(대[i])
        while 돈 > 1e-9 and 큐:
            뺄 = min(돈, 큐[0][1])
            큐[0][1] -= 뺄
            돈 -= 뺄
            if 큐[0][1] <= 1e-9:
                큐.pop(0)
        초과 += 돈                      # 올해 매출을 다 갚고도 남은 돈 → 기초채권에서

    줄 = []
    for n, 큐 in 남은.items():
        칸 = {c: 0.0 for c in 채권구간}
        for 날짜, 금액 in 큐:
            if 금액 <= 0.005:
                continue
            칸[_채권구간((기준일 - _만기일(날짜)).days)] += 금액
        if sum(칸.values()) > 0.5:
            줄.append({'거래처명': n, **칸, '채권총액': sum(칸.values())})

    기초남음 = max(float(기초채권) - 초과, 0.0)
    if 기초남음 > 0.5:                  # 25년에서 넘어온 몫 — 만기가 26년 1월말이라 121일↑
        칸 = {c: 0.0 for c in 채권구간}
        칸['121~'] = 기초남음
        줄.append({'거래처명': '기초채권 (거래처 없음)', **칸, '채권총액': 기초남음})

    표 = pd.DataFrame(줄, columns=['거래처명'] + 채권구간 + ['채권총액'])
    if len(표):
        표 = 표.sort_values('채권총액', ascending=False).reset_index(drop=True)
    return 표, 초과


def 자금변동(df, 매핑=None):
    """기초 현금에서 출발해 현금이 무엇 때문에 늘고 줄었는지 월별로 가릅니다.

    원장은 복식부기라, 현금 계정이 들어간 전표(TransactionID)에서 현금이 아닌
    상대 계정의 (대변−차변)이 곧 그 성격의 현금 증감이 됩니다. 은행↔은행 이체처럼
    현금끼리 오간 전표는 상대 계정이 없어 자동으로 상계됩니다.
    """
    매핑 = 매핑 or 매핑표      # 평소엔 화면이 읽어 둔 매핑표, 엑셀작성 화면에선 넘겨받은 것
    현금 = {k for k, v in 매핑['kor'].items()
            if 매핑['bsis'].get(k) == 'BS' and isinstance(v, str)
            and ('현금' in v or '예금' in v)}
    키 = df['계정영문'].astype(str).str.strip().str.lower()
    현금행 = 키.isin(현금)
    if 'TransactionID' not in df.columns:
        return None, None
    전표 = df['TransactionID']
    현금전표 = set(전표[현금행].dropna())
    상대 = df[전표.isin(현금전표) & ~현금행].copy()
    if 상대.empty:
        return None, None
    상대['영향'] = 상대['Debit'].mul(-1).add(상대['Credit'])      # 대변−차변 = 현금 증감

    이름 = 상대['계정과목'].astype(str)

    def 구분(n):
        for k, s in 자금분류.items():
            if n in s:
                return k
        return 'OTHOP'
    상대['구분'] = [구분(n) for n in 이름]
    상대.loc[이름.str.contains('감가상각누계', na=False), '구분'] = 'CAPEX'
    # 매출채권 회수가 아닌 특별 입금은 기타입금으로 갈라 냅니다
    같은전표 = set(전표[_기타입금줄(df, 현금행, df.get('Debit'))].dropna())
    상대.loc[상대['TransactionID'].isin(같은전표), '구분'] = 'ETCIN'

    월들 = list(range(1, 13))
    표 = (상대.pivot_table(index='구분', columns='월', values='영향', aggfunc='sum')
          .reindex(columns=월들).fillna(0.0))

    def 줄(k):
        return 표.loc[k] if k in 표.index else pd.Series(0.0, index=월들)

    def 양(k):
        return 줄(k).clip(lower=0)

    def 음(k):
        return 줄(k).clip(upper=0)

    o = {}
    o['AR'], o['AP'], o['PAY'], o['OTHOP'] = 줄('AR'), 줄('AP'), 줄('PAY'), 줄('OTHOP')
    o['ETCIN'] = 줄('ETCIN')          # 정부지원금 환급 등 매출채권 회수가 아닌 입금
    o['지급계'] = o['AP'] + o['PAY'] + o['OTHOP']
    o['영업'] = o['AR'] + o['ETCIN'] + o['지급계']
    o['CAPEX'] = 줄('CAPEX')
    o['투자'] = o['CAPEX']
    o['EQ'] = 줄('EQ')
    # 차입은 한 줄로 봅니다 — 빌리면 +, 갚으면 − (같은 달에 둘 다 있으면 서로 상계)
    o['차입'] = 양('DEBT')
    o['상환'] = 음('DEBT')
    o['차입증감'] = 줄('DEBT')
    o['리스'] = 줄('LEASE')
    o['재무'] = o['EQ'] + o['차입증감'] + o['리스']
    o['FX'] = 줄('FX')
    o['순증감'] = o['영업'] + o['투자'] + o['재무'] + o['FX']
    t = pd.DataFrame(o).T
    t.columns = 월들
    return t, 상대[['구분', '영향', '월', 'Name']]






# ══════════════════════════════════════════════════════════════
# 분석 엑셀 만들기 — 올린 원장 하나로 보고용 엑셀을 즉석에서 생성
# ══════════════════════════════════════════════════════════════
def 분석엑셀만들기(원장, 전년, 손익표본, 활동표본, 자금, 잔고, 기초현금액, 연도, 월끝, 매핑,
                   원천시트=None):
    """대시보드가 보여 주는 표를 그대로 담은 엑셀을 만들어 bytes 로 돌려줍니다.

    ★피벗·외부링크를 쓰지 않습니다★ — 엑셀에서 「복구됨」이 뜰 여지를 없앴습니다.
    「원장_처리본」 시트가 원천이고, 손익 표는 그 시트를 SUMIFS 로 집계하므로
    원장을 고치면 엑셀 안에서 숫자가 다시 계산됩니다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY, OR_, GREY, RED = 'FF12294D', 'FFE2611C', 'FF8A8880', 'FFD03B3B'
    글꼴 = '맑은 고딕'
    머리채움 = PatternFill('solid', fgColor=NAVY)
    띠채움 = PatternFill('solid', fgColor='FFF7F8FA')
    누적채움 = PatternFill('solid', fgColor='FFF1F4F9')
    가는 = Side(style='thin', color='FFE7E7E1')
    테두리 = Border(left=가는, right=가는, top=가는, bottom=가는)
    돈 = '#,##0;(#,##0);-'
    월들 = list(range(1, 월끝 + 1))

    wb = Workbook()

    def 표머리(ws, 제목, 부제, 열들, 시작행=1):
        c = ws.cell(시작행, 1, 제목)
        c.font = Font(글꼴, size=15, bold=True, color=NAVY)
        ws.cell(시작행 + 1, 1, 부제).font = Font(글꼴, size=9, color=GREY)
        hr = 시작행 + 3
        for i, (이름, 폭) in enumerate(열들, 1):
            h = ws.cell(hr, i, 이름)
            h.font = Font(글꼴, size=10, bold=True, color='FFFFFFFF')
            h.fill = 머리채움
            h.alignment = Alignment('center', 'center', wrap_text=True)
            h.border = 테두리
            ws.column_dimensions[get_column_letter(i)].width = 폭
        ws.freeze_panes = ws.cell(hr + 1, 2)
        return hr + 1

    # ── 1. 읽어보기 ────────────────────────────────────────
    ws = wb.active
    ws.title = '읽어보기'
    ws.sheet_view.showGridLines = False
    for col, w in zip('AB', (26, 104)):
        ws.column_dimensions[col].width = w
    ws['A1'] = f'CTK OTC LAB {연도}년 1~{월끝}월 실적분석'
    ws['A1'].font = Font(글꼴, size=17, bold=True, color=NAVY)
    ws['A2'] = '이 파일은 대시보드가 원장에서 자동으로 계산해 만든 것입니다 (클로드 작성)'
    ws['A2'].font = Font(글꼴, size=10, color=OR_, bold=True)
    안내 = [
        ('만든 날', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('만든 주체', '클로드(Claude) — CTK OTC LAB 실적보고 대시보드'),
        ('원천 자료', f'업로드한 원장 {len(원장):,}건 (당해 {연도}년) · 전년 {len(전년):,}건'),
        ('기준 기간', f'{연도}년 1월 ~ {월끝}월'),
        ('통화', '전부 USD (환산 없음)'),
        ('', ''),
        ('■ 시트 구성', ''),
        ('원장_처리본', '업로드한 원장에 계정분류·계정과목·월을 붙인 것. 아래 표들의 원천입니다'),
        ('월별 손익계산서', '손익계산서 계정과목 기준. 원장_처리본을 SUMIFS 로 집계 — 원장을 고치면 다시 계산됩니다'),
        ('월별 손익(활동기준)', '활동분류 기준으로 본 같은 손익'),
        ('당월 실적집계', f'{월끝}월을 전월·전년동월·누적과 나란히 비교'),
        ('현금흐름', '기초 현금에서 출발해 월별로 무엇 때문에 늘고 줄었는지'),
        ('미수채권', '거래처별 매출채권 잔액'),
    ] + ([
        ('○○년 원장_raw', '원장 원본 + 활동분류. ★대시보드가 읽는 시트라 지우면 안 됩니다★'),
        ('BS_IS_매핑', '영문 계정 → 한글 계정과목·BS/IS 구분. 이 시트가 있어야 대시보드가 그려집니다'),
        ('월간보고_입력', '월간실적 보고서에 손으로 넣는 값'),
        ('◆ 다시 쓰는 법', '이 파일을 대시보드 왼쪽 「⚙ 설정 · 파일 업로드」에 그대로 올리면 화면이 그려집니다. '
                           '다음 달에는 이 파일을 「전월 실적보고자료」로 올려 주세요'),
    ] if 원천시트 else []) + [
        ('', ''),
        ('■ 계산 기준', ''),
        ('영업이익', '매출총이익 − 판매비와관리비'),
        ('EBITDA', '영업이익 + 감가상각비 − 리스(RoU) 상각비'),
        ('현금흐름', '현금 계정이 든 전표에서 현금이 아닌 상대 계정의 (대변−차변)을 그 성격의 현금 증감으로 봅니다'),
        ('', ''),
        ('■ 꼭 알아두실 점', ''),
        ('활동분류', '판관비의 활동분류(대분류·활동세부)는 원장에 이미 적혀 있는 것을 그대로 씁니다. '
                     '원장에 없으면 「기타관리경비」로 잡히므로, 새 거래는 원장에서 먼저 분류해 주셔야 합니다'),
        ('피벗 없음', '이 파일에는 피벗테이블과 외부 파일 링크가 없습니다 — 「복구됨」이 뜨지 않습니다'),
        ('25년 자료', '전년 비교는 업로드한 파일의 25년 원장을 그대로 씁니다'),
    ]
    r = 4
    for 이름, 값 in 안내:
        if 이름.startswith('■'):
            ws.cell(r, 1, 이름).font = Font(글꼴, size=11, bold=True, color=NAVY)
        elif 이름:
            ws.cell(r, 1, 이름).font = Font(글꼴, size=10, bold=True)
            c = ws.cell(r, 2, 값)
            c.font = Font(글꼴, size=10)
            c.alignment = Alignment(vertical='top', wrap_text=True)
        r += 1

    # ── 2. 원장_처리본 ─────────────────────────────────────
    원 = 원장.copy()
    원['거래일'] = pd.to_datetime(원['거래일']).dt.strftime('%Y-%m-%d')
    보낼열 = ['거래일', '계정영문', '계정과목', '계정분류', '분류', '보고항목',
              '활동분류(대분류)', '활동세부', 'Name', 'Description',
              'Debit', 'Credit', '금액', '월']
    보낼열 = [c for c in 보낼열 if c in 원.columns]
    ws2 = wb.create_sheet('원장_처리본')
    ws2.sheet_view.showGridLines = False
    폭 = {'거래일': 12, '계정영문': 30, '계정과목': 16, '계정분류': 9, '분류': 10,
          '보고항목': 15, '활동분류(대분류)': 18, '활동세부': 20, 'Name': 28,
          'Description': 40, 'Debit': 14, 'Credit': 14, '금액': 14, '월': 6}
    hr2 = 표머리(ws2, '원장 처리본',
                 f'업로드한 원장에 계정분류·계정과목·월을 붙였습니다 · {len(원):,}건 · 아래 표들의 원천',
                 [(c, 폭.get(c, 14)) for c in 보낼열])
    for i, 행 in enumerate(원[보낼열].itertuples(index=False), hr2):
        for j, v in enumerate(행, 1):
            c = ws2.cell(i, j, None if (isinstance(v, float) and pd.isna(v)) else v)
            c.font = Font(글꼴, size=9)
            if 보낼열[j - 1] in ('Debit', 'Credit', '금액'):
                c.number_format = '#,##0.00'
    끝행 = hr2 + len(원) - 1
    열자리 = {c: get_column_letter(i + 1) for i, c in enumerate(보낼열)}
    범위 = lambda c: f"'원장_처리본'!${열자리[c]}${hr2}:${열자리[c]}${끝행}"

    # ── 3. 월별 손익계산서 (SUMIFS 로 다시 계산됨) ──────────
    def 손익시트(제목, 표, 기준열):
        ws3 = wb.create_sheet(제목)
        ws3.sheet_view.showGridLines = False
        열들 = [('구분', 30)] + [(f'{m}월', 14) for m in 월들] + [('누적', 15)]
        hr3 = 표머리(ws3, 제목,
                     f'{연도}년 · 단위 USD · 비용은 양수 · 원장_처리본을 SUMIFS 로 집계합니다', 열들)
        수준 = 표.attrs['수준']
        r3 = hr3
        for 이름 in 표.index:
            if 이름 in 집계제외:
                continue
            lv = 수준.get(이름, 1)
            c = ws3.cell(r3, 1, ('    ' * lv) + 이름)
            c.font = Font(글꼴, size=10, bold=(lv == 0))
            c.border = 테두리
            if lv == 0:
                c.fill = 띠채움
            계정식 = None
            if 기준열 and lv >= 1 and 이름 in set(원[기준열].dropna().astype(str)):
                계정식 = 이름
            for k, m in enumerate(월들):
                cc = ws3.cell(r3, k + 2)
                if 계정식:
                    cc.value = (f'=-SUMIFS({범위("금액")},{범위(기준열)},$A{r3}&"",'
                                f'{범위("월")},{get_column_letter(k + 2)}${hr3 - 1})')
                else:
                    cc.value = float(표.loc[이름, m])
                cc.number_format = 돈
                cc.font = Font(글꼴, size=10, bold=(lv == 0),
                               color=(RED if float(표.loc[이름, m]) < 0 else 'FF000000'))
                cc.border = 테두리
                if lv == 0:
                    cc.fill = 띠채움
            L = get_column_letter(2)
            R = get_column_letter(len(월들) + 1)
            t = ws3.cell(r3, len(월들) + 2, f'=SUM({L}{r3}:{R}{r3})')
            t.number_format = 돈
            t.font = Font(글꼴, size=10, bold=True)
            t.fill = 누적채움
            t.border = 테두리
            r3 += 1
        # 머리글 줄에 월 숫자를 숨겨 두어 SUMIFS 가 참조하게 함
        for k, m in enumerate(월들):
            h = ws3.cell(hr3 - 1, k + 2, m)
            h.font = Font(글꼴, size=10, bold=True, color='FFFFFFFF')
            h.fill = 머리채움
            h.alignment = Alignment('center', 'center')
            h.border = 테두리
        return ws3

    손익시트('월별 손익계산서', 손익표본, '계정과목')
    손익시트('월별 손익(활동기준)', 활동표본, '보고항목')

    # ── 4. 당월 실적집계 ───────────────────────────────────
    ws4 = wb.create_sheet('당월 실적집계')
    ws4.sheet_view.showGridLines = False
    이전 = 월끝 - 1 if 월끝 > 1 else None
    열들 = [('구분', 30), (f'당월({월끝}월)', 15), (f'전월({이전}월)' if 이전 else '전월', 15),
            ('전월 증감', 15), (f'전년동월({str(연도-1)[2:]}년 {월끝}월)', 17), ('전년동월 증감', 15),
            (f'당해누적(1~{월끝}월)', 17), (f'전년동기(1~{월끝}월)', 17), ('누적 증감', 15)]
    hr4 = 표머리(ws4, f'{연도}년 {월끝}월 실적집계',
                 '전월 · 전년동월 · 누적을 나란히 비교 · 단위 USD · 비용은 양수', 열들)
    수준4 = 손익표본.attrs['수준']
    전년표 = 손익표(전년, '손익') if len(전년) else None
    r4 = hr4
    for 이름 in 손익표본.index:
        if 이름 in 집계제외:
            continue
        lv = 수준4.get(이름, 1)
        당월 = float(손익표본.loc[이름, 월끝])
        전월 = float(손익표본.loc[이름, 이전]) if 이전 else 0.0
        전동 = float(전년표.loc[이름, 월끝]) if (전년표 is not None and 이름 in 전년표.index) else 0.0
        누적 = float(손익표본.loc[이름, 월들].sum())
        전누 = float(전년표.loc[이름, 월들].sum()) if (전년표 is not None and 이름 in 전년표.index) else 0.0
        값들 = [당월, 전월, 당월 - 전월, 전동, 당월 - 전동, 누적, 전누, 누적 - 전누]
        c = ws4.cell(r4, 1, ('    ' * lv) + 이름)
        c.font = Font(글꼴, size=10, bold=(lv == 0))
        c.border = 테두리
        if lv == 0:
            c.fill = 띠채움
        for j, v in enumerate(값들, 2):
            cc = ws4.cell(r4, j, v)
            cc.number_format = 돈
            cc.font = Font(글꼴, size=10, bold=(lv == 0), color=(RED if v < 0 else 'FF000000'))
            cc.border = 테두리
            if lv == 0:
                cc.fill = 띠채움
        r4 += 1

    # ── 5. 현금흐름 ────────────────────────────────────────
    if 자금 is not None:
        ws5 = wb.create_sheet('현금흐름')
        ws5.sheet_view.showGridLines = False
        열들 = [('구분', 32)] + [(f'{m}월', 14) for m in 월들] + [('누적', 15)]
        hr5 = 표머리(ws5, '현금흐름',
                     '기초 현금에서 출발해 월별로 무엇 때문에 늘고 줄었는지 · 단위 USD', 열들)
        기초 = [float(잔고.get(m - 1, 기초현금액)) if m > 1 else float(기초현금액) for m in 월들]
        기말 = [float(잔고.get(m, 0.0)) for m in 월들]
        ㅈ = lambda k: [float(자금.loc[k, m]) for m in 월들]
        줄들 = [('기초 현금', 0, 기초, 기초[0]),
                ('영업활동', 0, ㅈ('영업'), sum(ㅈ('영업'))),
                ('    매출채권 회수', 1, ㅈ('AR'), sum(ㅈ('AR'))),
                ('    기타입금 (정부지원금 환급 등)', 1, ㅈ('ETCIN'), sum(ㅈ('ETCIN'))),
                ('    매입채무 · 비용 지급', 1, ㅈ('지급계'), sum(ㅈ('지급계'))),
                ('        매입채무', 2, ㅈ('AP'), sum(ㅈ('AP'))),
                ('        급여', 2, ㅈ('PAY'), sum(ㅈ('PAY'))),
                ('        기타 비용', 2, ㅈ('OTHOP'), sum(ㅈ('OTHOP'))),
                ('투자활동', 0, ㅈ('투자'), sum(ㅈ('투자'))),
                ('    설비 투자', 1, ㅈ('CAPEX'), sum(ㅈ('CAPEX'))),
                ('재무활동', 0, ㅈ('재무'), sum(ㅈ('재무'))),
                ('    유상증자', 1, ㅈ('EQ'), sum(ㅈ('EQ'))),
                ('    차입증감', 1, ㅈ('차입증감'), sum(ㅈ('차입증감'))),
                ('    리스료 지급', 1, ㅈ('리스'), sum(ㅈ('리스'))),
                ('환율변동 효과', 1, ㅈ('FX'), sum(ㅈ('FX'))),
                ('순증감', 0, ㅈ('순증감'), sum(ㅈ('순증감'))),
                ('기말 현금', 0, 기말, 기말[-1])]
        r5 = hr5
        for 이름, lv, 값들, 총 in 줄들:
            c = ws5.cell(r5, 1, 이름)
            c.font = Font(글꼴, size=10, bold=(lv == 0))
            c.border = 테두리
            if lv == 0:
                c.fill = 띠채움
            for j, v in enumerate(값들 + [총], 2):
                cc = ws5.cell(r5, j, float(v))
                cc.number_format = 돈
                cc.font = Font(글꼴, size=10, bold=(lv == 0 or j == len(값들) + 2),
                               color=(RED if v < 0 else 'FF000000'))
                cc.border = 테두리
                cc.fill = 누적채움 if j == len(값들) + 2 else (띠채움 if lv == 0 else PatternFill())
            r5 += 1
        ws5.cell(r5 + 1, 1, '계산 방법 — 원장은 복식부기라, 현금이 든 전표에서 현금이 아닌 상대 계정의 '
                            '(대변−차변)이 그 돈의 성격별 현금 증감이 됩니다. 은행↔은행 이체는 서로 상계됩니다.'
                 ).font = Font(글꼴, size=9, color=GREY)

    # ── 6. 미수채권 ────────────────────────────────────────
    ar = 원장[원장['계정과목'].eq('매출채권') & 원장['Name'].notna()].copy()
    if len(ar):
        ws6 = wb.create_sheet('미수채권')
        ws6.sheet_view.showGridLines = False
        hr6 = 표머리(ws6, '거래처별 매출채권',
                     f'{연도}년 발생·회수 기준 · 단위 USD · ★기초잔액은 포함되어 있지 않습니다★',
                     [('거래처', 40), ('발생액', 16), ('회수액', 16), ('잔액', 16),
                      ('최근 거래일', 14), ('건수', 10)])
        g = ar.groupby('Name').agg(발생=('Debit', 'sum'), 회수=('Credit', 'sum'),
                                   최근=('거래일', 'max'), 건수=('Debit', 'size'))
        g['잔액'] = g['발생'] - g['회수']
        g = g.sort_values('잔액', ascending=False)
        r6 = hr6
        for 이름, x in g.iterrows():
            for j, v in enumerate([이름, x['발생'], x['회수'], x['잔액'],
                                   pd.Timestamp(x['최근']).strftime('%Y-%m-%d'), int(x['건수'])], 1):
                cc = ws6.cell(r6, j, v)
                cc.font = Font(글꼴, size=10,
                               color=(RED if (j == 4 and float(x['잔액']) < 0) else 'FF000000'))
                cc.border = 테두리
                if j in (2, 3, 4):
                    cc.number_format = '#,##0.00'
                if j in (5, 6):
                    cc.alignment = Alignment('center')
            r6 += 1

    # ── 7. 원본 시트 덧붙이기 (실적보고 엑셀작성 화면에서만 씁니다) ──
    #    이 시트들이 있어야 만든 파일을 다시 사이드바에 올렸을 때 대시보드가 그려집니다.
    if 원천시트:
        for 이름, (부제, d) in 원천시트.items():
            wsx = wb.create_sheet(이름)
            wsx.sheet_view.showGridLines = False
            머리 = 5 if 부제 is not None else 1
            if 부제 is not None:
                wsx.cell(1, 1, 'CTK OTC LAB').font = Font(글꼴, size=12, bold=True, color=NAVY)
                wsx.cell(2, 1, 'General Ledger List').font = Font(글꼴, size=9, color=GREY)
                wsx.cell(3, 1, 부제).font = Font(글꼴, size=9, color=GREY)
                for j, c in enumerate(d.columns, 1):
                    h = wsx.cell(머리, j, str(c))
                    h.font = Font(글꼴, size=10, bold=True, color='FFFFFFFF')
                    h.fill = 머리채움
                    h.alignment = Alignment('center', 'center')
                    wsx.column_dimensions[get_column_letter(j)].width = 16
                wsx.freeze_panes = wsx.cell(머리 + 1, 1)
            else:
                for j in range(1, min(d.shape[1], 20) + 1):
                    wsx.column_dimensions[get_column_letter(j)].width = 22
            for 행 in d.itertuples(index=False):
                wsx.append(['' if (isinstance(v, float) and pd.isna(v)) or v is pd.NaT else v
                            for v in 행])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# 실적보고 엑셀 만들기 — 원장 원본 + 전월 실적보고자료 → 당월 실적보고 엑셀
# ══════════════════════════════════════════════════════════════
# ★방식★ 전월 파일을 zip 그대로 열어 「○○년 원장_raw」 시트 하나만 갈아끼웁니다.
#        피벗·차트·메모·나머지 시트는 손도 대지 않으므로 그대로 살아 있고,
#        당월/월별 실적집계는 이미 SUMIFS·OFFSET 수식이라 엑셀이 열릴 때 스스로 계산합니다.
#        (파이썬으로 파일을 새로 쓰면 피벗과 차트가 사라집니다)
#
# 원장에서 저절로 나오는 것 : K 금액 · L 년 · M 월 · N 계정분류 · O 분류 · P 계정과목 ·
#                             U 보고금액 · V 월(숫자)   ← 전부 원본과 같은 수식을 채웁니다
# 전월에서 이어받는 것      : Q 활동분류 · R 활동세부 (사람이 손으로 나눈 값) · T 검토메모
엑셀기준일 = datetime.date(1899, 12, 30)
_행패턴 = re.compile(r'<row\b[^>]*\br="(\d+)"[^>]*>.*?</row>|<row\b[^>]*\br="(\d+)"[^>]*/>', re.S)
_칸패턴 = re.compile(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>(.*?)</c>)', re.S)
_수식패턴 = re.compile(r'<f([^>]*?)>(.*?)</f>', re.S)

원장열자리 = {'A': 'Unnamed: 0', 'B': 'Distribution account', 'C': 'Transaction date',
              'D': 'Description', 'E': 'Name', 'F': 'Transaction ID', 'G': '#',
              'H': 'Balance', 'I': 'Debit', 'J': 'Credit'}
전체열 = list('ABCDEFGHIJKLMNOPQRSTUV')
자동수식열 = ['K', 'L', 'M', 'N', 'O', 'P', 'U', 'V']
이어받기열 = {'Q': '활동분류', 'R': '활동세부',
              'S': '클로드 검토', 'T': '보고금액'}


def _xml(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def _날짜일련(v):
    """엑셀이 쓰는 날짜 일련번호로 바꿉니다 (YEAR·MONTH 수식이 돌아가려면 숫자여야 합니다)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        p = v.strip().replace('-', '/').split('/')
        if len(p) != 3:
            return None
        try:
            d = datetime.date(int(p[2]), int(p[1]), int(p[0]))
        except Exception:
            return None
    else:
        try:
            t = pd.Timestamp(v)
            if pd.isna(t):
                return None
            d = t.date()
        except Exception:
            return None
    return (d - 엑셀기준일).days


def _이어받기키(df):
    """전월 자료와 이번 원장의 같은 거래를 짝지어 주는 「지문」.

    ※ 같은 전표·같은 금액이 여러 줄인 경우가 있어(같은 날 같은 금액의 급여 두 줄 등)
      적요(Description)까지 넣고, 그래도 겹치면 그 안에서의 순번을 붙여 갈라 냅니다.
      순번을 안 붙이면 둘 중 하나의 분류가 엉뚱한 줄에 붙습니다.
    """
    k = df['Distribution account'].fillna('∅').astype(str).str.strip()
    for c in ('Transaction ID', 'Debit', 'Credit'):
        k = k + '|' + pd.to_numeric(df.get(c), errors='coerce').round(2).fillna(-1).astype(str)
    for c in ('Name', 'Description'):
        칸 = df[c] if c in df.columns else pd.Series(index=df.index, dtype=object)
        k = k + '|' + 칸.fillna('∅').astype(str).str.strip()
    k = k + '|' + df['Transaction date'].map(lambda v: str(_날짜일련(v)))
    return k + '#' + k.groupby(k).cumcount().astype(str)


# ── 새 거래를 스스로 나눠 주는 부분 ────────────────────────────
#    전월 자료에 없던 거래는 활동분류·활동세부가 비어 있어서, 그대로 두면
#    「월별 실적집계」의 SUMIFS 가 그 줄을 못 세고 이번 달 칸이 0 으로 남습니다.
#    그래서 지난달까지 사람이 나눠 놓은 것을 배워서 새 줄에 붙이고,
#    확실하지 않은 것은 T열 「클로드 검토」에 왜 그렇게 붙였는지 적어 둡니다.
_숫자토막 = re.compile(r'\d+')
_공백토막 = re.compile(r'\s+')


def _글자정리(v):
    """적요·거래처를 견주기 좋게 다듬습니다 (대소문자·여러 칸 공백·전표번호 무시)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ''
    return _공백토막.sub(' ', _숫자토막.sub('#', str(v).lower().strip()))


def _글자칸(df, 이름):
    """한 열을 「빈칸이면 '' 인 글자 목록」으로 바꿉니다.

    ※ .astype(str) 만 쓰면 빈칸이 float nan 으로 남아 「nan」이라는 계정이 생깁니다.
    """
    if 이름 not in df.columns:
        return [''] * len(df)
    return ['' if (v is None or (isinstance(v, float) and pd.isna(v))
                   or v is pd.NA) else str(v).strip() for v in df[이름]]


def _열찾기(df, 이름):
    for c in df.columns:
        if str(c).strip() == 이름:
            return c
    for c in df.columns:
        if str(c).startswith(이름):
            return c
    return None


def _분류학습(전):
    """전월 자료에서 「이 계정·이 적요는 어떤 활동으로 나눴는지」를 배웁니다."""
    적요, 거래처, 계정, 본줄 = {}, {}, {}, collections.Counter()
    q열, r열 = _열찾기(전, '활동분류'), _열찾기(전, '활동세부')
    if q열 is None or r열 is None:
        return 적요, 거래처, 계정, 본줄
    계정칸 = _글자칸(전, 'Distribution account')
    적요칸 = [_글자정리(v) for v in _글자칸(전, 'Description')]
    이름칸 = [_글자정리(v) for v in _글자칸(전, 'Name')]
    for a, dn, nn, q, r in zip(계정칸, 적요칸, 이름칸, 전[q열], 전[r열]):
        if not a:
            continue
        본줄[a] += 1
        r = '' if (r is None or (isinstance(r, float) and pd.isna(r))) else str(r).strip()
        if not r:
            continue
        짝 = ('' if (q is None or (isinstance(q, float) and pd.isna(q)))
              else str(q).strip(), r)
        계정.setdefault(a, collections.Counter())[짝] += 1
        if dn:
            적요.setdefault((a, dn), collections.Counter())[짝] += 1
        if nn:
            거래처.setdefault((a, nn), collections.Counter())[짝] += 1
    return 적요, 거래처, 계정, 본줄


def _자동분류(새, 이어받음, 배움, 유효세부, 매핑분류):
    """이어받지 못한 줄의 Q·R 을 채우고, 그 근거를 T 에 적습니다.

    유효세부 : 「월별 실적집계」가 실제로 세는 활동세부 이름 모음
    매핑분류 : BS_IS_매핑의 계정 → 'BS' / 'IS'
    돌려주는 값 : 근거별 건수 dict
    """
    적요, 거래처, 계정, 본줄 = 배움
    계정칸 = _글자칸(새, 'Distribution account')
    적요칸 = [_글자정리(v) for v in _글자칸(새, 'Description')]
    이름칸 = [_글자정리(v) for v in _글자칸(새, 'Name')]
    날짜칸 = 새['Transaction date'].tolist()

    센것 = collections.Counter()
    메모 = 이어받음.setdefault('T', [None] * len(새))
    Q칸 = 이어받음.setdefault('Q', [None] * len(새))
    R칸 = 이어받음.setdefault('R', [None] * len(새))
    for i in range(len(새)):
        a = 계정칸[i]
        if not a:
            continue
        if _날짜일련(날짜칸[i]) is None:      # 이월잔액 · 소계 줄은 건드리지 않습니다
            continue
        아는계정 = (not 매핑분류) or (a in 매핑분류)
        if not 아는계정 and not 메모[i]:
            메모[i] = f'BS_IS_매핑에 없는 계정 「{a}」 — 계정과목·계정분류가 빈칸이 됩니다'
            센것['매핑밖'] += 1
        if R칸[i]:
            센것['이어받음'] += 1
            continue

        짝, 근거, 급함 = None, '', False
        if (a, 적요칸[i]) in 적요:
            짝, 근거 = 적요[(a, 적요칸[i])].most_common(1)[0][0], '적요일치'
        elif (a, 이름칸[i]) in 거래처:
            짝, 근거 = 거래처[(a, 이름칸[i])].most_common(1)[0][0], '거래처일치'
        elif a in 계정 and len(계정[a]) == 1:
            짝, 근거 = next(iter(계정[a])), '계정과목대표값'
        elif a in 계정:
            짝, 근거, 급함 = 계정[a].most_common(1)[0][0], '계정과목최빈값', True
        elif a in 본줄:
            센것['원래빈칸'] += 1        # 이 계정은 원래 활동분류를 안 붙이는 계정입니다
            continue
        elif 매핑분류.get(a) == 'BS':
            센것['신규BS계정'] += 1      # 재무상태표 계정이라 활동분류가 필요 없습니다
            continue
        else:
            센것['신규계정'] += 1
            if not 메모[i]:
                메모[i] = f'신규 계정 「{a}」 — 활동분류·활동세부를 새로 정해 주셔야 합니다'
            continue

        # 같은 계정인데 지난달까지 두 갈래 이상으로 나뉜 적이 있으면 더 세게 표시합니다
        if len(계정.get(a, ())) > 1:
            급함 = True
        Q칸[i] = 짝[0] or None
        R칸[i] = 짝[1]
        센것[근거] += 1

        말 = f'자동분류-{근거}({"확인필요" if 급함 else "검토요망"})'
        if 유효세부 and 짝[1] not in 유효세부:
            말 += ' ※월별 실적집계에 없는 활동세부'
            센것['집계밖'] += 1
            급함 = True
        센것['확인필요' if 급함 else '검토요망'] += 1
        if not 메모[i]:
            메모[i] = 말
    return dict(센것)


def _풀기(s):
    return (s.replace('&lt;', '<').replace('&gt;', '>')
            .replace('&quot;', '"').replace('&apos;', "'").replace('&amp;', '&'))


def _공유글자(z):
    try:
        sx = z.read('xl/sharedStrings.xml').decode('utf-8')
    except KeyError:
        return []
    return [_풀기(re.sub(r'<[^>]+>', '', m.group(1)))
            for m in re.finditer(r'<si>(.*?)</si>', sx, re.S)]


def _시트글자(z, 파일, 열, 공유):
    """시트 XML 에서 한 열의 글자값만 빠르게 뽑습니다 (openpyxl 은 이 파일에 너무 느립니다)."""
    try:
        x = z.read(파일).decode('utf-8')
    except KeyError:
        return []
    if '<sheetData>' not in x:
        return []
    본문 = x.split('<sheetData>', 1)[1].rsplit('</sheetData>', 1)[0]
    값 = []
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>(.*?)</c>)', 본문, re.S):
        if cm.group(1) != 열:
            continue
        속성, 몸 = cm.group(2), cm.group(3) or ''
        if '<f' in 몸:
            continue
        vm = re.search(r'<v>(.*?)</v>', 몸, re.S)
        if vm:
            v = vm.group(1)
            값.append(공유[int(v)] if ' t="s"' in 속성 and v.isdigit()
                      and int(v) < len(공유) else _풀기(v))
        elif '<is>' in 몸:
            값.append(_풀기(re.sub(r'<[^>]+>', '', 몸)))
    return [str(v).strip() for v in 값 if str(v).strip()]


def _시트칸읽기(z, 파일, 열들, 공유):
    """시트 XML 에서 원하는 열들의 값을 {행번호: {열: 값}} 으로 읽습니다."""
    try:
        x = z.read(파일).decode('utf-8')
    except KeyError:
        return {}
    if '<sheetData>' not in x:
        return {}
    본문 = x.split('<sheetData>', 1)[1].rsplit('</sheetData>', 1)[0]
    표 = {}
    for rm in re.finditer(r'<row[^>]*\br="(\d+)"[^>]*>(.*?)</row>', 본문, re.S):
        행 = int(rm.group(1))
        칸 = {}
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>(.*?)</c>)',
                              rm.group(2), re.S):
            열, 속성, 몸 = cm.group(1), cm.group(2), cm.group(3) or ''
            if 열 not in 열들:
                continue
            vm = re.search(r'<v>(.*?)</v>', 몸, re.S)
            if vm:
                v = vm.group(1)
                칸[열] = (공유[int(v)] if ' t="s"' in 속성 and v.isdigit()
                          and int(v) < len(공유) else _풀기(v))
            elif '<is>' in 몸:
                칸[열] = _풀기(re.sub(r'<[^>]+>', '', 몸))
        if 칸:
            표[행] = 칸
    return 표


def _매핑표읽기(z, 파일, 공유):
    """BS_IS_매핑 시트에서 「계정 이름 → (BS/IS, 한글 계정과목)」을 읽습니다."""
    지도 = {}
    for _행, 칸 in _시트칸읽기(z, 파일, ('A', 'B', 'C'), 공유).items():
        이름 = str(칸.get('A', '')).strip()
        if 이름:
            지도[이름] = (str(칸.get('B', '')).strip().upper(),
                          str(칸.get('C', '')).strip())
    return 지도


def _시트본뜨기(원본xml):
    """원본 원장 시트에서 1~5행 · 칸 서식번호 · 수식 본을 뽑아냅니다."""
    앞, 나머지 = 원본xml.split('<sheetData>', 1)
    본문, 뒤 = 나머지.rsplit('</sheetData>', 1)
    머리행, 그룹서식, 자료서식, 수식본 = {}, {}, {}, {}
    for m in _행패턴.finditer(본문):
        r = int(m.group(1) or m.group(2))
        조각 = m.group(0)
        if r <= 5:
            머리행[r] = 조각
            continue
        칸 = {}
        for cm in _칸패턴.finditer(조각):
            col, attrs, body = cm.group(1), cm.group(2), cm.group(3) or ''
            if col not in 전체열:
                continue
            s = re.search(r's="(\d+)"', attrs)
            칸[col] = (s.group(1) if s else None, body)
            fm = _수식패턴.search(body)
            if fm and col not in 수식본 and fm.group(2).strip():
                수식본[col] = (r, fm.group(2))
        표 = 그룹서식 if 칸.get('B', (None, ''))[1] in ('', None) else 자료서식
        if not 표:
            for col, (s, _b) in 칸.items():
                if s:
                    표[col] = s
        if 그룹서식 and 자료서식 and len(수식본) >= 10:
            break
    return 앞, 뒤, 머리행, 그룹서식, 자료서식, 수식본


def _수식옮기기(본, 원행, 새행):
    """C6·J6 처럼 본이 된 줄 번호를 새 줄 번호로 바꿉니다."""
    return re.sub(r'(\$?[A-Z]{1,3}\$?)' + str(원행) + r'(?![0-9])',
                  lambda m: m.group(1) + str(새행), 본)


# 「26년 원장_raw」 K~V 칸의 본디 수식 (6번째 줄 기준).
#  ※ 올려 주신 전월 파일에서 어느 칸의 수식이 지워져 있으면 여기 것으로 되살립니다.
#    (한 번 값만 남은 파일을 다시 넣으면 그 칸이 통째로 비어 버리던 것을 막습니다)
_되살린열 = []          # 이번에 되살린 칸 (화면에 알려 주려고 남깁니다)
기본수식본 = {
    'K': (6, 'IF(C6="","",J6-I6)'),
    'L': (6, 'IF(C6="","",YEAR(C6)&amp;"년")'),
    'M': (6, 'IF(C6="","",MONTH(C6)&amp;"월")'),
    'N': (6, 'IFERROR(INDEX(BS_IS_매핑!$B:$B,MATCH(B6,BS_IS_매핑!$A:$A,0)),"")'),
    'O': (6, 'IF(N6&lt;&gt;"IS","",IF(OR(P6="제품매출",P6="기타매출"),"매출",'
             'IF(P6="제품매출원가","매출원가",IF(OR(P6="이자비용",P6="잡손실",P6="잡이익",'
             'P6="외환차익",P6="외환차손",P6="외화환산이익",P6="외화환산손실",'
             'P6="유형자산처분손실"),"영업외손익","판관비"))))'),
    'P': (6, 'IFERROR(INDEX(BS_IS_매핑!$C:$C,MATCH(B6,BS_IS_매핑!$A:$A,0)),"")'),
    'Q': (6, 'IF(P6="제품매출","친환경 바이오 소재",'
             'IF(P6="기타매출","기술 서비스 및 자문",""))'),
    'R': (6, 'IF(P6="제품매출","친환경 빨대/원자재 및 부자재 공급",'
             'IF(P6="기타매출","R&amp;D용역",""))'),
    'U': (8, 'IF($C8="","",IF($O8="매출",$J8-$I8,$I8-$J8))'),
    'V': (8, 'IF($C8="","",MONTH($C8))'),
}


def _수식값미리계산(새, 이어받음, 매핑표, 날짜):
    """K~V 수식이 낼 값을 파이썬으로 미리 구해 칸에 같이 적어 둡니다.

    ※ 엑셀은 수식만 있고 계산값이 없으면, 다시 계산하기 전까지 그 칸을 빈칸으로 보여 줍니다.
      원장이 1만 줄이라 계산이 오래 걸려서, 내려받아 열면 K~T 가 한동안 비어 보였습니다.
      그래서 값을 미리 넣어 둡니다 (수식은 그대로라 나중에 다시 계산해도 같은 값입니다).
    """
    영업외 = {'이자비용', '잡손실', '잡이익', '외환차익', '외환차손',
              '외화환산이익', '외화환산손실', '유형자산처분손실'}
    계정 = _글자칸(새, 'Distribution account')
    차변 = pd.to_numeric(새.get('Debit'), errors='coerce').fillna(0).tolist()
    대변 = pd.to_numeric(새.get('Credit'), errors='coerce').fillna(0).tolist()
    값 = {c: [] for c in 자동수식열}
    for i in range(len(새)):
        d = 날짜[i]
        비었다 = d is None
        날 = None if 비었다 else 엑셀기준일 + datetime.timedelta(days=d)
        b, k = 계정[i], 계정[i].lower()
        bs, kor = 매핑표.get(b, ('', ''))
        분류 = ''
        if bs == 'IS':
            분류 = ('매출' if kor in ('제품매출', '기타매출')
                    else '매출원가' if kor == '제품매출원가'
                    else '영업외손익' if kor in 영업외 else '판관비')
        값['K'].append('' if 비었다 else 대변[i] - 차변[i])
        값['L'].append('' if 비었다 else f'{날.year}년')
        값['M'].append('' if 비었다 else f'{날.month}월')
        값['N'].append(bs if b else '')
        값['O'].append(분류)
        값['P'].append(kor if b else '')
        값['U'].append('' if 비었다 else
                       (대변[i] - 차변[i] if 분류 == '매출' else 차변[i] - 대변[i]))
        값['V'].append('' if 비었다 else 날.month)
    # Q·R 은 매출 계정이면 수식이 값을 만들어 냅니다 — 그 값도 같이 적어 둡니다
    Q, R = [], []
    for i in range(len(새)):
        kor = 매핑표.get(계정[i], ('', ''))[1]
        Q.append('친환경 바이오 소재' if kor == '제품매출'
                 else '기술 서비스 및 자문' if kor == '기타매출' else '')
        R.append('친환경 빨대/원자재 및 부자재 공급' if kor == '제품매출'
                 else 'R&D용역' if kor == '기타매출' else '')
    return 값, Q, R


def _값칸(열, r, sa, 수식, v):
    """수식과 그 계산값을 함께 적습니다 (글자 결과는 t="str")."""
    if isinstance(v, str):
        return (f'<c r="{열}{r}"{sa} t="str">{수식}<v>{_xml(v)}</v></c>'
                if v else f'<c r="{열}{r}"{sa} t="str">{수식}<v></v></c>')
    if isinstance(v, int) and not isinstance(v, bool):
        return f'<c r="{열}{r}"{sa}>{수식}<v>{v}</v></c>'
    return f'<c r="{열}{r}"{sa}>{수식}<v>{float(v)!r}</v></c>'


def _원장시트XML(원본xml, 새, 이어받음, 기간글, 매핑표=None):
    앞, 뒤, 머리행, 그룹서식, 자료서식, 수식본 = _시트본뜨기(원본xml)
    # 전월 파일에서 수식이 지워진 칸은 본디 수식으로 되살립니다
    되살림 = [c for c in 기본수식본 if c not in 수식본]
    수식본 = {**{c: 기본수식본[c] for c in 되살림}, **수식본}
    _되살린열.clear()
    _되살린열.extend(sorted(되살림))

    앞부분 = []
    for r in sorted(머리행):
        조각 = 머리행[r]
        if r == 3 and 기간글:                     # 맨 위 기간 표시만 새로 씁니다
            조각 = re.sub(
                r'<c r="A3"([^>]*?)(?:/>|>.*?</c>)',
                lambda m: '<c r="A3"' + re.sub(r'\s*t="\w+"', '', m.group(1))
                + f' t="inlineStr"><is><t>{_xml(기간글)}</t></is></c>',
                조각, count=1, flags=re.S)
        앞부분.append(조각)

    시작행, 끝행 = 6, 6 + len(새) - 1
    si = {c: i for i, c in enumerate(자동수식열)}
    칸값 = {col: (새[이름].tolist() if 이름 in 새.columns else [None] * len(새))
            for col, 이름 in 원장열자리.items()}
    날짜 = [_날짜일련(v) for v in 칸값['C']]
    미리값, 미리Q, 미리R = ({}, None, None)
    if 매핑표:
        미리값, 미리Q, 미리R = _수식값미리계산(새, 이어받음, 매핑표, 날짜)

    줄들, 첫줄 = [], {c: True for c in 자동수식열}
    for i in range(len(새)):
        r = 시작행 + i
        b = 칸값['B'][i]
        그룹행 = (b is None or (isinstance(b, float) and pd.isna(b))
                  or str(b).strip() in ('', 'nan'))
        서식 = 그룹서식 if 그룹행 else 자료서식
        칸 = []
        for col in 전체열:
            s = 서식.get(col) or 자료서식.get(col) or 그룹서식.get(col)
            sa = f' s="{s}"' if s else ''
            if col in 원장열자리:
                v = 칸값[col][i]
                if col == 'C':
                    d = 날짜[i]
                    칸.append(f'<c r="{col}{r}"{sa}><v>{d}</v></c>' if d is not None
                              else f'<c r="{col}{r}"{sa}/>')
                elif v is None or (isinstance(v, float) and pd.isna(v)):
                    칸.append(f'<c r="{col}{r}"{sa}/>')
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    칸.append(f'<c r="{col}{r}"{sa}><v>{v!r}</v></c>')
                else:
                    칸.append(f'<c r="{col}{r}"{sa} t="inlineStr">'
                              f'<is><t xml:space="preserve">{_xml(v)}</t></is></c>')
            elif col in 자동수식열 and col in 수식본:
                원행, 본 = 수식본[col]
                if 첫줄[col]:
                    첫줄[col] = False
                    식 = (f'<f t="shared" ref="{col}{시작행}:{col}{끝행}" '
                          f'si="{si[col]}">{_수식옮기기(본, 원행, r)}</f>')
                else:
                    식 = f'<f t="shared" si="{si[col]}"/>'
                if col in 미리값:
                    칸.append(_값칸(col, r, sa, 식, 미리값[col][i]))
                else:
                    칸.append(f'<c r="{col}{r}"{sa}>{식}</c>')
            elif col in 이어받기열:
                v = 이어받음.get(col, [None] * len(새))[i]
                if v:
                    칸.append(f'<c r="{col}{r}"{sa} t="inlineStr">'
                              f'<is><t xml:space="preserve">{_xml(v)}</t></is></c>')
                elif col in 수식본:
                    원행, 본 = 수식본[col]
                    식 = f'<f>{_수식옮기기(본, 원행, r)}</f>'
                    미리 = ({'Q': 미리Q, 'R': 미리R}.get(col) or [None] * len(새))[i]
                    칸.append(_값칸(col, r, sa, 식, 미리)
                              if 미리 is not None
                              else f'<c r="{col}{r}"{sa}>{식}</c>')
                else:
                    칸.append(f'<c r="{col}{r}"{sa}/>')
            else:
                칸.append(f'<c r="{col}{r}"{sa}/>')
        줄들.append(f'<row r="{r}" spans="1:22" ht="15" customHeight="1">'
                    + ''.join(칸) + '</row>')

    앞 = re.sub(r'<dimension ref="[^"]*"/>', f'<dimension ref="A1:V{끝행}"/>', 앞, count=1)
    뒤 = re.sub(r'(<autoFilter ref="A5:)[A-Z]+\d+(")', rf'\g<1>T{끝행}\g<2>', 뒤, count=1)
    return 앞 + '<sheetData>' + ''.join(앞부분) + ''.join(줄들) + '</sheetData>' + 뒤


def _시트지도(z):
    """시트 이름 → XML 파일 경로 (엑셀 안에서의 순서 그대로)."""
    wb = z.read('xl/workbook.xml').decode('utf-8')
    rel = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid = {m.group(1): m.group(2) for m in
           re.finditer(r'<Relationship Id="([^"]+)"[^>]*Target="([^"]+)"', rel)}
    시트 = []
    for m in re.finditer(r'<sheet\b([^>]*?)/>', wb):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        시트.append((_풀기(str(a.get('name'))), 'xl/' + rid.get(a.get('r:id'), '').lstrip('/')))
    return 시트


def _전월원장시트(z):
    """전월 파일에서 「○○년 원장_raw」 시트의 이름과 XML 파일 경로를 찾습니다."""
    시트 = _시트지도(z)
    원장 = sorted((n, f) for n, f in 시트 if '원장_raw' in str(n))
    if not 원장:
        raise ValueError('전월 실적보고자료에서 「○○년 원장_raw」 시트를 찾지 못했습니다. '
                         '지난달에 쓰시던 실적보고 엑셀이 맞는지 확인해 주세요.')
    return 원장[-1], [n for n, _f in 시트]


def _관계표(z, 파트):
    """파트 하나가 딸린 파트들을 무엇무엇 쓰고 있는지 → (관계파일 이름, {rId: 파트경로})."""
    관계파일 = posixpath.join(posixpath.dirname(파트), '_rels',
                              posixpath.basename(파트) + '.rels')
    try:
        글 = z.read(관계파일).decode('utf-8')
    except KeyError:
        return 관계파일, {}
    나온것 = {}
    for m in re.finditer(r'<Relationship\b[^>]*?/>', 글):
        조각 = m.group(0)
        if 'TargetMode="External"' in 조각:
            continue
        아이디 = re.search(r'Id="([^"]+)"', 조각)
        대상 = re.search(r'Target="([^"]+)"', 조각)
        if not (아이디 and 대상):
            continue
        나온것[아이디.group(1)] = posixpath.normpath(
            posixpath.join(posixpath.dirname(파트), 대상.group(1).lstrip('/')))
    return 관계파일, 나온것


def _딸린파트(z, 시작):
    """그 파트와, 그 파트만 타고 내려가 만나는 모든 파트 (피벗 → 캐시 → 기록, 그림 → 그림파일)."""
    본것, 남은것 = set(), [시작]
    while 남은것:
        p = 남은것.pop()
        if p in 본것:
            continue
        본것.add(p)
        남은것.extend(_관계표(z, p)[1].values())
    return 본것


def _시트빼기(z, 뺄이름):
    """실적보고 엑셀에서 시트 하나를 통째로 들어냅니다 → (지울 파트 set, {파트: 새 바이트}).

    ※ 「당월 실적집계」처럼 이제 쓰지 않는 시트를 내려받는 파일에서 빼기 위한 것입니다.
      그 시트에만 딸린 피벗표·피벗캐시·그림·메모까지 같이 빼되, 다른 시트도 같이 쓰는
      파트는 그대로 둡니다. 워크북 목차·관계·파일목록·문서정보도 함께 맞춰 고칩니다.
      시트가 없으면 (set(), {}) 를 돌려주므로 그냥 지나갑니다.
    """
    wb = z.read('xl/workbook.xml').decode('utf-8')
    wbrel = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    묶음 = {m.group(1): 'xl/' + m.group(2).lstrip('/') for m in
            re.finditer(r'<Relationship Id="([^"]+)"[^>]*Target="([^"]+)"', wbrel)}

    시트조각, 자리, 뺄rId, 뺄파트 = None, -1, None, None
    for i, m in enumerate(re.finditer(r'<sheet\b[^>]*?/>', wb)):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(0)))
        if _풀기(str(a.get('name'))) == 뺄이름:
            시트조각, 자리, 뺄rId = m.group(0), i, a.get('r:id')
            뺄파트 = posixpath.normpath(묶음.get(뺄rId, ''))
            break
    if 시트조각 is None or not 뺄파트:
        return set(), {}

    모든파트 = set(z.namelist())
    후보 = _딸린파트(z, 뺄파트) & 모든파트

    # 살아남는 시트가 같이 쓰는 파트는 되살립니다 (워크북 자신은 아래에서 따로 고칩니다)
    바뀜 = True
    while 바뀜:
        바뀜 = False
        for 파트 in sorted(모든파트 - 후보):
            if 파트 == 'xl/workbook.xml' or 파트.endswith('.rels'):
                continue
            for t in _관계표(z, 파트)[1].values():
                if t in 후보:
                    후보 -= _딸린파트(z, t)
                    바뀜 = True
                    break
            if 바뀜:
                break

    지울파트 = set()
    for 파트 in 후보:
        지울파트.add(파트)
        관계파일 = _관계표(z, 파트)[0]
        if 관계파일 in 모든파트:
            지울파트.add(관계파일)

    # ── 워크북 목차 고치기 ────────────────────────────────
    새wb = wb.replace(시트조각, '', 1)

    뺄rId들 = {뺄rId}
    for m in re.finditer(r'<pivotCache\b[^>]*?/>', 새wb):
        아이디 = re.search(r'r:id="([^"]+)"', m.group(0))
        if 아이디 and posixpath.normpath(묶음.get(아이디.group(1), '')) in 지울파트:
            새wb = 새wb.replace(m.group(0), '', 1)
            뺄rId들.add(아이디.group(1))
    새wb = re.sub(r'<pivotCaches>\s*</pivotCaches>', '', 새wb)

    # 시트 번호로 매인 이름(정의된 이름)은 앞시트가 빠진 만큼 번호를 당겨 줍니다
    def _이름고침(m):
        조각 = m.group(0)
        번호 = int(re.search(r'localSheetId="(\d+)"', 조각).group(1))
        if 번호 == 자리:
            return ''
        return 조각.replace(f'localSheetId="{번호}"', f'localSheetId="{번호 - 1}"') \
            if 번호 > 자리 else 조각
    새wb = re.sub(r'<definedName\b[^>]*localSheetId="\d+"[^>]*>.*?</definedName>',
                  _이름고침, 새wb, flags=re.S)
    새wb = re.sub(r'<definedNames>\s*</definedNames>', '', 새wb)

    def _보기고침(m):
        조각 = m.group(0)
        for 이름 in ('activeTab', 'firstSheet'):
            n = re.search(rf'{이름}="(\d+)"', 조각)
            if n:
                v = max(int(n.group(1)) - (1 if int(n.group(1)) > 자리 else 0), 0)
                조각 = 조각.replace(n.group(0), f'{이름}="{v}"')
        return 조각
    새wb = re.sub(r'<workbookView\b[^>]*?/>', _보기고침, 새wb, count=1)

    새rel = wbrel
    for rid in 뺄rId들:
        새rel = re.sub(rf'<Relationship[^>]*Id="{re.escape(rid)}"[^>]*/>', '', 새rel, count=1)

    고칠것 = {'xl/workbook.xml': 새wb.encode('utf-8'),
              'xl/_rels/workbook.xml.rels': 새rel.encode('utf-8')}

    # ── 파일목록([Content_Types]) 에서 빠진 파트를 지웁니다 ──
    ct = z.read('[Content_Types].xml').decode('utf-8')
    for 파트 in 지울파트:
        ct = re.sub(rf'<Override PartName="/{re.escape(파트)}"[^>]*/>', '', ct)
    고칠것['[Content_Types].xml'] = ct.encode('utf-8')

    # ── 문서정보(시트 이름 목록)도 한 칸 줄입니다 ─────────
    try:
        app = z.read('docProps/app.xml').decode('utf-8')
        표기 = 시트조각 and re.search(r'name="([^"]*)"', 시트조각).group(1)
        낱말 = f'<vt:lpstr>{표기}</vt:lpstr>'
        if 낱말 in app:
            app = app.replace(낱말, '', 1)
            app = re.sub(r'(<TitlesOfParts><vt:vector size=")(\d+)(")',
                         lambda m: m.group(1) + str(max(int(m.group(2)) - 1, 0)) + m.group(3),
                         app, count=1)
            app = re.sub(r'(<vt:lpstr>워크시트</vt:lpstr></vt:variant><vt:variant><vt:i4>)(\d+)(</vt:i4>)',
                         lambda m: m.group(1) + str(max(int(m.group(2)) - 1, 0)) + m.group(3),
                         app, count=1)
            고칠것['docProps/app.xml'] = app.encode('utf-8')
    except KeyError:
        pass

    return 지울파트, 고칠것


def _칸값바꾸기(시트xml, 칸, 값):
    """시트 XML 에서 칸 하나의 값만 갈아끼웁니다 (서식 s= 는 그대로 둡니다).

    ※ 「당월 실적집계」의 H2 가 몇 월 보고서인지 정하는 칸입니다.
      이 칸이 지난달 그대로면 표 전체가 지난달 숫자로 남습니다.
    """
    행 = re.search(r'\d+', 칸).group(0)
    새칸 = None

    def _고치기(m):
        속성 = re.sub(r'\s*t="\w+"', '', m.group(1))
        return f'<c r="{칸}"{속성}><v>{값}</v></c>'

    시트xml, 바뀜 = re.subn(rf'<c r="{칸}"([^>]*?)(?:/>|>.*?</c>)', _고치기,
                            시트xml, count=1, flags=re.S)
    if 바뀜:
        return 시트xml, True

    # 그 칸이 아예 없으면 해당 행의 「열 순서에 맞는 자리」에 새로 끼워 넣습니다
    열 = re.match(r'[A-Z]+', 칸).group(0)
    새칸 = f'<c r="{칸}"><v>{값}</v></c>'

    def _열번호(c):
        n = 0
        for ch in c:
            n = n * 26 + (ord(ch) - 64)
        return n

    행틀 = re.search(rf'<row[^>]*\br="{행}"[^>]*>(.*?)</row>', 시트xml, re.S)
    if not 행틀:
        return 시트xml, False
    속 = 행틀.group(1)
    자리 = len(속)
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"', 속):
        if _열번호(cm.group(1)) > _열번호(열):
            자리 = cm.start()
            break
    새속 = 속[:자리] + 새칸 + 속[자리:]
    return (시트xml[:행틀.start(1)] + 새속 + 시트xml[행틀.end(1):]), True


# ══════════════════════════════════════════════════════════════
#  「26년 자금」 시트 — 왼쪽은 원장에서 뽑은 실제 현금흐름, 오른쪽은 남은 달의 예상 출금
#  ※ 보고 월이 바뀌면 왼쪽은 그달까지 다시 쓰고, 오른쪽은 그다음 달부터만 남깁니다.
# ══════════════════════════════════════════════════════════════
자금항목이름 = {'인건비성 항목': '인건비성 항목', '커미션 비용': '커미션 비용',
                '외부용역': '외부용역', '건물': '건물관리비', '차량': '차량유지비',
                'ICT': 'ICT', '출장비': '출장비', '접대비': '접대비',
                '소모성경비': '소모성경비', '기타관리경비': '기타관리경비',
                '기타': '기타관리경비', '샘플 및 금형관련비': '샘플 및 금형관련비',
                '금형관련비': '샘플 및 금형관련비', '운반 및 창고료': '운반 및 창고료',
                '정부지원': '정부지원'}
자금분류표 = {'AR': {'매출채권', '미수금', '선수금', '기타채권(Chargeback)'},
              'AP': {'매입채무', '미지급금', '미지급비용'},
              'PAY': {'예수금', '예수금(보험료)', '예수금(401k)', '미지급급여',
                      '급여', '복리후생비', '퇴직급여'},
              'EQ': {'보통주자본금', '주식발행초과금', '자본잉여금', '기타자본'},
              'DEBT': {'단기차입금', '장기차입금', '전환사채', '단기대여금', '장기대여금'},
              'LEASE': {'리스부채(유동)', '리스부채(비유동)'},
              'CAPEX': {'기계장치', '건설중인자산', '시설장치', '비품', '공구와기구',
                        '차량운반구', '소프트웨어', '기타의무형자산', '장기선급금',
                        '무형자산', '특허권', '사용권자산', '사용권자산(비유동)'},
              'FX': {'외환차손', '외화환산이익', '외환차익', '외화환산손실',
                     '외환차손익', '외화환산손익'}}
자금구분 = {'AR': '영업활동', 'ETCIN': '영업활동', 'AP': '영업활동', 'PAY': '영업활동',
            'OTHOP': '영업활동', 'CAPEX': '투자활동', 'EQ': '재무활동', 'DEBT': '재무활동',
            'LEASE': '재무활동', 'FX': '환율변동 효과'}
자금이름 = {'AR': '매출채권 회수', 'ETCIN': '기타입금', 'AP': '매입채무 지급',
            'PAY': '급여', 'OTHOP': '기타 비용',
            'CAPEX': '설비 투자', 'EQ': '유상증자', 'DEBT': '차입증감',
            'LEASE': '리스료 지급', 'FX': '환율변동 효과'}
자금표시 = 252          # 「이 파일에는 자금 시트 서식이 이미 들어 있다」는 표시로 쓰는 번호


def _열이름(n):
    이름 = ''
    while n:
        n, 나머지 = divmod(n - 1, 26)
        이름 = chr(65 + 나머지) + 이름
    return 이름


def _열번호(이름):
    n = 0
    for 글 in str(이름):
        n = n * 26 + (ord(글) - 64)
    return n


def _자금칸그대로(z, 파일):
    """옛 자금 시트의 칸을 XML 그대로 {(행, 열번호): 조각} 으로 읽습니다.

    ※ 담당자가 손으로 넣은 3번 표(예상 입금)를 한 글자도 건드리지 않고
      그대로 옮기기 위한 것입니다.
    """
    try:
        x = z.read(파일).decode('utf-8')
    except (KeyError, AttributeError):
        return {}
    if '<sheetData>' not in x:
        return {}
    본문 = x.split('<sheetData>', 1)[1].rsplit('</sheetData>', 1)[0]
    표 = {}
    for m in re.finditer(r'<c r="([A-Z]+)(\d+)"[^>]*?(?:/>|>.*?</c>)', 본문, re.S):
        표[(int(m.group(2)), _열번호(m.group(1)))] = m.group(0)
    return 표


def _자금코드(계정과목):
    for k, s in 자금분류표.items():
        if 계정과목 in s:
            return k
    return 'OTHOP'


def _자금서식(스타일):
    """자금 시트에 필요한 서식을 styles.xml 뒤에 덧붙입니다.

    ※ 반드시 목록의 「끝」에 붙여야 합니다. 앞에 끼워 넣으면 기존 번호가 전부
      한 칸씩 밀려 파일 전체 서식이 망가집니다.
    ※ 이미 붙어 있으면(자금표시 번호를 쓰는 xf 가 있으면) 그 번호를 그대로 씁니다.
    """
    xf목록 = re.search(r'<cellXfs count="\d+">(.*?)</cellXfs>', 스타일, re.S)
    if xf목록:
        칸들 = re.findall(r'<xf\b[^>]*?(?:/>|>.*?</xf>)', xf목록.group(1), re.S)
        for i, x in enumerate(칸들):
            if f'numFmtId="{자금표시}"' in x:
                return 스타일, [i + k for k in range(17)], False

    글꼴수 = int(re.search(r'<fonts count="(\d+)"', 스타일).group(1))
    테두리수 = int(re.search(r'<borders count="(\d+)"', 스타일).group(1))
    xf수 = int(re.search(r'<cellXfs count="(\d+)"', 스타일).group(1))
    쓰인번호 = {int(v) for v in re.findall(r'<numFmt numFmtId="(\d+)"', 스타일)}
    F날짜 = next(n for n in range(자금표시 + 1, 400) if n not in 쓰인번호)
    F숫자 = next(n for n in range(F날짜 + 1, 400) if n not in 쓰인번호)

    새서식 = (f'<numFmt numFmtId="{F날짜}" formatCode="yyyy\\-mm\\-dd"/>'
              f'<numFmt numFmtId="{F숫자}" formatCode="#,##0.00;[Red]\\({{#,##0.00}}\\)"/>'
              f'<numFmt numFmtId="{자금표시}" formatCode="General"/>')
    if '<numFmts' in 스타일:
        서식수 = int(re.search(r'<numFmts count="(\d+)"', 스타일).group(1))
        스타일 = 스타일.replace('</numFmts>', 새서식 + '</numFmts>', 1)
        스타일 = 스타일.replace(f'<numFmts count="{서식수}"',
                                f'<numFmts count="{서식수 + 3}"', 1)
    else:
        스타일 = 스타일.replace('<fonts', f'<numFmts count="3">{새서식}</numFmts><fonts', 1)

    맑은 = '<name val="맑은 고딕"/><family val="2"/>'
    새글꼴 = (f'<font><sz val="12"/><b/><color rgb="FF000000"/>{맑은}</font>'
              f'<font><sz val="11"/><b/><color rgb="FF000000"/>{맑은}</font>'
              f'<font><sz val="9"/><color rgb="FF3B6AA0"/>{맑은}</font>'
              f'<font><sz val="10"/><color rgb="FF000000"/>{맑은}</font>'
              f'<font><sz val="10"/><b/><color rgb="FF000000"/>{맑은}</font>')
    스타일 = 스타일.replace('</fonts>', 새글꼴 + '</fonts>', 1)
    스타일 = 스타일.replace(f'<fonts count="{글꼴수}"', f'<fonts count="{글꼴수 + 5}"', 1)
    G제목, G머리, G설명, G본문, G굵게 = (글꼴수, 글꼴수 + 1, 글꼴수 + 2,
                                        글꼴수 + 3, 글꼴수 + 4)

    점 = '<left style="dotted"><color rgb="FF808080"/></left>'
    실 = '<bottom style="thin"><color rgb="FF000000"/></bottom>'
    위실 = '<top style="thin"><color rgb="FF000000"/></top>'
    새테두리 = ('<border><left/><right/><top/><bottom/><diagonal/></border>'
                f'<border>{점}<right/><top/><bottom/><diagonal/></border>'
                f'<border><left/><right/>{위실}{실}<diagonal/></border>'
                f'<border>{점}<right/>{위실}{실}<diagonal/></border>'
                f'<border><left/><right/><top/>{실}<diagonal/></border>'
                f'<border>{점}<right/><top/>{실}<diagonal/></border>')
    스타일 = 스타일.replace('</borders>', 새테두리 + '</borders>', 1)
    스타일 = 스타일.replace(f'<borders count="{테두리수}"',
                            f'<borders count="{테두리수 + 6}"', 1)
    B없음, B점, B머리0, B머리, B끝0, B끝 = (테두리수, 테두리수 + 1, 테두리수 + 2,
                                           테두리수 + 3, 테두리수 + 4, 테두리수 + 5)

    def xf(글꼴, 테두리, 서식=0, 가로=None):
        맞춤 = f'<alignment horizontal="{가로}" vertical="center"/>' if 가로 else ''
        속성 = (f'<xf numFmtId="{서식}" fontId="{글꼴}" fillId="0" borderId="{테두리}" '
                f'xfId="0" applyFont="1" applyBorder="1"'
                + (' applyNumberFormat="1"' if 서식 else '')
                + (' applyAlignment="1"' if 가로 else ''))
        return 속성 + (f'>{맞춤}</xf>' if 맞춤 else '/>')

    새xf = [
        xf(G제목, B없음, 자금표시),                    # 0 제목 (← 여기가 표시 번호)
        xf(G굵게, B없음), xf(G설명, B없음),            # 1 소제목 · 2 설명
        xf(G머리, B머리0, 가로='center'),              # 3 머리(첫 칸)
        xf(G머리, B머리, 가로='center'),               # 4 머리
        xf(G본문, B없음, 서식=F날짜, 가로='center'),   # 5 일자(첫 칸)
        xf(G본문, B점, 서식=F숫자),                    # 6 숫자
        xf(G본문, B점, 가로='center'),                 # 7 가운데
        xf(G본문, B점, 가로='left'),                   # 8 왼쪽
        xf(G본문, B없음, 서식=F날짜, 가로='center'),   # 9 (예비)
        xf(G본문, B끝0), xf(G본문, B끝),               # 10 끝줄(첫 칸) · 11 끝줄
        xf(G설명, B없음, 가로='left'),                 # 12 주석
        xf(G본문, B점, 서식=F날짜, 가로='center'),     # 13 일자
        xf(G본문, B없음, 가로='left'),                 # 14 근거
        xf(G굵게, B없음, 가로='left'),                 # 15 근거 제목
        xf(G본문, B없음, 서식=F숫자),                  # 16 근거 숫자
    ]
    스타일 = 스타일.replace('</cellXfs>', ''.join(새xf) + '</cellXfs>', 1)
    스타일 = 스타일.replace(f'<cellXfs count="{xf수}"',
                            f'<cellXfs count="{xf수 + len(새xf)}"', 1)
    return 스타일, [xf수 + i for i in range(len(새xf))], True


def _관리비항목찾기(z, 파일, 공유):
    """「월별 실적집계」 C열에서 관리비 하위 항목(1) … 13))과 그 행 번호를 찾습니다."""
    표 = _시트칸읽기(z, 파일, ('C',), 공유)
    글 = {행: str(칸.get('C', '')).strip() for 행, 칸 in 표.items()}
    시작 = next((r for r in sorted(글) if 글[r].startswith('관리비 합계')), None)
    끝 = next((r for r in sorted(글) if r > (시작 or 0)
               and 글[r].startswith('영업이익')), None)
    if 시작 is None:
        return []
    항목 = []
    for 행 in sorted(글):
        if 행 <= 시작 or (끝 and 행 >= 끝):    # 매출·매출원가와 아래쪽 요약표는 건너뜁니다
            continue
        m = re.match(r'^(\d+)\)\s*(.+)$', 글[행])
        if not m:
            continue
        이름 = m.group(2).strip()
        if '상각비' in 이름:                   # 현금이 나가지 않으니 예상 출금에서 뺍니다
            continue
        항목.append((이름, 행))
    return 항목


def _자금시트만들기(z, 시트칸, 공유, 새, 이어받음, 매핑표, 연도, 보고달):
    """「26년 자금」 시트 XML 을 통째로 다시 만듭니다. → (시트xml, 새styles, 안내)"""
    자금시트 = next((n for n in 시트칸 if '자금' in n), None)
    if not 자금시트:
        return None, None, None, {}
    현금계정 = {k.strip().lower() for k, (b, kor) in 매핑표.items()
                if b == 'BS' and ('현금' in kor or '예금' in kor)}
    한글 = {k.strip().lower(): v[1] for k, v in 매핑표.items()}
    비에스 = {k.strip().lower(): v[0] for k, v in 매핑표.items()}

    g = pd.DataFrame({
        '그룹': pd.Series(_글자칸(새, 원장열자리['A'])).replace('', None).ffill().fillna(''),
        '계정': _글자칸(새, 'Distribution account'),
        '전표': _글자칸(새, 'Transaction ID'),
        'Description': _글자칸(새, 'Description'),
        '#': _글자칸(새, '#'),
        'Name': _글자칸(새, 'Name'),
        '일련': [_날짜일련(v) for v in 새['Transaction date']],
        '차변': pd.to_numeric(새.get('Debit'), errors='coerce').fillna(0).tolist(),
        '대변': pd.to_numeric(새.get('Credit'), errors='coerce').fillna(0).tolist(),
        '잔액': pd.to_numeric(새.get('Balance'), errors='coerce').fillna(0).tolist(),
        '활동': [v or '' for v in 이어받음.get('Q', [None] * len(새))],
    })
    # 기초 현금 = 현금·예금 계정의 Beginning Balance 합계
    기초현금액 = float(g.loc[g['계정'].eq('Beginning Balance')
                             & g['그룹'].str.lower().isin(현금계정), '잔액'].sum())

    g = g[g['계정'].ne('') & g['일련'].notna()].copy()
    키 = g['계정'].str.lower()
    g['계정과목'] = 키.map(한글).fillna('')
    g['계정분류'] = 키.map(비에스).fillna('')
    g['현금'] = 키.isin(현금계정)

    상대 = g[~g['현금']].copy()
    상대['코드'] = [_자금코드(n) for n in 상대['계정과목']]
    상대.loc[상대['계정과목'].str.contains('감가상각누계', na=False), '코드'] = 'CAPEX'
    # 매출채권 회수가 아닌 특별 입금은 기타입금으로 갈라 냅니다
    지원전표 = set(g.loc[_기타입금줄(g, g['현금'], g['차변']), '전표'])
    상대.loc[상대['전표'].isin(지원전표), '코드'] = 'ETCIN'
    상대['크기'] = (상대['대변'] - 상대['차변']).abs()
    상대['환율아님'] = (~상대['코드'].eq('FX')).astype(int)
    상대['세부'] = np.where(상대['계정분류'].eq('IS') & 상대['활동'].ne(''),
                            상대['활동'].map(자금항목이름).fillna('기타관리경비'), '')
    대표 = (상대.sort_values(['환율아님', '크기'], ascending=False)
            .drop_duplicates('전표').set_index('전표')[['코드', '세부']])

    현 = g[g['현금']].sort_values(['일련', '전표'], kind='stable').reset_index(drop=True)
    현['코드'] = 현['전표'].map(대표['코드']).fillna('이체')
    현['세부'] = 현['전표'].map(대표['세부']).fillna('')
    현['구분'] = 현['코드'].map(자금구분).fillna('현금 간 이체')
    현['항목'] = np.where(현['세부'].ne(''), 현['세부'],
                          현['코드'].map(자금이름).fillna('현금 계정 간 이체'))
    기말액 = 기초현금액 + float((현['차변'] - 현['대변']).sum())

    # ── 담당자가 손으로 채운 3번 표(예상 입금)는 그대로 옮깁니다 ──
    옛칸 = _자금칸그대로(z, 시트칸[자금시트])
    옛글 = _시트칸읽기(z, 시트칸[자금시트],
                       tuple(_열이름(i) for i in range(1, 31)), 공유)
    옛제목 = {}
    보존시작 = None
    for 행, 칸 in 옛글.items():
        if 행 > 8:
            continue
        for 열, v in 칸.items():
            m = re.match(r'^\s*([23])\s*\.', str(v))
            if m:
                옛제목.setdefault(m.group(1), (행, _열번호(열), str(v)))
                if m.group(1) == '3':
                    보존시작 = _열번호(열) - 1      # 3번 표 앞 빈 칸부터 오른쪽 전부
    보존칸 = ({칸: 조각 for 칸, 조각 in 옛칸.items() if 칸[1] >= 보존시작}
              if 보존시작 else {})

    스타일, S, 새서식 = _자금서식(z.read('xl/styles.xml').decode('utf-8'))
    (S제목, S소제목, S설명, S머리0, S머리, S일자0, S숫자, S가운데, S왼쪽,
     _예비, S끝0, S끝, _주석, S일자, S근거, S근거제목, S근거숫자) = S

    칸들 = []

    def 글(ref, s, v):
        칸들.append(f'<c r="{ref}" s="{s}" t="inlineStr">'
                    f'<is><t xml:space="preserve">{_xml(v)}</t></is></c>')

    def 숫(ref, s, v):
        # ※ numpy 숫자를 그대로 쓰면 'np.float64(…)' 로 적혀 파일이 깨집니다
        칸들.append(f'<c r="{ref}" s="{s}"><v>{float(v)!r}</v></c>')

    def 식(ref, s, f):
        칸들.append(f'<c r="{ref}" s="{s}"><f>{_xml(f)}</f></c>')

    def 빈(ref, s):
        칸들.append(f'<c r="{ref}" s="{s}"/>')

    머리 = ['일자', '차변', '대변', '기말', '구분', '활동성항목']   # 1. 실제 (기말 있음)
    예상머리 = ['일자', '차변', '대변', '구분', '활동성항목']        # 2·3. 예상 (기말 없음)
    왼열 = ['B', 'C', 'D', 'E', 'F', 'G']
    오열 = ['I', 'J', 'K', 'L', 'M']               # 2. 예상 출금 (대변만 클로드가 채웁니다)
    입열 = ['O', 'P', 'Q', 'R', 'S']               # 3. 예상 입금 (담당자가 직접 넣습니다)
    글('A1', S제목, '● 원장 작성 현금성 자산 계정별원장')
    글('B3', S소제목, '1. 실제 현금성 자금 흐름')
    글('B4', S설명, f'- 원장상 현금흐름을 기준으로 클로드가 작성한 현금흐름 '
                    f'({연도}년 1~{보고달}월)')
    글('I3', S소제목, 옛제목.get('2', (0, 0, '2. 예상 현금성 자금 흐름(출금)'))[2])
    글('I4', S설명, f'- 실적 월평균을 기준으로 클로드가 채운 예상 출금 '
                    f'({보고달 + 1}~12월)' if 보고달 < 12 else '- 12월까지 실적이 다 들어와 예상분은 없습니다')
    for 열, 이름 in zip(왼열, 머리):
        글(f'{열}5', S머리0 if 열 == 'B' else S머리, 이름)
    for 열, 이름 in zip(오열, 예상머리):
        글(f'{열}5', S머리0 if 열 == 'I' else S머리, 이름)

    # ── 왼쪽 : 원장에서 뽑은 실제 현금 계정별원장
    r = 6
    빈(f'B{r}', S일자0)
    빈(f'C{r}', S숫자)
    빈(f'D{r}', S숫자)
    숫(f'E{r}', S숫자, round(기초현금액, 2))
    글(f'F{r}', S가운데, '기초')
    글(f'G{r}', S왼쪽, f'{연도}년 1월 1일 현금·예금 잔액')
    for _i, t in 현.iterrows():
        r += 1
        숫(f'B{r}', S일자, int(t['일련']))
        차, 대 = float(t['차변']), float(t['대변'])
        숫(f'C{r}', S숫자, round(차, 2)) if 차 else 빈(f'C{r}', S숫자)
        숫(f'D{r}', S숫자, round(대, 2)) if 대 else 빈(f'D{r}', S숫자)
        식(f'E{r}', S숫자, f'E{r - 1}+C{r}-D{r}')
        글(f'F{r}', S가운데, str(t['구분']))
        글(f'G{r}', S왼쪽, str(t['항목']))
    왼끝 = r + 1
    for 열 in 왼열:
        빈(f'{열}{왼끝}', S끝0 if 열 == 'B' else S끝)

    # ── 오른쪽 : 남은 달의 예상 출금 (보고월 다음 달부터 12월까지)
    집계시트 = next((n for n in 시트칸 if '월별' in n and '실적집계' in n), None)
    관리비행 = _관리비항목찾기(z, 시트칸[집계시트], 공유) if 집계시트 else []
    예상월 = list(range(보고달 + 1, 13))
    끝열 = _열이름(23 + 보고달)          # X=1월 … 1~보고달 평균을 냅니다
    r = 6
    빈(f'I{r}', S일자0)
    빈(f'J{r}', S숫자)
    빈(f'K{r}', S숫자)
    글(f'L{r}', S가운데, '기초')
    글(f'M{r}', S왼쪽, f'{연도}년 {보고달}월말 실제 잔액')
    # 수식이 낼 값을 미리 구해 같이 적어 둡니다 (엑셀을 열자마자 보이고, 대시보드도 읽습니다)
    월평균 = {}
    if 집계시트 and 관리비행:
        달칸 = [_열이름(24 + k) for k in range(보고달)]      # X=1월 …
        읽은 = _시트칸읽기(z, 시트칸[집계시트], tuple(달칸), 공유)
        for 이름, 시트행 in 관리비행:
            값 = [pd.to_numeric(pd.Series([읽은.get(시트행, {}).get(c)]),
                                errors='coerce').fillna(0).iat[0] for c in 달칸]
            월평균[이름] = round(float(sum(값)) / max(보고달, 1), 0)

    def 값식(ref, s, f, v):
        """수식과 그 값을 함께 적습니다."""
        칸들.append(f'<c r="{ref}" s="{s}">'
                    f'<f>{_xml(f)}</f><v>{float(v)!r}</v></c>')

    for 이름, 시트행 in 관리비행:
        for 월 in 예상월:
            r += 1
            숫(f'I{r}', S일자, _날짜일련(datetime.date(
                연도 + (1 if 월 > 12 else 0), 월,
                calendar.monthrange(연도, 월)[1])))
            빈(f'J{r}', S숫자)
            값식(f'K{r}', S숫자,
                 f"ROUND(AVERAGE('{집계시트}'!$X${시트행}:${끝열}${시트행}),0)",
                 월평균.get(이름, 0.0))
            글(f'L{r}', S가운데, '영업활동')
            글(f'M{r}', S왼쪽, 이름)
    오끝 = r + 1
    for 열 in 오열:
        빈(f'{열}{오끝}', S끝0 if 열 == 'I' else S끝)

    # ── 3. 예상 입금 ─────────────────────────────────────────
    #    지난 파일에 이미 있으면 담당자가 채운 그대로 옮깁니다 (한 칸도 안 건드립니다).
    입금항목 = [('영업활동', '매출채권 회수'), ('영업활동', '기타입금'),
                ('재무활동', '유상증자'), ('재무활동', '차입증감')]
    if 보존칸:
        입끝 = max(r for r, _c in 보존칸)
        for (행, 열), 조각 in sorted(보존칸.items()):
            칸들.append(조각)
    else:
        글('O3', S소제목, '3. 예상 현금성 자금 흐름(입금)')
        글('O4', S설명, f'- 담당자가 직접 채워 넣는 예상 입금 ({보고달 + 1}~12월)'
                        if 보고달 < 12 else '- 12월까지 실적이 다 들어와 예상분은 없습니다')
        for 열, 이름 in zip(입열, 예상머리):
            글(f'{열}5', S머리0 if 열 == 'O' else S머리, 이름)
        r = 6
        빈(f'O{r}', S일자0)
        빈(f'P{r}', S숫자)
        빈(f'Q{r}', S숫자)
        글(f'R{r}', S가운데, '기초')
        글(f'S{r}', S왼쪽, f'{연도}년 {보고달}월말 실제 잔액')
        for 구분이름, 이름 in 입금항목:
            for 월 in 예상월:
                r += 1
                숫(f'O{r}', S일자, _날짜일련(datetime.date(
                    연도, 월, calendar.monthrange(연도, 월)[1])))
                빈(f'P{r}', S숫자)           # ← 담당자가 여기에 금액을 넣습니다
                빈(f'Q{r}', S숫자)
                글(f'R{r}', S가운데, 구분이름)
                글(f'S{r}', S왼쪽, 이름)
        입끝 = r + 1
        for 열 in 입열:
            빈(f'{열}{입끝}', S끝0 if 열 == 'O' else S끝)

    # ── 작성 근거
    순 = (현['차변'] - 현['대변'])
    구분합 = (현.assign(순=순).groupby('항목')['순'].agg(['size', 'sum'])
              .sort_values('sum'))
    근거 = [
        ('■ 작성 근거', ''),
        ('', ''),
        ('[1] 실제 현금성 자금 흐름 (왼쪽)', ''),
        ('무엇을 옮겼나', f'{연도 % 100}년 원장_raw 에서 현금·예금 계정에 찍힌 줄만 날짜순으로 '
                         f'뽑았습니다 ({len(현):,}줄, {연도}년 1~{보고달}월). 현금 계정은 '
                         f'BS_IS_매핑에서 「현금」·「예금」이 들어간 재무상태표 계정으로 잡았습니다.'),
        ('차변 / 대변', '원장의 Debit / Credit 을 그대로 옮겼습니다. 손대지 않았습니다.'),
        ('기말', f'바로 윗줄 기말 + 차변 − 대변 (수식). 기초 {기초현금액:,.2f} 는 원장의 '
                 f'Beginning Balance 합계입니다.'),
        ('구분 · 활동성항목', '원장은 복식부기라, 현금이 든 전표에서 현금이 아닌 상대 계정이 그 돈의 '
                            '성격을 알려 줍니다. 전표(Transaction ID)가 같은 줄 중 현금이 아닌 것을 '
                            '찾아 그 계정으로 정했습니다. 상대가 여럿이면 금액이 가장 큰 것을 쓰고, '
                            '외환차손익은 곁다리라 뒤로 미뤘습니다.'),
        ('맞는지 확인', f'기초 {기초현금액:,.2f} + 순증감 {float(순.sum()):,.2f} '
                       f'= {기말액:,.2f} — 대시보드 현금흐름의 {보고달}월말 잔액과 같습니다.'),
        ('', ''),
        ('[2] 예상 현금성 자금 흐름 (오른쪽)', ''),
        ('무엇을 넣었나', f'「{집계시트}」 시트에서 관리비 아래에 있는 항목을 그 순서 그대로 놓고, '
                         f'항목마다 {보고달 + 1}월부터 12월까지 {len(예상월)}줄씩 넣었습니다. '
                         f'이미 실적이 들어온 {보고달}월까지는 왼쪽 표에 있으니 여기서 뺐습니다.'
                         if 예상월 else '12월까지 실적이 다 들어와 예상분은 없습니다.'),
        ('금액을 어떻게 잡았나', f'그 항목의 1~{보고달}월 실적을 {보고달}달로 나눈 월평균입니다. '
                               f"칸에는 =ROUND(AVERAGE('{집계시트}'!X:{끝열}),0) 수식이 들어 있어 "
                               f'실적이 바뀌면 예상액도 같이 바뀝니다.'),
        ('뺀 것', '상각비는 현금이 나가지 않는 비용이라 넣지 않았습니다.'),
        ('기말', f'{보고달}월말 실제 잔액에서 출발해 한 줄씩 빼 내려갑니다. 줄 순서가 '
                 f'활동성항목 순서라 기말도 그 순서대로 쌓입니다.'),
        ('', ''),
        ('[3] 예상 현금성 자금 흐름(입금)', ''),
        ('무엇을 깔았나', '매출채권 회수 · 기타입금 · 유상증자 · 차입증감 네 항목을 '
                        f'{보고달 + 1}월부터 12월까지 줄로 깔아 두었습니다. '
                        '금액은 담당자가 「차변」 칸에 직접 넣으시면 됩니다.'
                        if 예상월 else '12월까지 실적이 다 들어와 예상분은 없습니다.'),
        ('넣으면 어떻게 되나', '대시보드 현금흐름 표의 「○월(예상)」 칸에 그대로 반영됩니다 — '
                            '출금(2번)과 입금(3번)을 합쳐 연말까지 잔액이 이어집니다.'),
    ]
    r2 = max(왼끝, 오끝, 입끝) + 2
    for 제목, 본문 in 근거:
        if 제목.startswith(('■', '[')):
            글(f'I{r2}', S근거제목, 제목)
        elif 제목:
            글(f'I{r2}', S근거, '· ' + 제목)
            글(f'K{r2}', S근거, 본문)
        r2 += 1
    if 예상월:
        글(f'I{r2}', S근거, '· 한 달 예상 출금 합계 (USD)')
        식(f'L{r2}', S근거숫자, f'SUM($K$7:$K${오끝 - 1})/{len(예상월)}')
        r2 += 2
    글(f'I{r2}', S근거제목,
      f'[3] 참고 — 실제 현금이 어디로 오갔나 ({연도}년 1~{보고달}월, USD)')
    r2 += 1
    for 이름, 줄 in 구분합.iterrows():
        글(f'I{r2}', S근거, '· ' + str(이름))
        글(f'K{r2}', S근거, f'{int(줄["size"]):,}줄')
        숫(f'L{r2}', S근거숫자, round(float(줄['sum']), 2))
        r2 += 1
    글(f'I{r2}', S근거, f'· 합계 (= {보고달}개월 순증감)')
    숫(f'L{r2}', S근거숫자, round(float(순.sum()), 2))

    폭 = ('<cols><col min="1" max="1" width="2.5" customWidth="1"/>'
          '<col min="2" max="2" width="13" customWidth="1"/>'
          '<col min="3" max="5" width="15.5" customWidth="1"/>'
          '<col min="6" max="6" width="14" customWidth="1"/>'
          '<col min="7" max="7" width="24" customWidth="1"/>'
          '<col min="8" max="8" width="3" customWidth="1"/>'
          '<col min="9" max="9" width="13" customWidth="1"/>'
          '<col min="10" max="11" width="15.5" customWidth="1"/>'
          '<col min="12" max="12" width="14" customWidth="1"/>'
          '<col min="13" max="13" width="24" customWidth="1"/>'
          '<col min="14" max="14" width="3" customWidth="1"/>'
          '<col min="15" max="15" width="13" customWidth="1"/>'
          '<col min="16" max="17" width="15.5" customWidth="1"/>'
          '<col min="18" max="18" width="14" customWidth="1"/>'
          '<col min="19" max="19" width="24" customWidth="1"/></cols>')
    행모음 = {}
    for c in 칸들:
        ref = re.match(r'<c r="([A-Z]+)(\d+)"', c)
        행모음.setdefault(int(ref.group(2)), []).append((ref.group(1), c))
    본문 = ''.join(
        f'<row r="{n}" spans="1:{끝열}">'
        + ''.join(x for _c, x in sorted(행모음[n], key=lambda t: (len(t[0]), t[0])))
        + '</row>' for n in sorted(행모음))
    끝행 = max(행모음)
    끝열 = max((_열번호(re.match(r'<c r="([A-Z]+)', x).group(1))
                for 줄 in 행모음.values() for _c, x in 줄), default=19)
    시트 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheetPr><tabColor theme="9" tint="0.79998168889431442"/></sheetPr>'
            f'<dimension ref="A1:{_열이름(끝열)}{끝행}"/>'
            '<sheetViews><sheetView showGridLines="0" workbookViewId="0">'
            '<pane ySplit="5" topLeftCell="A6" activePane="bottomLeft" state="frozen"/>'
            '<selection pane="bottomLeft" activeCell="B6" sqref="B6"/></sheetView></sheetViews>'
            '<sheetFormatPr defaultRowHeight="17.4"/>'
            + 폭 + '<sheetData>' + 본문 + '</sheetData>'
            '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
            'header="0.3" footer="0.3"/></worksheet>')
    안내 = {'시트': 자금시트, '현금줄': len(현), '기초': 기초현금액, '기말': 기말액,
            '예상월': 예상월, '항목수': len(관리비행), '서식추가': 새서식}
    return 시트, (스타일 if 새서식 else None), 시트칸[자금시트], 안내


#  「당월 실적집계」에서 기준 월(H2) 대신 빈칸을 보고 있던 수식들.
#  Q18·R18 은 매출원가 누적인데 G5·H5(빈칸)를 참조해 OFFSET 이 0칸이 되어,
#  1월부터 이번 달까지가 아니라 「연간실적 칸 + 1월」을 더하고 있었습니다.
당월수식고침 = {
    'Q18': ("SUM(OFFSET('월별 실적집계'!$W$10,0,1):"
            "(OFFSET('월별 실적집계'!$W$10,0,G5)))",
            "SUM(OFFSET('월별 실적집계'!$W$10,0,1):"
            "(OFFSET('월별 실적집계'!$W$10,0,$H$2)))"),
    'R18': ("SUM(OFFSET('월별 실적집계'!$J$10,0,1):"
            "(OFFSET('월별 실적집계'!$J$10,0,H5)))",
            "SUM(OFFSET('월별 실적집계'!$J$10,0,1):"
            "(OFFSET('월별 실적집계'!$J$10,0,$H$2)))"),
}


def _당월수식고치기(시트xml):
    """잘못된 참조를 쓰던 칸의 수식을 바로잡습니다. 돌려주는 값 : (xml, 고친 칸 목록)"""
    고침 = []
    for 칸, (옛, 새) in 당월수식고침.items():
        def _바꿔(m, 칸=칸, 옛=옛, 새=새):
            속성, 몸 = m.group(1), m.group(2)
            fm = re.search(r'<f([^>]*)>(.*?)</f>', 몸, re.S)
            if not fm or _풀기(fm.group(2)).replace(' ', '') != 옛.replace(' ', ''):
                return m.group(0)
            고침.append(칸)
            return f'<c r="{칸}"{속성}><f{fm.group(1)}>{_xml(새)}</f></c>'
        시트xml = re.sub(rf'<c r="{칸}"([^>]*)>(.*?)</c>', _바꿔,
                         시트xml, count=1, flags=re.S)
    return 시트xml, 고침


def 실적보고엑셀만들기(원장바이트, 전월바이트):
    """돌려주는 값 : (엑셀 bytes, 안내 정보 dict)"""
    새bio, 전bio = io.BytesIO(원장바이트), io.BytesIO(전월바이트)

    # ── 1. 이번 달 원장 원본
    새 = None
    for s in pd.ExcelFile(새bio).sheet_names:
        for h in (4, 0, 3, 5):
            try:
                d = pd.read_excel(새bio, sheet_name=s, header=h)
            except Exception:
                continue
            if 'Distribution account' in [str(c) for c in d.columns]:
                새 = d
                break
        if 새 is not None:
            break
    if 새 is None:
        raise ValueError('원장 파일에서 「Distribution account」 열을 찾지 못했습니다. '
                         'QuickBooks 에서 내려받은 General Ledger 원본이 맞는지 확인해 주세요.')

    # ── 2. 전월 실적보고자료
    z = zipfile.ZipFile(전bio)
    (올해시트, 올해파일), 시트목록 = _전월원장시트(z)
    전 = pd.read_excel(전bio, sheet_name=올해시트, header=4)

    # ── 3. 손으로 나눈 활동분류·검토메모 이어받기
    전키, 새키 = _이어받기키(전), _이어받기키(새)
    이어받음, 이어받은수 = {}, 0
    for col, 이름 in 이어받기열.items():
        후보 = _열찾기(전, 이름)
        if 후보 is None:
            이어받음[col] = [None] * len(새)
            continue
        사전 = pd.Series(전[후보].values, index=전키)
        사전 = 사전[~사전.index.duplicated()]
        옮김 = 새키.map(사전)
        이어받음[col] = [None if (v is None or (isinstance(v, float) and pd.isna(v))
                                 or str(v).strip() == '') else str(v) for v in 옮김]
        if col == 'Q':
            이어받은수 = sum(1 for v in 이어받음[col] if v)

    # ── 3-2. 전월에 없던 새 거래는 지난달까지의 분류를 배워서 스스로 채웁니다
    시트칸 = dict(_시트지도(z))
    공유 = _공유글자(z)
    집계시트 = next((n for n in 시트칸 if '월별' in n and '실적집계' in n), None)
    유효세부 = set(_시트글자(z, 시트칸[집계시트], 'E', 공유)) if 집계시트 else set()
    매핑시트 = next((n for n in 시트칸 if 'BS_IS_매핑' in n), None)
    매핑표 = _매핑표읽기(z, 시트칸[매핑시트], 공유) if 매핑시트 else {}
    매핑분류 = {k: v[0] for k, v in 매핑표.items()}
    분류센것 = _자동분류(새, 이어받음, _분류학습(전), 유효세부, 매핑분류)

    날짜들 = [d for d in (_날짜일련(v) for v in 새['Transaction date']) if d is not None]
    처음 = 엑셀기준일 + datetime.timedelta(days=min(날짜들))
    마지막 = 엑셀기준일 + datetime.timedelta(days=max(날짜들))
    기간글 = f'{처음:%Y.%m.%d} ~ {마지막:%Y.%m.%d}'
    보고달 = 마지막.month

    # ── 4. 원장 시트를 갈아끼웁니다
    새xml = _원장시트XML(z.read(올해파일).decode('utf-8'), 새, 이어받음, 기간글, 매핑표)
    끝행 = 5 + len(새)
    # 「당월 실적집계」는 이제 쓰지 않는 시트라 결과 파일에서 통째로 뺍니다.
    #   (담당자가 원본에서 지우셨습니다. 옛 파일을 올리셔도 결과에는 들어가지 않습니다.)
    당월시트 = next((n for n in 시트칸 if '당월' in n and '실적집계' in n), None)
    당월xml, 당월파일, 달바뀜, 수식고침 = None, None, False, []
    뺀파트, 뺀패치 = (_시트빼기(z, 당월시트) if 당월시트 else (set(), {}))
    뺀시트 = 당월시트 if 뺀파트 else None

    # ── 4-2. 「26년 자금」 시트도 이번 달까지로 다시 씁니다
    try:
        자금xml, 새스타일, 자금파일, 자금안내 = _자금시트만들기(
            z, 시트칸, 공유, 새, 이어받음, 매핑표, 마지막.year, 보고달)
    except Exception as e:                      # 자금 시트 하나 때문에 전체가 막히면 안 됩니다
        자금xml, 새스타일, 자금파일, 자금안내 = None, None, None, {'오류': str(e)}
    buf = io.BytesIO()
    out = zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED, compresslevel=6)
    for it in z.infolist():
        이름 = it.filename
        if 이름 == 'xl/calcChain.xml':      # 낡은 계산순서표는 빼야 「복구됨」이 안 뜹니다
            continue
        if 이름 in 뺀파트:                  # 「당월 실적집계」와 거기에만 딸린 파트들
            continue
        데이터 = 뺀패치.get(이름) or z.read(이름)
        if 이름 == 올해파일:
            데이터 = 새xml.encode('utf-8')
        elif 당월xml is not None and 이름 == 당월파일:
            데이터 = 당월xml.encode('utf-8')
        elif 자금xml is not None and 이름 == 자금파일:
            데이터 = 자금xml.encode('utf-8')
        elif 새스타일 is not None and 이름 == 'xl/styles.xml':
            데이터 = 새스타일.encode('utf-8')
        elif 이름 == 'xl/workbook.xml':
            데이터 = re.sub(r'<calcPr[^>]*/>',
                            '<calcPr calcId="191029" fullCalcOnLoad="1"/>',
                            데이터.decode('utf-8'), count=1).encode('utf-8')
        elif 이름 == '[Content_Types].xml':
            데이터 = re.sub(r'<Override PartName="/xl/calcChain\.xml"[^>]*/>', '',
                            데이터.decode('utf-8')).encode('utf-8')
        elif 이름 == 'xl/_rels/workbook.xml.rels':
            데이터 = re.sub(r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', '',
                            데이터.decode('utf-8')).encode('utf-8')
        elif 'pivotCacheDefinition' in 이름 and 이름.endswith('.xml'):
            t = 데이터.decode('utf-8')
            t = re.sub(r'(<pivotCacheDefinition\b[^>]*?)(\s*>)',
                       lambda m: re.sub(r'\s+refreshOnLoad="[^"]*"', '', m.group(1))
                       + ' refreshOnLoad="1"' + m.group(2), t, count=1)
            t = re.sub(r'(<worksheetSource ref="[A-Z]+5:[A-Z]+)\d+(")',
                       rf'\g<1>{max(끝행 + 3000, 30000)}\g<2>', t)
            데이터 = t.encode('utf-8')
        out.writestr(it, 데이터)
    out.close()

    연도, 월끝 = 마지막.year, 마지막.month
    if 뺀시트:                       # 뺀 시트는 결과 파일의 시트 목록에서도 지웁니다
        시트목록 = [n for n in 시트목록 if n != 뺀시트]
    거래 = int(새['Distribution account'].notna().sum())
    자동합 = sum(분류센것.get(k, 0) for k in
                 ('적요일치', '거래처일치', '계정과목대표값', '계정과목최빈값'))
    안내 = {'연도': 연도, '월끝': 월끝, '기간': 기간글, '올해시트': 올해시트,
            '행수': len(새), '거래건수': 거래, '이어받음': 이어받은수,
            '새거래': int(새키.isin(set(전키)).eq(False).sum()),
            '시트수': len(시트목록), '시트목록': 시트목록,
            '당월시트': 당월시트, '기준월맞춤': 달바뀜, '보고달': 보고달,
            '뺀시트': 뺀시트,
            '수식고침': 수식고침, '자금': 자금안내,
            '자동분류': 자동합, '분류내역': 분류센것,
            '확인필요': 분류센것.get('확인필요', 0) + 분류센것.get('신규계정', 0),
            '집계밖': 분류센것.get('집계밖', 0), '매핑밖': 분류센것.get('매핑밖', 0),
            '되살린열': list(_되살린열)}
    return buf.getvalue(), 안내


@st.cache_data(max_entries=1, show_spinner='실적보고 엑셀을 만드는 중입니다...')
def 실적보고엑셀캐시(원장바이트, 전월바이트):
    """같은 파일이면 한 번만 만듭니다 (내려받기를 눌러도 다시 안 만듦)."""
    return 실적보고엑셀만들기(원장바이트, 전월바이트)


자료실 = '자료실'
# 「마지막에 올린 자료」를 기억해 두는 파일들을 넣는 곳 —
# 대시보드 폴더가 지저분해지지 않도록 따로 모아 둡니다.
기억폴더 = '_기억자료'
try:
    os.makedirs(기억폴더, exist_ok=True)
except OSError:
    기억폴더 = '.'


def 기억자리(이름):
    return os.path.join(기억폴더, 이름)


def 오늘날짜():
    """보고 계신 분의 시계로 「오늘」 — 웹서버는 영국 시간이라 하루가 어긋납니다.

    ※ 한국에서 8월 29일 아침이어도 서버는 아직 8월 28일이라, 파일 이름이
      하루 뒤처져 붙던 것을 막습니다. 브라우저가 알려 주는 시간대를 씁니다.
    """
    try:
        이름 = getattr(st.context, 'timezone', None)
        if 이름:
            import zoneinfo
            return datetime.datetime.now(zoneinfo.ZoneInfo(str(이름))).date()
    except Exception:
        pass
    try:
        분 = getattr(st.context, 'timezone_offset', None)   # 한국이면 −540
        if 분 is not None:
            return (datetime.datetime.now(datetime.timezone.utc)
                    - datetime.timedelta(minutes=int(분))).date()
    except Exception:
        pass
    return datetime.date.today()
# 파일 이름 끝의 「2026.08.27_v1」 을 「만든 날짜 + 그날의 몇 번째」로 봅니다
날짜버전패턴 = re.compile(r'(\d{4})\.(\d{1,2})\.(\d{1,2})[_\s]*v\s*(\d+)\s*\.xlsx$', re.I)
# 예전 이름(…클로드작성_v24.xlsx)도 목록에는 그대로 보이게 합니다
버전패턴 = re.compile(r'v\s*(\d+)\s*\.xlsx$', re.I)
# 파일 이름 앞쪽의 「26년 6월」은 그 파일이 다루는 실적 월입니다
연월패턴 = re.compile(r'(\d{2})\s*년\s*(\d{1,2})\s*월')


def _이름속연월(이름):
    m = 연월패턴.search(str(이름) or '')
    return (int(m.group(1)), int(m.group(2))) if m else None


def _이름속날짜(이름):
    m = 날짜버전패턴.search(str(이름) or '')
    if not m:
        return None
    try:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def 자료실목록():
    """자료실 폴더에 쌓인 실적보고 엑셀을 새 것부터 돌려줍니다."""
    try:
        이름들 = [n for n in os.listdir(자료실) if n.lower().endswith('.xlsx')
                  and not n.startswith('~$') and 버전패턴.search(n)
                  and '연결패키지' not in n]
    except OSError:
        return []
    목록 = []
    for n in 이름들:
        경로 = os.path.join(자료실, n)
        try:
            st_ = os.stat(경로)
        except OSError:
            continue
        목록.append({'이름': n, '경로': 경로, '버전': int(버전패턴.search(n).group(1)),
                     '연월': _이름속연월(n), '만든날': _이름속날짜(n),
                     '시각': st_.st_mtime, '크기': st_.st_size})
    목록.sort(key=lambda d: (d['시각'], d['버전']), reverse=True)
    return 목록


def 다음버전(오늘=None, 처음부터=False):
    """번호는 「만든 날」로 셉니다.

    ※ 같은 날 또 만들면 v2 · v3 … 이고, 날이 바뀌면 다시 v1 부터입니다.
      (8월 27일에 두 번 만들면 2026.08.27_v1 · 2026.08.27_v2,
       28일에 만들면 2026.08.28_v1)
    """
    if 처음부터:
        return 1
    오늘 = 오늘 or 오늘날짜()
    번호 = [d['버전'] for d in 자료실목록() if d['만든날'] == 오늘]
    return (max(번호) + 1) if 번호 else 1


def 보고서파일이름(연도, 월끝, 버전, 오늘=None):
    오늘 = 오늘 or 오늘날짜()
    return (f'OTC실적분석_{연도 % 100}년 {월끝}월 누적_클로드작성_'
            f'{오늘:%Y.%m.%d}_v{버전}.xlsx')


def _달표시(d):
    """자료실 목록에 「6월 / 7월」처럼 그 파일이 다루는 실적 월을 보여 줍니다."""
    return f'{d["연월"][1]}월' if d.get('연월') else '—'


def _만든날표시(d):
    return f'{d["만든날"]:%Y.%m.%d}' if d.get('만든날') else '—'

# ══════════════════════════════════════════════════════════════
# 10. 연결패키지 만들기 — 원장(퀵북 BS·IS) + 지난번 패키지 양식 → 일곱 장짜리 엑셀
#     BS · IS · CE · MC · CF · CF정산서 · 결산조정분개 (수식 없이 값만)
# ══════════════════════════════════════════════════════════════
def 패_풀기(s):
    return (s.replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
            .replace('&apos;', "'").replace('&amp;', '&'))


def 패_엑스(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def 패_공유(z):
    try:
        sx = z.read('xl/sharedStrings.xml').decode('utf-8')
    except KeyError:
        return []
    return [패_풀기(re.sub(r'<[^>]+>', '', m.group(1)))
            for m in re.finditer(r'<si>(.*?)</si>', sx, re.S)]


def 패_시트지도(z):
    wb = z.read('xl/workbook.xml').decode('utf-8')
    rel = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid = {m.group(1): m.group(2) for m in
           re.finditer(r'<Relationship Id="([^"]+)"[^>]*Target="([^"]+)"', rel)}
    지도 = {}
    for m in re.finditer(r'<sheet\b([^>]*?)/>', wb):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        지도[패_풀기(str(a.get('name')))] = ('xl/' + rid.get(a.get('r:id'), '').lstrip('/'),
                                           a.get('r:id'))
    return 지도


def 패_칸읽기(z, 파일, 공유):
    """{행: {열: ('값'|'수식', 내용)}}

    ※ 엑셀은 같은 모양의 수식을 「공유 수식」으로 한 번만 적어 두고 나머지 칸은
      번호(si)만 남깁니다. 그대로 읽으면 수식이 빈 글자로 보여서, 대표 수식을 찾아
      줄 번호만 바꿔 채워 넣습니다.
    """
    x = z.read(파일).decode('utf-8')
    if '<sheetData>' not in x:
        return {}
    본문 = x.split('<sheetData>', 1)[1].rsplit('</sheetData>', 1)[0]
    대표 = {}
    for m in re.finditer(r'<c r="([A-Z]+)(\d+)"[^>]*?[^/]>\s*<f([^>]*?[^/])>(.*?)</f>',
                         본문, re.S):
        si = re.search(r'si="(\d+)"', m.group(3))
        if si and 'shared' in m.group(3) and m.group(4).strip():
            대표.setdefault(si.group(1), (int(m.group(2)), 패_풀기(m.group(4))))

    def _옮기기(글, 원행, 새행):
        if 원행 == 새행:
            return 글
        return re.sub(r'(\$?[A-Z]{1,3}\$?)' + str(원행) + r'(?![0-9])',
                      lambda mm: mm.group(1) + str(새행)
                      if '$' not in mm.group(1)[-1:] else mm.group(0), 글)

    표 = {}
    for rm in re.finditer(r'<row[^>]*?\br="(\d+)"[^>]*?(?:/>|>(.*?)</row>)', 본문, re.S):
        r = int(rm.group(1))
        칸 = {}
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>(.*?)</c>)',
                              rm.group(2) or '', re.S):
            c, 속, 몸 = cm.group(1), cm.group(2), cm.group(3) or ''
            fm = re.search(r'<f([^>]*?)(?:/>|>(.*?)</f>)', 몸, re.S)
            if fm:
                글 = 패_풀기(fm.group(2)) if fm.group(2) else ''
                if not 글.strip():
                    si = re.search(r'si="(\d+)"', fm.group(1))
                    if si and si.group(1) in 대표:
                        원행, 본 = 대표[si.group(1)]
                        글 = _옮기기(본, 원행, r)
                캐시 = re.search(r'<v>(.*?)</v>', 몸, re.S)
                칸[c] = ('수식', 글, 패_풀기(캐시.group(1)) if 캐시 else None)
                continue
            vm = re.search(r'<v>(.*?)</v>', 몸, re.S)
            if vm:
                v = vm.group(1)
                칸[c] = ('값', 공유[int(v)] if ' t="s"' in 속 and v.isdigit()
                         and int(v) < len(공유) else 패_풀기(v), None)
            elif '<is>' in 몸:
                칸[c] = ('값', 패_풀기(re.sub(r'<[^>]+>', '', 몸)), None)
        if 칸:
            표[r] = 칸
    return 표


def 패_숫자(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def 패_참조풀기(수식, 열, 내행, 있는행):
    """소계 수식을 [(행, +1/−1)] 로 풉니다. SUM(a:b) · SUM(a,b) · +A−B 를 다룹니다."""
    s = 수식.lstrip('=').replace(' ', '')
    쓸 = []

    def _덩어리(m):
        부호 = -1 if m.group(1) == '-' else 1
        속 = m.group(2)
        for a, b in re.findall(rf'{열}(\d+):{열}(\d+)', 속):
            쓸.extend((r, 부호) for r in range(int(a), int(b) + 1))
        for r in re.findall(rf'(?<![A-Z]){열}(\d+)',
                            re.sub(rf'{열}\d+:{열}\d+', '', 속)):
            쓸.append((int(r), 부호))
        return '#'

    남 = re.sub(rf'([+-]?)SUM\(([^()]*)\)', _덩어리, s)
    for m in re.finditer(rf'([+-]?)(?<![A-Z]){열}(\d+)', 남):
        쓸.append((int(m.group(2)), -1 if m.group(1) == '-' else 1))
    본, 본행 = [], set()
    for r, g in 쓸:
        if r != 내행 and r in 있는행 and r not in 본행:
            본.append((r, g))
            본행.add(r)
    return 본 or None



# ── 작은 수식 계산기 ────────────────────────────────────────────
#    이 파일의 BS·IS 시트는 SUM · ROUND · IFERROR · 더하기빼기만 씁니다.
#    (SUMIF 는 「계정과목으로 금액을 찾아오는 줄」이라 따로 다룹니다)
패_참조 = re.compile(r'\$?([A-Z]{1,3})\$?(\d+)')


def 패_펼치기(속, 잡기):
    쓸 = []

    def _열번호(c):
        n = 0
        for ch in c:
            n = n * 26 + (ord(ch) - 64)
        return n

    def _열이름(n):
        이름 = ''
        while n:
            n, 나 = divmod(n - 1, 26)
            이름 = chr(65 + 나) + 이름
        return 이름

    for a1, r1, a2, r2 in re.findall(r'\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)', 속):
        if a1 == a2:                               # 세로 범위 (C8:C20)
            쓸 += [잡기(f'{a1}{r}') for r in range(int(r1), int(r2) + 1)]
        elif r1 == r2:                             # 가로 범위 (C8:N8)
            쓸 += [잡기(f'{_열이름(n)}{r1}')
                   for n in range(_열번호(a1), _열번호(a2) + 1)]
        else:                                      # 네모 범위
            쓸 += [잡기(f'{_열이름(n)}{r}')
                   for n in range(_열번호(a1), _열번호(a2) + 1)
                   for r in range(int(r1), int(r2) + 1)]
    남 = re.sub(r'\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+', '', 속)
    for c, r in 패_참조.findall(남):
        쓸.append(잡기(f'{c}{r}'))
    return '(' + '+'.join(f'({v!r})' for v in 쓸) + ')' if 쓸 else '0'


def 패_계산(수식, 잡기):
    """수식 글자를 숫자로 바꿉니다. 못 풀면 None."""
    F = 수식.lstrip('=').replace(' ', '')
    if 'SUMIF' in F or 'OFFSET' in F or 'INDEX' in F:
        return None
    for _ in range(6):
        새 = re.sub(r'IFERROR\(([^()]*),[^()]*\)', r'(\1)', F)
        새 = re.sub(r'ROUND\(([^()]*),(-?\d+)\)', r'_r((\1),\2)', 새)
        새 = re.sub(r'SUM\(([^()]*)\)', lambda m: 패_펼치기(m.group(1), 잡기), 새)
        새 = re.sub(r'MAX\(([^()]*)\)', r'_mx(\1)', 새)
        새 = re.sub(r'MIN\(([^()]*)\)', r'_mn(\1)', 새)
        새 = re.sub(r'ABS\(([^()]*)\)', r'_ab(\1)', 새)
        if 새 == F:
            break
        F = 새
    if re.search(r'[A-Z]{2,}\(', F):          # 아직 모르는 함수가 남아 있으면 포기
        return None
    F = 패_참조.sub(lambda m: f'({잡기(m.group(1) + m.group(2))!r})', F)
    F = F.replace('=', '==').replace('<==', '<=').replace('>==', '>=').replace('!==', '!=')
    try:
        v = eval(F, {'__builtins__': {}},
                 {'_r': lambda x, n: round(x, n), '_mx': max, '_mn': min, '_ab': abs})
        return float(v)
    except Exception:
        return None


def 패_시트계산(칸들, 덮개=None):
    """시트 한 장의 모든 칸 값을 구합니다. 덮개 = {칸: 값} 은 그 값을 먼저 씁니다."""
    덮개 = 덮개 or {}
    메모, 진행 = {}, set()

    def 잡기(칸):
        if 칸 in 덮개:
            return float(덮개[칸])
        if 칸 in 메모:
            return 메모[칸]
        if 칸 in 진행:                      # 서로 물고 도는 수식은 0 으로 끊습니다
            return 0.0
        m = re.fullmatch(r'([A-Z]{1,3})(\d+)', 칸)
        if not m:
            return 0.0
        셀 = 칸들.get(int(m.group(2)), {}).get(m.group(1))
        if 셀 is None:
            메모[칸] = 0.0
            return 0.0
        if 셀[0] == '값':
            메모[칸] = 패_숫자(셀[1]) or 0.0
            return 메모[칸]
        진행.add(칸)
        v = 패_계산(셀[1], 잡기)
        진행.discard(칸)
        if v is None:                     # 못 푸는 수식은 엑셀이 계산해 둔 값을 씁니다
            v = 패_숫자(셀[2] if len(셀) > 2 else None)
        메모[칸] = 0.0 if v is None else v
        return 메모[칸]

    return 잡기


def 패_퀵북잔액(xls, 시트):
    """퀵북 BS/IS 시트 (A열 계정 · B열 금액) → {계정: 금액}. Total 줄은 뺍니다."""
    d = pd.read_excel(xls, sheet_name=시트, header=None)
    out = {}
    for _i, 줄 in d.iterrows():
        n = 줄.iloc[0]
        if n is None or (isinstance(n, float) and pd.isna(n)):
            continue
        n = str(n).strip()
        if not n or n.lower().startswith('total'):
            continue
        v = pd.to_numeric(pd.Series([줄.iloc[1] if len(줄) > 1 else None]),
                          errors='coerce').fillna(0).iat[0]
        out[n] = out.get(n, 0.0) + float(v)
    return out


# 퀵북 계정 이름이 패키지와 다른 것 (패키지 이름 → 퀵북 이름)
패_계정바꿔치기 = {'Prepayment': 'General Deposit',
                'LOC - LT': 'Loan payable (Long-term)',
                'CITI Bank LOC': 'Citi Bank LOC'}
# 자산 자리에 있지만 실제로는 부채라 퀵북 부호를 뒤집어야 표가 맞는 계정
패_부호뒤집기 = {'GST/HST Payable'}
# 퀵북 한 계정이 패키지에서 이익 줄·손실 줄로 갈라지는 것
패_FX갈라짐 = {'Exchange Gain or Loss':
            ('Gains Foreign Currency Transactions',
             'Losses Foreign Currency Transactions'),
            'Exchange Translation Gain or Loss':
            ('Gain on foreign exchagne translations',
             'Losses on foreign exchagne translations')}
패_남길시트 = ['BS', 'IS', 'CE', 'MC', 'CF', 'CF정산서', '결산조정분개']
패_IS월열 = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
# 맨 앞에 새로 만들어 붙이는 영문 시트
패_영문시트 = ['BS(영문)', 'IS(영문)']
패_꼬리줄 = re.compile(r'(accrual|cash)\s+basis|^\s*\w+day,\s', re.I)


def 패_이름맞추기(n):
    """「Total for Income」 과 「Total Income」 을 같은 줄로 봅니다."""
    n = re.sub(r'\s+', ' ', str(n or '')).strip().lower()
    return re.sub(r'^total for\s+', 'total ', n)


def 패_전기표(z, 지도, 공유, 시트, 열):
    """양식의 전기(작년) 칸을 계정 이름으로 찾아 쓸 수 있게 모읍니다."""
    if 시트 not in 지도:
        return {}
    칸들 = 패_칸읽기(z, 지도[시트][0], 공유)
    잡기 = 패_시트계산(칸들)
    표 = {}
    for r in sorted(칸들):
        a = str(칸들[r].get('A', ('값', ''))[1]).strip()
        if not a or 열 not in 칸들[r]:
            continue
        키 = 패_이름맞추기(a)
        if 키 not in 표:
            표[키] = 잡기(f'{열}{r}')
    return 표


def 패_영문줄(xls, 시트, 전기표):
    """퀵북 시트를 차례 그대로 읽어 (제목들, [(계정, 전기, 당기, 굵게)]) 로."""
    d = pd.read_excel(xls, sheet_name=시트, header=None)
    제목, 줄들 = [], []
    for i, 행 in d.iterrows():
        n = 행.iloc[0]
        n = '' if n is None or (isinstance(n, float) and pd.isna(n)) else str(n).strip()
        v = pd.to_numeric(pd.Series([행.iloc[1] if len(행) > 1 else None]),
                          errors='coerce').iat[0]
        if i < 5:
            if n and n.lower() != 'total':
                제목.append(n)
            continue
        if not n or 패_꼬리줄.search(n):
            continue
        키 = 패_이름맞추기(n)
        당 = None if pd.isna(v) else float(v)
        # 금액이 없는 줄은 구역 머리글이라 전기도 비워 둡니다
        줄들.append((n, None if 당 is None else 전기표.get(키), 당,
                     bool(re.match(r'^total\b', 키))))
    while len(제목) < 3:
        제목.append('')
    return 제목, 줄들


def 패_스타일더하기(sx):
    """서식은 반드시 「맨 뒤」에 덧붙입니다 (앞에 끼우면 번호가 밀려 파일이 깨집니다)."""
    번호 = 900
    while f'numFmtId="{번호}"' in sx:
        번호 += 1
    금액꼴 = f'<numFmt numFmtId="{번호}" formatCode="#,##0.00;\\(#,##0.00\\)"/>'
    if '<numFmts' in sx:
        sx = re.sub(r'<numFmts count="(\d+)">',
                    lambda m: f'<numFmts count="{int(m.group(1)) + 1}">', sx, count=1)
        sx = sx.replace('</numFmts>', 금액꼴 + '</numFmts>', 1)
    else:
        sx = sx.replace('<fonts', f'<numFmts count="1">{금액꼴}</numFmts><fonts', 1)

    m = re.search(r'<fonts count="(\d+)"[^>]*>(<font>.*?</font>)', sx, re.S)
    글꼴수 = int(m.group(1))
    굵은글꼴 = 글꼴수
    첫글꼴 = m.group(2)
    sx = re.sub(r'<fonts count="(\d+)"',
                lambda mm: f'<fonts count="{int(mm.group(1)) + 1}"', sx, count=1)
    sx = sx.replace('</fonts>', 첫글꼴.replace('<font>', '<font><b/>', 1) + '</fonts>', 1)

    m = re.search(r'<borders count="(\d+)"', sx)
    밑줄 = int(m.group(1))
    sx = re.sub(r'<borders count="(\d+)"',
                lambda mm: f'<borders count="{int(mm.group(1)) + 1}"', sx, count=1)
    sx = sx.replace('</borders>',
                    '<border><left/><right/><top/>'
                    '<bottom style="thin"><color indexed="64"/></bottom>'
                    '<diagonal/></border></borders>', 1)

    m = re.search(r'<cellXfs count="(\d+)"', sx)
    첫칸 = int(m.group(1))
    새것 = [
        f'<xf numFmtId="0" fontId="{굵은글꼴}" fillId="0" borderId="0" xfId="0" '
        f'applyFont="1"/>',                                                   # 제목
        f'<xf numFmtId="0" fontId="{굵은글꼴}" fillId="0" borderId="{밑줄}" xfId="0" '
        f'applyFont="1" applyBorder="1" applyAlignment="1">'
        f'<alignment horizontal="center"/></xf>',                             # 머리글
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>',     # 글자
        f'<xf numFmtId="{번호}" fontId="0" fillId="0" borderId="0" xfId="0" '
        f'applyNumberFormat="1"/>',                                           # 숫자
        f'<xf numFmtId="0" fontId="{굵은글꼴}" fillId="0" borderId="0" xfId="0" '
        f'applyFont="1"/>',                                                   # 굵은 글자
        f'<xf numFmtId="{번호}" fontId="{굵은글꼴}" fillId="0" borderId="0" xfId="0" '
        f'applyNumberFormat="1" applyFont="1"/>',                             # 굵은 숫자
    ]
    sx = re.sub(r'<cellXfs count="(\d+)"',
                lambda mm: f'<cellXfs count="{int(mm.group(1)) + len(새것)}"', sx, count=1)
    sx = sx.replace('</cellXfs>', ''.join(새것) + '</cellXfs>', 1)
    칸번호 = dict(zip(('제목', '머리', '글자', '숫자', '굵은글자', '굵은숫자'),
                      range(첫칸, 첫칸 + len(새것))))
    return sx, 칸번호


def 패_영문시트XML(제목, 줄들, 전기이름, 서식):
    """새 워크시트 한 장을 통째로 만들어 냅니다."""
    def 글(칸, s, sty):
        return (f'<c r="{칸}" s="{sty}" t="inlineStr"><is>'
                f'<t xml:space="preserve">{패_엑스(s)}</t></is></c>')

    def 수(칸, v, sty):
        return f'<c r="{칸}" s="{sty}"><v>{round(float(v), 2)!r}</v></c>'

    줄XML = []
    for i, s in enumerate(제목[:3], 1):
        if s:
            줄XML.append(f'<row r="{i}">{글(f"A{i}", s, 서식["제목"])}</row>')
    머리 = [('A', 'Account'), ('B', 전기이름), ('C', 제목[2] or 'Current'),
            ('D', 'Change')]
    줄XML.append('<row r="5">' + ''.join(글(f'{c}5', s, 서식['머리'])
                                        for c, s in 머리) + '</row>')
    r = 6
    for 이름, 전, 당, 굵게 in 줄들:
        칸 = [글(f'A{r}', 이름, 서식['굵은글자' if 굵게 else '글자'])]
        숫자꼴 = 서식['굵은숫자' if 굵게 else '숫자']
        if 전 is not None:
            칸.append(수(f'B{r}', 전, 숫자꼴))
        if 당 is not None:
            칸.append(수(f'C{r}', 당, 숫자꼴))
        if 전 is not None and 당 is not None:
            칸.append(수(f'D{r}', 당 - 전, 숫자꼴))
        줄XML.append(f'<row r="{r}">' + ''.join(칸) + '</row>')
        r += 1
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main">'
            f'<dimension ref="A1:D{max(r - 1, 6)}"/>'
            '<sheetViews><sheetView workbookViewId="0">'
            '<pane ySplit="5" topLeftCell="A6" activePane="bottomLeft" state="frozen"/>'
            '</sheetView></sheetViews>'
            '<sheetFormatPr defaultRowHeight="16.5"/>'
            '<cols><col min="1" max="1" width="46" customWidth="1"/>'
            '<col min="2" max="4" width="17" customWidth="1"/></cols>'
            '<sheetData>' + ''.join(줄XML) + '</sheetData>'
            '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
            'header="0.3" footer="0.3"/></worksheet>')


def 패_원장월(xls):
    """원장 파일에서 「○○년 N월(BS)」 같은 시트를 찾아 (월, BS시트, IS시트)."""
    bs = is_ = None
    월 = 0
    for s in xls.sheet_names:
        m = re.search(r'(\d+)\s*월\s*\(\s*(BS|IS)\s*\)', str(s), re.I)
        if not m:
            continue
        월 = max(월, int(m.group(1)))
        if m.group(2).upper() == 'BS':
            bs = s
        else:
            is_ = s
    return 월, bs, is_


def 패_연결패키지만들기(원장바이트, 양식바이트):
    """돌려주는 값 : (엑셀 bytes, 안내 dict)"""
    xls = pd.ExcelFile(io.BytesIO(원장바이트))
    보고월, BS시트, IS시트 = 패_원장월(xls)
    if not BS시트 or not IS시트:
        raise ValueError('원장 파일에서 「○월(BS)」·「○월(IS)」 시트를 찾지 못했습니다. '
                         '퀵북에서 그 달 재무상태표·손익계산서를 같이 담아 주세요.')
    QB_BS, QB_IS = 패_퀵북잔액(xls, BS시트), 패_퀵북잔액(xls, IS시트)
    소BS = {k.lower(): v for k, v in QB_BS.items()}
    소IS = {k.lower(): v for k, v in QB_IS.items()}

    z = zipfile.ZipFile(io.BytesIO(양식바이트))
    공유 = 패_공유(z)
    지도 = 패_시트지도(z)
    for 이름 in ('BS-당기', 'IS-당기', 'BS', 'IS'):
        if 이름 not in 지도:
            raise ValueError(
                f'패키지 양식에서 「{이름}」 시트를 찾지 못했습니다. '
                '②번에는 클로드가 만들어 준 파일이 아니라, 본사에서 받은 '
                '「연결 패키지」 원본(BS-당기·IS-당기 시트가 든 파일)을 올려 주세요.')
    안내 = {'보고월': 보고월, '확인': [], '없는계정': [], '새계정': []}

    # ── 1. BS-당기 : 퀵북 잔액을 넣고, 계정과목(국문)별로 모읍니다 ──
    bs당 = 패_칸읽기(z, 지도['BS-당기'][0], 공유)
    덮개, 쓴것 = {}, set()
    for r in sorted(bs당):
        칸 = bs당[r]
        a = str(칸.get('A', ('값', ''))[1]).strip()
        D = 칸.get('D')
        if not a or D is None or D[0] == '수식':
            continue
        원 = 패_계정바꿔치기.get(a, a)
        if 원 in QB_BS:
            v = QB_BS[원]
        elif 원.lower() in 소BS:
            v = 소BS[원.lower()]
        else:
            v = 0.0
            if abs(패_숫자(D[1]) or 0) > 0.005:
                안내['없는계정'].append(('BS', a, 패_숫자(D[1])))
        if 원 in 패_부호뒤집기:
            v = -v
            안내['확인'].append(f'BS 「{a}」 는 퀵북이 {-v:,.2f} 인데 자산 자리에 있는 '
                               f'부채라 {v:,.2f} 로 뒤집어 넣었습니다')
        쓴것.add(원.lower())
        덮개[f'D{r}'] = round(float(v), 2)
    # 이름이 바뀌어 옮겨 간 계정은 원래 자리를 0 으로 둡니다 (두 번 세지 않도록)
    for _패, 원 in 패_계정바꿔치기.items():
        for r in sorted(bs당):
            if str(bs당[r].get('A', ('값', ''))[1]).strip() == 원 and f'D{r}' in 덮개:
                덮개[f'D{r}'] = 0.0
    for n, v in QB_BS.items():
        if n.lower() not in 쓴것 and abs(v) > 0.005:
            안내['새계정'].append(('BS', n, v))
    잡기BS = 패_시트계산(bs당, 덮개)
    BS당기 = {}
    for r in sorted(bs당):
        과목 = str(bs당[r].get('H', ('값', ''))[1]).strip()
        if 과목:
            BS당기[과목] = BS당기.get(과목, 0.0) + 잡기BS(f'B{r}')

    # ── 2. IS-당기 : 이번 달 칸을 채우고, 계정과목(국문)별로 모읍니다 ──
    is당 = 패_칸읽기(z, 지도['IS-당기'][0], 공유)
    첫줄, 이름줄 = {}, {}
    for r in sorted(is당):
        a = str(is당[r].get('A', ('값', ''))[1]).strip()
        if a:
            첫줄.setdefault(a, r)
            이름줄.setdefault(a, []).append(r)

    def 앞달합(r):
        s = 0.0
        for c in 패_IS월열[:max(보고월 - 1, 0)]:
            칸 = is당.get(r, {}).get(c)
            if 칸 and 칸[0] == '값':
                s += 패_숫자(칸[1]) or 0.0
        return s

    갈라진 = {}
    for 퀵, (이익, 손실) in 패_FX갈라짐.items():
        갈라진[이익] = (퀵, '이익')
        갈라진[손실] = (퀵, '손실')

    이번칸 = 패_IS월열[보고월 - 1]
    덮개2 = {}
    for r in sorted(is당):
        a = str(is당[r].get('A', ('값', ''))[1]).strip()
        칸 = is당.get(r, {}).get(이번칸)
        if not a or (칸 is not None and 칸[0] == '수식'):
            continue
        if a in 갈라진:
            퀵, 쪽 = 갈라진[a]
            이i, 손i = 첫줄.get(패_FX갈라짐[퀵][0]), 첫줄.get(패_FX갈라짐[퀵][1])
            누적 = QB_IS.get(퀵, 소IS.get(퀵.lower(), 0.0))
            앞 = 앞달합(손i) - 앞달합(이i)
            늘 = 누적 - 앞
            v = max(늘, 0.0) if 쪽 == '손실' else max(-늘, 0.0)
            if r == 이i:
                안내['확인'].append(
                    f'IS 「{퀵}」 : 퀵북 누적 {누적:,.2f} − 양식의 1~{보고월 - 1}월 순액 '
                    f'{앞:,.2f} = {보고월}월 {늘:,.2f} → '
                    f'{"이익" if 늘 < 0 else "손실"} 줄에 넣었습니다')
        elif a in QB_IS or a.lower() in 소IS:
            if 첫줄.get(a) != r:
                v = 0.0
                안내['확인'].append(f'IS 「{a}」 는 양식에 {len(이름줄[a])}줄이라 위 줄에 모두 넣었습니다')
            else:
                # 같은 이름이 여러 줄이면 그 줄들의 앞달 금액을 모두 빼야 이중계산이 없습니다
                앞 = sum(앞달합(rr) for rr in 이름줄[a])
                v = QB_IS.get(a, 소IS.get(a.lower(), 0.0)) - 앞
        else:
            v = 0.0
        덮개2[f'{이번칸}{r}'] = round(float(v), 2)
    잡기IS = 패_시트계산(is당, 덮개2)
    IS당기 = {}
    for r in sorted(is당):
        과목 = str(is당[r].get('S', ('값', ''))[1]).strip()
        if 과목:
            IS당기[과목] = IS당기.get(과목, 0.0) + 잡기IS(f'R{r}')

    # ── 3. BS · IS 표 계산 (계정과목으로 금액을 찾아오는 줄만 넣어 주면 나머지는 수식대로) ──
    표값 = {}
    for 시트, 당기맵 in (('BS', BS당기), ('IS', IS당기)):
        칸들 = 패_칸읽기(z, 지도[시트][0], 공유)
        덮개3 = {}
        for r in sorted(칸들):
            국 = str(칸들[r].get('E', ('값', ''))[1]).strip()
            셀 = 칸들[r].get('I')
            if 국 and 셀 is not None and 셀[0] == '수식' and 'SUMIF' in 셀[1]:
                덮개3[f'I{r}'] = round(float(당기맵.get(국, 0.0)), 2)
        잡기 = 패_시트계산(칸들, 덮개3)
        값 = {}
        for r in sorted(칸들):
            for c in ('H', 'I', 'L', 'M'):
                셀 = 칸들[r].get(c)
                있음 = f'{c}{r}' in 덮개3 or (셀 is not None and 셀[0] == '수식')
                if not 있음:
                    continue
                # 글자를 내놓는 수식(머리글 등)은 건드리지 않습니다
                if (셀 is not None and 셀[0] == '수식' and len(셀) > 2
                        and 셀[2] is not None and 패_숫자(셀[2]) is None):
                    continue
                v = 잡기(f'{c}{r}')
                값[f'{c}{r}'] = round(v, 0 if c in ('H', 'M') else 2)
        표값[시트] = 값

    # ── 3-1. 표지의 기간·환율 이름표를 이번 달로 맞춥니다 ──────
    해 = datetime.date.today().year
    m5 = re.search(r'(20\d{2})', str(패_칸읽기(z, 지도['BS'][0], 공유)
                                     .get(2, {}).get('B', ('값', ''))[1]))
    if m5:
        해 = int(m5.group(1))
    다음달 = datetime.date(해 + (보고월 == 12), 보고월 % 12 + 1, 1)
    월말 = 다음달 - datetime.timedelta(days=1)
    글자값 = {
        'BS': {'B2': f'제? (당) 기 : {월말:%Y년 %m월 %d일}',
               'L5': f'{해 % 100}년 {보고월}월말 환율'},
        'IS': {'B2': f'제?(당)기 : {해}.01.01~{월말:%Y.%m.%d}',
               'L5': f"'{해 % 100}년 평균환율(01월~{보고월}월)"},
    }
    안내['확인'].append(
        f'환율칸(BS M5 · IS M5)은 양식에 있던 값 그대로입니다 — '
        f'{보고월}월말 환율과 1~{보고월}월 평균환율은 직접 넣어 주세요')

    # ── 4. 맨 앞에 붙일 영문 BS · IS (전년도 + 원장상 퀵북 시트) ──
    전기BS = 패_전기표(z, 지도, 공유, 'BS-당기', 'E')
    전기IS = 패_전기표(z, 지도, 공유, 'IS(QB)-기초 검증', 'B')
    영문 = []
    for 시트이름, 퀵북시트, 전기표, 전기말 in (
            (패_영문시트[0], BS시트, 전기BS, 'As of Dec 31, {}'),
            (패_영문시트[1], IS시트, 전기IS, 'January - December {}')):
        제목, 줄들 = 패_영문줄(xls, 퀵북시트, 전기표)
        해 = re.search(r'(20\d{2})', 제목[2] or '')
        해 = int(해.group(1)) if 해 else datetime.date.today().year
        영문.append((시트이름, 제목, 줄들, 전기말.format(해 - 1)))
    안내['영문줄수'] = {n: len(줄) for n, _제, 줄, _전 in 영문}

    # ── 5. 엑셀 만들기 (영문 두 장 + 원래 일곱 장, 수식은 값으로) ──
    결과 = 패_추리기(z, 지도, 표값, 보고월, 영문, 글자값)
    안내['BS계정수'] = len(BS당기)
    안내['IS계정수'] = len(IS당기)
    return 결과, 안내


def 패_추리기(z, 지도, 표값, 보고월, 영문=(), 글자값=None):
    """필요한 일곱 장만 남기고, 수식은 계산된 값으로 바꿔 넣습니다.
       영문 = [(시트이름, 제목들, 줄들, 전기 이름표)] — 맨 앞에 새로 만들어 붙입니다."""
    공유 = 패_공유(z)
    남은시트 = [n for n in 패_남길시트 if n in 지도]
    남길부품, 새시트 = set(), {}
    for 이름 in 패_남길시트:
        if 이름 not in 지도:
            continue
        경로 = 지도[이름][0]
        원xml = z.read(경로).decode('utf-8')
        수식표 = {f'{c}{r}': 셀[1]
                  for r, 줄 in 패_칸읽기(z, 경로, 공유).items()
                  for c, 셀 in 줄.items() if 셀[0] == '수식'}
        새시트[경로] = 패_글자칸(
            패_수식빼기(원xml, 공유, 표값.get(이름, {}), 수식표, 남은시트),
            (글자값 or {}).get(이름, {}))
        남길부품.add(경로)
        rels = 경로.replace('xl/worksheets/', 'xl/worksheets/_rels/') + '.rels'
        if rels in z.namelist():
            남길부품.add(rels)
            for m in re.finditer(r'Target="([^"]+)"', z.read(rels).decode('utf-8')):
                남길부품.add('xl/' + m.group(1).replace('../', '').lstrip('/'))

    빼는것 = set()
    for n, (경로, _r) in 지도.items():
        if n in 패_남길시트:
            continue
        빼는것.add(경로)
        rels = 경로.replace('xl/worksheets/', 'xl/worksheets/_rels/') + '.rels'
        빼는것.add(rels)
        if rels in z.namelist():
            for m in re.finditer(r'Target="([^"]+)"', z.read(rels).decode('utf-8')):
                빼는것.add('xl/' + m.group(1).replace('../', '').lstrip('/'))
    빼는것 -= 남길부품

    wb = z.read('xl/workbook.xml').decode('utf-8')
    쓸rid = set()

    def _시트줄(m):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        if 패_풀기(str(a.get('name'))) not in 패_남길시트:
            return ''
        쓸rid.add(a.get('r:id'))
        return re.sub(r'\s*state="\w+"', '', m.group(0))

    쓰던시트id = [int(x) for x in re.findall(r'<sheet\b[^>]*sheetId="(\d+)"',
                                             z.read('xl/workbook.xml').decode('utf-8'))]
    wb = re.sub(r'<sheet\b([^>]*?)/>', _시트줄, wb)
    wb = re.sub(r'<definedNames>.*?</definedNames>', '', wb, flags=re.S)
    # 수식이 살아 있으므로 엑셀이 열릴 때 스스로 다시 계산하도록 표시해 둡니다
    wb = re.sub(r'<calcPr[^>]*/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>',
                wb, count=1)
    wb = re.sub(r'\sactiveTab="\d+"', ' activeTab="0"', wb)

    # 시트 차례는 언제나 BS → IS → CE → MC → CF → CF정산서 → 결산조정분개
    def _차례(m):
        줄들 = re.findall(r'<sheet\b[^>]*?/>', m.group(1))

        def 순서(s):
            a = dict(re.findall(r'([\w:]+)="([^"]*)"', s))
            이름 = 패_풀기(str(a.get('name')))
            return 패_남길시트.index(이름) if 이름 in 패_남길시트 else len(패_남길시트)
        return '<sheets>' + ''.join(sorted(줄들, key=순서)) + '</sheets>'

    wb = re.sub(r'<sheets>(.*?)</sheets>', _차례, wb, count=1, flags=re.S)

    rel = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

    def _관계줄(m):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        t = 'xl/' + a.get('Target', '').replace('../', '').lstrip('/')
        if t.startswith('xl/worksheets/') and a.get('Id') not in 쓸rid:
            return ''
        return '' if t in 빼는것 else m.group(0)

    rel = re.sub(r'<Relationship\b([^>]*?)/>', _관계줄, rel)
    ct = z.read('[Content_Types].xml').decode('utf-8')
    ct = re.sub(r'<Override PartName="([^"]+)"[^>]*/>',
                lambda m: '' if (m.group(1).lstrip('/') in 빼는것
                                 or m.group(1).lstrip('/') == 'xl/calcChain.xml')
                else m.group(0), ct)

    # ── 영문 두 장을 새로 만들어 맨 앞에 끼웁니다 ──────────────
    붙임, 스타일 = {}, None
    if 영문:
        스타일xml = z.read('xl/styles.xml').decode('utf-8')
        스타일xml, 서식 = 패_스타일더하기(스타일xml)
        스타일 = 스타일xml
        있는이름 = set(z.namelist())
        쓰던rid = set(re.findall(r'Id="(rId\d+)"', rel))
        시트id = max(쓰던시트id or [0])
        머리, 꼬리 = [], []
        for k, (시트이름, 제목, 줄들, 전기이름) in enumerate(영문, 1):
            부품 = f'xl/worksheets/sheetEng{k}.xml'
            while 부품 in 있는이름:
                부품 = 부품.replace('.xml', 'x.xml')
            있는이름.add(부품)
            rid = f'rIdEng{k}'
            while rid in 쓰던rid:
                rid += 'x'
            쓰던rid.add(rid)
            시트id += 1
            붙임[부품] = 패_영문시트XML(제목, 줄들, 전기이름, 서식)
            머리.append(f'<sheet name="{패_엑스(시트이름)}" sheetId="{시트id}" '
                        f'r:id="{rid}"/>')
            꼬리.append(f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats'
                        f'.org/officeDocument/2006/relationships/worksheet" '
                        f'Target="worksheets/{부품.split("/")[-1]}"/>')
            ct = ct.replace('</Types>',
                            f'<Override PartName="/{부품}" ContentType="application/'
                            f'vnd.openxmlformats-officedocument.spreadsheetml.'
                            f'worksheet+xml"/></Types>', 1)
        wb = wb.replace('<sheets>', '<sheets>' + ''.join(머리), 1)
        rel = rel.replace('</Relationships>', ''.join(꼬리) + '</Relationships>', 1)

    buf = io.BytesIO()
    out = zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED, compresslevel=6)
    for it in z.infolist():
        n = it.filename
        if n in 빼는것 or n == 'xl/calcChain.xml':
            continue
        if n == 'xl/workbook.xml':
            데이터 = wb.encode('utf-8')
        elif n == 'xl/_rels/workbook.xml.rels':
            데이터 = rel.encode('utf-8')
        elif n == '[Content_Types].xml':
            데이터 = ct.encode('utf-8')
        elif n == 'xl/styles.xml' and 스타일 is not None:
            데이터 = 스타일.encode('utf-8')
        elif n in 새시트:
            데이터 = 새시트[n].encode('utf-8')
        else:
            데이터 = z.read(n)
        out.writestr(it, 데이터)
    for 부품, xml in 붙임.items():
        out.writestr(부품, xml.encode('utf-8'))
    out.close()
    return buf.getvalue()


def 패_글자칸(시트xml, 새글):
    """머리글 같은 글자칸을 갈아 끼웁니다 (서식 번호는 그대로 둡니다)."""
    if not 새글:
        return 시트xml
    앞, 나머지 = 시트xml.split('<sheetData>', 1)
    본문, 뒤 = 나머지.rsplit('</sheetData>', 1)

    def _고치기(m):
        칸, 속 = m.group(1), m.group(2)
        if 칸 not in 새글:
            return m.group(0)
        s = re.search(r'\ss="(\d+)"', 속)
        s = f' s="{s.group(1)}"' if s else ''
        return (f'<c r="{칸}"{s} t="inlineStr"><is>'
                f'<t xml:space="preserve">{패_엑스(새글[칸])}</t></is></c>')

    본문 = re.sub(r'<c r="([A-Z]+\d+)"([^>]*?)(?:/>|>(.*?)</c>)', _고치기, 본문, flags=re.S)
    return 앞 + '<sheetData>' + 본문 + '</sheetData>' + 뒤


패_시트참조 = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_가-힣.]+))\s*!")


def 패_바깥본다(수식, 남은시트):
    """수식이 「빼 버리는 시트」를 보고 있으면 True (그런 수식만 값으로 바꿉니다)."""
    for a, b in 패_시트참조.findall(수식 or ''):
        이름 = (a or b).strip()
        if 이름 and 이름 not in 남은시트:
            return True
    return False


def 패_수식빼기(시트xml, 공유, 새값, 수식표=None, 남은시트=()):
    """빼는 시트를 보던 수식만 값으로 바꾸고, 시트 안에서 도는 수식은 살려 둡니다.

      · 수식표 = 공유수식까지 풀어 놓은 {칸: 수식}
      · 새값   = 원장으로 다시 계산한 금액 (수식이 살아남으면 「계산해 둔 값」으로 넣습니다)
    """
    수식표 = 수식표 or {}
    남은 = set(남은시트)
    앞, 나머지 = 시트xml.split('<sheetData>', 1)
    본문, 뒤 = 나머지.rsplit('</sheetData>', 1)

    def _고치기(m):
        칸, 속, 몸 = m.group(1), m.group(2), m.group(3) or ''
        if '<f' not in 몸:
            return m.group(0)
        s = re.search(r'\ss="(\d+)"', 속)
        s = f' s="{s.group(1)}"' if s else ''
        t = re.search(r'\st="(\w+)"', 속)
        t = t.group(1) if t else ''
        vm = re.search(r'<v>(.*?)</v>', 몸, re.S)
        수식 = (수식표.get(칸) or '').strip()
        # TRUE/FALSE·글자를 내놓는 수식은 새로 계산한 숫자를 덮어쓰지 않습니다
        새 = 새값.get(칸) if (칸 in 새값 and t not in ('s', 'str', 'e', 'b')) else None

        if 수식 and not 패_바깥본다(수식, 남은):
            식 = f'<f>{패_엑스(수식.lstrip("="))}</f>'
            if 새 is not None:
                return f'<c r="{칸}"{s}>{식}<v>{float(새)!r}</v></c>'
            if vm is None:
                return f'<c r="{칸}"{s}>{식}</c>'
            꼬리 = f' t="{t}"' if t else ''
            return f'<c r="{칸}"{s}{꼬리}>{식}<v>{vm.group(1)}</v></c>'

        if 새 is not None:
            return f'<c r="{칸}"{s}><v>{float(새)!r}</v></c>'
        if vm is None:
            return f'<c r="{칸}"{s}/>'
        v = vm.group(1)
        if t == 's' and v.isdigit() and int(v) < len(공유):
            v = 공유[int(v)]
        elif t in ('str',):
            v = 패_풀기(v)
        elif t == 'e':
            return f'<c r="{칸}"{s} t="e"><v>{패_엑스(패_풀기(v))}</v></c>'
        elif t == 'b':
            return f'<c r="{칸}"{s} t="b"><v>{v}</v></c>'
        else:
            return f'<c r="{칸}"{s}><v>{v}</v></c>'
        return (f'<c r="{칸}"{s} t="inlineStr">'
                f'<is><t xml:space="preserve">{패_엑스(v)}</t></is></c>')

    본문 = re.sub(r'<c r="([A-Z]+\d+)"([^>]*?)(?:/>|>(.*?)</c>)', _고치기, 본문, flags=re.S)
    return 앞 + '<sheetData>' + 본문 + '</sheetData>' + 뒤



# ── 연결패키지 파일 이름 · 버전 (실적 월마다 v1 부터) ─────────────
패키지패턴 = re.compile(r'연결패키지.*?v\s*(\d+)\s*\.xlsx$', re.I)


def 패키지파일이름(연도, 월, 버전):
    return (f'{연도}. {(월 - 1) // 3 + 1}Q_연결패키지_CTK OTC LAB_'
            f'{월}월말_v{버전}.xlsx')


def 패키지목록():
    """자료실에 쌓인 연결패키지 엑셀을 새 것부터 돌려줍니다."""
    try:
        이름들 = [n for n in os.listdir(자료실) if n.lower().endswith('.xlsx')
                  and not n.startswith('~$') and 패키지패턴.search(n)]
    except OSError:
        return []
    목록 = []
    for n in 이름들:
        경로 = os.path.join(자료실, n)
        try:
            st_ = os.stat(경로)
        except OSError:
            continue
        m = re.search(r'(\d{1,2})\s*월말', n)
        목록.append({'이름': n, '경로': 경로,
                     '버전': int(패키지패턴.search(n).group(1)),
                     '월': int(m.group(1)) if m else None,
                     '시각': st_.st_mtime, '크기': st_.st_size})
    목록.sort(key=lambda d: (d['시각'], d['버전']), reverse=True)
    return 목록


def 패키지다음버전(월):
    """같은 실적 월 안에서만 번호를 이어 갑니다 — 달이 바뀌면 저절로 v1 부터입니다."""
    번호 = [d['버전'] for d in 패키지목록() if d['월'] == 월]
    return (max(번호) + 1) if 번호 else 1


@st.cache_data(max_entries=1, show_spinner='연결패키지를 만드는 중입니다...')
def 연결패키지캐시(원장바이트, 양식바이트):
    return 패_연결패키지만들기(원장바이트, 양식바이트)



양식보관 = 기억자리('최근패키지양식.xlsx')    # 본사에서 받은 연결패키지 원본
패키지보관 = 기억자리('최근패키지.xlsx')      # 마지막으로 만든 연결패키지 재무제표


def 연결패키지양식칸():
    """재무제표 화면 가운데 — 연결패키지 양식만 올립니다 (원장은 올려 둔 것을 씁니다)."""
    st.markdown('**연결패키지 양식** (지난번 본사 패키지)')
    올림 = st.file_uploader('연결패키지 양식', type=['xlsx'], key='pkg_tpl',
                            label_visibility='collapsed',
                            help='BS-당기 · IS-당기 시트가 든 본사 원본 패키지를 올려 주세요')
    자료남기기(양식보관, 올림)
    양식있음, 양식이름, _크기, _때 = 보관자료(양식보관)
    원장있음, 원장이름, _크기2, _때2 = 보관자료(원장보관)
    if 양식있음 and 올림 is None:
        st.caption(f'양식 : 지난번에 올린 **{양식이름}** 을 그대로 씁니다.')
        바이트 = 보관바이트(양식보관)
        if 바이트:
            st.download_button('⤓  올려 둔 양식 내려받기', data=바이트, file_name=양식이름,
                               width='stretch', key='양식받기',
                               mime=('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'))
    if 원장있음:
        st.caption(f'원장 : 「실적보고 엑셀작성」 의 **{원장이름}** 을 그대로 씁니다.')
    else:
        st.caption('⚠ 원장이 아직 없습니다 — 「실적보고 엑셀작성」 에서 ① 원장 원본을 '
                   '한 번 올려 주세요.')


def 연결패키지칸():
    """재무제표 화면 오른쪽 — 연결패키지 엑셀 내려받기 · 만들기."""
    보관 = 패키지목록()
    최신 = 보관[0] if 보관 else None
    st.caption('연결패키지 재무제표')
    if 최신 is not None:
        try:
            with open(최신['경로'], 'rb') as f:
                st.download_button(
                    f"⤓  {최신['월'] or ''}월말 v{최신['버전']} 내려받기", data=f.read(),
                    file_name=최신['이름'], width='stretch', type='primary',
                    mime=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'),
                    help='BS(영문) · IS(영문) · BS · IS · CE · MC · CF · '
                         'CF정산서 · 결산조정분개 아홉 장')
        except OSError:
            st.button('⤓  내려받기', width='stretch', disabled=True)
    elif 보관자료(패키지보관)[0]:
        # 자료실이 비어 있어도(웹사이트가 다시 켜진 뒤) 마지막으로 만든 것은 드립니다
        _있, 팩이름, 팩크기, 팩때 = 보관자료(패키지보관)
        바이트 = 보관바이트(패키지보관)
        if 바이트:
            st.download_button(
                '⤓  마지막 패키지 내려받기', data=바이트, file_name=팩이름,
                width='stretch', type='primary', key='마지막패키지받기',
                mime=('application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'),
                help=f'{팩이름}  ·  {팩크기/1e6:,.1f}MB  ·  '
                     f'{datetime.datetime.fromtimestamp(팩때):%Y.%m.%d %H:%M} 에 만듦')
        else:
            st.button('⤓  내려받기', width='stretch', disabled=True)
    else:
        st.button('⤓  내려받기', width='stretch', disabled=True,
                  help='아직 만든 패키지가 없습니다. 가운데에 양식을 올리고 만들어 주세요')

    원장있음 = 보관자료(원장보관)[0]
    양식있음 = 보관자료(양식보관)[0]
    준비 = 원장있음 and 양식있음

    def _누름():
        st.session_state['패키지생성요청'] = True

    st.button('⚙  연결패키지 재무제표 만들기', width='stretch', disabled=not 준비,
              on_click=_누름, key='pkg_make', type='primary' if 준비 else 'secondary',
              help=('가운데에 연결패키지 양식을 올리면 켜집니다' if 원장있음
                    else '「실적보고 엑셀작성」 에서 원장을 먼저 올려 주세요'))
    if st.session_state.pop('패키지생성요청', False) and 준비:
        try:
            바이트, 안내 = 연결패키지캐시(보관바이트(원장보관), 보관바이트(양식보관))
            월 = 안내['보고월']
            버전 = 패키지다음버전(월)
            이름 = 패키지파일이름(당해연도, 월, 버전)
            저장 = ''
            try:
                os.makedirs(자료실, exist_ok=True)
                with open(os.path.join(자료실, 이름), 'wb') as f:
                    f.write(바이트)
                저장 = os.path.abspath(os.path.join(자료실, 이름))
            except OSError as e:
                st.caption(f'※ 자료실에 저장하지 못했습니다({e})')
            _자리에쓰기(패키지보관, 바이트, 이름)      # 언제든 다시 받을 수 있게
            st.session_state['패키지결과'] = (이름, 바이트, 안내, 저장)
        except Exception as e:
            st.session_state['패키지오류'] = str(e)

    오류 = st.session_state.pop('패키지오류', None)
    if 오류:
        st.caption(f'⚠ 만들지 못했습니다 — {오류}')
    결과 = st.session_state.get('패키지결과')
    if 결과:
        이름, 바이트, 안내, 저장 = 결과
        st.download_button(f'⤓  {이름[:28]}… 내려받기', data=바이트, file_name=이름,
                           width='stretch', key='pkg_dl_new',
                           mime=('application/vnd.openxmlformats-officedocument'
                                 '.spreadsheetml.sheet'))
        st.caption(f"{안내['보고월']}월말 기준 · 계정과목 BS {안내['BS계정수']}개 · "
                   f"IS {안내['IS계정수']}개" + (f' · 저장 {저장}' if 저장 else ''))
        for 말 in 안내.get('확인', [])[:8]:
            st.caption(f'※ {말}')


def 음수표기(글자):
    """마이너스 금액은 회계 방식대로 (괄호) 안에 빨간 글씨로 보여줍니다."""
    s = str(글자)
    if s[:1] in ('-', '−'):
        return f'<span class="neg">({s[1:]})</span>'
    return s


def 금액(v, 자릿수=0, 괄호=True):
    """괄호=False 면 캡션처럼 HTML 을 못 쓰는 곳에서 쓸 수 있게 그냥 글자만 돌려줍니다."""
    if v is None or pd.isna(v):
        return '-'
    if abs(v) < 0.5 / (10 ** 자릿수):        # 반올림하면 0 이 되는 금액은 '-' 로
        return '-'
    t = f'{v:,.{자릿수}f}'
    return 음수표기(t) if 괄호 else t


def 증감HTML(현재, 이전, 라벨='전월대비'):
    if 이전 is None or pd.isna(이전) or 이전 == 0:
        return '<span class="muted">- 비교불가</span>'
    diff = 현재 - 이전
    pct = diff / abs(이전) * 100
    화살표 = '▲' if diff >= 0 else '▼'
    색 = 'up' if diff >= 0 else 'down'
    return f'<span class="{색}">{화살표} {라벨} {abs(pct):,.1f}%</span>'


def 한줄설명(행들):
    """그 거래처에서 무슨 일이 있었는지 20자 안팎으로.

    원장의 한글 분류(활동세부 → 대분류)만 씁니다. 영문 적요를 그대로 내보내면
    브라우저 번역 기능이 엉뚱한 말로 바꿔 버려서, 한글 분류가 없으면 비워 둡니다.
    """
    for 열 in ('활동세부', '활동분류(대분류)'):
        if 열 in 행들.columns:
            값 = 행들[열].dropna().astype(str).str.strip()
            값 = 값[값.ne('') & 값.ne('nan') & 값.ne('0')]
            값 = 값[값.apply(lambda t: any('가' <= c <= '힣' for c in t))]
            if len(값):
                t = 값.value_counts().index[0]
                return t if len(t) <= 22 else t[:20] + '…'
    return ''


def 세부패널(대상, 값컬럼, 부호=1, 상위N=7):
    """항목 하나를 골랐을 때 거래처별 누적 내역을 보여주는 카드 내용(표 + 거래 내용)을 만듭니다.
    매출이든 비용이든 금액이 큰 상위 N곳만 적고 나머지는 「기타」로 한 줄에 모읍니다.
    당월 실적집계·대시보드 화면에서 공통으로 씁니다."""
    if 대상 is None or 대상.empty:
        return '<div class="muted">해당 기간에 거래가 없습니다.</div>'
    집계 = (대상.groupby(대상['Name'].fillna('(거래처 미기재)'))[값컬럼].sum() * 부호)
    집계 = 집계[집계.abs() > 0.5].sort_values(ascending=False)
    if 집계.empty:
        return '<div class="muted">해당 기간에 거래가 없습니다.</div>'
    앞 = 집계.iloc[:상위N]
    기타합 = float(집계.iloc[상위N:].sum())
    총액 = float(집계.sum()) or 1.0
    줄 = ''
    for 이름2, v in 앞.items():
        줄 += (f'<tr><td class="lft notranslate" translate="no">{이름2}</td>'
               f'<td>{금액(v)}</td><td class="muted">{v / 총액 * 100:,.1f}%</td></tr>')
    if abs(기타합) > 0.5:
        줄 += (f'<tr><td class="lft">기타 ({len(집계) - 상위N}곳 합계)</td>'
               f'<td>{금액(기타합)}</td><td class="muted">{기타합 / 총액 * 100:,.1f}%</td></tr>')
    줄 += (f'<tr style="font-weight:700"><td class="lft">합계</td>'
           f'<td>{금액(float(집계.sum()))}</td><td>100.0%</td></tr>')
    # ── 거래 내용 : 거래처마다 「무슨 거래인지」 + 월별 금액 ─────────────
    #    어느 달에 늘었는지 가로로 훑어볼 수 있게, 달을 열로 놓고 옆으로 밀리는 상자에 담습니다.
    이름칸 = 대상['Name'].fillna('(거래처 미기재)')
    월들 = []
    if '월' in 대상.columns:
        월들 = sorted({int(m) for m in pd.to_numeric(대상['월'], errors='coerce').dropna()})

    def 칸값(v):
        return 금액(v) if abs(v) > 0.5 else '<span class="muted">-</span>'

    if 월들:
        피벗 = (대상.assign(_이름=이름칸)
                .pivot_table(index='_이름', columns='월', values=값컬럼,
                             aggfunc='sum', fill_value=0.0)
                .reindex(columns=월들, fill_value=0.0) * 부호)
        머리 = ''.join(f'<th>{m}월</th>' for m in 월들)
        본문 = ''
        for 이름2 in 앞.index:
            묶음 = 대상[이름칸.eq(이름2)]
            구분 = 한줄설명(묶음)
            달값 = [float(피벗.loc[이름2, m]) if 이름2 in 피벗.index else 0.0 for m in 월들]
            본문 += (f'<tr><td class="lft"><b class="notranslate" translate="no">{이름2}</b>'
                     + (f'<span class="cap">{구분}</span>' if 구분 else '')
                     + '</td>'
                     + ''.join(f'<td>{칸값(v)}</td>' for v in 달값)
                     + f'<td class="sum">{금액(float(앞[이름2]))}</td></tr>')
        나머지 = [i for i in 피벗.index if i not in set(앞.index)]
        if 나머지 and abs(기타합) > 0.5:
            달값 = [float(피벗.loc[나머지, m].sum()) for m in 월들]
            본문 += (f'<tr><td class="lft"><b>기타 ({len(나머지)}곳 합계)</b></td>'
                     + ''.join(f'<td>{칸값(v)}</td>' for v in 달값)
                     + f'<td class="sum">{금액(기타합)}</td></tr>')
        달합 = [float(피벗[m].sum()) for m in 월들]
        본문 += ('<tr class="tot"><td class="lft"><b>합계</b></td>'
                 + ''.join(f'<td>{칸값(v)}</td>' for v in 달합)
                 + f'<td class="sum">{금액(float(집계.sum()))}</td></tr>')
        설명 = (f'<div class="mscroll"><table class="mtab">'
                f'<thead><tr><th class="lft">거래처</th>{머리}<th class="sum">누적</th></tr></thead>'
                f'<tbody>{본문}</tbody></table></div>')
    else:
        설명 = ''
        for 이름2 in 앞.index:
            묶음 = 대상[이름칸.eq(이름2)]
            설명 += (f'<div class="memo"><b class="notranslate" translate="no">{이름2}</b>'
                     f'<span>{한줄설명(묶음)}</span></div>')

    # 화면이 좁아도 카드 밖으로 잘려 나가지 않도록 스크롤 상자에 담습니다
    return (f'<div class="scrollx"><table class="detail">'
            f'<colgroup><col style="width:47%"><col style="width:32%"><col style="width:21%"></colgroup>'
            f'<thead><tr><th class="lft">거래처</th><th>누적 금액</th>'
            f'<th>비중</th></tr></thead><tbody>{줄}</tbody></table></div>'
            f'<div class="sub" style="margin:14px 0 6px">거래 내용 '
            f'<span class="calc">(달마다 얼마인지 · 옆으로 밀어 보세요)</span></div>{설명}')


# ══════════════════════════════════════════════════════════════
# 2. 사이드바
# ══════════════════════════════════════════════════════════════
if 'menu' not in st.session_state:
    st.session_state['menu'] = '누적 실적보고'

st.sidebar.markdown("""
<div class="brand-box">
    <div class="brand-title">CTK OTC LAB 실적보고</div>
    <div class="brand-sub">GROUP FINANCE VIEW</div>
</div>
""", unsafe_allow_html=True)

메뉴목록 = ['실적보고', '미수채권 관리', '계정과목 상세', '데이터 점검',
            '연결재무제표 패키지', '실적보고 엑셀작성', '월간실적']
# 「실적보고」를 누르면 아래 세 화면이 펼쳐집니다 (맨 처음은 누적 실적보고)
실적하위 = ['누적 실적보고', '당월 실적보고', '월별 실적보고']
# 「연결재무제표 패키지」도 같은 방식으로 네 화면을 거느립니다
패키지하위 = ['재무제표', '주석사항']
# 재무제표 화면 안에서 고르는 세 가지 (자동 번역을 막으려고 영문 약칭을 앞에 답니다)
재무제표탭 = ['BS  재무상태표', 'IS  손익계산서', 'MC  제조원가명세서']
# 연결패키지 BS·IS 시트에 국문·영문 두 가지 이름이 나란히 있어, 화면도 같게 나눕니다
표기구분 = ['국문 (한글 계정과목)', 'English (account names)']

# ── 연결패키지 BS 서식 : (번호, 국문, 영문, 들여쓰기, 더할줄) ──
#    더할줄이 None 이면 계정과목 이름으로 금액을 찾아오는 줄,
#    아니면 [(줄번호, 부호)] 대로 위 줄들을 더하고 빼는 줄입니다.
BS서식 = [
    ('', '자산', '', 3, None),
    ('Ⅰ.', '유동자산', 'Current assets', 0, [(2, 1), (21, 1)]),
    ('(1)', '당좌자산', 'Quick Assets', 1, [(3, 1), (4, 1), (5, 1), (6, 1), (7, 1), (8, 1), (9, 1), (10, 1), (11, 1), (12, 1), (13, 1), (14, 1), (15, 1), (16, 1), (17, 1), (18, 1), (19, 1), (20, 1)]),
    ('1.', '현금및현금성자산', 'Cash and cash equivalents', 2, None),
    ('2.', '단기금융예치금', 'Short-term Financial Instruments', 2, None),
    ('3.', '기타금융자산(유동)', 'Current fair value financial asset', 2, None),
    ('4.', '매출채권', 'Trade Receivable, gross', 2, None),
    ('', '매출채권-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('5.', '단기대여금', 'Short-term loans', 2, None),
    ('', '단기대여금-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('6.', '미수금', 'Other Receivables, gross', 2, None),
    ('', '미수금-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('7.', '미수수익', 'Accrued Income, gross', 2, None),
    ('', '미수수익-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('8.', '선급금', 'Advance payments, gross', 2, None),
    ('9.', '선급비용', 'Prepaid Expense', 2, None),
    ('10.', '당기법인세자산', 'Current tax assets', 2, None),
    ('11.', '부가세대급금', 'Prepaid Value Added Tax', 2, None),
    ('12.', '파생상품금융자산', 'Current derivative financial assets', 2, None),
    ('13.', '단기보증금(유동)', 'Short-term deposits Provided', 2, None),
    ('14.', '매각예정자산', 'disposal groups classified as held for sale', 2, None),
    ('(2)', '재고자산', 'Inventory', 1, [(22, 1), (23, 1), (24, 1), (25, 1), (26, 1), (27, 1), (28, 1), (29, 1), (30, 1)]),
    ('1.', '상품', 'Merchandise, gross', 2, None),
    ('', '평가충당금(상품)', 'Accumulated Impairment loss', 3, None),
    ('2.', '제품', 'Finished goods, gross', 2, None),
    ('', '평가충당금(제품)', 'Accumulated Impairment loss', 3, None),
    ('3.', '재공품', 'Work in progress, gross', 2, None),
    ('', '평가충당금(재공품)', 'Accumulated Impairment loss', 3, None),
    ('4.', '원재료', 'Raw materials, gross', 2, None),
    ('', '평가충당금(원재료)', 'Accumulated Impairment loss', 3, None),
    ('5.', '미착품', 'Goods in transit, gross', 2, None),
    ('Ⅱ.', '비유동자산', 'Non-current assets', 0, [(32, 1), (38, 1), (58, 1), (62, 1), (70, 1), (73, 1), (75, 1)]),
    ('(1)', '투자자산', 'Investments', 1, [(33, 1), (34, 1), (35, 1), (36, 1), (37, 1)]),
    ('1.', '장기금융예치금', 'Llong-term financial instruments', 2, None),
    ('2.', '기타금융자산(비유동)', 'Non-current fair value financial asset', 2, None),
    ('3.', '관계기업투자자산', 'Investments in associates', 2, None),
    ('4.', '종속기업투자자산', 'Investments in subsidiaries', 2, None),
    ('5.', '장기대여금', 'Long-term loans', 2, None),
    ('(2)', '유형자산', 'Property, plant and equipment', 1, [(39, 1), (40, 1), (41, 1), (42, 1), (43, 1), (44, 1), (45, 1), (46, 1), (47, 1), (48, 1), (49, 1), (50, 1), (51, 1), (52, 1), (53, 1), (54, 1), (55, 1), (56, 1), (57, 1)]),
    ('1.', '토지', 'Land, gross', 2, None),
    ('2.', '건물', 'Buildings, gross', 2, None),
    ('', '건물-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('3.', '시설장치', 'facilities', 2, None),
    ('', '시설장치-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('4.', '기계장치', 'Machinery', 2, None),
    ('', '기계장치-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('5.', '차량운반구', 'Vehicles, gross', 2, None),
    ('', '차량운반구-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('6.', '공구와기구', 'Tools, gross', 2, None),
    ('', '공구와기구-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('7.', '비품', 'Office equipment, gross', 2, None),
    ('', '비품-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('8.', '금형', 'Molds, gross', 2, None),
    ('', '금형-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('9.', '건설중인자산', 'Construction in progress, gross', 2, None),
    ('', '정부보조금-건설중인자산', 'Government grants_Construction in progress, gross', 3, None),
    ('10.', '정부보조금-기계장치', 'Government grants_Machinery', 2, None),
    ('', '정부보조금-기계장치-감가상각누계액', 'Accumulated depreciation', 3, None),
    ('(3)', '투자부동산', 'Investment property', 1, [(59, 1), (60, 1), (61, 1)]),
    ('1.', '투자부동산-토지', 'Investment property - land, gross', 2, None),
    ('2.', '투자부동산-건물', 'Investment property - buildings, gross', 2, None),
    ('', '투자부동산-건물감가상각누계액', 'Accumulated depreciation', 3, None),
    ('(4)', '무형자산', 'Intangible assets', 1, [(63, 1), (64, 1), (65, 1), (66, 1), (67, 1), (68, 1), (69, 1)]),
    ('1.', '영업권', 'Goodwill', 2, None),
    ('2.', '특허권', 'Patents', 2, None),
    ('3.', '상표권', 'Trademarks', 2, None),
    ('4.', '의장권', 'Patents-Design', 2, None),
    ('5.', '소프트웨어', 'Software', 2, None),
    ('6.', '기타의무형자산', 'Other intangible assets', 2, None),
    ('7.', '시설이용권', 'Facility usage rights', 2, None),
    ('(5)', '사용권자산', 'Lease assets', 1, [(71, 1), (72, 1)]),
    ('', '사용권자산(비유동)', 'Non-currment lease assets', 3, None),
    ('', '사용권자산-상각누계액', 'Accumulated depreciation', 3, None),
    ('(6)', '순확정급여자산', 'Defined benefit assets', 1, [(74, 1)]),
    ('', '사외적립자산(자산)', 'Fair value of plan assets', 3, None),
    ('(7)', '기타비유동자산', 'Other non-current assets', 1, [(76, 1), (77, 1), (78, 1), (79, 1), (80, 1), (81, 1), (82, 1), (83, 1), (84, 1)]),
    ('1.', '장기보증금', 'Long-term deposits provided', 2, None),
    ('', '보증금현재가치할인차금', 'Present Value Discount', 3, None),
    ('', '장기보증금-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('2.', '장기선급금', 'Advance payments, gross', 2, None),
    ('3.', '장기선급비용', 'Prepaid Expense', 2, None),
    ('4.', '파생상품금융자산(비유동)', 'Non-currment derivative financial assets', 2, None),
    ('5.', '이연법인세자산(비유동)', 'Non-current deferred tax assets', 2, None),
    ('6.', '장기미수수익', 'Long-term accrued Income, gross', 2, None),
    ('', '장기미수수익-대손충당금', 'Allowance for Doubtful Accounts', 3, None),
    ('', '자산총계', 'Total Assets', 3, [(1, 1), (31, 1)]),
    ('', '부채', '', 3, None),
    ('Ⅰ.', '유동부채', 'Current liabilities', 0, [(88, 1), (89, 1), (90, 1), (91, 1), (92, 1), (93, 1), (94, 1), (95, 1), (96, 1), (97, 1), (98, 1), (99, 1), (100, 1), (101, 1), (102, 1), (103, 1)]),
    ('1.', '매입채무', 'Trade Payables, gross', 2, None),
    ('2.', '단기차입금', 'Short term borrowings', 2, None),
    ('3.', '미지급금', 'Other Payables', 2, None),
    ('4.', '미지급비용', 'Accrued expenses', 2, None),
    ('5.', '리스부채(유동)', 'Current finance lease liabilities', 2, None),
    ('6.', '예수금', 'Withholdings banks', 2, None),
    ('7.', '부가세예수금', 'Value Added Tax Withheld', 2, None),
    ('8.', '선수금', 'Advances Customers', 2, None),
    ('9.', '금융보증부채(유동)', 'Current financial guarantee liabilities', 2, None),
    ('10.', '유동성장기부채', 'Current Portion of Long-term Liabilities', 2, None),
    ('11.', '미지급배당금', 'Dividends Payable', 2, None),
    ('12.', '당기법인세부채', 'Current tax liabilities', 2, None),
    ('13.', '당기손익인식금융부채', 'Fair value financial liabilities', 2, None),
    ('14.', '파생상품금융부채(유동)', 'Currment derivative financial liabilities', 2, None),
    ('15.', '전환사채(유동)', '', 2, None),
    ('16.', '하자보수충당부채', 'Provision for Construction Warranties', 2, None),
    ('Ⅱ.', '비유동부채', 'Non-current liabilities', 0, [(105, 1), (106, 1), (107, 1), (108, 1), (109, 1), (110, 1), (111, 1), (112, 1), (113, 1), (114, 1), (115, 1)]),
    ('1.', '장기차입금', 'Long-term borrowings, gross', 2, None),
    ('2.', '임대보증금', 'Guarantee Deposit Rent, gross', 2, None),
    ('3.', '리스부채(비유동)', 'Non-current finance lease liabilities', 2, None),
    ('4.', '확정급여채무', 'Defined benefit obligation', 2, None),
    ('', '사외적립자산', 'Fair value of plan assets', 3, None),
    ('5.', '이연법인세부채', 'Deferred tax liabilities', 2, None),
    ('6.', '장기미지급금', 'Long-term other payables', 2, None),
    ('7.', '기타장기급여채무', 'Long-term benefit obligation', 2, None),
    ('8.', '금융보증부채(비유동)', 'Non-current financial guarantee liabilities', 2, None),
    ('9.', '파생상품금융부채(비유동)', 'Non-currment derivative financial liabilities', 2, None),
    ('10.', '전환사채(비유동)', 'Convertible bonds', 2, None),
    ('', '부채총계', 'Total liabilities', 3, [(87, 1), (104, 1)]),
    ('', '자본', '', 3, None),
    ('Ⅰ.', '자본금', 'Issued capital', 0, [(119, 1), (120, 1)]),
    ('1.', '보통주자본금', 'Issued capital of common stock', 2, None),
    ('2.', '우선주자본금', 'Issued capital of preferred stock', 2, None),
    ('Ⅱ.', '기타불입자본', 'Capital surplus', 0, [(122, 1), (123, 1), (124, 1)]),
    ('1.', '주식발행초과금', 'Share premium', 2, None),
    ('2.', '자기주식처분이익', 'Gains on Sale of Treasury Stock', 2, None),
    ('3.', '기타자본잉여금', 'Other capital surplus', 2, None),
    ('Ⅲ.', '자본조정', 'Capital Adjustments', 0, [(126, 1), (127, 1), (128, 1), (129, 1), (130, 1), (131, 1)]),
    ('1.', '주식할인발행차금', 'Discount on Stock Issuance', 2, None),
    ('2.', '기타자본조정', 'Other capital adjustments', 2, None),
    ('3.', '자기주식', 'Treasury Stock', 2, None),
    ('4.', '주식매수선택권', 'Stock Option', 2, None),
    ('5.', '자기주식처분이익', 'Gains sale treasury stock', 2, None),
    ('6.', '자기주식처분손실', 'Losses sale treasury stock', 2, None),
    ('IV.', '기타포괄손익누계액', 'Other Comprehensive income/loss accumulated amount', 0, [(133, 1), (134, 1), (135, 1), (136, 1), (137, 1), (138, 1)]),
    ('1.', '매도가능증권평가이익', 'Gains on Valuation of available-for-sale financial assets', 2, None),
    ('2.', '매도가능증권평가손실', 'Losses on Valuation of available-for-sale financial assets', 2, None),
    ('3.', '해외사업환산이익', 'Cumulative gains of Foreign operation', 2, None),
    ('4.', '해외사업환산손실', 'Cumulative losses of Foreign operation', 2, None),
    ('5.', '지분법자본변동', 'Gains on financial assets as measured at fair value', 2, None),
    ('6.', '부의지분법자본변동', 'Losses on financial assets as measured at fair value', 2, None),
    ('IV.', '이익잉여금', 'Retained earnings', 0, [(140, 1), (141, 1), (142, 1)]),
    ('1.', '이익준비금', 'Legal Reserve', 2, None),
    ('2.', '미처분이익잉여금', 'Unappropriated Retained Earnings', 2, None),
    ('3.', '확정급여부채재측정요소', 'Reclassification factor on defined benefit plans', 2, None),
    ('IV.', '비지배지분', 'Non-controlling interests', 0, None),
    ('', '자본총계', 'Total Equities', 3, [(118, 1), (121, 1), (125, 1), (132, 1), (139, 1), (143, 1)]),
    ('', '부채및자본총계', 'Total equity and liabilities', 3, [(144, 1), (116, 1)]),
]

# BS 서식에서 「부채」 머리줄 아래에 있는 계정과목 (대변이 늘어나는 쪽)
BS부채자본 = set()
_부채구역 = False
for _번, _국, _영, _수준, _더할 in BS서식:
    if not _영 and _국 in ('부채', '자본'):
        _부채구역 = True
    elif not _영 and _국 == '자산':
        _부채구역 = False
    if _부채구역 and _국:
        BS부채자본.add(_국)

# ── 연결패키지 IS 서식 : (번호, 국문, 영문, 들여쓰기, 더할줄) ──
#    더할줄이 None 이면 계정과목 이름으로 금액을 찾아오는 줄,
#    아니면 [(줄번호, 부호)] 대로 위 줄들을 더하고 빼는 줄입니다.
IS서식 = [
    ('Ⅰ.', '매 출 액', 'Sales', 0, [(1, 1), (2, 1), (3, 1), (4, 1), (5, 1)]),
    ('1.', '상품매출', 'Sales of Merchandise', 2, None),
    ('2.', '제품매출', 'Sales of Finished Goods', 2, None),
    ('3.', '패키지매출', 'Sales of Package', 2, None),
    ('4.', '용역매출', 'Revenues-Services', 2, None),
    ('5.', '기타매출', 'Sales of Other', 2, None),
    ('Ⅱ.', '매 출 원 가', 'Cost of sales', 0, [(7, 1), (8, 1), (9, 1), (10, 1), (11, 1), (12, 1), (13, 1), (14, 1), (15, 1), (16, 1), (17, 1), (18, 1)]),
    ('1.', '상품매출원가', 'Cost of Merchandise Sold', 2, None),
    ('', '기초상품재고액', '', 3, None),
    ('', '당기상품입고액', '', 3, None),
    ('', '기말상품재고액', '', 3, None),
    ('2.', '제품매출원가', 'Cost of Finished Goods Sold', 2, None),
    ('', '기초제품재고액', '', 3, None),
    ('', '당기제품제조원가', '', 3, None),
    ('', '기말제품재고액', '', 3, None),
    ('', '타계정출고액', '', 3, None),
    ('3.', '패키지매출원가', '', 2, None),
    ('4.', '용역매출원가', 'Cost of Service', 2, None),
    ('5.', '기타매출원가', 'Cost of Other', 2, None),
    ('Ⅲ.', '매출총이익', 'Gross Profits', 0, [(0, 1), (6, -1)]),
    ('Ⅳ.', '판매비와관리비', 'Selling and Administrative Expenses', 0, [(21, 1), (22, 1), (23, 1), (24, 1), (25, 1), (26, 1), (27, 1), (28, 1), (29, 1), (30, 1), (31, 1), (32, 1), (33, 1), (34, 1), (35, 1), (36, 1), (37, 1), (38, 1), (39, 1), (40, 1), (41, 1), (42, 1), (43, 1), (44, 1), (45, 1), (46, 1), (47, 1), (48, 1), (49, 1), (50, 1), (51, 1), (52, 1)]),
    ('1.', '급여', 'Salaries wages', 2, None),
    ('2.', '퇴직급여', 'Provision for severance indemnities', 2, None),
    ('3.', '복리후생비', 'Employee benefits', 2, None),
    ('4.', '주식보상비용', 'Stock option expenses', 2, None),
    ('5.', '여비교통비', 'Travel expenses', 2, None),
    ('6.', '접대비', 'Entertainment expenses', 2, None),
    ('7.', '통신비', 'Communication expenses', 2, None),
    ('8.', '수도광열비', 'Utility expenses', 2, None),
    ('9.', '전력비', 'Electricity expenses', 2, None),
    ('10.', '세금과공과', 'Taxes dues', 2, None),
    ('11.', '감가상각비', 'Depreciation expense', 2, None),
    ('12.', '지급임차료', 'Rental expenses', 2, None),
    ('13.', '수선비', 'Repair expenses', 2, None),
    ('14.', '보험료', 'Insurance premiums', 2, None),
    ('15.', '차량유지비', 'Vehicle maintenance expenses', 2, None),
    ('16.', '경상연구개발비', 'Ordinary development expense', 2, None),
    ('17.', '운반비', 'Freight expenses', 2, None),
    ('18.', '교육훈련비', 'Training Expenses', 2, None),
    ('19.', '도서인쇄비', 'Publication expenses', 2, None),
    ('20.', '사무용품비', 'office supply expenses', 2, None),
    ('21.', '소모품비', 'Supply expenses', 2, None),
    ('22.', '지급수수료', 'Commissions & Professional/Outside service Fees', 2, None),
    ('23.', '광고선전비', 'Advertising expenses', 2, None),
    ('24.', '판매촉진비', 'Sales Promotional Expenses', 2, None),
    ('25.', '대손상각비', 'Bad Debt Expenses', 2, None),
    ('26.', '건물관리비', 'Maintenance expenses on buildings', 2, None),
    ('27.', '수출제비용', 'Export expenses', 2, None),
    ('28.', '무형자산상각비', 'Amortisation expense', 2, None),
    ('29.', '견본비', 'Samples Expenses', 2, None),
    ('30.', '해외시장개척비', 'Overseas Marketing Expenses', 2, None),
    ('31.', '하자보수비', 'Warranty Expenses', 2, None),
    ('32.', '잡비', 'Miscellaneous Expenses', 2, None),
    ('Ⅴ.', '영업이익', 'Operating Profits', 0, [(19, 1), (20, -1)]),
    ('Ⅵ.', '기타수익', 'Other Revenues', 0, [(55, 1), (56, 1), (57, 1), (58, 1), (59, 1), (60, 1), (61, 1), (62, 1)]),
    ('1.', '유형자산처분이익', 'Gains on disposals of property, plant and equipment', 2, None),
    ('2.', '사용권자산처분이익', 'Gains on disposals of lease assets', 2, None),
    ('3.', '대손충당금환입', 'Reversal Allowance Doubtful Accounts', 2, None),
    ('4.', '잡이익', 'Miscellaneous Income', 2, None),
    ('5.', '염가매수이익', 'Amortization on negative goodwill', 2, None),
    ('6.', '관계기업투자주식처분이익', 'Gains on sale of investments in associates financial income', 2, None),
    ('7.', '종속기업투자자산처분이익', 'Gains on sale of investments in subsidiaries', 2, None),
    ('8.', '전기오류수정이익', '', 2, None),
    ('Ⅶ.', '기타비용', 'Other Expenses', 0, [(64, 1), (65, 1), (66, 1), (67, 1), (68, 1), (69, 1), (70, 1), (71, 1), (72, 1), (73, 1), (74, 1)]),
    ('1.', '유형자산처분손실', 'Losses on disposals of property, plant and equipment', 2, None),
    ('2.', '유형자산폐기손실', '', 2, None),
    ('3.', '유형자산손상차손', 'Impairment loss of property, plant and equipment', 2, None),
    ('4.', '무형자산처분손실', '', 2, None),
    ('5.', '관계기업투자주식처분손실', '', 2, None),
    ('6.', '종속기업투자자산처분손실', '', 2, None),
    ('7.', '종속기업투자자산손상차손', '', 2, None),
    ('8.', '관계기업투자자산손상차손', '', 2, None),
    ('9.', '기타의대손상각비', 'Other Bad Debt Expenses', 2, None),
    ('10.', '기부금', 'Donations', 2, None),
    ('11.', '잡손실', 'Miscellaneous Losses', 2, None),
    ('Ⅷ.', '금융수익', 'Finance income', 0, [(76, 1), (77, 1), (78, 1), (79, 1), (80, 1), (81, 1), (82, 1), (83, 1), (84, 1), (85, 1)]),
    ('1.', '이자수익', 'Interest income', 2, None),
    ('2.', '배당금수익', 'Dividend Income', 2, None),
    ('3.', '외환차익', 'Gains Foreign Currency Transactions', 2, None),
    ('4.', '외화환산이익', 'Gain on foreign exchagne translations', 2, None),
    ('5.', '보증수익', '', 2, None),
    ('6.', '기타금융자산평가이익', 'Gains on valuation of fair value financial asset', 2, None),
    ('7.', '기타금융자산처분이익', 'Gains on disposal of fair value financial asset', 2, None),
    ('8.', '매도가능증권처분이익', 'Gains on disposal of available-for-sale financial assets', 2, None),
    ('9.', '파생상품평가이익', 'Gains on valuation of derivative financial assets', 2, None),
    ('10.', '파생상품거래이익', 'Gains on disposal of derivative financial assets', 2, None),
    ('Ⅸ.', '금융비용', 'Finance expense', 0, [(87, 1), (88, 1), (89, 1), (90, 1), (91, 1), (92, 1), (93, 1), (94, 1), (95, 1), (96, 1), (97, 1)]),
    ('1.', '이자비용', 'Interest expenses', 2, None),
    ('2.', '외환차손', 'Losses Foreign Currency Transactions', 2, None),
    ('3.', '외화환산손실', 'Losses on foreign exchagne translations', 2, None),
    ('4.', '보증비용', '', 2, None),
    ('5.', '투자자산처분손실', 'Losses on disposal of investmens', 2, None),
    ('6.', '매도가능증권손상차손', 'Losses on valuation of available-for-sale financial assets', 2, None),
    ('7.', '기타금융자산평가손실', 'Losses on valuation of fair value financial asset', 2, None),
    ('8.', '기타금융자산처분손실', 'Losses on disposal of fair value financial asset', 2, None),
    ('9.', '매도가능증권처분손실', 'Losses on disposal of available-for-sale financial assets', 2, None),
    ('10.', '파생상품평가손실', 'Losses on valuation of derivative financial assets', 2, None),
    ('11.', '파생상품거래손실', 'Losses on disposal of derivative financial assets', 2, None),
    ('Ⅹ.', '지분법손익', '', 0, [(99, 1), (100, 1)]),
    ('1', '관계기업투자자산평가이익', '', 2, None),
    ('2', '관계기업투자자산평가손실', '', 2, None),
    ('XI.', '법인세비용차감전 계속사업이익', 'Profit (loss) before tax', 0, [(53, 1), (54, 1), (75, 1), (63, -1), (86, -1), (98, 1)]),
    ('XⅡ.', '법인세비용', 'Income tax expense', 0, None),
    ('ⅩⅢ.', '계속사업이익', 'Profit (loss) from continuing operations', 0, [(101, 1), (102, -1)]),
    ('XⅣ.', '중단사업이익', 'Profit (loss) from discontinued operations', 0, None),
    ('XⅤ.', '당기순이익(손실)', 'Profit (loss)', 0, [(103, 1), (104, -1)]),
    ('XⅥ.', '기타포괄손익', '', 0, None),
    ('가.', '당기손익재분류항목', '', 1, [(108, 1), (109, 1), (110, 1), (111, 1), (112, 1), (113, 1)]),
    ('1.', '기타포괄손익-기타금융자산평가이익', '', 2, None),
    ('2.', '기타포괄손익-기타금융자산평가손실', '', 2, None),
    ('3.', '지분법자본변동', '', 2, None),
    ('4.', '부의지분법자본변동', '', 2, None),
    ('5.', '해외사업환산이익', '', 2, None),
    ('6.', '해외사업환산손실', '', 2, None),
    ('나.', '당기손익비분류항목', '', 1, [(115, 1), (116, 1), (117, 1), (118, 1), (119, 1), (120, 1), (121, 1), (122, 1)]),
    ('1.', '재측정요소-이익', '', 2, None),
    ('2.', '재측정요소-손실', '', 2, None),
    ('3.', '기타포괄손익-기타금융자산평가이익', '', 2, None),
    ('4.', '기타포괄손익-기타금융자산평가손실', '', 2, None),
    ('5.', '지분법자본변동', '', 2, None),
    ('6.', '부의지분법자본변동', '', 2, None),
    ('7.', '해외사업환산이익', '', 2, None),
    ('8.', '해외사업환산손실', '', 2, None),
    ('XⅦ.', '총포괄손익', '', 0, [(105, 1), (107, 1), (114, 1)]),
    ('XⅧ.', '주당손익', '', 0, None),
    ('1.', '기본주당순손익', '', 2, None),
]

접이메뉴 = {'실적보고': 실적하위, '연결재무제표 패키지': 패키지하위}
보고서페이지 = ['1. Executive Summary', '2. P&L | 손익계산서', '3. SG&A | 판관비',
                '4. 운영 KPI', '5. Sales Forecast']
if 'rep_page' not in st.session_state:
    st.session_state['rep_page'] = 보고서페이지[0]

def _메뉴이동(n):
    # 상위 메뉴를 누르면 그 아래 첫 화면으로 갑니다 (실적보고 → 누적 실적보고)
    st.session_state['menu'] = 접이메뉴[n][0] if n in 접이메뉴 else n


def _보고서이동(n):
    st.session_state['menu'] = '월간실적'
    st.session_state['rep_page'] = n


with st.sidebar.container(key='nav_container'):
    st.caption('보기')
    for name in 메뉴목록:
        하위 = 접이메뉴.get(name, [])
        펼침 = bool(하위) and st.session_state['menu'] in 하위
        st.button(name, key=f'nav_{name}', width='stretch',
                  type='primary' if (st.session_state['menu'] == name or 펼침) else 'secondary',
                  on_click=_메뉴이동, args=(name,))
        # 들여쓰기는 글자(공백) 대신 CSS 로 줍니다 — 브라우저 번역기가
        # 앞 공백을 지워도 줄이 어긋나지 않습니다
        if 펼침:
            with st.container(key='pl_sub'):
                for sp in 하위:
                    st.button(sp, key=f'pl_{sp}', width='stretch',
                              type='primary' if st.session_state['menu'] == sp else 'secondary',
                              on_click=_메뉴이동, args=(sp,))
        if name == '월간실적' and st.session_state['menu'] == '월간실적':
            with st.container(key='rep_sub'):
                for sp in 보고서페이지:
                    st.button(sp, key=f'rep_{sp}', width='stretch',
                              type='primary' if st.session_state['rep_page'] == sp else 'secondary',
                              on_click=_보고서이동, args=(sp,))

메뉴 = st.session_state['menu']
st.sidebar.write('---')

CACHE_PATH = 기억자리('최근업로드.xlsx')
이름기록 = 기억자리('최근업로드_파일이름.txt')
자료있음 = os.path.exists(CACHE_PATH)
# 한 번 올리면 그대로 남겨 두는 자료 — 다음에 또 올리지 않아도 됩니다
원장보관 = 기억자리('최근원장.xlsx')          # ① 퀵북 원장 원본
전월보관 = 기억자리('최근전월실적.xlsx')      # ② 지난달 실적보고 엑셀


def 자료남기기(자리, 올림):
    """올려 주신 파일을 그대로 남겨 둡니다 (같은 파일이면 다시 쓰지 않습니다)."""
    if 올림 is None:
        return
    도장 = f'{올림.name}|{올림.size}'
    if st.session_state.get('보관도장_' + 자리) == 도장 and os.path.exists(자리):
        return
    try:
        with open(자리, 'wb') as f:
            f.write(올림.getvalue())
        with open(자리 + '.txt', 'w', encoding='utf-8') as f:
            f.write(올림.name)
        st.session_state['보관도장_' + 자리] = 도장
    except OSError:
        pass


def 보관자료(자리):
    """(있나, 원래 파일이름, 크기, 남긴 때) — 없으면 (False, '', 0, 0)."""
    try:
        s = os.stat(자리)
    except OSError:
        return False, '', 0, 0
    try:
        with open(자리 + '.txt', encoding='utf-8') as f:
            이름 = f.read().strip()
    except OSError:
        이름 = 자리
    return True, 이름, s.st_size, s.st_mtime


def 보관바이트(자리):
    try:
        with open(자리, 'rb') as f:
            return f.read()
    except OSError:
        return None


def 보관원장정보():
    """(있나, 파일이름, 남긴 때) — 예전 이름 그대로 쓰는 곳을 위해 남겨 둡니다."""
    있음, 이름, _크기, 때 = 보관자료(원장보관)
    return 있음, 이름, 때


def _자리에쓰기(자리, 바이트, 이름):
    try:
        with open(자리, 'wb') as f:
            f.write(바이트)
        with open(자리 + '.txt', 'w', encoding='utf-8') as f:
            f.write(이름)
        return True
    except OSError:
        return False


def 자료종류(바이트):
    """파일 안 시트 이름을 보고 어떤 자료인지 스스로 알아냅니다."""
    try:
        시트 = [str(s) for s in pd.ExcelFile(io.BytesIO(바이트)).sheet_names]
    except Exception:
        return ''
    if any('원장_raw' in s for s in 시트):
        return '실적'                       # 실적보고 엑셀 (대시보드가 그리는 자료)
    if any(s.strip() in ('BS-당기', 'IS-당기') for s in 시트):
        return '양식'                       # 본사 연결패키지 원본
    if any(re.search(r'\d+\s*월\s*\(', s) for s in 시트):
        return '원장'                       # 퀵북 원장 원본
    if any(re.search(r'ledger|원장', s, re.I) for s in 시트):
        return '원장'
    return ''


종류이름 = {'실적': '실적 엑셀', '원장': '원장 원본', '양식': '연결패키지 양식'}


def 자료받기(파일들):
    """올려 주신 파일을 안을 보고 제자리에 넣습니다 → (실적 엑셀 파일, [(이름, 종류)])."""
    실적, 목록 = None, []
    for f in 파일들 or []:
        바이트 = f.getvalue()
        종 = 자료종류(바이트)
        if 종 == '실적':
            실적 = f
            _자리에쓰기(전월보관, 바이트, f.name)
        elif 종 == '원장':
            _자리에쓰기(원장보관, 바이트, f.name)
        elif 종 == '양식':
            _자리에쓰기(양식보관, 바이트, f.name)
        목록.append((f.name, 종))
    return 실적, 목록


def 쓰던파일이름():
    """지금 화면이 쓰고 있는 실적 엑셀의 원래 파일 이름 (마지막으로 올린 것)."""
    try:
        with open(이름기록, encoding='utf-8') as f:
            return f.read().strip()
    except OSError:
        return ''

# 설정은 평소엔 접어 두고, 필요할 때만 펴서 씁니다 (사이드바를 깔끔하게).
# ※ 올려 둔 원장이 없으면(웹사이트가 다시 켜지면 지워집니다) 저절로 펼쳐서 바로 올릴 수 있게 합니다.
설정칸 = st.sidebar.expander('⚙  설정 · 파일 업로드', expanded=not 자료있음)
설정칸.caption('화면 밝기')
설정칸.radio('화면 밝기', ['밝은 화면', '어두운 화면'], horizontal=True,
             key='ui_theme', label_visibility='collapsed')

설정칸.caption('실적 엑셀 · 원장 · 패키지 양식을 한꺼번에 올리셔도 됩니다 — '
               '파일 안을 보고 알아서 제자리에 넣습니다.')
올린것들 = 설정칸.file_uploader('자료 올리기 (.xlsx · 여러 개 가능)', type=['xlsx'],
                                accept_multiple_files=True, key='설정업로드')
# 방금 올리신 것인지 확인해 둡니다 (이름+크기) — 새 자료일 때만 화면을 딱 한 번 새로 그립니다.
_올린도장 = tuple(sorted((str(f.name), int(getattr(f, 'size', 0) or 0))
                         for f in (올린것들 or [])))
_새로올림 = bool(올린것들) and st.session_state.get('_올린자료도장') != _올린도장
if not 올린것들:                       # 목록을 비우시면 다음에 같은 파일을 올려도 다시 읽습니다
    st.session_state.pop('_올린자료도장', None)

업로드파일, 가른것 = 자료받기(올린것들)

# 실적 엑셀은 어느 화면에서 올리셔도 곧바로 보관해 둡니다
# (실적보고·연결패키지 화면은 아래 본문까지 내려가지 않고 끝나기 때문입니다).
if 업로드파일 is not None:
    try:
        with open(CACHE_PATH, 'wb') as f:
            f.write(업로드파일.getvalue())
        with open(이름기록, 'w', encoding='utf-8') as f:
            f.write(업로드파일.name)
    except OSError:
        pass

for _이름, _종 in 가른것:
    설정칸.caption(f'{"✔" if _종 else "?"} {_이름[:34]} → '
                   + 종류이름.get(_종, '무슨 자료인지 모르겠습니다'))

if _새로올림:
    st.session_state['_올린자료도장'] = _올린도장
    st.rerun()          # 올리자마자 새 자료로 화면을 다시 그립니다

# ※ 사이드바의 「이 대시보드 공유」 칸은 없앴습니다.
#   공유는 화면 오른쪽 아래 Streamlit 의 「Manage app」 → ⋮ → Settings → Sharing 에서 합니다.
#   그 단추는 웹사이트(streamlit.app)가 우리 화면 바깥에 직접 그리는 것이라
#   app.py 에서는 손댈 수 없습니다 — 브라우저에서 Streamlit 에 로그인해야 나타납니다.

# ══════════════════════════════════════════════════════════════
# 실적보고 엑셀작성 — 원장이 아직 없어도 열 수 있어야 하므로 여기서 처리합니다
# ══════════════════════════════════════════════════════════════
if 메뉴 == '실적보고 엑셀작성':
    st.html("""<div class="wrap" translate="no" style="margin-bottom:0">
  <div class="page-head"><span class="t">실적보고 엑셀작성</span>
    <span class="pill">원장 + 전월 자료 → 당월 실적보고</span></div>
</div>""")
    st.html(f"""<div class="wrap" translate="no">
  <div style="font-size:14.5px; line-height:1.9; color:{T['ink2']}; margin:2px 0 14px">
    <b style="color:{T['ink']}">①</b> QuickBooks 원장 원본과
    <b style="color:{T['ink']}">②</b> 지난달 실적보고 엑셀을 올리고
    <b style="color:{T['ink']}">⚙ 생성</b> 을 누르면 이번 달 실적보고 엑셀이 만들어집니다.<br>
    올려주신 파일을 <b style="color:{T['ink']}">그대로 두고 「원장_raw」 시트만 새 원장으로 갈아끼우므로</b>,
    피벗·차트·나머지 시트가 손상 없이 남고 당월·월별 실적집계는 엑셀이 열릴 때 스스로 다시 계산합니다.
  </div>
</div>""")

    작성좌, 작성우 = st.columns(2, gap='medium')
    with 작성좌:
        st.markdown('**① 원장 원본** (QuickBooks General Ledger)')
        원장올림 = st.file_uploader('원장 원본', type=['xlsx'], key='gen_gl',
                                    label_visibility='collapsed')
    with 작성우:
        st.markdown('**② 전월 실적보고자료** (지난달 실적보고 엑셀)')
        전월올림 = st.file_uploader('전월 실적보고자료', type=['xlsx'], key='gen_prev',
                                    label_visibility='collapsed')
    # 올려 주신 파일은 그대로 남겨 두어, 다음에는 다시 올리지 않아도 됩니다
    자료남기기(원장보관, 원장올림)
    자료남기기(전월보관, 전월올림)
    원장있음, 원장이름, 원장크기, _t1 = 보관자료(원장보관)
    전월있음, 전월이름, 전월크기, _t2 = 보관자료(전월보관)

    보관목록 = 자료실목록()
    최신 = 보관목록[0] if 보관목록 else None
    둘다 = 원장있음 and 전월있음

    # ── 어떤 자료로 만드는지 눈으로 확인할 수 있게 보여 줍니다
    def _올린줄(번호, 파일, 설명, 있음, 이름, 크기):
        if 파일 is None and not 있음:
            return (f'<tr class="sub"><td class="name">{번호} {설명}</td>'
                    f'<td style="color:{T["ink3"]}">아직 올리지 않음</td></tr>')
        새로 = 파일 is not None
        이름 = 파일.name if 새로 else 이름
        크기 = 파일.size if 새로 else 크기
        표 = ('' if 새로 else
              f'<span style="color:{T["accent"]}; font-weight:700">  ·  지난번에 올린 자료</span>')
        return (f'<tr class="sub"><td class="name">{번호} {설명}</td>'
                f'<td class="lft" style="font-weight:700">{이름}'
                f'<span style="color:{T["ink3"]}; font-weight:400">'
                f'  ·  {크기/1e6:,.1f}MB</span>{표}</td></tr>')

    st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="margin-top:6px">
    <h3>이 자료로 만듭니다</h3>
    <div class="sub">한 번 올린 파일은 그대로 기억합니다 — 바뀐 것만 새로 올리시면 됩니다</div>
    <table class="lined" style="margin-top:8px">
      <colgroup><col style="width:32%"><col style="width:68%"></colgroup>
      <tbody>
        {_올린줄('①', 원장올림, '원장 원본', 원장있음, 원장이름, 원장크기)}
        {_올린줄('②', 전월올림, '전월 실적보고자료', 전월있음, 전월이름, 전월크기)}
      </tbody>
    </table>
  </div>
</div>""")

    # ── 기억해 둔 자료는 언제든 다시 내려받을 수 있게 합니다
    보관줄 = [('② 전월 실적보고자료', 전월보관), ('① 원장 원본', 원장보관),
              ('연결패키지 양식', 양식보관)]
    있는것 = [(라벨, 자리) for 라벨, 자리 in 보관줄 if 보관자료(자리)[0]]
    if 있는것:
        st.caption('보관 중인 자료 내려받기 — 언제 올리셨든 마지막 것을 그대로 드립니다')
        받는칸 = st.columns(len(있는것) + max(0, 3 - len(있는것)), gap='small')
        for 칸, (라벨, 자리) in zip(받는칸, 있는것):
            _있, 이름, 크기, 때 = 보관자료(자리)
            바이트 = 보관바이트(자리)
            if 바이트 is None:
                continue
            칸.download_button(
                f'⤓  {라벨}', data=바이트, file_name=이름 or os.path.basename(자리),
                width='stretch', key='보관받기_' + 자리,
                mime=('application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'),
                help=f'{이름}  ·  {크기/1e6:,.1f}MB  ·  '
                     f'{datetime.datetime.fromtimestamp(때):%Y.%m.%d %H:%M} 에 올림')

    # ── 생성 · 내려받기
    def _생성누름():
        st.session_state['엑셀생성요청'] = True

    만들칸, 받을칸, 설명칸2 = st.columns([1, 1, 2], gap='small')
    with 만들칸:
        st.button('⚙  생성', width='stretch', type='primary', key='보고서생성',
                  disabled=not 둘다, on_click=_생성누름,
                  help='위 「이 자료로 만듭니다」 의 두 파일로 오늘 날짜의 다음 번호를 붙여 만듭니다')

    만든것 = None
    if st.session_state.pop('엑셀생성요청', False) and 둘다:
        try:
            만든바이트, 만든안내 = 실적보고엑셀캐시(
                원장올림.getvalue() if 원장올림 is not None else 보관바이트(원장보관),
                전월올림.getvalue() if 전월올림 is not None else 보관바이트(전월보관))
            버전 = 다음버전()
            이름 = 보고서파일이름(만든안내['연도'], 만든안내['월끝'], 버전)
            저장됨 = ''
            try:
                os.makedirs(자료실, exist_ok=True)
                with open(os.path.join(자료실, 이름), 'wb') as f:
                    f.write(만든바이트)
                저장됨 = os.path.abspath(os.path.join(자료실, 이름))
            except OSError as e:
                st.session_state['엑셀저장오류'] = str(e)
            만든것 = (이름, 만든바이트, 만든안내, 버전, 저장됨)
            st.session_state['엑셀마지막안내'] = 만든안내
            보관목록 = 자료실목록()
            최신 = 보관목록[0] if 보관목록 else None
        except Exception as e:
            st.session_state['엑셀생성오류'] = str(e)

    오류 = st.session_state.pop('엑셀생성오류', None)
    if 오류:
        st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="max-width:860px; margin-top:12px; border-left:4px solid {ROSE}">
    <h3>만들지 못했습니다</h3>
    <div style="font-size:14px; line-height:1.9; color:{T['ink2']}; margin-top:8px">{오류}</div>
  </div>
</div>""")

    with 받을칸:
        if 만든것 is not None:
            st.download_button(f'⤓  {오늘날짜():%Y.%m.%d}_v{만든것[3]} 내려받기',
                               data=만든것[1], file_name=만든것[0],
                               mime=('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'), width='stretch')
        elif 최신 is not None:
            try:
                날 = 최신.get('만든날')
                이름표 = (f"{날:%Y.%m.%d}_v{최신['버전']}" if 날 else f"v{최신['버전']}")
                with open(최신['경로'], 'rb') as f:
                    st.download_button(f'⤓  {이름표} 내려받기', data=f.read(),
                                       file_name=최신['이름'],
                                       mime=('application/vnd.openxmlformats-officedocument'
                                             '.spreadsheetml.sheet'), width='stretch')
            except OSError:
                st.button('⤓  내려받기', width='stretch', disabled=True)
        elif 전월있음:
            # 자료실이 비어 있어도(웹사이트가 다시 켜진 뒤) 기억해 둔 마지막 자료는 드립니다
            바이트 = 보관바이트(전월보관)
            if 바이트:
                st.download_button('⤓  마지막 자료 내려받기', data=바이트,
                                   file_name=전월이름 or '실적보고.xlsx', width='stretch',
                                   key='마지막자료받기',
                                   mime=('application/vnd.openxmlformats-officedocument'
                                         '.spreadsheetml.sheet'),
                                   help=f'기억해 둔 「{전월이름}」 입니다')
            else:
                st.button('⤓  내려받기', width='stretch', disabled=True)
        else:
            st.button('⤓  내려받기', width='stretch', disabled=True,
                      help='아직 만든 자료가 없습니다. 두 파일을 올리고 「생성」을 눌러 주세요')
    with 설명칸2:
        if 최신 is not None or 만든것 is not None or 둘다:
            st.caption('내려받은 파일을 왼쪽 사이드바 「⚙ 설정 · 파일 업로드」에 올리면 '
                       '대시보드가 이번 달 숫자로 바뀝니다.')
        else:
            st.caption('두 파일을 모두 올리면 「생성」 단추가 켜집니다.')
        st.caption(f'파일 이름은 「…클로드작성_{오늘날짜():%Y.%m.%d}_v번호」 입니다 — '
                   f'번호는 만든 날마다 v1 부터 셉니다 (같은 날 또 만들면 v2).')

    저장오류 = st.session_state.pop('엑셀저장오류', None)
    if 저장오류:
        st.caption(f'※ 자료실 폴더에 저장하지 못했습니다({저장오류}). '
                   f'위 내려받기 단추로 받아 주세요.')

    # ── 방금 만든 결과 요약
    if 만든것 is not None:
        이름, _b, 안내, 버전, 저장됨 = 만든것
        st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="margin-top:14px">
    <h3>v{버전} 파일을 만들었습니다</h3>
    <div class="sub">{이름}</div>
    <table class="lined" style="margin-top:10px">
      <colgroup><col style="width:46%"><col style="width:54%"></colgroup>
      <tbody>
        <tr class="sub"><td class="name">대상 기간</td><td class="lft">{안내['기간']}</td></tr>
        <tr class="sub"><td class="name">갈아끼운 시트</td>
          <td class="lft">{안내['올해시트']} · {안내['행수']:,}줄 (거래 {안내['거래건수']:,}건)</td></tr>
        <tr class="sub"><td class="name">그대로 옮긴 시트</td>
          <td class="lft">{안내['시트수'] - 1}개 (피벗·차트 포함)</td></tr>
        <tr class="sub"><td class="name">보고 기준 월</td>
          <td class="lft">{안내['보고달']}월
            {'— 「' + str(안내['뺀시트']) + '」 시트는 이제 쓰지 않으므로 '
             '결과 파일에서 뺐습니다' if 안내.get('뺀시트')
             else '— 원장 마지막 거래일 기준'}</td></tr>
        <tr class="sub"><td class="name">26년 자금 시트</td>
          <td class="lft">{(
            f"실제 현금흐름 {안내['자금']['현금줄']:,}줄 "
            f"(기초 {안내['자금']['기초']:,.2f} → {안내['보고달']}월말 "
            f"{안내['자금']['기말']:,.2f}) · 예상은 "
            + (f"{안내['자금']['예상월'][0]}~12월 "
               f"{안내['자금']['항목수']}개 항목"
               if 안내['자금'].get('예상월') else '없음')
            ) if 안내['자금'].get('현금줄') is not None
            else '다시 쓰지 못했습니다 — ' + str(안내['자금'].get('오류', '시트를 찾지 못함'))}
          </td></tr>
        <tr class="sub"><td class="name">전월에서 이어받은 활동분류</td>
          <td class="lft">{안내['이어받음']:,}건</td></tr>
        <tr class="sub"><td class="name">새 거래를 스스로 나눈 것</td>
          <td class="lft">{안내['자동분류']:,}건
            <span style="color:{T['ink3']}">
              (적요 {안내['분류내역'].get('적요일치', 0):,} ·
               거래처 {안내['분류내역'].get('거래처일치', 0):,} ·
               계정 {안내['분류내역'].get('계정과목대표값', 0):,} ·
               계정 최빈값 {안내['분류내역'].get('계정과목최빈값', 0):,})</span></td></tr>
        <tr class="sub"><td class="name">사람이 꼭 봐야 할 줄</td>
          <td class="lft" style="{'font-weight:800; color:' + ROSE
                                  if 안내['확인필요'] else ''}">{안내['확인필요']:,}건
            <span style="color:{T['ink3']}">
              (여러 갈래로 나뉘던 계정 {안내['분류내역'].get('확인필요', 0):,} ·
               처음 보는 계정 {안내['분류내역'].get('신규계정', 0):,})
              — T열에서 「확인필요」로 걸러 보세요</span></td></tr>
        <tr class="sub"><td class="name">전월 자료에 없던 새 거래</td>
          <td class="lft">{안내['새거래']:,}건</td></tr>
        <tr class="total"><td class="name">저장 위치</td>
          <td class="lft">{저장됨 or '(저장 안 됨 — 내려받기로 받아 주세요)'}</td></tr>
      </tbody>
    </table>
    <div style="margin-top:14px; padding-top:12px; border-top:1px solid {T['line']};
                font-size:13px; line-height:1.85; color:{T['ink3']}">
      원장에서 저절로 계산되는 것 — 금액 · 년 · 월 · 계정분류 · 분류 · 계정과목 · 보고금액 · 월(숫자)
      <b style="color:{T['ink2']}">(K~V열 수식 그대로)</b><br>
      전월에서 이어받는 것 — 활동분류 · 활동세부 · 정부지원 · 검토메모<br>
      새 거래는 지난달까지 나눠 두신 기준을 배워서 채우고, 그 근거를
      <b style="color:{T['ink2']}">T열 「클로드 검토」</b> 에 적어 둡니다
      (「확인필요」가 붙은 줄부터 봐 주세요)
    </div>
  </div>
</div>""")
        경고 = []
        if 안내['확인필요']:
            경고.append(f"활동분류를 확실하게 정하지 못한 줄이 {안내['확인필요']:,}건 있습니다. "
                        f"엑셀 「{안내['올해시트']}」 시트에서 T열을 「확인필요」로 걸러 보시면 "
                        f"바로 찾으실 수 있습니다.")
        if 안내['집계밖']:
            경고.append(f"「월별 실적집계」에 없는 활동세부가 붙은 줄이 {안내['집계밖']:,}건 있습니다. "
                        f"그대로 두면 그 금액이 집계표에서 빠집니다.")
        if 안내['매핑밖']:
            경고.append(f"BS_IS_매핑 표에 없는 계정이 {안내['매핑밖']:,}건 있습니다. "
                        f"계정과목·계정분류가 빈칸으로 나옵니다.")
        if 안내.get('수식고침'):
            경고.append(f"「{안내['당월시트']}」 {' · '.join(안내['수식고침'])} 칸이 "
                        f"빈칸(G5·H5)을 보고 있어 매출원가 누적이 잘못 나오던 것을 "
                        f"기준 월($H$2)을 보도록 고쳤습니다.")
        for 말 in 경고:
            st.caption(f'※ {말}')

    # ── 자료실에 쌓인 버전 목록
    if 보관목록:
        줄 = ''.join(
            f'<tr class="{"total" if i == 0 else "sub"}">'
            f'<td class="name">{_달표시(d)}</td>'
            f'<td class="name">{_만든날표시(d)}</td>'
            f'<td class="name">v{d["버전"]}{"  ← 최신" if i == 0 else ""}</td>'
            f'<td class="lft">{d["이름"]}</td>'
            f'<td>{datetime.datetime.fromtimestamp(d["시각"]):%Y-%m-%d %H:%M}</td>'
            f'<td>{d["크기"]/1e6:,.1f}MB</td></tr>'
            for i, d in enumerate(보관목록[:12]))
        st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="margin-top:14px">
    <h3>자료실에 보관된 실적보고 엑셀</h3>
    <div class="sub">{os.path.abspath(자료실)} · 새 것부터 ·
      번호는 만든 날마다 v1 부터 셉니다 (웹사이트를 업데이트하면 이 목록은 비워집니다)</div>
    <table class="lined" style="margin-top:10px">
      <colgroup><col style="width:8%"><col style="width:12%"><col style="width:8%">
                <col style="width:41%"><col style="width:19%"><col style="width:12%"></colgroup>
      <thead><tr><th>실적 월</th><th>만든 날</th><th>버전</th><th>파일 이름</th>
                 <th>만든 때</th><th>크기</th></tr></thead>
      <tbody>{줄}</tbody>
    </table>
  </div>
</div>""")
    st.stop()


# ★화면은 언제나 「마지막으로 올린 실적 엑셀」을 기준으로 그립니다★
if 업로드파일 is not None:
    데이터바이트 = 업로드파일.getvalue()
    with open(CACHE_PATH, 'wb') as f:
        f.write(데이터바이트)
    try:
        with open(이름기록, 'w', encoding='utf-8') as f:
            f.write(업로드파일.name)
    except OSError:
        pass
    쓰는파일 = 업로드파일.name
    설정칸.caption(f'✔ 방금 올림: {업로드파일.name}')
elif os.path.exists(CACHE_PATH):
    with open(CACHE_PATH, 'rb') as f:
        데이터바이트 = f.read()
    쓰는파일 = 쓰던파일이름()
    설정칸.caption(f'📁 사용 중: {쓰는파일}' if 쓰는파일 else '📁 마지막으로 올린 자료를 쓰고 있습니다')
else:
    # 자료가 없을 때 — 여기서 바로 올릴 수 있게 합니다 (사이드바까지 가지 않도록)
    st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="max-width:820px; margin-top:8px">
    <div class="page-head" style="margin-bottom:6px">
      <span class="t">자료를 올려 주세요</span></div>
    <div style="font-size:15px; line-height:1.85; color:{T['ink2']}; margin-top:10px">
      아래 상자에 파일을 <b style="color:{T['ink']}">끌어다 놓기만</b> 하시면 됩니다.
      여러 개를 한꺼번에 올리셔도 파일 안을 보고 알아서 제자리에 넣습니다.
      <div style="margin-top:10px; font-size:14px; line-height:1.9">
        · <b style="color:{T['ink']}">실적보고 엑셀</b> (「26년 원장_raw」 시트가 든 파일)
          — 이 화면의 표와 그래프를 그립니다<br>
        · <b style="color:{T['ink']}">퀵북 원장 원본</b> (General Ledger)
          — 실적보고 엑셀작성 · 연결패키지에 씁니다<br>
        · <b style="color:{T['ink']}">연결패키지 양식</b> (본사에서 받은 원본)
          — 연결패키지 재무제표를 만듭니다
      </div>
    </div>
  </div>
</div>""")
    첫칸, _빈 = st.columns([2, 1], gap='small')
    with 첫칸:
        올린것 = st.file_uploader('자료 올리기 (.xlsx · 여러 개 가능)', type=['xlsx'],
                                  accept_multiple_files=True, key='첫화면업로드')
    if 올린것:
        실적, 목록 = 자료받기(올린것)
        for _n, _k in 목록:
            st.caption(f'{"✔" if _k else "?"} {_n} → '
                       + 종류이름.get(_k, '무슨 자료인지 모르겠습니다'))
        if 실적 is not None:
            try:
                with open(CACHE_PATH, 'wb') as f:
                    f.write(실적.getvalue())
                with open(이름기록, 'w', encoding='utf-8') as f:
                    f.write(실적.name)
            except OSError as e:
                st.error(f'저장하지 못했습니다: {e}')
            else:
                st.rerun()
        else:
            st.caption('※ 표를 그리려면 「26년 원장_raw」 시트가 든 '
                       '실적보고 엑셀이 하나는 있어야 합니다.')
    st.html(f"""<div class="wrap" translate="no">
  <div class="card" style="max-width:820px; margin-top:4px">
    <div style="font-size:13.5px; line-height:1.85; color:{T['ink3']}">
      <b style="color:{T['ink2']}">왜 다시 올려야 하나요?</b><br>
      회사 재무자료는 인터넷(GitHub)에 올리지 않기 때문에, 웹사이트에는 올려 주신 파일이
      <b style="color:{T['ink2']}">잠시 머물다 지워집니다</b> — 웹사이트를 업데이트했을 때,
      그리고 며칠 아무도 안 들어와 서버가 잠들었다 깨어날 때 지워집니다.
      평소 쓰실 때는 지워지지 않습니다.<br>
      늘 켜 두고 쓰시려면 <b style="color:{T['ink2']}">「1. 대시보드_실행.bat」</b> 로
      회사 PC에서 여시면 됩니다 — 그때는 파일이 D:\\otc-Dashboard 에 그대로 남아
      다시 올리실 필요가 없습니다.
    </div>
  </div>
</div>""")
    st.stop()

try:
    (당해원장, 전년원장, 당해연도, 매핑표, 기초현금,
     입력값, BS밑자료, 차입리스) = 원장읽기(데이터바이트)
except Exception as e:
    st.error(f'파일을 읽지 못했습니다: {e}')
    st.stop()

전년연도 = 당해연도 - 1
P = 손익표(당해원장, '활동')          # 대시보드·월간실적은 활동분류 기준을 씁니다
P전 = 손익표(전년원장, '활동')
P손익 = 손익표(당해원장, '손익')      # 실적집계 화면에서 고를 수 있는 손익계산서 기준
P손익전 = 손익표(전년원장, '손익')
보고월 = int(당해원장['월'].max())
기준일 = 당해원장['거래일'].max()

_쓰는파일줄 = (f'<div title="{쓰는파일}">자료 <b>{쓰는파일[:26]}'
               f'{"…" if len(쓰는파일) > 26 else ""}</b></div>') if 쓰는파일 else ''
st.sidebar.markdown(
    f"""<div class="side-foot">
      <div>기준 <b>{당해연도}년 1~{보고월}월</b></div>
      <div>원장 {len(당해원장):,}건</div>
      {_쓰는파일줄}
    </div>""", unsafe_allow_html=True)

미분류건수 = int((당해원장['계정분류'].eq('IS') & 당해원장['분류'].eq('판관비')
                & 당해원장['활동세부'].isna()).sum())


def 누적(t, 연도표=None, upto=None):
    upto = upto or 보고월
    return t[list(range(1, upto + 1))].sum(axis=1)


당해누적 = 누적(P)
전년동기 = 누적(P전)
전년연간 = P전[list(range(1, 13))].sum(axis=1)




# 오른쪽 세부내역 패널에서 항목 하나를 고르면 어떤 자금분류(구분) 코드를 찾아볼지 매핑.
# 부호구분 'DEBT'는 같은 코드를 부호로 다시 나눕니다 — '+' 는 차입(들어온 돈), '−' 는 상환(나간 돈).
현금흐름목록 = [
    ('매출채권 회수', 'AR', None),
    ('기타입금', 'ETCIN', None),
    ('매입채무', 'AP', None),
    ('급여', 'PAY', None),
    ('기타 비용', 'OTHOP', None),
    ('설비 투자', 'CAPEX', None),
    ('유상증자', 'EQ', None),
    ('차입', 'DEBT', '+'),
    ('차입 상환', 'DEBT', '-'),
    ('리스료 지급', 'LEASE', None),
    ('환율변동 효과', 'FX', None),
]

# ══════════════════════════════════════════════════════════════
# 환율 — OTC법인은 장부 통화가 USD 라 화면 환산이 없습니다.
# 원화(KRW) 환산은 「연결재무제표 패키지」 화면에서 직접 넣는 두 칸으로만 씁니다.
SRED = {}          # OTC법인은 정부 R&D 세액공제가 없습니다
현금잔고, 현금증감 = 현금잔고월별(당해원장, 기초현금)
자금표, 자금상세 = 자금변동(당해원장)


# OTC법인은 장부 통화가 이미 USD 라 화면 환산이 없습니다
통화, 환율, 배율 = 'USD', 1.0, 1.0
if 메뉴 in ('월간실적', '누적 실적보고'):
    설정칸.caption('표시 통화 — USD (OTC법인 장부 통화)')
    if not 입력값['단일'] and not 입력값['서술']:
        설정칸.caption('ℹ️ 실적파일에 「월간보고_입력」 시트가 없어 '
                       '목표·진척률·코멘트는 비어 있습니다.')


def C(v, 자릿수=0, 괄호=True):
    """표시 통화로 바꾼 금액 문자열. 음수는 (괄호) 빨간 글씨."""
    if v is None or pd.isna(v):
        return '-'
    t = f'{v * 배율:,.{자릿수}f}'
    return 음수표기(t) if 괄호 else t


def CK(v, 괄호=True):
    """천 단위 표기 (카드용) — 값만 돌려주고 단위는 K USD 로 따로 붙입니다."""
    if v is None or pd.isna(v):
        return '-'
    t = f'{v * 배율 / 1000:,.0f}'
    if t in ('-0', '−0'):
        t = '0'
    return 음수표기(t) if 괄호 else t


def K(v):
    """그래프에 쓸 천 단위 숫자."""
    return None if v is None or pd.isna(v) else v * 배율 / 1000


단위K = f'K {통화}'
환율문구 = ''


def 차트제목(제목, 단위=None):
    """차트 위에 제목과 단위를 따로 그립니다 (Altair 제목은 글자가 잘려서 쓰지 않습니다)."""
    st.markdown(f"<div class='chart-head'><span class='t'>{제목}</span>"
                f"<span class='u'>(단위: {단위 or 단위K})</span></div>",
                unsafe_allow_html=True)


def 값라벨(차트, dy=-9, dx=0, align='center', 필드='금액'):
    """그래프에 천단위 쉼표 숫자를 붙입니다. 음수는 아래쪽에 표시합니다."""
    글꼴 = dict(fontSize=12, fontWeight=600, color=T['ink2'])
    양 = 차트.transform_filter(alt.datum[필드] >= 0).mark_text(
        dy=dy, dx=dx, align=align, **글꼴).encode(
        text=alt.Text(f'{필드}:Q', format=',.0f'), color=alt.value(T['ink2']))
    음 = 차트.transform_filter(alt.datum[필드] < 0).mark_text(
        dy=(-dy if dy < 0 else dy) + 4, dx=dx, align=align, **글꼴).encode(
        text=alt.Text(f'{필드}:Q', format=',.0f'), color=alt.value(T['ink2']))
    return 양 + 음


def X축(값들, 제목=None, 여유=0.20):
    """가로 막대 오른쪽 끝의 숫자가 잘리지 않도록 오른쪽 여백을 둡니다."""
    유효 = [float(v) for v in 값들 if v is not None and not pd.isna(v)]
    if not 유효:
        return alt.X('금액:Q', title=제목)
    lo, hi = min(유효 + [0.0]), max(유효 + [0.0])
    폭 = (hi - lo) or (abs(hi) or 1.0)
    return alt.X('금액:Q', title=제목,
                 scale=alt.Scale(domain=[lo - (폭 * 여유 if lo < 0 else 0), hi + 폭 * 여유],
                                 nice=False))


def Y축(값들, 제목=None, 여유=0.18):
    """막대 위·아래의 숫자가 축 글씨와 겹치지 않도록 위아래 여백을 넉넉히 둡니다."""
    유효 = [float(v) for v in 값들 if v is not None and not pd.isna(v)]
    if not 유효:
        return alt.Y('금액:Q', title=제목)
    lo, hi = min(유효 + [0.0]), max(유효 + [0.0])
    폭 = (hi - lo) or (abs(hi) or 1.0)
    return alt.Y('금액:Q', title=제목,
                 scale=alt.Scale(domain=[lo - 폭 * 여유, hi + 폭 * 여유], nice=False))

# ══════════════════════════════════════════════════════════════
# 3. 대시보드
# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# 분석 엑셀은 「만들기」를 누를 때에만 만듭니다.
#   ※ 화면을 열 때마다 만들면 클릭 한 번에 5초 넘게 멈추고 메모리도 크게 써서,
#     웹사이트(Streamlit Cloud)가 앱을 강제로 내려 화면이 하얗게 보였습니다.
#     한 번 만든 파일은 원장이 바뀌기 전까지 다시 만들지 않습니다.
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner='분석 엑셀을 만드는 중입니다... (30초쯤 걸립니다)', max_entries=1)
def 분석엑셀캐시(열쇠):
    return 분석엑셀만들기(당해원장, 전년원장, P손익, P, 자금표, 현금잔고,
                          기초현금, 당해연도, 보고월, 매핑표)


def 엑셀열쇠():
    """원장이 바뀌면 값이 달라지는 「지문」 — 이 값이 같으면 엑셀을 다시 만들지 않습니다."""
    return (당해연도, 보고월, len(당해원장), len(전년원장),
            round(float(당해원장['금액'].abs().sum()), 2), round(float(기초현금), 2))


if 메뉴 == '누적 실적보고':
    제목칸, 내려받기칸 = st.columns([4, 1], gap='small')
    with 내려받기칸:
        if st.session_state.get('엑셀만들기'):
            try:
                엑셀바이트 = 분석엑셀캐시(엑셀열쇠())
                st.download_button(
                    '⤓  분석 엑셀 내려받기', data=엑셀바이트,
                    file_name=f'CTK OTC_실적분석_{당해연도}년{보고월}월누적_클로드작성.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    width='stretch',
                    help='지금 화면의 손익·현금흐름·미수채권과 처리본 원장을 담은 엑셀입니다')
            except Exception as e:
                st.session_state['엑셀만들기'] = False
                st.caption(f'엑셀 생성 실패: {e}')
        else:
            st.button('⤓  분석 엑셀 만들기', width='stretch', key='엑셀만들기버튼',
                      on_click=lambda: st.session_state.__setitem__('엑셀만들기', True),
                      help='누르면 손익·현금흐름·미수채권과 처리본 원장을 담은 엑셀을 만듭니다')
    with 제목칸:
        st.html(f"""<div class="wrap" translate="no" style="margin-bottom:0">
      <div class="page-head"><span class="t">실적보고</span>
        <span class="pill">{당해연도}년 1~{보고월}월 누적</span></div>
    </div>""")

    이전월 = 보고월 - 1 if 보고월 > 1 else None
    # ★KPI 카드는 1~보고월 누적★ (아래 표·그래프와 같은 기준)
    매출, 원가 = 당해누적['매출액'], 당해누적['매출원가']
    총이익, 영익 = 당해누적['매출총이익'], 당해누적['영업이익(손실)']
    세전 = 당해누적.get('세전이익', 영익)
    원가율 = (원가 / 매출 * 100) if 매출 else 0
    영익률 = (영익 / 매출 * 100) if 매출 else 0
    세전률 = (세전 / 매출 * 100) if 매출 else 0

    def 전년(항목):
        """누적끼리 비교해야 하니 전년 같은 기간(1~보고월) 누적을 돌려줍니다."""
        v = float(전년동기.get(항목, 0.0))
        return v if abs(v) > 0.005 else None

    누적표기 = f'1~{보고월}월 누적'
    kpi = f"""
    <div class="kpi-row" style="--n:5">
      <div class="kpi-card"><div class="kpi-label">매출액 <span class="calc">({누적표기})</span></div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(매출)} <span class="unit">{단위K}</span></div>
        <div class="kpi-delta">{증감HTML(매출, 전년('매출액'), '전년동기')}</div></div>
      <div class="kpi-card"><div class="kpi-label">매출원가 <span class="calc">({누적표기})</span></div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(원가)} <span class="unit">{단위K}</span></div>
        <div class="kpi-delta">{증감HTML(원가, 전년('매출원가'), '전년동기')}</div></div>
      <div class="kpi-card"><div class="kpi-label">매출총이익 <span class="calc">({누적표기})</span></div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(총이익)} <span class="unit">{단위K}</span></div>
        <div class="kpi-delta"><span class="muted">원가율 {원가율:,.1f}%</span></div></div>
      <div class="kpi-card"><div class="kpi-label">영업이익(손실) <span class="calc">({누적표기})</span></div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(영익)} <span class="unit">{단위K}</span></div>
        <div class="kpi-delta"><span class="muted">영업이익률 {영익률:,.1f}%</span></div></div>
      <div class="kpi-card"><div class="kpi-label">세전이익(손실) <span class="calc">({누적표기})</span></div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(세전)} <span class="unit">{단위K}</span></div>
        <div class="kpi-delta"><span class="muted">세전이익률 {세전률:,.1f}%</span></div></div>
    </div>"""

    months = list(range(1, 보고월 + 1))
    매출26 = [P.loc['매출액', m] for m in months]
    매출25 = [P전.loc['매출액', m] for m in months]
    누적26 = float(np.nansum(매출26))                    # 당해 1~보고월 누적
    누적25 = float(np.nansum(매출25))                    # 전년 같은 기간(1~보고월) 누적
    월최대 = max([abs(v) for v in 매출26 + 매출25] + [1])
    누적최대 = max(abs(누적26), abs(누적25), 1)

    def 막대(v25, v26, 라벨, 기준, 강조=False):
        cls = ' cum' if 강조 else ''
        return (f'<div class="bar-group{cls}"><div class="bar-pair">'
                f'<div class="bar-col"><div class="bar-val">{CK(v25)}</div>'
                f'<div class="bar prev" style="height:{abs(v25)/기준*118:.0f}px;"></div></div>'
                f'<div class="bar-col"><div class="bar-val">{CK(v26)}</div>'
                f'<div class="bar cur" style="height:{abs(v26)/기준*118:.0f}px;"></div></div>'
                f'</div><div class="bar-month">{라벨}</div></div>')

    bars = ''.join(막대(매출25[i], 매출26[i], f'{m}월', 월최대) for i, m in enumerate(months))
    bars += '<div class="bar-sep"></div>' + 막대(누적25, 누적26, '누적', 누적최대, True)

    # ── 매출액 구성 (손익 기준) ─────────────────────────────
    # KPI 카드·손익표와 같게 1~보고월 누적으로 봅니다
    매출행 = 당해원장[당해원장['월'].le(보고월) & (당해원장['분류'] == '매출')]
    mix = 매출행.groupby('계정과목')['금액'].sum().sort_values(ascending=False)
    mix_html = ''
    if mix.sum():
        for 이름, 값 in mix.items():
            비율 = 값 / mix.sum() * 100
            mix_html += (f'<div class="mix-item"><div class="top"><span>{이름}</span>'
                         f'<span>{CK(값)} <span class="muted">{단위K} ({비율:,.0f}%)</span></span></div>'
                         f'<div class="mix-track">'
                         f'<div class="mix-fill" style="width:{max(0, min(100, 비율)):.0f}%;"></div></div></div>')
        mix_html += (f'<div class="mix-total">매출액 합계<span>{CK(mix.sum())} '
                     f'<span class="muted">{단위K}</span></span></div>')
    else:
        mix_html = '<div class="muted">해당 기간 매출 데이터가 없습니다.</div>'

    # ══ 손익계산서 표 ══════════════════════════════════════
    IS = 당해원장[당해원장['계정분류'].eq('IS')]

    def 월별(mask, 부호=1):
        s = IS[mask].groupby('월')['금액'].sum().reindex(months, fill_value=0.0) * 부호
        return [float(s.get(m, 0.0)) for m in months]

    # ★실적집계 화면과 같은 표(P)에서 그대로 가져옵니다 — 두 화면이 어긋날 수 없습니다★
    def 줄(이름):
        return [float(P.loc[이름, m]) for m in months]

    매출L, 원가L, 총이익L = 줄('매출액'), 줄('매출원가'), 줄('매출총이익')
    판관L = 줄('판매관리비')
    영익L, EBITDAL = 줄('영업이익(손실)'), 줄('EBITDA 이익(손실)')
    영외L, 세전L = 줄('영업외손익'), 줄('세전이익')

    리스상각 = 당해원장['계정영문'].astype(str).str.strip().str.lower().str.startswith('6623')
    상각L = 월별(IS['계정과목'].isin(['감가상각비', '무형자산상각비']), -1)
    리스L = 월별(IS['계정과목'].isin(['감가상각비', '무형자산상각비']) & 리스상각, -1)

    판관계정 = (IS[IS['분류'].eq('판관비')].groupby('계정과목')['금액'].sum().mul(-1)
                .sort_values(ascending=False))
    상위4 = list(판관계정.index[:4])
    세부 = []
    for 이름 in 상위4:
        마스크 = IS['분류'].eq('판관비') & IS['계정과목'].eq(이름)
        if 이름 in ('감가상각비', '무형자산상각비'):
            # 감가상각비는 「리스(사용권자산) 상각」과 「그 밖의 상각」으로 나눠 보여줍니다
            리스분 = 월별(마스크 & 리스상각, -1)
            본체 = [a - b for a, b in zip(월별(마스크, -1), 리스분)]
            세부.append((이름, 본체))
            if any(abs(v) > 0.005 for v in 리스분):
                세부.append(('리스상각비', 리스분))
        else:
            세부.append((이름, 월별(마스크, -1)))
    세부.append(('기타', [판관L[i] - sum(v[1][i] for v in 세부) for i in range(len(months))]))

    def 이름칸(이름, 산식=''):
        """산식은 어떻게 계산했는지 덧붙이는 설명이라 () 안에 한 포인트 작게 씁니다."""
        return 이름 + (f' <span class="calc">({산식})</span>' if 산식 else '')

    예상월들 = list(range(보고월 + 1, 13))        # 8월(예상) … 12월(예상)

    def 행(이름, 값들, 굵게=False, 들여=0, 부호표시=False, 색='', 산식=''):
        cls = ' class="total"' if 굵게 else (' class="sub"' if 들여 else '')
        def f(v):
            if abs(v * 배율 / 1000) < 0.5:   # 반올림하면 0 이 되는 금액은 '-' 로
                return '-'
            if 부호표시:
                return ('+' if v > 0 else '−') + CK(abs(v))
            return CK(v)
        cells = ''.join(f'<td class="{색}">{f(v)}</td>' for v in 값들)
        # 앞으로 올 달(예상) 칸 — 금액은 아직 비어 있습니다
        cells += ''.join(f'<td class="fc {색}">-</td>' for _ in 예상월들)
        총 = sum(값들)
        return (f'<tr{cls}><td class="name">{"&nbsp;&nbsp;&nbsp;&nbsp;" * 들여}'
                f'{이름칸(이름, 산식)}</td>{cells}'
                f'<td class="{색}"><b>{f(총)}</b></td></tr>')

    rows = 행('매출액', 매출L, True)
    rows += 행('매출원가', 원가L)
    rows += 행('매출총이익', 총이익L, True)
    rows += 행('판관비 합계', 판관L, True)
    for 이름, 값들 in 세부:
        rows += 행(이름, 값들, 들여=1)
    rows += 행('영업이익(손실)', 영익L, True)
    rows += 행('영업외손익<br><span class="calc">(이자 · 외환 등)</span>', 영외L)
    rows += 행('세전이익(손실)', 세전L, True)
    rows += 행('EBITDA 이익(손실)', EBITDAL, True)

    # 열너비를 %로 잡아 화면 크기에 맞춰 표가 같이 줄었다 늘었다 하게 합니다.
    # 손익계산서와 현금흐름이 똑같은 %와 똑같은 최대너비를 쓰므로, 화면이 크든 작든
    # 두 표의 월별 칸은 항상 같은 자리에서 시작합니다 (위아래 라인이 맞습니다).
    칸수 = len(months) + len(예상월들) + 1        # 월 칸들 + 예상 칸들 + 「누적」 칸
    이름칸, 숫자칸 = 215, 78                     # 글자를 키운 대신 칸을 좁혀 촘촘하게 봅니다
    이름비율 = 이름칸 / (이름칸 + 숫자칸 * 칸수) * 100
    숫자비율 = (100 - 이름비율) / 칸수
    표너비 = 이름칸 + 숫자칸 * 칸수               # 이보다 커지면 보기 나쁘니 최대너비로 씁니다
    콜그룹 = (f'<colgroup><col style="width:{이름비율:.4f}%">'
              + f'<col style="width:{숫자비율:.4f}%">' * 칸수 + '</colgroup>')
    표스타일 = f'style="width:100%; max-width:{표너비}px; table-layout:fixed;"'

    # ══ 오른쪽 세부내역 : 손익계산서 항목을 고르면 거래처별 누적 내역 ══════
    누적IS = IS[IS['월'].le(보고월)]

    def 항목조건_손익(이름):
        d = 누적IS
        if 이름 == '매출액':
            return d['분류'].eq('매출'), 1
        if 이름 == '매출원가':
            return d['분류'].eq('매출원가'), -1
        if 이름 == '판관비 합계':
            return d['분류'].eq('판관비'), -1
        if 이름 == '기타':
            return d['분류'].eq('판관비') & ~d['계정과목'].isin(상위4), -1
        return d['분류'].eq('판관비') & d['계정과목'].eq(이름), -1

    # 「판관비 합계」는 아래 상위4 + 기타로 쪼개지는 상위 줄이라 목록에서 뺍니다 (최하위 줄만)
    드릴목록_손익 = ['매출액', '매출원가'] + 상위4 + ['기타']

    # EBITDA 계산 근거 — 무엇을 더하고 무엇을 뺐는지
    상각합, 리스합 = sum(상각L), sum(리스L)
    리스계정 = sorted(set(
        당해원장.loc[리스상각 & 당해원장['계정분류'].eq('IS'), '계정영문']
        .astype(str).str.strip()))
    def CK1(v):
        """계산 근거는 덧셈·뺄셈 식이라 괄호 대신 부호를 두고 색만 붉게 씁니다."""
        t = f'{v * 배율 / 1000:,.1f}'
        return f'<span class="neg">{t}</span>' if t[:1] in ('-', '−') else t

    ebitda설명 = (
        f'<details class="calcnote">'
        f'<summary>EBITDA 계산 근거 <span class="hint">누르면 펼쳐집니다</span></summary>'
        f'<div class="body">'
        f'영업이익 {CK1(sum(영익L))} <b>+</b> 감가상각비 {CK1(상각합)} '
        f'<b>−</b> 리스(RoU) 상각비 {CK1(리스합)} <b>=</b> '
        f'<b>{CK1(sum(EBITDAL))} {단위K}</b> '
        f'<span class="muted">(표에는 반올림해 {CK(sum(EBITDAL))} 로 적었습니다)</span><br>'
        f'더한 감가상각비는 손익 계정 「감가상각비 · 무형자산상각비」 전체입니다. '
        f'그중 아래 원장 계정만 리스 상각으로 보아 EBITDA 에서 다시 뺐습니다 — '
        f'<b>{" · ".join(리스계정) if 리스계정 else "해당 없음"}</b>'
        f'{f" (1~{보고월}월 누적 {CK1(리스합)} {단위K})" if 리스합 else ""}</div></details>')

    # ══ 자금 변동 (Statement of Cash Flows) ══════════════════
    자금html = ''
    if 자금표 is not None:
        기초 = [float(현금잔고.get(m - 1, 기초현금)) if m > 1 else float(기초현금)
                for m in months]
        기말 = [float(현금잔고.get(m, 0.0)) for m in months]

        # ── 8월(예상) … 12월(예상) — 「26년 자금」 시트의 2번(출금)·3번(입금)에서
        예상 = 예상자금읽기(데이터바이트) or {}
        # 출금이 아직 안 들어 있으면(옛 파일) 월별 실적집계 평균으로 대신 채웁니다
        출금없음 = not any(m > 보고월 and v < 0
                           for 칸 in 예상.values() for m, v in 칸.items())
        예상밑값 = 예상출금기본(데이터바이트, 보고월) if 출금없음 else {}
        for 줄, 칸 in 예상밑값.items():
            for m, v in 칸.items():
                예상.setdefault(줄, {})
                예상[줄][m] = 예상[줄].get(m, 0.0) + v

        def 예(키, m):
            return float(예상.get(키, {}).get(m, 0.0))

        def 묶음(키, m):
            """소계 줄은 아래 줄을 더해서 냅니다."""
            if 키 == '지급계':
                return 예('AP', m) + 예('PAY', m) + 예('OTHOP', m)
            if 키 == '영업':
                return 예('AR', m) + 예('ETCIN', m) + 묶음('지급계', m)
            if 키 == '투자':
                return 예('CAPEX', m)
            if 키 == '재무':
                return 예('EQ', m) + 예('차입증감', m) + 예('리스', m)
            if 키 == '순증감':
                return (묶음('영업', m) + 묶음('투자', m) + 묶음('재무', m) + 예('FX', m))
            return 예(키, m)

        def ㅈ(키):
            v = [float(자금표.loc[키, m]) for m in months]
            앞 = [float(np.nansum(v))]           # 누적은 실적(1~보고월)만 셉니다
            return v + [묶음(키, m) for m in 예상월들] + 앞

        # 예상 달의 기말은 마지막 실적 잔액에서 이어 붙입니다
        예상기초, 예상기말, 앞잔액 = [], [], (기말[-1] if 기말 else 기초현금)
        for m in 예상월들:
            예상기초.append(앞잔액)
            앞잔액 = 앞잔액 + 묶음('순증감', m)
            예상기말.append(앞잔액)

        자금줄 = [
            ('기초 현금', 0, 기초 + 예상기초 + [기초[0]], True),
            ('영업활동', 0, ㅈ('영업'), True),
            ('매출채권 회수', 1, ㅈ('AR'), False),
            ('기타입금', 1, ㅈ('ETCIN'), False),
            ('매입채무 · 비용 지급', 1, ㅈ('지급계'), False),
            ('매입채무', 2, ㅈ('AP'), False),
            ('급여', 2, ㅈ('PAY'), False),
            ('기타 비용', 2, ㅈ('OTHOP'), False),
            ('투자활동', 0, ㅈ('투자'), True),
            ('설비 투자', 1, ㅈ('CAPEX'), False),
            ('재무활동', 0, ㅈ('재무'), True),
            ('유상증자', 1, ㅈ('EQ'), False),
            ('차입증감', 1, ㅈ('차입증감'), False),
            ('리스료 지급', 1, ㅈ('리스'), False),
            ('환율변동 효과', 0, ㅈ('FX'), False),
            ('순증감', 0, ㅈ('순증감'), True),
            ('기말 현금', 0, 기말 + 예상기말 + [(예상기말[-1] if 예상기말 else 기말[-1])], True),
        ]
        # 손익계산서와 똑같은 너비를 써서 두 표의 월 칸이 늘 같은 자리에 섭니다
        자금칸수, 자금이름비율, 자금숫자비율 = 칸수, 이름비율, 숫자비율
        자금콜그룹, 자금표스타일 = 콜그룹, 표스타일

        자금rows = ''
        for 국문, 깊이, 값들, 굵게 in 자금줄:
            cls = ' class="total"' if 굵게 else (' class="sub sub2"' if 깊이 == 2
                                                 else (' class="sub"' if 깊이 else ''))
            예상칸 = ' class="fc"'
            칸 = ''.join(
                '<td' + (예상칸 if len(months) <= i < len(months) + len(예상월들) else '')
                + '>' + ('-' if abs(v * 배율 / 1000) < 0.5 else CK(v)) + '</td>'
                for i, v in enumerate(값들))
            자금rows += f'<tr{cls}><td class="name">{국문}</td>{칸}</tr>'

        # ── 월말 잔액 꺾은선 : 아래 표의 월 칸 한가운데에 점이 오도록 %로 자리를 잡습니다
        #    (그림 영역도 표의 「1월~마지막 달」 칸과 정확히 같은 자리에 둡니다)
        #    맨 앞에는 기초 잔액을 얹어 「어디서 출발했는지」가 같이 보이게 합니다.
        #    기초 점은 표의 「구분」 머리칸 위쪽에 놓습니다.
        잔액K = ([기초현금 * 배율 / 1000] + [v * 배율 / 1000 for v in 기말]
                 + [v * 배율 / 1000 for v in 예상기말])
        낮, 높 = min(잔액K), max(잔액K)
        폭 = (높 - 낮) or (abs(높) or 1.0)

        def 눈금만들기(lo, hi, 칸수=3):
            거친 = (hi - lo) / 칸수
            자리 = 10 ** math.floor(math.log10(거친)) if 거친 > 0 else 1
            for 배 in (1, 2, 2.5, 5, 10):
                if 거친 <= 자리 * 배:
                    간격 = 자리 * 배
                    break
            시작 = math.floor(lo / 간격) * 간격
            끝 = math.ceil(hi / 간격) * 간격
            값 = []
            v = 시작
            while v <= 끝 + 간격 * 1e-6:
                값.append(v)
                v += 간격
            return 값, 시작, 끝

        눈금값, 축낮, 축높 = 눈금만들기(min(0.0, 낮 - 폭 * 0.10), 높 + 폭 * 0.28)
        기초x = 자금이름비율 * 0.58        # 「구분」 머리칸 위쪽 — 기초 잔액이 서는 자리
        왼끝 = 기초x
        오른끝 = 자금이름비율 + 자금숫자비율 * (len(months) + len(예상월들))
        # 선 아래 옅은 면 — 계열색을 10%쯤으로 깔아 그림에 무게를 줍니다
        _주 = T['accent'].lstrip('#')
        _r, _g, _b = (int(_주[i:i + 2], 16) for i in (0, 2, 4))
        면색 = (f'linear-gradient(to bottom, rgba({_r},{_g},{_b},0.16) 0%,'
                f' rgba({_r},{_g},{_b},0.02) 100%)')

        def 가로(i):
            # i = 0 이면 기초, 그다음부터 1월 · 2월 … (표의 월 칸 한가운데)
            return 기초x if i == 0 else 자금이름비율 + 자금숫자비율 * (i - 0.5)

        def 세로(v):
            return (축높 - v) / (축높 - 축낮) * 100

        # 선(사다리꼴로 오려 낸 띠)과 그 아래 옅은 면
        굵기 = 1.2                                     # 선 두께의 절반 (px)
        바닥 = 세로(축낮)
        구간 = ''
        for i in range(len(잔액K) - 1):
            x1, x2 = 가로(i), 가로(i + 1)
            y1, y2 = 세로(잔액K[i]), 세로(잔액K[i + 1])
            자리 = f'left:{x1:.4f}%; width:{x2 - x1:.4f}%'
            구간 += (f'<div class="cfill" style="{자리}; background:{면색}; '
                     f'clip-path:polygon(0% {y1:.4f}%, 100% {y2:.4f}%, '
                     f'100% {바닥:.4f}%, 0% {바닥:.4f}%)"></div>')
            구간 += (f'<div class="cseg" style="{자리}; '
                     f'clip-path:polygon(0% calc({y1:.4f}% - {굵기}px), '
                     f'100% calc({y2:.4f}% - {굵기}px), '
                     f'100% calc({y2:.4f}% + {굵기}px), '
                     f'0% calc({y1:.4f}% + {굵기}px))"></div>')
        마지막 = len(잔액K) - 1
        표시 = ''.join(
            f'<div class="cdot{" last" if i == 마지막 else ""}'
            f'{" first" if i == 0 else ""}" '
            f'style="left:{가로(i):.4f}%; top:{세로(v):.4f}%"></div>'
            f'<div class="cval" style="left:{가로(i):.4f}%; top:{세로(v):.4f}%; '
            f'margin-top:-11px">{음수표기(f"{v:,.0f}")}</div>'
            for i, v in enumerate(잔액K))
        표시 += (f'<div class="ctag" style="left:{기초x:.4f}%; '
                 f'top:{세로(잔액K[0]):.4f}%">기초</div>')
        눈금 = ''.join(
            f'<div class="cgrid{" zero" if abs(t) < 1e-9 else ""}" '
            f'style="left:{왼끝:.4f}%; width:{오른끝 - 왼끝:.4f}%; top:{세로(t):.4f}%"></div>'
            f'<div class="cylab" style="width:calc({왼끝:.4f}% - 20px); '
            f'top:{세로(t):.4f}%">{t:,.0f}</div>'
            for t in 눈금값)
        축 = (f'<div class="cdomain" style="left:{왼끝:.4f}%"></div>'
              f'<div class="cbase" style="left:{왼끝:.4f}%; '
              f'width:{오른끝 - 왼끝:.4f}%; top:{세로(축낮):.4f}%"></div>')
        꺾은선 = (f'<div class="cashcap"><span class="t">월말 현금성 자산</span>'
                  f'<span class="u">기초 {CK(기초현금)} → {보고월}월 말 '
                  f'{CK(기말[-1])} {단위K}'
                  + (f' → 12월 말 (예상) {CK(예상기말[-1])} {단위K}' if 예상기말 else '')
                  + '</span></div>'
                  f'<div class="cashline" style="max-width:{이름칸 + 숫자칸 * 자금칸수}px">'
                  f'{눈금}{축}{구간}{표시}</div>')

        예상안내 = (f' · <b>{예상월들[0]}~{예상월들[-1]}월(예상)</b> 은 앞으로 올 달입니다'
                    if 예상월들 else '')
        예상설명 = ('' if not 예상월들 else
                    f'<br><b>{예상월들[0]}~{예상월들[-1]}월(예상)</b> 은 실적 엑셀 「26년 자금」 시트의 '
                    f'<b>2. 예상 현금성 자금 흐름(출금)</b>(클로드가 채운 월평균)과 '
                    f'<b>3. 예상 현금성 자금 흐름(입금)</b>(담당자 작성)을 그대로 옮긴 것입니다. '
                    f'입금을 아직 안 넣으셨으면 잔액이 계속 줄어드는 것으로 보입니다.')
        자금html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>현금흐름 <span class="unitbadge">단위 {단위K}</span></h3>
        <div class="sub">기초 현금에서 출발해 월별로 무엇 때문에
          늘고 줄었는지 나눴습니다 · 마지막 칸은 1~{보고월}월 누적{예상안내}</div>
        {꺾은선}
        <div class="scrollx"><table class="lined" {자금표스타일}>{자금콜그룹}<thead><tr><th>구분</th>
        {''.join(f'<th>{m}월</th>' for m in months)}
        {''.join(f'<th class="fc">{m}월<br><span class="fcs">예상</span></th>' for m in 예상월들)}
        <th>누적</th></tr></thead>
        <tbody>{자금rows}</tbody></table></div>
        <details class="calcnote">
          <summary>보는 방법 <span class="hint">누르면 펼쳐집니다</span></summary>
          <div class="body">원장은 복식부기라, 현금이 들어간 전표에서
          <b>현금이 아닌 상대 계정</b>이 그 돈의 성격을 알려 줍니다. 은행↔은행 이체처럼
          현금끼리 오간 것은 서로 상계되어 잡히지 않습니다.<br>
          들어온 돈은 양수, 나간 돈은 <span class="neg">(괄호)</span>로 적었습니다.
          기초 {CK(기초현금)} → {보고월}월 말 {CK(기말[-1])} {단위K} 로 이어집니다.{예상설명}</div>
        </details></div></div>"""

    경고 = ''
    if 미분류건수:
        경고 = (f'<div class="warnbox">⚠ 판관비 {미분류건수:,}건이 활동분류 미입력 상태입니다. '
                f'"기타관리경비"로 집계되고 있습니다.</div>')

    상단html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="note">모든 금액은 천 단위 {단위K} 로 표시했습니다 · 원장에서 직접 계산한 실제 수치입니다{환율문구}</div>
      {경고}
      {kpi}
      <div style="display:grid; grid-template-columns:1.5fr 1fr; gap:14px;">
        <div class="card"><h3>월별 매출 추이 <span class="unitbadge">단위 {단위K}</span></h3>
          <div class="sub">1~{보고월}월 · {전년연도}년 동월 비교 &nbsp;|&nbsp;
            맨 오른쪽 <b>누적</b> = 1~{보고월}월 같은 기간 비교 · {전년연도}년 {CK(누적25)} ↔ {당해연도}년 {CK(누적26)}
            <span class="muted">(막대 높이는 월별과 다른 기준으로 그렸습니다)</span></div>
          <div class="bars">{bars}</div>
          <div class="legend"><span><span class="sw" style="background:{SLATE};"></span> {전년연도}년</span>
          <span><span class="sw" style="background:{T['accent']};"></span> {당해연도}년</span></div></div>
        <div class="card"><h3>매출액 구성 <span class="unitbadge">단위 {단위K}</span></h3>
          <div class="sub">1~{보고월}월 누적 · 손익 계정과목 기준</div>{mix_html}</div>
      </div>
    </div>"""
    st.html(상단html)

    # ── 손익계산서 (왼쪽 3) + 항목별 거래처 세부내역 (오른쪽 1) ──────────
    # 선택상자는 카드 위 별도 줄에 둡니다 — 그래야 왼쪽·오른쪽 카드의 제목 줄이 나란히 맞습니다
    _, 손익선택칸 = st.columns([2, 1], gap='small')
    with 손익선택칸:
        손익골라본 = st.selectbox('계정과목', 드릴목록_손익, key='dash_pl_drill',
                                 help=f'고른 항목의 1~{보고월}월 누적 금액을 거래처별로 보여줍니다')
    손익조건, 손익부호 = 항목조건_손익(손익골라본)
    손익세부html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{손익골라본} <span class="unitbadge">단위 USD</span></h3>
        <div class="sub">1~{보고월}월 누적 · 거래처별 · 상위 7곳 + 기타</div>
        {세부패널(누적IS[손익조건], '금액', 손익부호)}</div></div>"""
    손익예상안내 = (f' · <b>{예상월들[0]}~{예상월들[-1]}월(예상)</b> 칸은 만들어 두었고 '
                    f'금액은 아직 넣지 않았습니다' if 예상월들 else '')
    손익html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>손익계산서 <span class="unitbadge">단위 {단위K}</span></h3>
        <div class="sub">비용은 양수 · 판관비는 상위 4개 계정과 기타로 나눠 표시했습니다 ·
          영업이익 아래 <b>영업외손익 · 세전이익</b>까지 보여줍니다{손익예상안내}</div>
        <div class="scrollx"><table class="lined" {표스타일}>{콜그룹}<thead><tr><th>구분</th>
        {''.join(f'<th>{m}월</th>' for m in months)}
        {''.join(f'<th class="fc">{m}월<br><span class="fcs">예상</span></th>' for m in 예상월들)}
        <th>누적</th></tr></thead>
        <tbody>{rows}</tbody></table></div>{ebitda설명}</div></div>"""
    손익칸, 손익세부칸 = st.columns([2, 1], gap='small')   # 작은 화면에서도 오른쪽 세부내역이 읽히도록
    with 손익칸:
        st.html(손익html)
    with 손익세부칸:
        st.html(손익세부html)

    # ══ 공헌이익 분석 — 매출을 서비스별로 나누고 직접비를 빼서 봅니다 ══════
    # 배부기준: 매출액 − 직접원가 (실질 순매출) 비례 — 배송·상품의 대납성 청구액 제외 효과
    영문소 = 당해원장['계정영문'].astype(str).str.strip().str.lower()

    def 접두합(*접두들):
        마스크 = False
        for p in 접두들:
            마스크 = 마스크 | 영문소.str.startswith(p.lower())
        return float(당해원장.loc[마스크, '금액'].sum())

    # OTC 법인 — 제조(OEM/ODM)·상품·용역 세 갈래로 나눕니다.
    공헌매출 = {
        '제품': 접두합('4000'),
        '상품': 접두합('sales retail'),
        '용역': 접두합('services'),
    }
    # 직접원가 — 제품은 제조원가 계정, 상품은 퀵북 기본 매출원가 계정
    공헌직접 = {'제품': -접두합('5100', '5300', '5400', '5800', '5830', '5840',
                              '5850', '5860', 'cogs_1300'),
                '상품': -접두합('cost of goods sold', 'inventory shrinkage')}
    IS전 = 당해원장[당해원장['계정분류'].eq('IS')]

    def 과목합(*과목들):
        return -float(IS전.loc[IS전['계정과목'].isin(과목들), '금액'].sum())

    감가전 = 과목합('감가상각비', '무형자산상각비')
    임차전 = 과목합('지급임차료')
    급여전 = 과목합('급여')
    복리전 = 과목합('복리후생비')
    # 부분직접비 — 매출 규모에 따라 나눠 지는 고정성 비용(임차·상각)
    부분직접총 = 감가전 + 임차전
    판관비총 = -float(IS전.loc[IS전['분류'].eq('판관비'), '금액'].sum())
    공통비총 = 판관비총 - 감가전 - 임차전
    공헌기준 = {c: 공헌매출[c] - 공헌직접.get(c, 0.0) for c in 공헌매출}
    기준합 = sum(공헌기준.values())
    공헌매출합 = sum(공헌매출.values())
    if 기준합 > 1 and 공헌매출합 > 1:
        # 열 순서는 매출 규모 순으로 고정합니다
        고정순서 = ['제품', '상품', '용역']
        공헌순서 = ([c for c in 고정순서 if c in 공헌매출]
                    + [c for c in 공헌매출 if c not in 고정순서])

        def 공헌줄(이름, 값들, cls='', pct=False, calc=''):
            칸들 = ''
            for v in 값들:
                if pct:
                    t = f'<span class="neg">{v:.0f}%</span>' if v < -0.5 else f'{v:.0f}%'
                else:
                    t = CK(v) if abs(v * 배율 / 1000) >= 0.5 else '-'
                칸들 += f'<td>{t}</td>'
            설명 = f' <span class="calc">{calc}</span>' if calc else ''
            속성 = f' class="{cls}"' if cls else ''
            return f'<tr{속성}><td>{이름}{설명}</td>{칸들}</tr>'

        매출값 = [공헌매출[c] for c in 공헌순서]
        직접값 = [공헌직접.get(c, 0.0) for c in 공헌순서]
        부분값 = [부분직접총 * 공헌기준[c] / 기준합 for c in 공헌순서]
        공헌값 = [m - d - b for m, d, b in zip(매출값, 직접값, 부분값)]
        률값 = [(cm / m * 100 if m else 0.0) for cm, m in zip(공헌값, 매출값)]
        공통값 = [공통비총 * 공헌기준[c] / 기준합 for c in 공헌순서]
        영익값 = [cm - co for cm, co in zip(공헌값, 공통값)]
        공헌합, 공통합 = sum(공헌값), sum(공통값)
        공헌률 = 공헌합 / 공헌매출합 * 100
        고정창고 = 감가전 + 임차전

        def 합붙(값들, pct=False):
            return 값들 + [sum(값들)] if not pct else 값들

        머리 = ''.join(
            f'<th>{c}<br><span style="font-size:11px;font-weight:600;opacity:.8">'
            f'{공헌매출[c] / 공헌매출합 * 100:.1f}%</span></th>' for c in 공헌순서)
        칸수공헌 = len(공헌순서) + 1
        칸너비 = ('<colgroup><col style="width:19%">'
                  + f'<col style="width:{81 / 칸수공헌:.2f}%">' * 칸수공헌 + '</colgroup>')
        공헌본문 = 공헌줄('매출액', 합붙(매출값), 'total')
        공헌본문 += 공헌줄('직접원가', 합붙(직접값), calc='그 매출에만 붙는 비용')
        공헌본문 += 공헌줄('부분직접비 *', 합붙(부분값))
        공헌본문 += 공헌줄('공헌이익(손실)', 합붙(공헌값), 'cmhl')
        공헌본문 += 공헌줄('공헌이익률', 률값 + [공헌률], 'cmrate', pct=True)
        공헌본문 += 공헌줄('공통비 **', 합붙(공통값))
        공헌본문 += 공헌줄('영업이익(손실)', 합붙(영익값), 'total')
        로컬열림 = ' open' if os.name == 'nt' else ''
        공헌카드 = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>공헌이익 분석 <span class="unitbadge">1~{보고월}월 누적 · 단위 {단위K}</span></h3>
        <div class="sub">매출을 {len(공헌순서)}개 서비스로 나누고, 서비스에 직접 드는 비용(운송비 · 창고 인력 ·
          감가상각/리스상각 · 임차료)을 뺀 것이 공헌이익입니다 ·
          <b>영업이익 합계는 위 손익계산서와 정확히 일치합니다</b></div>
        <div class="kpi-row" style="--n:3">
          <div class="kpi-card"><div class="kpi-label">매출액 <span class="calc">(1~{보고월}월 누적)</span></div>
            <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(공헌매출합)} <span class="unit">{단위K}</span></div></div>
          <div class="kpi-card"><div class="kpi-label">공헌이익(손실)</div>
            <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(공헌합)} <span class="unit">{단위K}</span></div>
            <div class="kpi-delta"><span class="muted">공헌이익률 {공헌률:,.1f}%</span></div></div>
          <div class="kpi-card"><div class="kpi-label">영업이익(손실)</div>
            <div class="kpi-value" style="font-size:24px;font-weight:700">{CK(sum(영익값))} <span class="unit">{단위K}</span></div>
            <div class="kpi-delta"><span class="muted">공헌이익 − 공통비 {CK(공통합)}</span></div></div>
        </div>
        <div class="scrollx"><table class="cmtab">{칸너비}
        <thead><tr><th style="white-space:normal">구분<br><span style="font-size:11px;font-weight:600;opacity:.8">(아래 %는 매출 비중)</span></th>
        {머리}<th>합계</th></tr></thead>
        <tbody>{공헌본문}</tbody></table></div>
        <div class="cmfoot">* <b>부분직접비</b> — 성격은 직접비이지만 원장에서 매출 갈래별 추적이 안 되는 비용
          (공장·사무실 임차료 · 감가상각비) ·
          <b>배부기준 : 매출액 − 직접원가 (실질 순매출) 비례</b><br>
          ** <b>공통비</b> — 관리직 인건비(임원보수 · 사무직 급여 · 복리후생 · 급여세)와 그 밖의 판관비 · 배부기준 동일</div>
        <details class="calcnote"{로컬열림}>
          <summary>직접원가 · 부분직접비 · 공통비 구성 <span class="hint">원장 계정 합계 · 누르면 펼쳐집니다</span></summary>
          <div class="body">
          <b>직접원가 {CK(sum(직접값))}</b> — 제조원가(5100 원재료 · 5300 매입운임 · 5400 직접노무 ·
          5800대 시험·실험소모품·포장재) {CK(-접두합('5100', '5300', '5400', '5800', '5830', '5840', '5850', '5860'))} ·
          상품원가(Cost of Goods Sold) {CK(-접두합('cost of goods sold'))}<br>
          <b>부분직접비 {CK(부분직접총)}</b> — 감가상각비 {CK(감가전)} · 지급임차료(7900 · 7200) {CK(임차전)}<br>
          <b>공통비 {CK(공통비총)}</b> — 인건비(6000 급여) {CK(급여전)} ·
          복리후생 · 급여세 {CK(복리전)} · 그 밖의 판관비(지급수수료 · 보험료 · 전력비 · 건물관리비 등)
          {CK(공통비총 - 급여전 - 복리전)}</div></details>
        <div class="calcnote"><b>왜 공헌이익이 관리비를 못 넘나</b> —
          ① 공헌이익 {CK(공헌합)} ({공헌률:,.1f}%)로는 관리직 인건비를 포함한 공통비 {CK(공통합)}를
          다 갚지 못해 영업손실 {CK(sum(영익값))}가 됩니다.
          ② 가장 큰 원인은 공장 고정비 — 감가상각비 {CK(감가전)} + 임차료 {CK(임차전)}
          = {CK(고정창고)}, 매출의 {고정창고 / 공헌매출합 * 100:,.0f}% 입니다.
          ③ 배부는 「매출액 − 직접원가」(실질 순매출) 기준으로 했습니다 —
          제품 매출의 실질 기여는 {CK(공헌기준.get('제품', 0.0))} 수준입니다.
          ④ 감가상각 · 임차료 · 인건비는 고정비라, 물량이 늘면 새 매출의 약
          {(부분직접총 / 기준합 * -100 + 100):,.0f}%가 남아 적자가 빠르게 줄어드는 구조입니다.
          <span class="muted">(6000 급여의 생산직·관리직 구분, 5400 직접노무비와의 경계는 회계 담당자 확인 필요)</span></div>
      </div></div>"""

        # ── 왼쪽: 공헌이익 표 · 오른쪽: 고른 매출 구분의 거래처 세부 ──────
        구분앞 = [('4111b', '출고'), ('4111c', '출고'), ('4111a', '입고'), ('4120', '보관'),
                  ('4130', '배송'), ('4111d', '부가서비스'), ('4111e', '부가서비스'),
                  ('4111f', '부가서비스'), ('4100', '기타'), ('4160', '기타'), ('4190', '기타'),
                  ('4210', '상품'), ('4213', '상품'), ('4220', '상품'), ('sales of product', '상품')]

        def _매출구분(s):
            for p, c in 구분앞:
                if s.startswith(p):
                    return c
            return None

        구분열 = 영문소.map(_매출구분)
        # 이름 없는 재분류 전표(적요의 E0E0 = EGONGEGONG)는 그 거래처로 붙여서 상계되게 합니다
        보정이름 = 당해원장['Name'].copy()
        E0마스크 = 보정이름.isna() & 당해원장['Description'].astype(str).str.contains('E0E0', na=False)
        보정이름 = 보정이름.mask(E0마스크, 'EGONGEGONG, INC.')
        매출행들 = 당해원장[구분열.notna()].copy()
        매출행들['공헌구분'] = 구분열[구분열.notna()]
        매출행들['Name'] = 보정이름[구분열.notna()]
        # 상품 순액 재분류(E0E0) 전표는 거래처가 비어 있지만 EGONGEGONG 몫입니다 — 그쪽에서 차감
        매출행들['거래처'] = 매출행들['Name']
        재분류상품 = (매출행들['공헌구분'].eq('상품') & 매출행들['거래처'].isna()
                      & 매출행들['Description'].astype(str).str.contains('E0E0', case=False, na=False))
        매출행들.loc[재분류상품, '거래처'] = 'EGONGEGONG, INC.'

        _, 공헌선택칸 = st.columns([2, 1], gap='small')
        with 공헌선택칸:
            공헌골라본 = st.selectbox('매출 구분', 공헌순서, key='dash_cm_drill',
                                      help=f'고른 구분의 1~{보고월}월 누적 매출을 거래처별로 보여줍니다')
        공헌세부html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{공헌골라본} 매출 세부 <span class="unitbadge">단위 USD</span></h3>
        <div class="sub">1~{보고월}월 누적 · 거래처별 · 상위 7곳 + 기타</div>
        {세부패널(매출행들[매출행들['공헌구분'].eq(공헌골라본)], '금액', 1)}</div></div>"""
        공헌칸, 공헌세부칸 = st.columns([2, 1], gap='small')
        with 공헌칸:
            st.html(공헌카드)
        with 공헌세부칸:
            st.html(공헌세부html)

        # ── 거래처별 매출표 (누적 상위 5 + 기타) ──────────────────────
        거래처구분 = (매출행들.assign(거래처=매출행들['거래처'].fillna('(거래처 미기재)'))
                      .groupby(['거래처', '공헌구분'])['금액'].sum().unstack(fill_value=0.0)
                      .reindex(columns=공헌순서, fill_value=0.0))
        거래처합 = 거래처구분.sum(axis=1).sort_values(ascending=False)
        상위5 = list(거래처합.index[:5])
        if 상위5:
            def 고객줄(이름3, 값s, cls=''):
                칸 = ''.join(f'<td>{CK(v) if abs(v * 배율 / 1000) >= 0.5 else "-"}</td>' for v in 값s)
                속 = f' class="{cls}"' if cls else ''
                return (f'<tr{속}><td class="notranslate" translate="no">{이름3}</td>{칸}'
                        f'<td>{CK(sum(값s))}</td></tr>')

            고객본문 = ''
            for 이름3 in 상위5:
                고객본문 += 고객줄(이름3, [float(거래처구분.loc[이름3, c]) for c in 공헌순서])
            나머지 = 거래처구분.drop(index=상위5)
            if len(나머지):
                고객본문 += 고객줄(f'기타 ({len(나머지)}곳 합계)',
                                   [float(나머지[c].sum()) for c in 공헌순서])
            고객본문 += 고객줄('합계', [float(거래처구분[c].sum()) for c in 공헌순서], 'total')
            고객머리 = ''.join(f'<th>{c}</th>' for c in 공헌순서)
            st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>거래처별 매출 <span class="unitbadge">1~{보고월}월 누적 · 단위 {단위K}</span></h3>
        <div class="sub">누적 매출 상위 5개 거래처 + 기타 · 매출 구분별 ·
          상품 순액 재분류(E0E0)는 EGONGEGONG 에서 차감했습니다 ·
          아래에서 거래처를 고르면 그 업체의 손익이 나옵니다</div>
        <div class="scrollx"><table class="cmtab">{칸너비}
        <thead><tr><th>거래처</th>{고객머리}<th>합계</th></tr></thead>
        <tbody>{고객본문}</tbody></table></div></div></div>""")

            # ── 고른 거래처의 공헌이익 (같은 배부기준) ────────────────
            거래처골라본 = st.selectbox('거래처 (누적 매출 상위 5곳)', 상위5, key='dash_cm_cust',
                                        help='고른 거래처의 매출 구분별 손익을 보여줍니다')
            c매출 = [float(거래처구분.loc[거래처골라본, c]) for c in 공헌순서]
            # 상품원가(5200)는 무조건 상품매출원가 — 원장 거래처(재분류 E0E0 포함)로 추적하고,
            # 이름이 없는 잔여분만 거래처 총매출 비중으로 배부합니다. 나머지 직접원가(배송비·포워딩)는
            # 거래처 추적이 안 되므로 그 구분 매출에서 차지하는 비중으로 배부합니다.
            오공마스크 = 영문소.str.startswith('5200')
            상품원가별 = (-당해원장[오공마스크]
                          .groupby(보정이름[오공마스크].fillna('(거래처 미기재)'))['금액'].sum())
            상품원가잔여 = float(상품원가별.get('(거래처 미기재)', 0.0))
            거래처총합 = float(거래처합.sum()) or 1.0
            c직접 = []
            for i, c in enumerate(공헌순서):
                if c == '상품':
                    c직접.append(float(상품원가별.get(거래처골라본, 0.0))
                                 + 상품원가잔여 * float(거래처합.get(거래처골라본, 0.0)) / 거래처총합)
                else:
                    c직접.append(공헌직접.get(c, 0.0)
                                 * (c매출[i] / 공헌매출[c] if 공헌매출[c] else 0.0))
            c부분 = [부분직접총 * (m - d) / 기준합 for m, d in zip(c매출, c직접)]
            c공헌 = [m - d - b for m, d, b in zip(c매출, c직접, c부분)]
            c률 = [(cm / m * 100 if m else 0.0) for cm, m in zip(c공헌, c매출)]
            c공통 = [공통비총 * (m - d) / 기준합 for m, d in zip(c매출, c직접)]
            c영익 = [cm - co for cm, co in zip(c공헌, c공통)]
            c매출합 = sum(c매출) or 1.0
            c본문 = 공헌줄('매출액', 합붙(c매출), 'total')
            c본문 += 공헌줄('직접원가', 합붙(c직접),
                            calc='상품원가는 원장 추적 · 배송비 등은 매출 비중 추정')
            c본문 += 공헌줄('부분직접비 *', 합붙(c부분))
            c본문 += 공헌줄('공헌이익(손실)', 합붙(c공헌), 'cmhl')
            c본문 += 공헌줄('공헌이익률', c률 + [sum(c공헌) / c매출합 * 100], 'cmblue', pct=True)
            c본문 += 공헌줄('공통비 **', 합붙(c공통))
            c본문 += 공헌줄('영업이익(손실)', 합붙(c영익), 'total')
            st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3><span class="notranslate" translate="no">{거래처골라본}</span> 손익
        <span class="unitbadge">1~{보고월}월 누적 · 단위 {단위K}</span></h3>
        <div class="sub">위 공헌이익 분석과 같은 배부기준 — 상품원가(5200)는 원장 거래처로 추적(재분류
          E0E0 = EGONGEGONG 상계 포함), 배송비·포워딩은 그 구분 매출 비중으로, 부분직접비·공통비는
          실질 순매출 비례로 나눈 추정치입니다</div>
        <div class="scrollx"><table class="cmtab">{칸너비}
        <thead><tr><th>구분</th>{고객머리}<th>합계</th></tr></thead>
        <tbody>{c본문}</tbody></table></div>
        <div class="cmfoot">* <b>부분직접비</b> — 공장·사무실 임차료 · 감가상각비<br>
          ** <b>공통비</b> — 인건비(급여 · 복리후생 · 급여세)와 그 밖의
          판관비(지급수수료 · 보험료 · 전력비 · 건물관리비 · 수선비 등)</div></div></div>""")

    # ── 현금흐름 (왼쪽 3) + 항목별 거래처 세부내역 (오른쪽 1) ────────────
    if 자금표 is not None:
        누적자금상세 = 자금상세[자금상세['월'].le(보고월)] if 자금상세 is not None else None
        _, 자금선택칸 = st.columns([2, 1], gap='small')     # 카드 제목 줄을 맞추려고 선택상자만 따로
        with 자금선택칸:
            자금골라본 = st.selectbox('현금흐름 항목', [n for n, _, _ in 현금흐름목록], key='dash_cf_drill',
                                    help=f'고른 항목의 1~{보고월}월 누적 금액을 거래처별로 보여줍니다')
        _, 자금코드, 자금부호구분 = next(t for t in 현금흐름목록 if t[0] == 자금골라본)
        if 누적자금상세 is not None:
            자금대상 = 누적자금상세[누적자금상세['구분'].eq(자금코드)]
            if 자금부호구분 == '+':
                자금대상 = 자금대상[자금대상['영향'] > 0]
            elif 자금부호구분 == '-':
                자금대상 = 자금대상[자금대상['영향'] < 0]
        else:
            자금대상 = None
        자금세부html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{자금골라본} <span class="unitbadge">단위 USD</span></h3>
        <div class="sub">1~{보고월}월 누적 · 거래처별 · 상위 7곳 + 기타</div>
        {세부패널(자금대상, '영향', 1)}</div></div>"""
        자금칸, 자금세부칸 = st.columns([2, 1], gap='small')
        with 자금칸:
            st.html(자금html)
        with 자금세부칸:
            st.html(자금세부html)

    # ══ 차입금·리스부채 현황과 지급이자 ══════════════════════
    if 차입리스:
        잔고 = 차입리스['잔고']
        I25 = 차입리스['이자'].get('전년', {})
        I26 = 차입리스['이자'].get('당해', {})
        렌트합 = float(차입리스['이자'].get('렌트', 0.0))
        비고파일 = 기억자리('차입금비고.json')
        try:
            with open(비고파일, encoding='utf-8') as f:
                차입비고 = json.load(f)
        except Exception:
            차입비고 = {}
        # 엑셀 「차입금 비고」 시트 내용을 바탕에 깔고, 내 PC 에서 키인한 값이 있으면 덮어씁니다
        차입비고 = {**차입리스.get('엑셀비고', {}), **차입비고}

        def 원(v):
            """원단위 금액. 0 은 '-', 음수는 (빨간 괄호), 양수는 그냥 검정 숫자."""
            if v is None or abs(v) < 0.5:
                return '-'
            t = f'{abs(v):,.0f}'
            return f'<span class="neg">({t})</span>' if v < 0 else t

        def 차입라벨(key):
            if key in 차입이름:
                return f'{차입이름[key]} <span class="calc">({key})</span>'
            if key == 'auto':
                return '차량 할부금 <span class="calc">(단기 + 장기)</span>'
            if key == 'aro':
                return '복구충당부채 — 원상복구 <span class="calc">(ARO)</span>'
            for k2, 이름, _ in 리스이름:
                if key == k2:
                    return 이름
            if key == 'unknown':
                return ('어느 차입금 이자인지 적요에 은행명이 없는 것 '
                        '<span class="calc">(확인 필요)</span>')
            if key == 'etc':
                return '기타 이자 <span class="calc">(적요로 분류 안 된 것)</span>'
            return key.replace('loan_', '차입금 ')

        def 있음(key):
            b = 잔고.get(key, (0.0, 0.0))
            return (abs(b[0]) > 0.5 or abs(b[1]) > 0.5
                    or abs(I25.get(key, 0.0)) > 0.5 or abs(I26.get(key, 0.0)) > 0.5)

        모든키 = set(잔고) | set(I25) | set(I26)
        차입키 = sorted([k for k in 모든키 if not k.startswith('lease_')
                         and k not in ('aro', 'unknown', 'etc') and 있음(k)],
                        key=lambda k: -(잔고.get(k, (0.0, 0.0))[0] + 잔고.get(k, (0.0, 0.0))[1]))
        if 있음('unknown'):
            차입키.append('unknown')
        리스키 = [k for k, _, _ in 리스이름 if 있음(k)] + (['aro'] if 있음('aro') else [])
        기타키 = ['etc'] if 있음('etc') else []

        def 묶음합(키들):
            return (sum(잔고.get(k, (0.0, 0.0))[0] for k in 키들),
                    sum(잔고.get(k, (0.0, 0.0))[1] for k in 키들),
                    sum(I25.get(k, 0.0) for k in 키들),
                    sum(I26.get(k, 0.0) for k in 키들))

        def 비고글(key):
            t = str(차입비고.get(key, '') or '')
            return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        def 차입줄(key=None, 라벨텍스트=None, 합=None, 굵게=False):
            if 합 is None:
                기초v, 증감v = 잔고.get(key, (None, None)) if key in 잔고 else (None, None)
                i1, i2 = I25.get(key, 0.0), I26.get(key, 0.0)
            else:
                기초v, 증감v, i1, i2 = 합
            기말v = None if 기초v is None else 기초v + 증감v
            cls = ' class="total"' if 굵게 else ''
            return (f'<tr{cls}><td>{라벨텍스트 or 차입라벨(key)}</td>'
                    f'<td>{원(기초v)}</td><td>{원(증감v)}</td><td>{원(기말v)}</td>'
                    f'<td>{원(i1)}</td><td>{원(i2)}</td>'
                    f'<td class="lft">{비고글(key) if key else ""}</td></tr>')

        # 리스부채는 세부 대신 「한 줄 합산」으로만 보여줍니다 (담당자 요청)
        차입합 = 묶음합(차입키)
        리스합 = 묶음합(리스키)
        전체합 = tuple(a + b + c for a, b, c in zip(차입합, 리스합, 묶음합(기타키)))
        본문 = ''.join(차입줄(k) for k in 차입키)
        본문 += 차입줄(합=차입합, 라벨텍스트='① 차입금 (대출) 합계', 굵게=True)
        본문 += 차입줄(합=리스합, 라벨텍스트='② 리스부채 <span class="calc">'
                       '(OTC 법인은 현재 리스 계정이 없습니다)</span>')
        본문 += ''.join(차입줄(k) for k in 기타키)
        본문 += 차입줄(합=전체합, 라벨텍스트='합계 (①+②) = 원장 이자비용 총액', 굵게=True)

        작게 = 'style="font-size:11.5px;font-weight:600;opacity:.85"'
        미상안내 = ''
        if 있음('unknown'):
            미상안내 += (f'ⓘ 거래처 이름으로도 은행을 못 찾은 이자가 남아 있습니다'
                         f'({전년연도}년 {원(I25.get("unknown", 0.0))} / {당해연도}년 '
                         f'{원(I26.get("unknown", 0.0))}) — 회계 담당자 확인이 필요합니다.<br>')
        리스안내 = ''
        if abs(리스합[2]) > 0.5 or abs(리스합[3]) > 0.5:
            리스안내 = ('② 리스부채는 세부를 펼치지 않고 <b>한 줄로 합산</b>했습니다. '
                        '리스부채 증감에는 이자 가산과 월 납입이 섞여 있어 이자가 더 크면 잔액이 늘어납니다.<br>')
        차입카드 = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>차입금 현황과 지급이자 <span class="unitbadge">단위 USD (원단위)</span></h3>
        <div class="sub">잔액은 원장의 기초잔액과 {당해연도}년 1~{보고월}월 증감으로 계산 ·
          지급이자는 이자비용 전표를 적요·거래처 기준으로 나눈 금액 —
          <b>표의 이자 합계는 원장의 이자비용 총액과 일치합니다</b></div>
        <div class="scrollx"><table class="bigo">
        <thead><tr><th style="text-align:left">구분</th>
        <th>기초<br><span {작게}>{전년연도}년 말</span></th>
        <th>증감<br><span {작게}>{당해연도}년 1~{보고월}월</span></th>
        <th>기말<br><span {작게}>{당해연도}년 {보고월}월 말</span></th>
        <th>{전년연도}년 지급이자<br><span {작게}>1~12월</span></th>
        <th>{당해연도}년 지급이자<br><span {작게}>1~{보고월}월</span></th>
        <th>비고</th></tr></thead>
        <tbody>{본문}</tbody></table></div>
        <div class="calcnote"><b>읽을 때 주의</b> — {미상안내}{리스안내}
          ③ 잔액도 이자도 0 인 계정(우리은행 · SBA · Citibank 한도대출 등)은 줄을 만들지 않았고,
          차량 할부금의 장기 ↔ 단기 재분류는 합산하면서 서로 상계됩니다.<br>
          비고는 내 컴퓨터에서 연 대시보드에서만 아래 「비고 입력」으로 적을 수 있고(자동 저장 · 다음에 열어도 유지),
          웹(URL) 대시보드에서는 입력칸 없이 보기만 됩니다 — 웹에서도 비고가 보이려면
          실적 엑셀의 「차입금 비고」 시트에 적힌 것을 씁니다.</div></div></div>"""
        st.html(차입카드)

        # ── 비고 입력표 — 신한은행·한미은행만. 칸을 클릭해 적으면 자동 저장됩니다.
        #    평소에는 접혀 있고, 내 컴퓨터(윈도우)에서 열 때만 나타납니다 —
        #    웹(URL) 대시보드는 리눅스 서버라 입력표 없이 보기만 됩니다.
        입력키 = [k for k in ('2450', '2460') if k in 차입키]
        if 입력키 and os.name == 'nt':
          with st.expander('비고 입력 — 누르면 열립니다 (신한은행 · 한미은행 · 자동 저장)'):
            표라벨 = {k: re.sub(r'<[^>]+>', ' ', 차입라벨(k)).strip() for k in 입력키}
            비고입력표 = pd.DataFrame(
                {'차입처': [표라벨[k] for k in 입력키],
                 '비고 (담보 등 — 클릭해서 입력)': [str(차입비고.get(k, '') or '')
                                                    for k in 입력키]})
            고친표 = st.data_editor(비고입력표, hide_index=True, width='stretch',
                                    disabled=['차입처'], key='차입비고입력')
            새값 = dict(차입비고)
            바뀜 = False
            for n, k in enumerate(입력키):
                v = str(고친표.iloc[n, 1] or '').strip()
                if v != str(차입비고.get(k, '') or '').strip():
                    바뀜 = True
                if v:
                    새값[k] = v
                else:
                    새값.pop(k, None)
            if 바뀜:
                try:
                    os.makedirs(기억폴더, exist_ok=True)
                    with open(비고파일, 'w', encoding='utf-8') as f:
                        json.dump(새값, f, ensure_ascii=False, indent=1)
                    st.rerun()
                except Exception as e:
                    st.warning(f'저장하지 못했습니다: {e}')



# ══════════════════════════════════════════════════════════════
# 3-2. 월간실적보고 (보고서 5개 페이지)
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '월간실적':
    페이지 = st.session_state['rep_page']
    월목록 = list(range(1, 보고월 + 1))
    단위 = 통화
    환율안내 = ''

    # ── 공통 수치 ──────────────────────────────────────────
    매출N = [P.loc['매출액', m] for m in 월목록]
    원가N = [P.loc['매출원가', m] for m in 월목록]
    총이익N = [P.loc['매출총이익', m] for m in 월목록]
    판관N = [P.loc['판매관리비', m] for m in 월목록]
    세금N = [P.loc['법인세비용', m] for m in 월목록]
    영익N = [총이익N[i] - 판관N[i] for i in range(len(월목록))]
    영외N = [P.loc['영업외손익', m] for m in 월목록]

    누계 = lambda a: float(np.nansum(a))
    R, CG, GP = 누계(매출N), 누계(원가N), 누계(총이익N)
    SG, TAX = 누계(판관N), 누계(세금N)
    OP, NO = 누계(영익N), 누계(영외N)
    PBT = OP + NO
    NI = PBT - TAX
    SGP, OPP, PBTP = SG, OP, PBT      # 옛 이름 호환
    전매출 = float(전년동기.get('매출액', 0.0))
    전원가 = float(전년동기.get('매출원가', 0.0))
    전판관 = float(전년동기.get('판매관리비', 0.0))
    전총이익, 전영익 = 전매출 - 전원가, 전매출 - 전원가 - 전판관

    목표매출 = 입력값['월별'].get('매출 목표', [0.0] * 12)
    목표누계 = float(np.nansum(목표매출[:보고월]))

    def 카드(라벨, 값, 부가='', 델타='', 색='', 단위표기=None):
        """단위표기=None 이면 K USD, ''(빈칸)이면 단위 없이, 그 밖에는 그 글자로."""
        u = 단위K if 단위표기 is None else 단위표기
        단위html = f' <span class="unit">{u}</span>' if u else ''
        return (f'<div class="kpi-card"><div class="kpi-label">{라벨}</div>'
                f'<div class="kpi-value {색}" style="font-size:24px;font-weight:700">{값}{단위html}</div>'
                f'<div class="kpi-delta"><span class="muted">{부가}</span> {델타}</div></div>')

    def 표(제목, 부제, thead, rows, 높이=None):
        return (f'<div class="card"><h3>{제목}<span class="unitbadge">단위 {단위}</span></h3>'
                f'<div class="sub">{부제}</div>'
                f'<table class="lined"><thead><tr>{thead}</tr></thead>'
                f'<tbody>{rows}</tbody></table></div>')

    def 알림(제목, 본문):
        return (f'<div class="card"><h3>{제목}</h3><div class="sub" '
                f'style="margin:0;line-height:1.75;white-space:pre-line">{본문}</div></div>')

    def 서술(키, 기본=''):
        v = 입력값['서술'].get(키, '')
        return v if str(v).strip() else 기본

    st.write(f'### {페이지}')
    st.caption(f'{당해연도}년 {보고월}월 누적 · 원장(QuickBooks GL) 기준 · '
               f'표(금액) {단위} · 그래프 {단위K}{환율안내}')

    # ── 1. Executive Summary ───────────────────────────────
    if 페이지.startswith('1.'):
        달성 = f'목표 대비 {R / 목표누계 * 100:,.1f}%' if 목표누계 else '목표 미입력'
        html = f"""{CARD_CSS}<div class="wrap" translate="no">
          <div class="kpi-row" style="--n:4">
            {카드('매출액', CK(R), f'1~{보고월}월 누적',
                  증감HTML(R, 전매출, '전년동기') if 전매출 else '')}
            {카드('매출총이익', CK(GP), f'이익률 {GP / R * 100:,.1f}%' if R else '')}
            {카드('영업이익(손실)', CK(OP), f'이익률 {OP / R * 100:,.1f}%' if R else '')}
            {카드('현금 잔고', CK(현금잔고.get(보고월, 0.0)), f'{보고월}월 말 · 원장 계산',
                  증감HTML(현금잔고.get(보고월, 0.0), 현금잔고.get(보고월 - 1) if 보고월 > 1 else None))}
          </div>
          <div class="kpi-row" style="--n:4">
            {카드('판관비 (SG&A)', CK(SG), f'매출 대비 {SG / R * 100:,.1f}%' if R else '')}
            {카드('영업외손익', CK(-NO), '이자 · 외환', '', 'warn' if NO < 0 else '')}
            {카드('세전이익', CK(PBT), '영업이익 + 영업외손익')}
            {카드('당기순이익', CK(NI), f'법인세 {CK(TAX)} {단위K} 차감 후')}
          </div>
          {알림('🔴 Red Flag — 리스크 · 이슈',
                서술('Red Flag', '실적파일의 「월간보고_입력」 시트에 적어주시면 여기에 나옵니다.'))}
          {알림('🟢 Green Flag — 성과 · 기회',
                서술('Green Flag', '실적파일의 「월간보고_입력」 시트에 적어주시면 여기에 나옵니다.'))}
          <div class="note">매출·이익·판관비·현금 잔고는 원장에서 자동 계산됩니다.
            Flag와 진척률처럼 원장에 없는 항목만 「월간보고_입력」 시트에서 읽습니다. · {달성}</div>
        </div>"""
        st.html(html)

        차트제목('월별 기말 현금 잔고')
        현금df1 = pd.DataFrame({'월': [f'{m}월' for m in 월목록],
                               '금액': [K(현금잔고.get(m, 0.0)) for m in 월목록]})
        기본c = alt.Chart(현금df1).encode(
            x=alt.X('월:N', sort=[f'{m}월' for m in 월목록], title=None,
                    axis=alt.Axis(labelAngle=0)),
            y=Y축(현금df1['금액'], 단위K),
            tooltip=[alt.Tooltip('월:N'), alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
        st.altair_chart((기본c.mark_bar(cornerRadiusEnd=4, size=38, color=TEAL)
                         + 값라벨(기본c, dy=-9)).properties(height=320),
                        use_container_width=True)
        st.caption(f'원장의 현금·예금 계정에서 계산했습니다 · 기초 {CK(기초현금, 괄호=False)} '
                   f'{단위K} → {보고월}월 말 '
                   f'{CK(현금잔고.get(보고월, 0.0), 괄호=False)} {단위K}')

    # ── 2. P&L ─────────────────────────────────────────────
    elif 페이지.startswith('2.'):
        계획 = 입력값['월별']
        def 계획누계(k):
            v = 계획.get(k)
            if not v:
                return None
            t = float(np.nansum(v[:보고월]))
            return t if abs(t) > 0.005 else None

        행 = [('매출액', R, 계획누계('매출 목표'), 전매출, 0),
              ('매출원가', CG, 계획누계('매출원가 목표'), 전원가, 0),
              ('매출총이익', GP, None, 전총이익, 1),
              ('판관비 (SG&A)', SG, 계획누계('판관비 목표'), 전판관, 0),
              ('영업이익', OP, 계획누계('영업이익 목표'), 전영익, 1),
              ('영업외손익 (비용)', -NO, None, None, 0),
              ('세전이익', PBT, None, None, 1),
              ('법인세비용', TAX, None, None, 0),
              ('당기순이익', NI, None, None, 1)]
        rows = ''
        for 이름, 실적, 계획v, 전년v, 굵게 in 행:
            차 = ('-' if 계획v in (None, 0) else C(실적 - 계획v))
            전 = '-' if 전년v is None else C(전년v)
            증 = (f'{(실적 / 전년v - 1) * 100:,.1f}%'
                 if 전년v and 전년v > 0 else '-')
            cls = ' style="font-weight:700"' if 굵게 else ''
            rows += (f'<tr{cls}><td>{이름}</td><td>{C(실적)}</td>'
                     f'<td>{"-" if 계획v is None else C(계획v)}</td><td>{차}</td>'
                     f'<td>{전}</td><td>{증}</td></tr>')
        기본remark = ('연결패키지 손익계산서와 같은 기준입니다. '
                      '매출원가에는 원재료 · 직접노무비 · 시험분석 · 실험소모품 · 포장재가 들어갑니다.')
        머리 = ('<th>구분</th><th>실적</th><th>계획</th><th>계획 대비</th>'
                f'<th>{전년연도}년 동기</th><th>전년 대비</th>')
        html = f"""{CARD_CSS}<div class="wrap" translate="no">
          {표('P&L 요약', f'{당해연도}년 1~{보고월}월 누적 · 단위 {단위}', 머리, rows)}
          {알림('💡 P&L Remark', 서술('P&L Remark', 기본remark))}
        </div>"""
        st.html(html)

        긴 = pd.DataFrame({'월': [f'{m}월' for m in 월목록] * 2,
                           '구분': ['매출액'] * len(월목록) + ['영업이익'] * len(월목록),
                           '금액': [K(v) for v in 매출N + 영익N]})
        기본 = alt.Chart(긴).encode(
            x=alt.X('월:N', sort=[f'{m}월' for m in 월목록], title=None,
                    axis=alt.Axis(labelAngle=0)),
            xOffset=alt.XOffset('구분:N', sort=['매출액', '영업이익']),
            y=Y축(긴['금액'], 단위K),
            tooltip=[alt.Tooltip('월:N'), alt.Tooltip('구분:N'),
                     alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
        막대 = 기본.mark_bar(cornerRadiusEnd=3, size=17).encode(
            color=alt.Color('구분:N', sort=['매출액', '영업이익'], scale=alt.Scale(
                domain=['매출액', '영업이익'], range=[BLUE, AMBER]), title=None))
        차트제목('월별 매출 · 영업이익')
        st.altair_chart((막대 + 값라벨(기본)).properties(height=330),
                        use_container_width=True)

    # ── 3. SG&A ────────────────────────────────────────────
    elif 페이지.startswith('3.'):
        구성 = pd.DataFrame({
            '항목': 판관비항목,
            '금액': [K(float(누적(P).get(it, 0.0))) for it in 판관비항목]})
        구성 = 구성[구성['금액'].abs() > 0.0005].sort_values('금액', ascending=False)
        c1, c2 = st.columns([1, 1])
        with c1:
            b1 = alt.Chart(구성).encode(
                y=alt.Y('항목:N', sort=list(구성['항목']), title=None),
                x=X축(구성['금액'], 단위K),
                tooltip=[alt.Tooltip('항목:N'), alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
            차트제목('판관비 구성 — 활동분류(보고항목)')
            st.altair_chart((b1.mark_bar(color=TEAL, cornerRadiusEnd=4, size=16)
                             + 값라벨(b1, dy=0, dx=7, align='left')).properties(height=380),
                            use_container_width=True)
        with c2:
            계정 = (당해원장[당해원장['분류'].eq('판관비') & 당해원장['월'].le(보고월)]
                    .groupby('계정과목')['금액'].sum().mul(-배율 / 1000)
                    .sort_values(ascending=False).head(10).reset_index())
            b2 = alt.Chart(계정).encode(
                y=alt.Y('계정과목:N', sort=list(계정['계정과목']), title=None),
                x=X축(계정['금액'], 단위K),
                tooltip=[alt.Tooltip('계정과목:N'), alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
            차트제목('판관비 구성 — 계정과목 상위 10')
            st.altair_chart((b2.mark_bar(color=BLUE, cornerRadiusEnd=4, size=16)
                             + 값라벨(b2, dy=0, dx=7, align='left')).properties(height=380),
                            use_container_width=True)

        추이 = pd.DataFrame({
            '월': [f'{m}월' for m in range(1, 13)] * 2,
            '연도': [f'{당해연도}년'] * 12 + [f'{전년연도}년'] * 12,
            '금액': [(K(P.loc['판매관리비', m]) if m <= 보고월 else None) for m in range(1, 13)]
                    + [K(P전.loc['판매관리비', m]) for m in range(1, 13)]})
        기본2 = alt.Chart(추이.dropna()).encode(
            x=alt.X('월:N', sort=[f'{m}월' for m in range(1, 13)], title=None,
                    axis=alt.Axis(labelAngle=0)),
            y=Y축(추이['금액'], 단위K),
            color=alt.Color('연도:N', scale=alt.Scale(
                domain=[f'{당해연도}년', f'{전년연도}년'], range=[TEAL, SLATE]), title=None),
            tooltip=[alt.Tooltip('월:N'), alt.Tooltip('연도:N'),
                     alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
        올해선 = 기본2.transform_filter(alt.datum.연도 == f'{당해연도}년')
        차트제목('월별 판관비 추이')
        st.altair_chart((기본2.mark_line(point=alt.OverlayMarkDef(size=70), strokeWidth=2.5)
                         + 값라벨(올해선, dy=-13)).properties(height=310),
                        use_container_width=True)
        st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
          {알림('💡 SG&A Remark', 서술('SG&A Remark',
                '전월·계획 대비 변동 원인을 「월간보고_입력」 시트에 적어주시면 여기에 나옵니다.'))}
        </div>""")

    # ── 4. 운영 KPI ────────────────────────────────────────
    elif 페이지.startswith('4.'):
        단일 = 입력값['단일']
        번율 = 단일.get('Burn Rate (월, USD)') or 단일.get('Burn Rate (월, CAD)')
        if 번율 in (None, 0):
            번율 = -(현금잔고.get(보고월, 0.0) - 현금잔고.get(1, 0.0)) / max(보고월 - 1, 1)
            번율출처 = '2~%d월 평균 현금 감소 (자동 계산)' % 보고월
        else:
            번율출처 = '입력값'
        런웨이 = (현금잔고.get(보고월, 0.0) / 번율) if 번율 and 번율 > 0 else None

        def 값(k, 단위표시='%'):
            v = 단일.get(k)
            return f'{v:,.1f}{단위표시}' if isinstance(v, (int, float)) else '입력 필요'

        html = f"""{CARD_CSS}<div class="wrap" translate="no">
          <div class="kpi-row" style="--n:4">
            {카드('현금 잔고', CK(현금잔고.get(보고월, 0.0)), f'{보고월}월 말 · 원장 계산',
                  증감HTML(현금잔고.get(보고월, 0.0), 현금잔고.get(보고월 - 1) if 보고월 > 1 else None))}
            {카드('Burn Rate', CK(번율), 번율출처)}
            {카드('Runway', f'{런웨이:,.1f}' if 런웨이 else '-',
                  '현금 ÷ Burn Rate', 단위표기='개월')}
            {카드('R&D 집행률', 값('R&D 집행률 (%)'), '연간 예산 대비', 단위표기='')}
          </div>
          <div class="kpi-row" style="--n:3">
            {카드('파일럿 수율', 값('파일럿 수율 (%)'), f"목표 {값('목표 수율 (%)')}", 단위표기='')}
            {카드('R&D 마일스톤', 값('R&D 마일스톤 (%)'), '이번달 목표', 단위표기='')}
            {카드('판관비 중 인건비', f"{누적(P).get('인건비성 항목', 0.0) / SG * 100:,.1f}%" if SG else '-',
                  '원장 계산', 단위표기='')}
          </div>"""
        진척 = 입력값['진척']
        if 진척:
            rows = ''.join(f'<tr><td>{k}</td><td>{v:,.0f}%</td></tr>' for k, v in 진척.items())
            html += 표('개발 · 운영 진척', '「월간보고_입력」 시트의 값입니다',
                       '<th>항목</th><th>진척률</th>', rows)
        else:
            html += 알림('개발 · 운영 진척',
                         '원장에 없는 값입니다. 실적파일 「월간보고_입력」 시트의 '
                         '「■ 진척률」 구역에 항목과 % 를 적어주시면 여기에 표로 나옵니다.')
        html += 알림('💡 CTK OTC LAB Remark',
                     서술('KPI Remark', 'R&D 보조금·품질·인증 진행 상황을 '
                          '「월간보고_입력」 시트에 적어주시면 여기에 나옵니다.'))
        html += '</div>'
        st.html(html)

        현금df = pd.DataFrame({'월': [f'{m}월' for m in 월목록],
                              '금액': [K(현금잔고.get(m, 0.0)) for m in 월목록]})
        기본3 = alt.Chart(현금df).encode(
            x=alt.X('월:N', sort=[f'{m}월' for m in 월목록], title=None,
                    axis=alt.Axis(labelAngle=0)),
            y=Y축(현금df['금액'], 단위K),
            tooltip=[alt.Tooltip('월:N'), alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
        차트제목(f'월별 기말 현금 잔고  (기초 {CK(기초현금)} {단위K})')
        st.altair_chart((기본3.mark_bar(cornerRadiusEnd=4, size=34, color=TEAL)
                         + 값라벨(기본3, dy=-9)).properties(height=330),
                        use_container_width=True)

    # ── 5. Sales Forecast ──────────────────────────────────
    else:
        매출행 = 당해원장[당해원장['분류'].eq('매출') & 당해원장['월'].le(보고월)]
        거래처 = (매출행.groupby(매출행['Name'].fillna('(거래처 미기재)'))
                  .agg(누적=('금액', 'sum')).reset_index()
                  .rename(columns={'Name': '고객사'}).sort_values('누적', ascending=False))
        당월 = 매출행[매출행['월'].eq(보고월)].groupby(
            매출행['Name'].fillna('(거래처 미기재)'))['금액'].sum()
        전월s = 매출행[매출행['월'].eq(보고월 - 1)].groupby(
            매출행['Name'].fillna('(거래처 미기재)'))['금액'].sum() if 보고월 > 1 else pd.Series(dtype=float)
        고객목표 = 입력값['고객목표']

        rows = ''
        for i, r in enumerate(거래처.head(12).itertuples(), 1):
            이름 = getattr(r, '고객사')
            c, p = float(당월.get(이름, 0.0)), float(전월s.get(이름, 0.0))
            t = float(고객목표.get(이름, 0) or 0)
            증감 = '-' if abs(c - p) < 0.5 else (
                ('▲ ' if c - p > 0 else '▼ ') + C(abs(c - p)))
            색 = '' if abs(c - p) < 0.5 else ('up' if c - p > 0 else 'down')
            rows += (f'<tr><td>{i}</td>'
                     f'<td class="lft notranslate" translate="no">{이름}</td><td>{C(c)}</td>'
                     f'<td>{C(t) if t else "-"}</td>'
                     f'<td>{f"{c / t * 100:,.0f}%" if t else "-"}</td>'
                     f'<td class="{색}">{증감}</td>'
                     f'<td>{C(r.누적)}</td>'
                     f'<td>{r.누적 / R * 100:,.1f}%</td></tr>')
        rows += (f'<tr style="font-weight:700"><td></td><td class="lft">합계 (전체 거래처)</td>'
                 f'<td>{C(P.loc["매출액", 보고월])}</td><td>-</td><td>-</td>'
                 f'<td>-</td><td>{C(R)}</td><td>100.0%</td></tr>')
        html = f"""{CARD_CSS}<div class="wrap" translate="no">
          {표('주요 고객사별 실적', f'{보고월}월 · 원장의 거래처(Name) 기준 · 단위 {단위}',
              '<th>No</th><th class="lft">고객사</th><th>이번달</th><th>이번달 목표</th><th>달성률</th>'
              '<th>전월 대비</th><th>누적</th><th>비중</th>', rows)}
        </div>"""
        st.html(html)

        계열 = []
        for m in range(1, 13):
            if m <= 보고월:
                계열.append({'월': f'{m}월', '구분': f'{당해연도}년 실적',
                            '금액': K(P.loc['매출액', m])})
            if 목표매출 and 목표매출[m - 1]:
                계열.append({'월': f'{m}월', '구분': '목표', '금액': K(목표매출[m - 1])})
            계열.append({'월': f'{m}월', '구분': f'{전년연도}년',
                        '금액': K(P전.loc['매출액', m])})
        추이 = pd.DataFrame(계열)
        도메인 = [f'{당해연도}년 실적', f'{전년연도}년']
        색상 = [BLUE, SLATE]
        if (추이['구분'] == '목표').any():
            도메인 = [f'{당해연도}년 실적', '목표', f'{전년연도}년']
            색상 = [BLUE, AMBER, SLATE]
        기본4 = alt.Chart(추이).encode(
            x=alt.X('월:N', sort=[f'{m}월' for m in range(1, 13)], title=None,
                    axis=alt.Axis(labelAngle=0)),
            xOffset=alt.XOffset('구분:N', sort=도메인),
            y=Y축(추이['금액'], 단위K),
            tooltip=[alt.Tooltip('월:N'), alt.Tooltip('구분:N'),
                     alt.Tooltip('금액:Q', format=',.0f', title=단위K)])
        실적만 = 기본4.transform_filter(alt.datum.구분 == f'{당해연도}년 실적')
        차트제목('월별 매출 — 실적 · 목표 · 전년')
        st.altair_chart((기본4.mark_bar(cornerRadiusEnd=3, size=13).encode(
            color=alt.Color('구분:N', sort=도메인, scale=alt.Scale(
                domain=도메인, range=색상), title=None))
            + 값라벨(실적만, dy=-8)).properties(height=330),
            use_container_width=True)
        if not 목표매출 or not any(목표매출):
            st.caption('목표(Plan)는 실적파일 「월간보고_입력」 시트의 '
                       '「■ 월별 목표」 구역에 넣으시면 자동으로 표시됩니다.')

# ══════════════════════════════════════════════════════════════
# 4. 미수채권 관리
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '미수채권 관리':
    st.write(f'### 미수채권 관리 — {기준일.date()} 기준')
    st.caption('통화 단위: USD · 만기는 「매출한 달의 다음 달 말일」 · '
               '거래처마다 오래된 매출부터 회수된 것으로 봅니다 (선입선출)')

    ar_all = 당해원장[당해원장['계정과목'].eq('매출채권')].copy()
    if ar_all.empty:
        st.warning('매출채권(Accounts Receivable) 계정을 찾지 못했습니다.')
        st.stop()
    기초채권 = float((BS밑자료 or {}).get('전기', {}).get('매출채권', 0.0) or 0.0)
    나이표, 초과회수 = 채권나이표(ar_all, 기초채권, 기준일)
    if 나이표.empty:
        st.success('현재 미수 잔액이 있는 거래처가 없습니다.')
        st.stop()

    총미수 = float(나이표['채권총액'].sum())
    정상 = float(나이표['within Due'].sum())
    연체 = 총미수 - 정상
    장기 = float(나이표[['~120', '121~']].sum().sum())

    kpi = f"""{CARD_CSS}<div class="wrap" translate="no"><div class="kpi-row" style="--n:5">
      <div class="kpi-card"><div class="kpi-label">총 미수채권</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{금액(총미수)} <span class="unit">USD</span></div></div>
      <div class="kpi-card"><div class="kpi-label">만기 전 (within Due)</div>
        <div class="kpi-value ok" style="font-size:24px;font-weight:700">{금액(정상)} <span class="unit">({정상/총미수*100:,.1f}%)</span></div></div>
      <div class="kpi-card"><div class="kpi-label">만기 지난 것</div>
        <div class="kpi-value warn" style="font-size:24px;font-weight:700">{금액(연체)} <span class="unit">({연체/총미수*100:,.1f}%)</span></div></div>
      <div class="kpi-card"><div class="kpi-label">91일 넘은 것</div>
        <div class="kpi-value warn" style="font-size:24px;font-weight:700">{금액(장기)} <span class="unit">({장기/총미수*100:,.1f}%)</span></div></div>
      <div class="kpi-card"><div class="kpi-label">미수거래처 수</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{len(나이표)}개</div></div>
    </div></div>"""
    st.html(kpi)

    # ── 거래처별 채권 연령표 (금액이 큰 곳부터)
    구간합 = {c: float(나이표[c].sum()) for c in 채권구간}
    머리 = ''.join(f'<th>{c}</th>' for c in 채권구간)
    합계줄 = ''.join(f'<td>{금액(구간합[c])}</td>' for c in 채권구간)
    줄들 = ''
    기초색 = f' style="color:{T["ink3"]}"'
    for _i, 행 in 나이표.iterrows():
        칸 = ''.join(f'<td>{금액(float(행[c]))}</td>' for c in 채권구간)
        이름꼴 = 기초색 if str(행['거래처명']).startswith('기초채권') else ''
        총 = 금액(float(행['채권총액']))
        줄들 += ('<tr class="sub"><td class="lft"' + 이름꼴 + '>'
                 + str(행['거래처명']) + '</td>' + 칸
                 + '<td><b>' + 총 + '</b></td><td></td></tr>')
    st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>거래처별 채권 연령표 <span class="unitbadge">단위 USD</span></h3>
        <div class="sub">채권총액이 큰 곳부터 · 만기 = 매출한 달의 다음 달 말일 ·
          경과일은 만기일에서 {기준일.date()} 까지</div>
        <div class="stick"><table class="lined" style="min-width:980px; margin-top:8px">
          <colgroup><col style="width:24%">{'<col style="width:9.5%">' * 6}
            <col style="width:11%"><col style="width:8%"></colgroup>
          <thead><tr><th>거래처명</th>{머리}<th>채권총액</th><th>대손충당금</th></tr></thead>
          <tbody>{줄들}
            <tr class="total"><td class="lft">합계</td>{합계줄}
              <td><b>{금액(총미수)}</b></td><td></td></tr></tbody></table></div>
        <div class="calcnote"><b>어떻게 셌나</b> — 거래처마다 올해 매출(차변)을 날짜순으로 쌓아 두고,
          들어온 돈(대변)으로 <b>오래된 것부터</b> 지웠습니다. 남은 매출만 만기일 기준으로 칸을 나눴습니다.<br>
          <b>기초채권</b> {금액(기초채권)} 은 25년에서 넘어온 몫이라 거래처를 붙이지 않았습니다.
          올해 매출보다 더 받은 돈 {금액(초과회수)} 은 이 기초채권을 갚은 것으로 보아 빼고
          <b>{금액(max(기초채권 - 초과회수, 0))}</b> 만 남겼습니다.<br>
          합계 <b>{금액(총미수)}</b> 은 재무상태표의 {보고월}월말 매출채권과 같습니다.</div>
      </div></div>""")

    c1, c2 = st.columns(2)
    with c1:
        정상비 = 정상 / 총미수 * 100 if 총미수 else 0
        st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
          <div class="card"><h3>만기 전 vs 만기 지난 것 <span class="unitbadge">단위 USD</span></h3>
            <div class="sub">만기는 매출한 달의 다음 달 말일입니다</div>
            <div class="split">
              <i class="ok" style="flex:{max(정상비, 0.5):.2f}">만기 전 {정상비:,.1f}%</i>
              <i class="bad" style="flex:{max(100 - 정상비, 0.5):.2f}">{100 - 정상비:,.1f}%</i>
            </div>
            <table><tbody>
              <tr><td class="lft"><span class="sw" style="background:{T['navy7']}"></span>
                  만기 전 (within Due)</td><td>{금액(정상)}</td>
                  <td class="muted">{정상비:,.1f}%</td></tr>
              <tr><td class="lft"><span class="sw" style="background:{ROSE}"></span>
                  만기 지난 것</td><td>{금액(연체)}</td>
                  <td class="muted">{100 - 정상비:,.1f}%</td></tr>
            </tbody></table></div></div>""")
    with c2:
        차트제목('경과 구간별 미수채권', 'USD')
        구간df = pd.DataFrame({'구간': 채권구간,
                               '미수잔액': [구간합[c] for c in 채권구간]})
        st.altair_chart(alt.Chart(구간df).mark_bar(color=BLUE).encode(
            x=alt.X('구간:N', sort=채권구간, title=None, axis=alt.Axis(labelAngle=0)),
            y=alt.Y('미수잔액:Q', title=None)),
            width='stretch', key='ar_bucket')

    차트제목('월별 매출 대비 미수채권 잔액', 'USD')
    월말잔액 = ar_all.sort_values('거래일').groupby('월')['Balance'].last()
    ms = [m for m in range(1, 보고월 + 1)]
    bar_df = pd.DataFrame({
        '월': [f'{m}월' for m in ms] * 2,
        '구분': ['매출 발생액'] * len(ms) + ['미수채권 잔액'] * len(ms),
        '금액': [P.loc['매출액', m] for m in ms] + [월말잔액.get(m, np.nan) for m in ms]})
    순서 = [f'{m}월' for m in ms]
    st.altair_chart(alt.Chart(bar_df).mark_bar().encode(
        x=alt.X('월:N', sort=순서, title=None, axis=alt.Axis(labelAngle=0)),
        xOffset='구분:N', y=alt.Y('금액:Q', title=None),
        color=alt.Color('구분:N', title=None, scale=alt.Scale(
            domain=['매출 발생액', '미수채권 잔액'], range=[BLUE, TEAL]))),
        width='stretch', key='ar_trend')

    st.write('**만기가 지난 채권 (거래처별)**')
    연체표 = 나이표[나이표[채권구간[1:]].sum(axis=1) > 0.5].copy()
    if len(연체표):
        연체표['만기 지난 금액'] = 연체표[채권구간[1:]].sum(axis=1)
        보기 = 연체표[['거래처명', '만기 지난 금액'] + 채권구간[1:] + ['채권총액']]
        st.dataframe(보기, width='stretch', hide_index=True,
                     column_config={c: st.column_config.NumberColumn(format='%,.0f')
                                    for c in ['만기 지난 금액', '채권총액'] + 채권구간[1:]})
    else:
        st.success('만기가 지난 채권이 없습니다.')


# ══════════════════════════════════════════════════════════════
# 5. 당월 실적집계
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '당월 실적보고':
    # ── 볼 달 고르기 : ◀ ▶ 로 옮기거나 숫자를 바로 골라도 됩니다 ─────
    if 'mo_month' not in st.session_state:
        st.session_state['mo_month'] = 보고월
    st.session_state['mo_month'] = min(max(int(st.session_state['mo_month']), 1), 보고월)

    def _달이동(걸음):
        st.session_state['mo_month'] = min(max(st.session_state['mo_month'] + 걸음, 1), 보고월)

    def _이번달로():
        st.session_state['mo_month'] = 보고월

    선택월 = st.session_state['mo_month']
    st.write(f'### {당해연도}년 {선택월}월 실적보고')
    st.caption(f'단위: USD · 비용 항목은 양수로 표시 · 원장 직접 집계 '
               f'· 원장에 {당해연도}년 1~{보고월}월 자료가 있습니다')

    앞칸, 달칸, 뒤칸, 오늘칸, 남은칸 = st.columns([0.5, 1.1, 0.5, 0.9, 5], gap='small')
    with 앞칸:
        st.button('◀', key='mo_prev', width='stretch', disabled=선택월 <= 1,
                  on_click=_달이동, args=(-1,), help='한 달 앞으로')
    with 달칸:
        st.selectbox('보는 달', list(range(1, 보고월 + 1)), key='mo_month',
                     format_func=lambda m: f'{m}월', label_visibility='collapsed')
    with 뒤칸:
        st.button('▶', key='mo_next', width='stretch', disabled=선택월 >= 보고월,
                  on_click=_달이동, args=(1,), help='한 달 뒤로')
    with 오늘칸:
        st.button('최종월', key='mo_last', width='stretch', disabled=선택월 == 보고월,
                  on_click=_이번달로, help=f'가장 마지막 달({보고월}월)로')
    선택월 = st.session_state['mo_month']

    이전월 = 선택월 - 1 if 선택월 > 1 else None
    기준선택 = st.radio('판관비 구분', ['손익계산서 비용', '활동성 비용'],
                        horizontal=True, key='mo_basis',
                        help='판매관리비를 손익계산서 계정과목으로 볼지, 활동분류로 볼지 고릅니다')
    손익기준_당월 = 기준선택.startswith('손익')
    표기준, 표기준전 = ((P손익, P손익전) if 손익기준_당월 else (P, P전))
    수준표 = 표기준.attrs['수준']
    보일줄 = [i for i in 표기준.index if i not in 집계제외]
    P보기, P전보기 = 표기준.loc[보일줄], 표기준전.reindex(표기준.index).loc[보일줄]
    표 = pd.DataFrame({
        '구분': 보일줄,
        f'당월({선택월}월)': P보기[선택월].values,
        f'전월({이전월}월)' if 이전월 else '전월': P보기[이전월].values if 이전월 else np.nan,
        f'전년동월({str(전년연도)[2:]}년 {선택월}월)': P전보기[선택월].values,
        f'당해누적({str(당해연도)[2:]}년 1월~{선택월}월)': 누적(P보기, upto=선택월).values,
        f'전년동기({str(전년연도)[2:]}년 1월~{선택월}월)': 누적(P전보기, upto=선택월).values,
    })
    표['전월 증감'] = 표.iloc[:, 1] - 표.iloc[:, 2]
    표['전년동월 증감'] = 표.iloc[:, 1] - 표.iloc[:, 3]
    표['누적 증감'] = 표.iloc[:, 4] - 표.iloc[:, 5]

    # ── 세 덩어리로 묶어 보여줍니다 (셋트마다 왼쪽에 구분선) ──────
    셋트 = [('전월대비', [표.columns[1], 표.columns[2], '전월 증감']),
            ('전년 동월대비', [표.columns[3], '전년동월 증감']),
            (f'{당해연도}년 {선택월}월 (누적)', [표.columns[4], 표.columns[5], '누적 증감'])]
    열순서 = [c for _, 열들 in 셋트 for c in 열들]
    증감열 = {'전월 증감', '전년동월 증감', '누적 증감'}
    첫열 = {열들[0] for _, 열들 in 셋트}          # 셋트가 시작되는 열 → 구분선

    def 두줄(이름):
        """「전년동월(25년 6월)」 → 「전년동월」 + 줄바꿈 + 「(25년 6월)」"""
        if 이름 in 증감열:
            return '증감액'
        i = 이름.find('(')
        return f'{이름[:i]}<br>{이름[i:]}' if i > 0 else 이름

    def 증감칸(v):
        """증감은 색으로 구분합니다 — 늘면 붉게, 줄면 푸르게, 변화 없으면 흐리게."""
        if v is None or pd.isna(v) or abs(v) < 0.5:
            return 'd zero', '-'
        return ('d plus', f'▲ {금액(abs(v))}') if v > 0 else ('d minus', f'▼ {금액(abs(v))}')

    그룹머리 = '<th class="gh gh0 name" rowspan="2">구분</th>'
    for 제목, 열들 in 셋트:
        그룹머리 += f'<th class="gh s1" colspan="{len(열들)}">{제목}</th>'
    소머리 = ''
    for c in 열순서:
        cls = ('d' if c in 증감열 else '') + (' s1' if c in 첫열 else '')
        소머리 += f'<th class="{cls.strip()}">{두줄(c)}</th>'

    rows = ''
    def 줄분류(이름):
        lv = 수준표.get(이름, 1)
        return 'total' if lv == 0 else ('sub' if lv == 1 else 'sub sub2')

    for _, r in 표.iterrows():
        cls = 줄분류(r['구분'])
        vals = ''
        for c in 열순서:
            선 = ' s1' if c in 첫열 else ''
            if c in 증감열:
                k, t = 증감칸(r[c])
                vals += f'<td class="{k}{선}">{t}</td>'
            else:
                vals += f'<td class="{선.strip()}">{금액(r[c])}</td>'
        rows += f'<tr class="{cls}"><td class="name">{r["구분"]}</td>{vals}</tr>'

    매출증감 = 표.loc[표['구분'] == '매출액', '전월 증감'].iat[0]
    매출전월 = P.loc['매출액', 이전월] if 이전월 else 0
    원가율 = (P.loc['매출원가', 선택월] / P.loc['매출액', 선택월] * 100) if P.loc['매출액', 선택월] else 0

    판관 = P.loc[판관비항목, [선택월]].copy()
    판관['전월'] = P.loc[판관비항목, 이전월] if 이전월 else 0
    판관['증감'] = 판관[선택월] - 판관['전월']
    상위 = 판관.reindex(판관['증감'].abs().sort_values(ascending=False).index).head(4)
    panel = ''
    for 항목, r in 상위.iterrows():
        방향 = '증가' if r['증감'] >= 0 else '감소'
        색 = 'up' if r['증감'] >= 0 else 'down'
        panel += (f'<tr><td>{항목}</td><td>{금액(r[선택월])}</td><td>{금액(r["전월"])}</td>'
                  f'<td class="{색}">{금액(r["증감"])} ({방향})</td>'
                  f'<td style="text-align:left; color:#94A3B8;">사유 입력 필요 — 회계 담당자 확인</td></tr>')

    표1 = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>1. {당해연도}년 {선택월}월 실적현황</h3>
        <div class="sub">전월 · 전년동월 · 누적을 셋트로 묶어 비교합니다 · 증감액은 색으로 구분</div>
        <div class="scrollx"><table class="fixed">
          <thead><tr>{그룹머리}</tr><tr>{소머리}</tr></thead>
          <tbody>{rows}</tbody></table></div></div></div>"""

    # ══ 오른쪽 : 항목을 고르면 거래처별 누적 내역 ═══════════════
    누적행 = 당해원장[당해원장['월'].le(선택월)]

    def 항목조건(이름):
        """표의 한 줄이 원장의 어떤 거래를 모은 것인지 알려 줍니다. (조건, 부호)"""
        d = 누적행
        if 이름 == '매출액':
            return d['분류'].eq('매출'), 1
        if 이름 == '제품 매출':
            return d['계정과목'].eq('제품매출'), 1
        if 이름 == '상품 매출':
            return d['계정과목'].eq('상품매출'), 1
        if 이름 == '용역 매출':
            return d['계정과목'].eq('용역매출'), 1
        if 이름 == '매출원가':
            return d['분류'].eq('매출원가'), -1
        if 이름 == '판매관리비':
            return d['분류'].eq('판관비'), -1
        if 이름 == '법인세비용':
            return d['분류'].eq('법인세'), -1
        리스 = d['계정영문'].astype(str).str.strip().str.lower().str.startswith('6623')
        상각 = d['분류'].eq('판관비') & (d['보고항목'].eq('상각비')
                                        | d['계정과목'].isin(['감가상각비', '무형자산상각비']))
        if 이름 == '리스상각비':
            return 상각 & 리스, -1
        if 이름 == '감가상각비':          # 리스분을 뺀 감가상각비
            return 상각 & ~리스, -1
        if 손익기준_당월:
            return d['분류'].eq('판관비') & d['계정과목'].eq(이름), -1
        return d['분류'].eq('판관비') & d['보고항목'].eq(이름), -1

    # 하위 항목을 거느린 줄(매출액 · 판매관리비 · 감가상각비)은 빼고 최하위 줄만 고르게 합니다
    드릴목록 = 최하위줄(보일줄, 수준표, 드릴제외)
    _, 선택칸 = st.columns([2, 1], gap='small')   # 카드 제목 줄을 맞추려고 선택상자만 따로 한 줄
    with 선택칸:
        골라본항목 = st.selectbox('계정과목', 드릴목록, key='mo_drill',
                                 help=f'고른 항목의 1~{선택월}월 누적 금액을 거래처별로 보여줍니다')

    조건, 부호 = 항목조건(골라본항목)
    대상 = 누적행[조건]

    옆 = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{골라본항목} <span class="unitbadge">단위 USD</span></h3>
        <div class="sub">1~{선택월}월 누적 · 거래처별 · 상위 7곳 + 기타</div>
        {세부패널(대상, '금액', 부호)}</div></div>"""

    왼, 오 = st.columns([2, 1], gap='small')   # 작은 화면에서도 오른쪽 세부내역이 읽히도록
    with 왼:
        st.html(표1)
    with 오:
        st.html(옆)

    html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>2. 매출 및 원가 요약</h3><div class="sub">자동 계산</div>
        <table><tbody>
          <tr><td>전월 대비 매출액</td><td>{금액(매출증감)} USD</td>
              <td>{증감HTML(P.loc['매출액', 선택월], 매출전월)}</td></tr>
          <tr><td>당월 원가율</td><td>{원가율:,.1f}%</td><td class="muted">매출원가 ÷ 매출액</td></tr>
          <tr><td>당월 영업이익률</td>
              <td>{(P.loc['영업이익(손실)',선택월]/P.loc['매출액',선택월]*100 if P.loc['매출액',선택월] else 0):,.1f}%</td>
              <td class="muted">영업이익 ÷ 매출액</td></tr>
        </tbody></table></div>
      <div class="card"><h3>3. 전월 대비 판관비 증감 상위 4개 항목</h3>
        <div class="sub">증감 사유는 시스템에서 알 수 없어 회계 담당자 입력이 필요합니다</div>
        <table><thead><tr><th>항목</th><th>당월</th><th>전월</th><th>증감</th>
        <th style="text-align:left;">증감 사유</th></tr></thead><tbody>{panel}</tbody></table></div>
    </div>"""
    st.html(html)

# ══════════════════════════════════════════════════════════════
# 6. 월별 실적집계
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '월별 실적보고':
    st.write('### 월별 실적집계')

    ㄱ, ㄴ = st.columns(2)
    with ㄱ:
        연도선택 = st.radio('연도', [f'{당해연도}년 (당해)', f'{전년연도}년 (전년)'],
                           horizontal=True, key='mm_year')
    with ㄴ:
        기준선택 = st.radio('판관비 구분', ['손익계산서 비용', '활동성 비용'],
                           horizontal=True, key='mm_basis',
                           help='판매관리비를 손익계산서 계정과목으로 볼지, 활동분류로 볼지 고릅니다')
    손익기준 = 기준선택.startswith('손익')
    당해표 = P손익 if 손익기준 else P
    전년표 = P손익전 if 손익기준 else P전
    표원본 = 당해표 if 연도선택.startswith(str(당해연도)) else 전년표
    수준표 = 표원본.attrs['수준']
    당해보기 = 연도선택.startswith(str(당해연도))
    최대월 = 보고월 if 당해보기 else int(전년원장['월'].max() or 12)

    ms = list(range(1, 최대월 + 1))
    rows = ''
    def 줄분류(이름):
        lv = 수준표.get(이름, 1)
        return 'total' if lv == 0 else ('sub' if lv == 1 else 'sub sub2')

    for 항목 in [i for i in 표원본.index if i not in 집계제외]:
        cls = 줄분류(항목)
        cells = ''.join('<td class="{}">{}</td>'.format(
            's1' if m == ms[0] else '', 금액(표원본.loc[항목, m])) for m in ms)
        rows += (f'<tr class="{cls}"><td class="name">{항목}</td>{cells}'
                 f'<td class="s1">{금액(표원본.loc[항목, ms].sum())}</td></tr>')
    head = ''.join('<th class="{}">{}월</th>'.format(
        's1' if m == ms[0] else '', m) for m in ms)
    html = f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="note">단위: USD · 비용 항목은 양수로 표시</div>
      <div class="card"><h3>{연도선택} 월별 손익</h3>
        <div class="sub">원장 직접 집계 · 열너비를 모두 같게 맞췄습니다</div>
      <div class="scrollx"><table class="fixed"><thead><tr><th class="name">구분</th>{head}
        <th class="s1">누적</th></tr></thead><tbody>{rows}</tbody></table></div></div>
    </div>"""
    st.html(html)

    차트제목('최근 12개월 롤링 매출액 · 영업이익 추이', 'USD')
    시리즈 = []
    for i in range(11, -1, -1):
        m = 보고월 - i
        y, mm = (당해연도, m) if m >= 1 else (전년연도, m + 12)
        t = P if y == 당해연도 else P전
        시리즈.append({'라벨': f'{str(y)[2:]}.{mm:02d}', '매출액': t.loc['매출액', mm],
                       '영업이익': t.loc['영업이익(손실)', mm]})
    롤링 = pd.DataFrame(시리즈)
    긴 = 롤링.melt('라벨', var_name='구분', value_name='금액')
    st.altair_chart(alt.Chart(긴).mark_line(point=True).encode(
        x=alt.X('라벨:N', sort=list(롤링['라벨']), title=None,
                axis=alt.Axis(labelAngle=0)),
        y=alt.Y('금액:Q', title=None),
        color=alt.Color('구분:N', title=None, scale=alt.Scale(
            domain=['매출액', '영업이익'], range=[TEAL, AMBER]))),
        width='stretch', key='rolling12')

# ══════════════════════════════════════════════════════════════
# 7. 계정과목 상세  (신규)
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '계정과목 상세':
    st.write('### 계정과목 상세')
    st.caption('계정과목을 고르면 월별 추이와 개별 거래내역까지 볼 수 있습니다.')

    c1, c2, c3 = st.columns([1, 1.4, 1])
    with c1:
        구분선택 = st.selectbox('구분', ['손익(IS) 전체', '매출', '매출원가', '판관비', '영업외손익',
                                       '재무상태(BS) 전체'], key='ad_cat')
    대상 = 당해원장.copy()
    if 구분선택 == '손익(IS) 전체':
        대상 = 대상[대상['계정분류'].eq('IS')]
    elif 구분선택 == '재무상태(BS) 전체':
        대상 = 대상[대상['계정분류'].eq('BS')]
    else:
        대상 = 대상[대상['분류'].eq(구분선택)]

    계정목록 = sorted(대상['계정과목'].dropna().unique())
    if not 계정목록:
        st.warning('해당 구분에 데이터가 없습니다.')
        st.stop()
    with c2:
        계정 = st.selectbox('계정과목', 계정목록, key='ad_acct')
    with c3:
        월범위 = st.slider('월 범위', 1, 보고월, (1, 보고월), key='ad_month') if 보고월 > 1 else (1, 1)

    sel = 대상[(대상['계정과목'] == 계정) &
              (대상['월'] >= 월범위[0]) & (대상['월'] <= 월범위[1])].copy()
    부호 = 1 if 구분선택 == '매출' or 계정 in 매출계정 else -1
    sel['표시금액'] = sel['금액'] * 부호

    ms = list(range(월범위[0], 월범위[1] + 1))
    월합 = sel.groupby('월')['표시금액'].sum().reindex(ms, fill_value=0)
    전년sel = 전년원장[전년원장['계정과목'] == 계정]
    전년월합 = (전년sel.groupby('월')['금액'].sum() * 부호).reindex(ms, fill_value=0)

    합계 = 월합.sum()
    평균 = 월합.mean()
    최대월 = int(월합.abs().idxmax()) if len(월합) else 0
    전년합 = 전년월합.sum()

    kpi = f"""{CARD_CSS}<div class="wrap" translate="no"><div class="kpi-row" style="--n:4">
      <div class="kpi-card"><div class="kpi-label">기간 합계</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{금액(합계)} <span class="unit">USD</span></div>
        <div class="kpi-delta">{증감HTML(합계, 전년합, f'{전년연도}년 동기대비')}</div></div>
      <div class="kpi-card"><div class="kpi-label">월 평균</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{금액(평균)} <span class="unit">USD</span></div></div>
      <div class="kpi-card"><div class="kpi-label">최대 발생월</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{최대월}월</div>
        <div class="kpi-delta"><span class="muted">{금액(월합.get(최대월, 0))} USD</span></div></div>
      <div class="kpi-card"><div class="kpi-label">거래 건수</div>
        <div class="kpi-value" style="font-size:24px;font-weight:700">{len(sel):,}건</div></div>
    </div></div>"""
    st.html(kpi)

    추이 = pd.concat([
        pd.DataFrame({'월': [f'{m}월' for m in ms], '금액': 월합.values, '연도': f'{당해연도}년'}),
        pd.DataFrame({'월': [f'{m}월' for m in ms], '금액': 전년월합.values, '연도': f'{전년연도}년'}),
    ])
    차트제목(f'{계정} 월별 추이 — {당해연도}년 · {전년연도}년', 'USD')
    st.altair_chart(alt.Chart(추이).mark_bar().encode(
        x=alt.X('월:N', sort=[f'{m}월' for m in ms], title=None, axis=alt.Axis(labelAngle=0)),
        xOffset='연도:N', y=alt.Y('금액:Q', title=None),
        color=alt.Color('연도:N', title=None, scale=alt.Scale(
            domain=[f'{당해연도}년', f'{전년연도}년'], range=[TEAL, SLATE]))),
        width='stretch', key='ad_trend')

    g1, g2 = st.columns(2)
    with g1:
        차트제목('활동세부별 구성', 'USD')
        구성 = sel.groupby('활동세부')['표시금액'].sum().sort_values(ascending=False).reset_index()
        구성 = 구성[구성['표시금액'].abs() > 0]
        if len(구성):
            st.altair_chart(alt.Chart(구성.head(12)).mark_bar(color=TEAL).encode(
                y=alt.Y('활동세부:N', sort='-x', title=None),
                x=alt.X('표시금액:Q', title=None)), width='stretch', key='ad_detail')
        else:
            st.caption('활동세부 분류가 입력되지 않은 계정입니다.')
    with g2:
        차트제목('거래처(Name)별 상위 10', 'USD')
        거래처 = (sel[sel['Name'].notna()].groupby('Name')['표시금액'].sum()
                 .abs().sort_values(ascending=False).head(10).reset_index())
        if len(거래처):
            st.altair_chart(alt.Chart(거래처).mark_bar(color=BLUE).encode(
                y=alt.Y('Name:N', sort='-x', title=None),
                x=alt.X('표시금액:Q', title=None)), width='stretch', key='ad_vendor')
        else:
            st.caption('거래처명이 기록되지 않은 계정입니다.')

    st.write(f'**{계정} 거래내역 ({월범위[0]}~{월범위[1]}월, {len(sel):,}건)**')
    상세 = sel[['거래일', '계정영문', 'Description', 'Name', '표시금액',
               '활동분류(대분류)', '활동세부', '분류상태', 'TransactionID']].copy()
    상세['거래일'] = 상세['거래일'].dt.date
    상세 = 상세.rename(columns={'계정영문': '원장 계정(영문)', 'Description': '적요',
                              'Name': '거래처', '표시금액': '금액(USD)'})
    st.dataframe(상세.sort_values('거래일'), width='stretch', hide_index=True,
                 column_config={'금액(USD)': st.column_config.NumberColumn(format='%,.2f')})
    st.download_button('이 내역 CSV로 내려받기',
                       상세.to_csv(index=False).encode('utf-8-sig'),
                       file_name=f'{당해연도}_{계정}_{월범위[0]}-{월범위[1]}월.csv',
                       mime='text/csv', key='ad_dl')

# ══════════════════════════════════════════════════════════════
# 8. 데이터 점검  (신규)
# ══════════════════════════════════════════════════════════════
elif 메뉴 == '데이터 점검':
    st.write('### 데이터 점검')
    st.caption('숫자를 보고에 쓰기 전에 이 페이지에서 원장이 제대로 읽혔는지 확인하세요.')

    검증 = []
    합계 = 당해원장['금액'].sum()
    검증.append(('차변 = 대변 (전체 합계 0)',
                 '0.00' if abs(합계) < 0.005 else f'{합계:,.2f}', abs(합계) < 1))
    미매핑 = int(당해원장['계정과목'].isna().sum())
    검증.append(('계정 매핑 누락 건수', f'{미매핑:,}건', 미매핑 == 0))
    범위밖 = int((~당해원장['월'].between(1, 12)).sum())
    검증.append(('월 범위(1~12) 이탈 건수', f'{범위밖:,}건', 범위밖 == 0))
    미래 = int((당해원장['월'] > 보고월).sum())
    검증.append((f'보고월({보고월}월) 이후 거래 건수 — 날짜 오류 신호', f'{미래:,}건', 미래 == 0))
    검증.append(('판관비 활동분류 미입력 건수', f'{미분류건수:,}건', 미분류건수 == 0))

    줄들 = ''.join(
        f'<div class="chk {"ok" if ok else "wn"}">'
        f'<span class="ic">{"✓" if ok else "!"}</span>'
        f'<span class="t">{이름}</span><span class="v">{값}</span></div>'
        for 이름, 값, ok in 검증)
    통과 = sum(1 for *_, ok in 검증 if ok)
    st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>점검 결과
        <span class="unitbadge">{통과}/{len(검증)}개 항목 정상</span></h3>
        <div class="sub">색을 못 알아보셔도 되도록 아이콘(✓ · !)을 함께 넣었습니다</div>
        {줄들}</div></div>""")

    st.write('---')
    st.write('**계정 매핑이 안 된 원장 계정**')
    미매핑표 = (당해원장[당해원장['계정과목'].isna()]
               .groupby('계정영문')['금액'].agg(건수='size', 합계='sum').reset_index())
    if len(미매핑표):
        st.dataframe(미매핑표, width='stretch', hide_index=True)
        st.warning('위 계정은 BS_IS_매핑 시트에 추가해야 손익에 반영됩니다. 회계 담당자 확인이 필요합니다.')
    else:
        st.success('모든 원장 계정이 매핑되어 있습니다.')

    차트제목('월별 거래 건수 분포', '건')
    분포 = 당해원장.groupby('월').size().reindex(range(1, 13), fill_value=0).reset_index()
    분포.columns = ['월', '건수']
    분포['월'] = 분포['월'].astype(str) + '월'
    st.altair_chart(alt.Chart(분포).mark_bar(color=TEAL).encode(
        x=alt.X('월:N', sort=[f'{m}월' for m in range(1, 13)], title=None,
                axis=alt.Axis(labelAngle=0)),
        y=alt.Y('건수:Q', title=None)), width='stretch', key='chk_dist')
    st.caption('보고월 이후 달에 막대가 서 있으면 날짜(일/월) 파싱 오류를 의심해야 합니다.')

    st.write('**분류 상태 요약**')
    상태 = (당해원장[당해원장['계정분류'].eq('IS')]['분류상태']
            .fillna('(표시 없음)').value_counts().reset_index())
    상태.columns = ['분류상태', '건수']
    st.dataframe(상태, width='stretch', hide_index=True)


# ══════════════════════════════════════════════════════════════
# 9. 연결재무제표 패키지  (틀만 만들어 둔 자리)
# ══════════════════════════════════════════════════════════════
elif 메뉴 in 패키지하위:
    # ── 연결재무제표 패키지 : 재무제표(BS · IS · MC) · 주석사항 (지금은 자리만 잡아 둡니다)
    설명 = {
        'BS':   ('BS  재무상태표', '연결 기준 자산·부채·자본',
                 [('원장 계정 → 패키지 계정 매핑', 'BS_IS_매핑 시트의 「계정분류 BS」'),
                  ('기초잔액 + 당기 증감 → 기말잔액', '26년 원장_raw · 26년 BS기초'),
                  ('연결 조정분개 반영', '대표님이 주실 연결 분개장'),
                  ('전분기 대비 증감 확인', '직전 분기 패키지')]),
        'IS':   ('IS  손익계산서', '연결 기준 수익·비용',
                 [('원장 계정 → 패키지 계정 매핑', 'BS_IS_매핑 시트의 「계정분류 IS」'),
                  ('매출 · 매출원가 · 판관비 집계', '월별 실적집계와 같은 기준'),
                  ('내부거래 제거 · 연결 조정', '대표님이 주실 연결 분개장'),
                  ('전년 동기 대비 확인', '25년 원장_raw')]),
        'MC':   ('MC  제조원가명세서', '재료비 · 노무비 · 경비 → 당기제품제조원가',
                 [('원장 계정 → 원가 요소 매핑', '재료비 · 노무비 · 제조경비'),
                  ('재공품 기초 · 기말 반영', '26년 BS기초 · 26년 원장_raw'),
                  ('당기제품제조원가 산출', 'IS 화면의 매출원가와 연결'),
                  ('연결 조정분개 반영', '대표님이 주실 연결 분개장')]),
        '주석사항': ('주석사항', '패키지에 붙는 주석',
                 [('현금및현금성자산 구성', '26년 자금 시트'),
                  ('매출채권 · 대손 현황', '미수채권 관리 화면'),
                  ('차입금 · 리스부채 명세', '26년 원장_raw'),
                  ('특수관계자 거래', '대표님 확인 필요')]),
    }
    if 메뉴 == '재무제표':
        머리 = '재무제표'
        st.html(f"""<div class="wrap" translate="no" style="margin-bottom:0">
  <div class="page-head"><span class="t">{머리}</span>
    <span class="pill">연결재무제표 패키지</span></div>
</div>""")
        왼칸, 가운데칸, 오른칸 = st.columns([1.5, 1.5, 1], gap='medium')
        with 왼칸:
            # 첫 칸(BS)이 처음부터 체크되어 있습니다
            고른탭 = st.radio('재무제표 구분', 재무제표탭, horizontal=True, key='pkg_tab',
                              help='보고 싶은 표를 고르세요 '
                                   '(BS 재무상태표 · IS 손익계산서 · MC 제조원가명세서)')
            구분 = 고른탭.split()[0]
            말 = st.radio('표기', 표기구분, horizontal=True, key='pkg_lang',
                          help='계정과목을 국문으로 볼지 영문으로 볼지 고르세요 '
                               '(연결패키지 BS·IS 시트의 국문·영문 열과 같습니다)')
            국문으로 = 말.startswith('국문')
        with 가운데칸:
            연결패키지양식칸()
        with 오른칸:
            연결패키지칸()
    else:
        국문으로 = True
        구분 = 메뉴
        st.html(f"""<div class="wrap" translate="no" style="margin-bottom:0">
  <div class="page-head"><span class="t">{메뉴}</span>
    <span class="pill">연결재무제표 패키지</span></div>
</div>""")
    이름, 한줄, 할일 = 설명[구분]
    st.caption(f'{한줄} · {당해연도}년 {(보고월 - 1) // 3 + 1}분기 기준 '
               f'(원장 1~{보고월}월) · 아직 안쪽 내용은 비어 있습니다')

    # ── 재무제표 표 그리기 ─────────────────────────────────────
    #    국문 : 연결패키지 BS·IS 서식 그대로
    #    영문 : 퀵북 계정 이름 그대로 (전기 · 당기 · 증감)
    그렸다 = False
    if 구분 in ('BS', 'IS') and BS밑자료:
        환칸1, 환칸2, _여백 = st.columns([1, 1, 2], gap='small')
        전기환율 = 환칸1.number_input(f'{당해연도 - 1}년 말 환율 (원/USD)', value=1434.90,
                                      step=0.01, format='%.2f', key='bs_fx0')
        당기환율 = 환칸2.number_input(f'{당해연도}년 {보고월}월말 환율 (원/USD)', value=1541.50,
                                      step=0.01, format='%.2f', key='bs_fx1')

        def _돈(v, 자리=2):
            if abs(v) < 0.005:
                return '0.00' if 자리 else '0'
            글 = f'{abs(v):,.{자리}f}'
            return f'<span class="neg">({글})</span>' if v < 0 else 글

        # ── 국문 서식 값 — 영문 화면에서도 「국문과 맞는지」 대사에 씁니다 ──
        서식 = BS서식 if 구분 == 'BS' else IS서식
        if 구분 == 'BS':
            전기값 = dict(BS밑자료['전기'])
            당기값 = dict(BS밑자료['당기'])
        else:
            전기값 = dict(BS밑자료['IS']['전기'])
            당기값 = dict(BS밑자료['IS']['당기'])
        푼값 = [None] * len(서식)

        def _금액(i, 어느):
            if 푼값[i] is not None and 어느 in 푼값[i]:
                return 푼값[i][어느]
            _번, 국, _영, _수준, 더할 = 서식[i]
            if 더할:
                v = sum(g * _금액(j, 어느) for j, g in 더할)
            else:
                v = float((전기값 if 어느 == '전기' else 당기값).get(국, 0.0))
            푼값[i] = (푼값[i] or {})
            푼값[i][어느] = v
            return v

        맞춘값 = {}
        if 구분 == 'BS':
            # 미처분이익잉여금은 원장에 없는 「맞추는 숫자」라 양쪽이 맞도록 채웁니다
            자산i = next((i for i, r in enumerate(서식) if r[1] == '자산총계'), None)
            합계i = next((i for i, r in enumerate(서식) if r[1] == '부채및자본총계'), None)
            잉여i = next((i for i, r in enumerate(서식) if r[1] == '미처분이익잉여금'), None)
            if None not in (자산i, 합계i, 잉여i):
                for 어느, 값들 in (('전기', 전기값), ('당기', 당기값)):
                    값들['미처분이익잉여금'] = 0.0
                    푼값[:] = [None] * len(서식)
                    차이 = _금액(자산i, 어느) - _금액(합계i, 어느)
                    값들['미처분이익잉여금'] = round(차이, 2)
                    맞춘값[어느] = round(차이, 2)
                푼값[:] = [None] * len(서식)

        def _서식줄(이름):
            i = next((k for k, r in enumerate(서식) if r[1] == 이름), None)
            return None if i is None else _금액(i, '당기')

        국문합 = ({'자산총계': _서식줄('자산총계'),
                   '부채및자본총계': _서식줄('부채및자본총계')} if 구분 == 'BS'
                  else {'매출액': _서식줄('매 출 액'), '매출원가': _서식줄('매 출 원 가'),
                        '당기순이익': _서식줄('당기순이익(손실)')})

        # ── 퀵북 표 (실적보고 엑셀작성에서 올려 둔 원장) ────────
        퀵 = None
        원장있음, 원장이름, 원장도장 = 보관원장정보()
        if 원장있음:
            try:
                퀵 = 퀵북표읽기(원장도장, 구분)
            except Exception:
                퀵 = None
        영문합 = {}
        if 퀵:
            찾기표 = dict(BS밑자료['영문BS'].get('전기맞춤', {})) if 구분 == 'BS' else {
                _맞춤이름(n): float(v)
                for n, v in BS밑자료['IS']['전기계정'].items()}
            줄들 = 퀵['줄']
            전기칸 = [None] * len(줄들)
            for i, x in enumerate(줄들):
                if x['당기'] is None:       # 금액이 없는 구역 머리글은 전기도 비워 둡니다
                    continue
                전기칸[i] = 찾기표.get(_맞춤이름(x['이름']))
            for i, x in enumerate(줄들):    # 표에 없는 Total 줄은 안쪽 줄을 더해 채웁니다
                if not x['총계'] or 전기칸[i] is not None:
                    continue
                깊 = x['깊이']
                모아, 봤다 = 0.0, False
                j = i - 1
                while j >= 0 and 줄들[j]['깊이'] > 깊:
                    if 줄들[j]['깊이'] == 깊 + 1 and 전기칸[j] is not None:
                        모아 += 전기칸[j]
                        봤다 = True
                    j -= 1
                if j >= 0 and 줄들[j]['깊이'] == 깊 and not 줄들[j]['총계'] \
                        and 전기칸[j] is not None:
                    모아 += 전기칸[j]
                    봤다 = True
                전기칸[i] = 모아 if 봤다 else None
            찾자 = ({'자산총계': 'total assets',
                     '부채및자본총계': 'total liabilities and equity'} if 구분 == 'BS'
                    else {'매출액': 'total income', '매출원가': 'total cost of goods sold',
                          '당기순이익': 'profit'})
            for 라벨, 키 in 찾자.items():
                for x in 줄들:
                    if _맞춤이름(x['이름']) == 키 and x['당기'] is not None:
                        영문합[라벨] = x['당기']
                        break

        # ── 국문 ↔ 영문 대사 (같은 원장에서 나온 숫자인지 확인) ──
        대사칸 = ''
        if 퀵 and 퀵.get('월') and 퀵['월'] != 보고월:
            대사칸 = (
                f'<div class="reconc"><b>국문 ↔ 영문 대사</b> — 기간이 달라 견주지 않았습니다.'
                f'<div class="chks"><span class="chk no">⚠ 화면 자료는 {보고월}월 · '
                f'올려 둔 원장은 {퀵["월"]}월</span></div>'
                f'같은 달로 맞추시려면 「실적보고 엑셀작성」 에서 {퀵["월"]}월 실적 엑셀을 만든 뒤, '
                f'그 파일을 왼쪽 「⚙ 설정 · 파일 업로드」 에 올려 주세요.</div>')
        elif 영문합:
            조각 = []
            for 라벨 in 국문합:
                국, 영 = 국문합.get(라벨), 영문합.get(라벨)
                if 국 is None or 영 is None:
                    continue
                차 = round(국 - 영, 2)
                맞다 = abs(차) < 1.0
                조각.append(
                    f'<span class="chk {"ok" if 맞다 else "no"}">'
                    f'{"✔" if 맞다 else "⚠"} {라벨} · 국문 {국:,.2f} / 영문 {영:,.2f}'
                    f'{"" if 맞다 else f" · 차이 {차:,.2f}"}</span>')
            if 조각:
                모두맞다 = all('chk ok' in s for s in 조각)
                대사칸 = (
                    f'<div class="reconc"><b>국문 ↔ 영문 대사</b> — '
                    f'{"두 표의 합계가 같습니다" if 모두맞다 else "차이가 있습니다"} '
                    f'(원장 「{퀵["시트"]}」 기준)'
                    f'<div class="chks">{"".join(조각)}</div></div>')

        if 국문으로:
            칸들 = ''
            for i, (번호, 국, 영, 수준, 더할) in enumerate(서식):
                이름 = (국 if 국문으로 else 영) or (영 if 국문으로 else 국)
                딴이름 = (영 if 국문으로 else 국)
                if not 영 and not 더할 and 국 in ('자산', '부채', '자본'):
                    칸들 += (f'<tr class="total"><td class="lft" colspan="8" '
                             f'style="background:{T["panel2"]}">{이름}</td></tr>')
                    continue
                전 = _금액(i, '전기')
                당 = _금액(i, '당기')
                굵게 = ' class="total"' if 수준 <= 1 else (' class="sub"' if 수준 == 2 else '')
                들여 = 0 if 수준 == 0 else (12 if 수준 == 1 else (24 if 수준 == 2 else 34))
                칸들 += (
                    f'<tr{굵게}><td class="lft" style="padding-left:{들여 + 11}px">'
                    f'{("<b>" + 번호 + "</b> " if 번호 else "")}{이름}'
                    f'<span style="color:{T["ink3"]}; font-size:12px"> · {딴이름}</span></td>'
                    f'<td>{_돈(전)}</td><td>{_돈(전 * 전기환율, 0)}</td>'
                    f'<td>{_돈(당)}</td><td></td><td></td>'
                    f'<td>{_돈(당)}</td><td>{_돈(당 * 당기환율, 0)}</td></tr>')
            제목 = '재 무 상 태 표' if 구분 == 'BS' else '손 익 계 산 서'
            기간 = (f'제 (당) 기 : {당해연도}년 {보고월:02d}월말 · '
                    f'제 (전) 기 : {당해연도 - 1}년 12월 31일' if 구분 == 'BS'
                    else f'제 (당) 기 : {당해연도}.01.01~{당해연도}.{보고월:02d} · '
                         f'제 (전) 기 : {당해연도 - 1}.01.01~{당해연도 - 1}.12.31')
            맞춤말 = ('' if not 맞춘값 else
                      f'<br><b>미처분이익잉여금</b> 은 원장에 없는 「맞추는 숫자」라 '
                      f'자산총계와 부채및자본총계가 맞도록 채웠습니다 '
                      f'(전기 {_돈(맞춘값.get("전기", 0))} · 당기 {_돈(맞춘값.get("당기", 0))}).')
            st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{제목} <span class="unitbadge">국문</span></h3>
        <div class="sub">{기간} &nbsp;·&nbsp; 연결패키지 {구분} 시트와 같은 모양입니다</div>
        {대사칸}
        <div class="stick"><table class="lined" style="min-width:1080px; margin-top:8px">
          <colgroup><col style="width:26%"><col style="width:11%"><col style="width:12%">
            <col style="width:11%"><col style="width:9%"><col style="width:9%">
            <col style="width:11%"><col style="width:11%"></colgroup>
          <thead>
            <tr><th>과목</th>
                <th>전기 ({당해연도 - 1}년)<br><span class="th2">현지통화(USD)</span></th>
                <th>전기 ({당해연도 - 1}년)<br><span class="th2">KRW</span></th>
                <th>당기 ({당해연도 % 100}년-수정전)<br><span class="th2">현지통화(USD)</span></th>
                <th>수정분개<br><span class="th2">Debit</span></th>
                <th>수정분개<br><span class="th2">Credit</span></th>
                <th>당기 ({당해연도 % 100}년-수정후)<br><span class="th2">현지통화(USD)</span></th>
                <th>당기 ({당해연도 % 100}년-수정후)<br><span class="th2">KRW</span></th></tr>
          </thead>
          <tbody>{칸들}</tbody></table></div>
        <div class="calcnote"><b>어디서 온 숫자인가</b> —
          전기는 올려 주신 실적 엑셀의
          <b>{'「26년 BS기초」' if 구분 == 'BS' else '「25년 원장_raw」'}</b>,
          당기는 <b>「26년 원장_raw」</b> 에서 계정별로 계산했습니다.
          <b>수정분개</b> 는 연결 결산조정 칸이라 아직 비어 있습니다.{맞춤말}</div>
      </div>
    </div>""")
            그렸다 = True
        elif 퀵:
            # ── 영문 : 퀵북 표를 차례·들여쓰기 그대로 옮겨 놓습니다 ──
            줄들 = 퀵['줄']
            줄 = ''
            for i, x in enumerate(줄들):
                전, 당 = 전기칸[i], x['당기']
                굵게 = ' class="total"' if x['총계'] else (
                    ' class="sub"' if 당 is not None else ' class="head"')
                들여 = 11 + x['깊이'] * 16
                줄 += (f'<tr{굵게}><td class="lft" style="padding-left:{들여}px">'
                       f'{x["이름"]}</td>'
                       f'<td>{"" if 전 is None else _돈(전)}</td>'
                       f'<td>{"" if 당 is None else _돈(당)}</td>'
                       f'<td>{"" if (전 is None or 당 is None) else _돈(당 - 전)}</td></tr>')
            머리 = 퀵['제목']
            st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{머리[1] or ('Balance Sheet' if 구분 == 'BS'
                                         else 'Profit and Loss')}
          <span class="unitbadge">English</span></h3>
        <div class="sub">{머리[0]} &nbsp;·&nbsp; {머리[2]}
          &nbsp;·&nbsp; prior year {당해연도 - 1} &nbsp;·&nbsp; USD</div>
        {대사칸}
        <div class="stick"><table class="lined" style="min-width:760px; margin-top:8px">
          <colgroup><col style="width:46%"><col style="width:18%">
                    <col style="width:18%"><col style="width:18%"></colgroup>
          <thead><tr><th>Account</th><th>Prior year ({당해연도 - 1})</th>
                     <th>Current ({당해연도})</th><th>Change</th></tr></thead>
          <tbody>{줄}</tbody></table></div>
        <div class="calcnote"><b>어디서 온 표인가</b> — 올려 두신 원장의
          <b>「{퀵['시트']}」</b> 시트를 줄 차례·들여쓰기 그대로 옮겼습니다
          (구역 머리글과 Total 줄 포함 {len(줄들)}줄).
          전기는
          {'실적 엑셀의 「26년 BS기초」' if 구분 == 'BS' else '「25년 원장_raw」 계정별 누계'}
          에서 같은 계정 이름을 찾아 넣었고, Total 줄은 안쪽 줄을 더해 채웠습니다.</div>
      </div>
    </div>""")
            그렸다 = True
        else:
            st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>영문 재무제표 <span class="unitbadge">English</span></h3>
        <div class="calcnote"><b>원장 원본이 아직 없습니다</b> —
          왼쪽 메뉴 <b>「실적보고 엑셀작성」</b> 에서 ① 원장 원본(퀵북 · 그 달 BS·IS 시트가
          들어 있는 파일)을 한 번 올려 주시면, 그 안의
          <b>「{당해연도 % 100}년 {보고월}월(BS)」·「{당해연도 % 100}년 {보고월}월(IS)」</b>
          시트를 차례 그대로 여기에 보여 드립니다.</div>
      </div>
    </div>""")
            그렸다 = True

    # 연결패키지 BS·IS 시트가 쓰는 계정과목 이름 (국문 · 영문)
    미리보기 = {
        'BS': [('Ⅰ. 유동자산', 'Current assets'),
               ('　1. 현금및현금성자산', 'Cash and cash equivalents'),
               ('　2. 매출채권', 'Trade Receivable, gross'),
               ('　3. 재고자산', 'Inventory'),
               ('Ⅱ. 비유동자산', 'Non-current assets'),
               ('　1. 유형자산', 'Property, plant and equipment'),
               ('　2. 사용권자산', 'Right-of-use assets'),
               ('Ⅲ. 부채', 'Liabilities'),
               ('　1. 매입채무', 'Trade payables'),
               ('　2. 차입금', 'Borrowings'),
               ('Ⅳ. 자본', 'Equity')],
        'IS': [('Ⅰ. 매출액', 'Revenue'),
               ('Ⅱ. 매출원가', 'Cost of sales'),
               ('Ⅲ. 매출총이익', 'Gross profit'),
               ('Ⅳ. 판매비와관리비', 'Selling, general and administrative expenses'),
               ('Ⅴ. 영업이익', 'Operating profit'),
               ('Ⅵ. 영업외손익', 'Other income and expenses'),
               ('Ⅶ. 법인세비용차감전순손익', 'Profit before income tax'),
               ('Ⅷ. 당기순손익', 'Profit for the period')],
        'MC': [('Ⅰ. 재료비', 'Raw materials used'),
               ('Ⅱ. 노무비', 'Labour costs'),
               ('Ⅲ. 경비', 'Manufacturing overhead'),
               ('Ⅳ. 당기총제조비용', 'Total manufacturing costs'),
               ('Ⅴ. 기초재공품재고액', 'Work in progress, beginning'),
               ('Ⅵ. 기말재공품재고액', 'Work in progress, ending'),
               ('Ⅶ. 당기제품제조원가', 'Cost of goods manufactured')],
    }.get(구분)
    if 그렸다:
        미리보기 = None
    if 미리보기:
        칸 = ''.join(
            f'<tr class="sub"><td class="name">{국 if 국문으로 else 영}</td>'
            f'<td class="lft" style="color:{T["ink3"]}">{영 if 국문으로 else 국}</td></tr>'
            for 국, 영 in 미리보기)
        st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="card"><h3>{'계정과목 (국문)' if 국문으로 else 'Line items (English)'}
        <span class="unitbadge">{'국문 기준' if 국문으로 else 'English'}</span></h3>
        <div class="sub">연결패키지 「{구분}」 시트가 쓰는 이름입니다 ·
          숫자는 원장이 연결되면 이 자리에 채워집니다</div>
        <table class="lined" style="margin-top:8px">
          <colgroup><col style="width:44%"><col style="width:56%"></colgroup>
          <thead><tr><th>{'국문' if 국문으로 else 'English'}</th>
                     <th>{'English' if 국문으로 else '국문'}</th></tr></thead>
          <tbody>{칸}</tbody></table></div>
    </div>""")

    줄 = ''.join(
        f'<tr class="sub"><td class="name">{i}</td><td class="lft">{무엇}</td>'
        f'<td class="lft">{자료}</td></tr>'
        for i, (무엇, 자료) in enumerate(할일, 1))
    st.html(f"""{CARD_CSS}<div class="wrap" translate="no">
      <div class="kpi-row" style="--n:3">
        <div class="kpi-card"><div class="kpi-label">기준 분기</div>
          <div class="kpi-value" style="font-size:24px;font-weight:700">{당해연도}년 {(보고월 - 1) // 3 + 1}Q</div>
          <div class="kpi-delta"><span class="muted">원장 1~{보고월}월 기준</span></div></div>
        <div class="kpi-card"><div class="kpi-label">화면</div>
          <div class="kpi-value" style="font-size:24px;font-weight:700">{이름}</div>
          <div class="kpi-delta"><span class="muted">{한줄}</span></div></div>
        <div class="kpi-card"><div class="kpi-label">준비 상태</div>
          <div class="kpi-value" style="font-size:24px;font-weight:700;color:{T['accent']}">준비 중</div>
          <div class="kpi-delta"><span class="muted">연결 분개장을 받으면 채웁니다</span></div></div>
      </div>
      <div class="card"><h3>이 화면에 들어갈 것</h3>
        <div class="sub">쓸 자료가 이미 있는 것부터 순서대로 채워 나갑니다</div>
        <table class="lined">
          <colgroup><col style="width:8%"><col style="width:50%"><col style="width:42%"></colgroup>
          <thead><tr><th>순서</th><th>내용</th><th>쓰는 자료</th></tr></thead>
          <tbody>{줄}</tbody></table></div>
    </div>""")

    받을칸, 여백 = st.columns([1, 3], gap='small')
    with 받을칸:
        st.download_button(f'⤓  {이름} 엑셀 내려받기', data=b'', file_name='준비중.xlsx',
                           disabled=True, key=f'pkg_{구분}', width='stretch',
                           help='연결 분개장을 받아 채우면 여기서 내려받을 수 있게 됩니다')
    with 여백:
        st.caption('※ 검증은 OTC 원장의 BS · IS 기초자료로 진행합니다.')
